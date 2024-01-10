import os
import os.path
from datetime import datetime
from io import BytesIO          

import PySimpleGUI as sg
import aspose.pdf as ap
from PyPDF2 import PdfReader, PdfWriter
from openpyxl import Workbook, load_workbook


def create_output_excel(output_xlsx):
    # Create a new workbook
    workbook = Workbook()

    # Get the active sheet (default sheet created with the workbook)
    sheet = workbook.active

    # Add data to the sheet
    sheet['A1'] = ''

    # Save the workbook
    workbook.save(output_xlsx)


def join_data(temp_xlsx, output_xlsx):
    source_workbook_1 = load_workbook(output_xlsx)
    sheet1 = source_workbook_1.active

    source_workbook_2 = load_workbook(temp_xlsx)

    sheet2 = source_workbook_2['Sheet1']

    for row in sheet2.iter_rows(min_row=2, values_only=True):
        # print (row)
        sheet1.append(row)
    source_workbook_1.save(output_xlsx)
    # print(f"added {page} data")


def conver_excel(pdf, temp_xlsx, output_xlsx):
    with open(temp_xlsx, 'wb') as f:
        document = ap.Document(pdf)
        # print(dir(document))
        # exit()
        save_option = ap.ExcelSaveOptions()
        # save_option.format = ap.ExcelSaveOptions.ExcelFormat
        save_option.minimize_the_number_of_worksheets = True
        # Save the file into MS Excel format
        # document.save(output_pdf, save_option)
        document.save(f, options=save_option)

    join_data(temp_xlsx, output_xlsx)


def split_pdf(input_pdf, temp_xlsx, output_xlsx):
    with open(input_pdf, "rb") as pdf:
        bytes_stream = BytesIO(pdf.read())

    reader = PdfReader(bytes_stream)
    count = 0
    for page in reader.pages:
        writer = PdfWriter()
        writer.add_page(page)

        with BytesIO() as bytes_stream:
            writer.write(bytes_stream)
            bytes_stream.seek(0)
            conver_excel(bytes_stream, temp_xlsx, output_xlsx)
        count = count + 1
        print(f"page {count} done")
    return count


def main(input_pdf):
    now = datetime.now()

    t = now.strftime("__%d-%m-%Y-%H-%M-%S")
    output_xlsx = input_pdf.replace('.pdf', t) + '.xlsx'
    temp_xlsx = input_pdf.replace('.pdf', '_temp.xlsx')
    create_output_excel(output_xlsx)
    count = split_pdf(input_pdf, temp_xlsx, output_xlsx)
    os.remove(temp_xlsx)
    #
    return 'Process Completed', f"{count} pages converted"


file_list_column = [
    [
        sg.Text("File Folder"),
        sg.In(size=(25, 1), enable_events=True, key="-FOLDER-"),
        sg.FolderBrowse(),
    ],
    [
        sg.Listbox(
            values=[], enable_events=True, size=(40, 20), key="-FILE LIST-"
        )
    ],
]

# Right side of the GUI
# Shows the name of the file that was chosen from folder to be converted
convert_column = [
    [sg.Text("Choose an .pdf file from list on left:")],
    [sg.Text(size=(100, 1), key="-TOUT-")],
    [sg.Button("Convert pdf to excel")],
    [sg.Text(size=(100, 1), key="-RESP-")],
    [sg.Text(size=(100, 1), key="-CHAP-")]
]

# Combining the left and right layout
layout = [
    [
        sg.Column(file_list_column),
        sg.VSeperator(),
        sg.Column(convert_column),
    ]
]

window = sg.Window("KSV PDF converter", layout)  ## App name

# Run the Event Loop
while True:
    event, values = window.read()
    if event == "Exit" or event == sg.WIN_CLOSED:
        break
    # Folder name was filled in, make a list of files in the folder
    if event == "-FOLDER-":
        folder = values["-FOLDER-"]
        try:
            # Get list of files in folder
            file_list = os.listdir(folder)

        except:
            file_list = []

        fnames = [
            f
            for f in file_list
            if os.path.isfile(os.path.join(folder, f))
               and f.lower().endswith((".pdf"))  # Filtering the .xlsx files in the folder
        ]
        window["-FILE LIST-"].update(fnames)  # Display the .xlsx files in the list of left column

    elif event == "-FILE LIST-":  # A file was chosen from the listbox
        try:
            filename = os.path.join(
                values["-FOLDER-"], values["-FILE LIST-"][0]
            )
            window["-RESP-"].update(" ")
            window["-TOUT-"].update(filename)  # Display the full name in the convert column of GUI

        except:
            pass

    elif event == "Convert pdf to excel":  # When the convert button is pressed
        try:
            print("convert process")
            # chapter_data = chapter_data + "  " + converter(filename)  ##Send the full file name to the converter block
            return_state = main(filename)
            window["-RESP-"].update(return_state[0])
            window["-CHAP-"].update(return_state[1])

        except:
            pass

window.close()
