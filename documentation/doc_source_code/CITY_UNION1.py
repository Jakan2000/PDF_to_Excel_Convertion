from datetime import datetime

import openpyxl

from CommonClass import Excel


def dateConversion(wb, start, end, column):
    """
        Convert date values in a specified column of an Excel workbook to a standard date format.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the worksheet to be processed.
        - start (int): The starting row index (inclusive) from which date conversion should begin.
        - end (int): The ending row index (exclusive) until which date conversion should occur.
        - column (str): The column letter (e.g., 'A', 'B') representing the column where date conversion should be applied.

        Returns:
        - openpyxl.Workbook: The modified workbook after date conversion.

        Notes:
        - This function assumes the date values in the specified column are in the format '%d/%m/%Y'.
        - The converted dates are stored in the same column as standard Python date objects.
        - The 'start' parameter denotes the first row to start converting dates.
        - The 'end' parameter denotes the row where date conversion stops (exclusive).
        - The original workbook is modified in place.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through start and end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d/%m/%Y").date()  # converting to standard date formate
    return wb


def deleteHeader(wb, start):
    """
        Delete rows above the specified start index row in the active sheet of an Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the worksheet to be processed.
        - start (int): The row index from which rows above will be deleted (inclusive).

        Returns:
        - openpyxl.Workbook: The modified workbook after deleting the header rows.

        Notes:
        - This function removes rows above the specified 'start' index in the active sheet.
        - The 'start' parameter denotes the row index from which rows above will be deleted.
        - If 'start' is 1, it will remove all rows above the first row, effectively clearing the sheet.
        - The original workbook is modified in place.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data header text to sheet's 1st row
        sheet.delete_rows(x)  # delete row
    return wb


def deleteFooter(wb, end):
    """
        Delete all rows below the specified end (last) row in the active sheet of an Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the worksheet to be processed.
        - end (int): The row index until which rows below will be deleted (exclusive).

        Returns:
        - openpyxl.Workbook: The modified workbook after deleting the footer rows.

        Notes:
        - This function removes rows below the specified 'end' index in the active sheet.
        - The 'end' parameter denotes the row index until which rows below will be deleted.
        - If 'end' is the last row index, it will effectively clear the sheet below that row.
        - The original workbook is modified in place.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through max row in sheet to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def cityunion1_validation(wb):
    """
        Validate columns for the core logic in an Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook to be validated.

        Returns:
        - bool: True if the number of columns is not equal to 6 (core logic count), False otherwise.

        Notes:
        - This function checks if the number of columns in the active sheet of the workbook is equal to 6.
        - It is designed to validate whether the workbook adheres to the expected structure for the core logic.
        - The 'countOfColumn' variable is set to 6, representing the expected number of columns for the core logic.
        - If the actual number of columns is different from the expected count, the function returns True.
        - If the actual number of columns matches the expected count, the function returns False.

    """
    sheet = wb.active
    max_column = sheet.max_column  # getting the max column using in build function
    countOfColumn = 6  # the column count of our core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def cityunion1_main(wb):
    """
        Perform data processing for City Union Bank format 1 in an Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing data to be processed.

        Returns:
        - dict: A dictionary with keys:
        - 'data': The modified Excel workbook after processing.
        - 'msg': A message indicating the processing result.

        Notes:
        - This function processes the City Union Bank format 1 data in the provided workbook.
        - It performs various operations including validation, header deletion, footer deletion, and column standardization.
        - The function utilizes several utility functions from the 'Excel' module (assumed to be imported).
        - The standardization includes altering header names, creating a serial number column, and creating a new transaction type column.
        - The 'data' key in the return dictionary contains the modified workbook.
        - The 'msg' key provides a message indicating the processing result.

    """
    sheet = wb.active
    if cityunion1_validation(wb):  # validate columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response
    else:
        startText = "DATE"  # header text in column A
        endText = "TOTAL"  # text define the end of the table data
        startEndRefColumn = "A"  # column containing the text to define start and end -> data within
        deleteFlagStartText = "Regd. Office"  # starting row reference text to delete the rows by range
        deleteFlagStopText = "DATE"  # ending row reference text to delete the rows by range
        deleteFlagRefColumn = "A"  # column containing starting row reference text and ending row reference text to delete the rows by range
        dateConversionColumn = "A"    # column to convert date to the standard date formate
        refHeaderText1 = "DATE"  # header text to replace with standardised column name
        refHeaderText2 = "DESCRIPTION"  # header text to replace with standardised column name
        refHeaderText3 = "CHEQUE NO"  # header text to replace with standardised column name
        refHeaderText4 = "DEBIT"  # header text to replace with standardised column name
        refHeaderText5 = "CREDIT"  # header text to replace with standardised column name
        refHeaderText6 = "BALANCE"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Narration"  # standard column name
        headerText3 = "ChequeNo_RefNo"  # standard column name
        headerText4 = "Withdrawal"  # standard column name
        headerText5 = "Deposit"  # standard column name
        headerText6 = "Balance"  # standard column name
        negativeValueColumnRefText1 = "Withdrawal"  # no need of removing negative values
        headerTextToReplaceNone1 = "ChequeNo_RefNo"  # reference header text to make empty cells in a column to none
        headerTextToReplaceNone2 = "Withdrawal"  # reference header text to make empty cells in a column to none
        headerTextToReplaceNone3 = "Value_Date"  # reference header text to make empty cells in a column to none
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        dupHeadersRemoved = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagStopText, deleteFlagRefColumn)  # deleting the rows by range
        start, end = Excel.get_start_end_row_index(dupHeadersRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        footerDeleted = deleteFooter(dupHeadersRemoved, end - 1)  # end-1 to Include Last Row, delete all the rows below the end(last) row
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # start-1 to Skip Header, delete rows above the start index row
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        convertedToDateA = dateConversion(headerDeleted, start + 1, end + 1, dateConversionColumn)  # start+1 to Skip Header, end+1 to include last row
        lastCol = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        transdate = Excel.alter_header_name(convertedToDateA, refHeaderText1, headerText1, lastCol)  # alter header name by standard column name
        narration = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name by standard column name
        chqno = Excel.alter_header_name(narration, refHeaderText3, headerText3, lastCol)  # alter header name by standard column name
        debit = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)  # alter header name by standard column name
        credit = Excel.alter_header_name(debit, refHeaderText5, headerText5, lastCol)  # alter header name by standard column name
        balance = Excel.alter_header_name(credit, refHeaderText6, headerText6, lastCol)  # alter header name by standard column name
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slnoCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))  # created new slno column
        columnFinalised = Excel.finalise_column(wb, columns)  # standardizing count of column
        negativeColumnChecked = Excel.check_neagativeValue_by_column(columnFinalised, negativeValueColumnRefText1)  # no need to convert the negative value to positive
        replacedNoneCHQNO = Excel.empty_cell_to_none(wb, start, end + 1, headerTextToReplaceNone1)  # making empty cells in a column to none by using the header text as reference
        replacedNoneWITHDRAWAL = Excel.empty_cell_to_none(replacedNoneCHQNO, start, end + 1, headerTextToReplaceNone2)  # making empty cells in a column to none by using the header text as reference
        replacedNoneVALUEDATE = Excel.empty_cell_to_none(replacedNoneWITHDRAWAL, start, end + 1, headerTextToReplaceNone3)  # making empty cells in a column to none by using the header text as reference
        createdTransTypeColumn = Excel.transaction_type_column(replacedNoneVALUEDATE)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/CITY_UNION_BANK_-_SB-500101012199098__23-09-2023-18-18-25.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    result = cityunion1_main(wb)
    # result["data"].save("C:/Users/Admin/Desktop/FinalOutput/CITY_UNION1output.xlsx")
