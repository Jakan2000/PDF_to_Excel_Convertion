import tempfile
from datetime import datetime
from io import BytesIO

import camelot
import pandas as pd
import requests
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

from CommonClass import Excel


def dateConversion(wb, start, end, column):
    """
        Convert date values in the specified column to a standard date format.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - start (int): The start row index for the table data.
        - end (int): The end row index for the table data.
        - column (str): The column letter where date values need to be converted.

        Returns:
        - Workbook: The modified Excel workbook object.

        Note:
        The function iterates through the specified range of rows in the given column, checks if the cell is not empty,
        and converts the date to a standard date format ('%d/%m/%Y'). The converted date is then assigned back to the cell.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        original_date = sheet[f"{column}{i}"].value
        if original_date:  # Check if the cell is not empty
            # Correct the format based on the actual date format in your Excel sheet
            new_date = datetime.strptime(original_date, '%d/%m/%Y').date()  # converting date to standard date formate
            sheet[f"{column}{i}"].value = new_date  # assigning converted date to date cell value
    return wb


def removeNoneRows(wb, start, end, column):
    """
        Remove rows where the reference column cell value is None.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - start (int): The start row index for the table data.
        - end (int): The end row index for the table data.
        - column (str): The column letter representing the reference column.

        Returns:
        - Workbook: The modified Excel workbook object.

        Note:
        The function iterates through the specified range of rows in the given column, and if the cell value is None,
        it deletes the entire row. Rows are processed in reverse order to avoid index issues during deletion.

    """
    sheet = wb.active
    for x in range(end - 1, start, -1):  # iterating through table data end row to table data start row
        if sheet[f"{column}{x}"].value is None:  # if column cell value is none
            sheet.delete_rows(x)  # delete row
    return wb


def remove_header(wb, start):
    """
        Delete rows above the start index row.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - start (int): The start row index for the table data.

        Returns:
        - Workbook: The modified Excel workbook object.

        Note:
        The function iterates through rows from the specified start index to the first row (row 1) and deletes each row.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merge the rows of the desired column based on a reference column.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - start (int): The start row index for the table data.
        - end (int): The end row index for the table data.
        - refColumn (str): The reference column for identifying starting rows.
        - mergingColumn (str): The column to merge rows.

        Returns:
        - Workbook: The modified Excel workbook object.

        Note:
        The function iterates through the specified range of rows, identifies starting rows based on the reference column,
        and merges the corresponding rows in the specified merging column.

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through table data start row to table data end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def pandas_df_to_openpyxl(df):
    """
        Convert a Pandas DataFrame to an Openpyxl Workbook.

        Parameters:
        - df (DataFrame): The input Pandas DataFrame.

        Returns:
        - Workbook: The Openpyxl Workbook containing the DataFrame data.

        Note:
        The function creates a new Openpyxl Workbook and worksheet, then appends the DataFrame data to the worksheet.

    """
    workbook = Workbook()  # Create a new Openpyxl Workbook
    worksheet = workbook.active  # Create a new worksheet
    # Append the DataFrame data to the worksheet
    for row in dataframe_to_rows(df, index=False, header=False):
        worksheet.append(row)
    return workbook  # returning openpyxl work book


def icici4_validation(wb):
    """
        Validate columns for the core logic in the ICICI4 format.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook containing the data.

        Returns:
        - bool: True if the column count doesn't match the expected count, False otherwise.

        Note:
        The function checks if the number of columns in the active sheet of the workbook matches the expected count
        for the ICICI4 format.

    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 8  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def icici4_main(pdf_url):
    """
        Process ICICI4 formatted PDF and return the result.

        Parameters:
        - pdf_url (str): URL of the PDF file in ICICI4 format.

        Returns:
        - dict: A dictionary containing the processed data or an error message.

        Note:
        This function processes an ICICI4 formatted PDF file. It downloads the PDF from the provided URL,
        extracts tables using Camelot, and converts the data into an Openpyxl Workbook. It then performs
        various operations on the data to standardize the format. The result is returned as a dictionary
        containing the processed data or an error message.

    """
    startText = "Value Date"  # header text to define table data start row
    endText = ""  # reference text to define table data end row -> hear if empty string means max row of sheet
    startEndRefColumn = "B"  # reference column contains the start and end text
    headerTextTOMakeEmptyCellTONone = "Value Date"  # column header text to make empty cells to none
    refColumnToMerg = "B"  # reference colum to merge misaligned rows of column
    ColumnToMerg1 = "E"  # column to merge misaligned rows
    refTextToDelteColumn = "S No."  # reference header text to delete column
    refColumnToRemoveNoneRows = "A"  # reference column to delete empty rows
    refHeaderText1 = "Value Date"  # header text to replace with standardised column name
    refHeaderText2 = "Transaction Date"  # header text to replace with standardised column name
    refHeaderText3 = "Cheque Number"  # header text to replace with standardised column name
    refHeaderText4 = "Transaction Remarks"  # header text to replace with standardised column name
    refHeaderText5 = "Withdrawal Amount"  # header text to replace with standardised column name
    refHeaderText6 = "Deposit Amount ( )"  # header text to replace with standardised column name
    refHeaderText7 = "Balance ( )"  # header text to replace with standardised column name
    headerText1 = "Value_Date"  # standard column name
    headerText2 = "Transaction_Date"  # standard column name
    headerText3 = "ChequeNo_RefNo"  # standard column name
    headerText4 = "Narration"  # standard column name
    headerText5 = "Withdrawal"  # standard column name
    headerText6 = "Deposit"  # standard column name
    headerText7 = "Balance"  # standard column name
    dateConversionColumn1 = "A"  # column to convert date to standard date
    dateConversionColumn2 = "B"  # column to convert date to standard date formate
    refTextToMakeCellNone = "-"  # reference text to make cell none
    refColumnToMakeCellNone = "C"  # column to make cells with refTextToMakeCellNone to none
    columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
    # Download the PDF file from the URL
    response = requests.get(pdf_url)  # getting pdf url
    pdf_data = BytesIO(response.content)  # converting it to bites
    # Save the BytesIO content to a temporary PDF file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
        temp_pdf.write(pdf_data.getvalue())
        temp_pdf_path = temp_pdf.name
    # Extract tables from PDF using Camelot
    tables = camelot.read_pdf(temp_pdf_path, flavor='stream', pages='all')
    # Concatenate DataFrames for each page into a single DataFrame
    df = pd.concat([table.df for table in tables])
    # Convert DataFrame to Openpyxl Workbook
    wb = pandas_df_to_openpyxl(df)
    # Remove the temporary PDF file
    temp_pdf.close()
    sheet = wb.active
    if icici4_validation(wb):  # validate columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        responce = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return responce  # return response with error msg
    else:
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # taking max row in sheet as end row
        removedHeader = remove_header(wb, start - 1)  # delete rows above table data start index row
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # taking max row in sheet as end row
        Excel.empty_cell_to_none(wb, start, end + 1, headerTextTOMakeEmptyCellTONone)  # making empty cells in a column to none by using the header text as reference
        mergedColumnE = mergingRows(wb, start, end+1, refColumnToMerg, ColumnToMerg1)  # merging the rows of desired column
        Excel.delete_column(wb, refTextToDelteColumn)   # deleting column using reference header name
        startEndRefColumn = "A"  # now we using column A as start row and end row reference column
        removeNoneRows(wb, start, end + 1, refColumnToRemoveNoneRows)  # removing unwanted rows when reference column cell value is none
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # taking max row in sheet as end row
        lastCol = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        valuedate = Excel.alter_header_name(wb, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
        transdate = Excel.alter_header_name(valuedate, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        refno = Excel.alter_header_name(transdate, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        naration = Excel.alter_header_name(refno, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        debit = Excel.alter_header_name(naration, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)  # alter header name from the excel file to the standard column name
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        dateConversion(wb, start + 1, end + 1, dateConversionColumn1)  # converting date to standard date formate
        dateConversion(wb, start + 1, end + 1, dateConversionColumn2)  # converting date to standard date formate
        Excel.replace_to_none(wb, start, end + 1, refTextToMakeCellNone, refColumnToMakeCellNone)  # replace to None when reference text is in cell of a column
        slCreated = Excel.create_slno_column(wb, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
        Excel.finalise_column(slCreated, columns)  # standardizing count of column
        Excel.transaction_type_column(wb)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/2._ICICI_-_4642__27-11-2023-17-43-28.xlsx"
    # path = "C:/Users/Admin/Downloads/2._ICICI_-_4642.pdf"
    # path = "http://ksvca-server-01:3502/ksv/%2Funlock_pdf/2._ICICI_-_4642.pdf"
    path = ""
    # wb = openpyxl.load_workbook(path)
    result = icici4_main(path)
    # if result["data"] is not None:
    #     result["data"].save('C:/Users/Admin/Desktop/ICICI4output.xlsx')
    # else:
    #     print(result["msg"])
