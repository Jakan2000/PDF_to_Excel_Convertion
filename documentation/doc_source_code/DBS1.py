from datetime import datetime

import openpyxl

from CommonClass import Excel


def removeNone(wb, start, end, column):  # removing "None" string from a column cells
    """
        Remove the string "None" from cells in a specified column within the given Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the worksheet to be processed.
        - start (int): The starting row index (inclusive) from which removal should begin.
        - end (int): The ending row index (exclusive) until which removal should occur.
        - column (str): The column letter (e.g., 'A', 'B') representing the column where string removal should be applied.

        Returns:
        - openpyxl.Workbook: The modified workbook after removing the "None" strings.

        Notes:
        - This function iterates through the specified range of rows in the specified column.
        - For each cell in the column within the range, it checks if the cell is not None and if the cell value contains the string "None".
        - If both conditions are met, it replaces the "None" string with an empty string in the cell.
        - The original workbook is modified in place.
    """
    sheet = wb.active
    for x in range(start, end):  # iterating from start to end row
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):  # if cell in the column is not none and cell value contains the string "None"
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")  # replace the string with empty string
    return wb


def deleteHeader(wb, start):
    """
        Delete rows above the specified start index row in the active sheet of an Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the worksheet to be processed.
        - start (int): The row index from which rows above will be deleted (inclusive).

        Returns:
        - openpyxl.Workbook: The modified workbook after deleting the header rows.

        Notes:
        - This function removes rows above the specified 'start' index in the active sheet.
        - The 'start' parameter denotes the row index from which rows above will be deleted.
        - If 'start' is 1, it will remove all rows above the first row, effectively clearing the sheet.
        - The original workbook is modified in place.
    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating from table data header row to sheet's 1st row
        sheet.delete_rows(x)  # delete row
    return wb


def dateConversion(wb, start, end, column):
    """
        Convert date values in a specified column of an Excel workbook to a standard date format.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the worksheet to be processed.
        - start (int): The starting row index (inclusive) from which date conversion should begin.
        - end (int): The ending row index (exclusive) until which date conversion should occur.
        - column (str): The column letter (e.g., 'A', 'B') representing the column where date conversion should be applied.

        Returns:
        - openpyxl.Workbook: The modified workbook after date conversion.

        Notes:
        - This function assumes the date values in the specified column are in the format '%d/%m/%Y'.
        - The converted dates are stored in the same column as standard Python date objects.
        - The 'start' parameter denotes the first row to start converting dates.
        - The 'end' parameter denotes the row where date conversion stops (exclusive).
        - The original workbook is modified in place.
    """
    sheet = wb.active
    for i in range(start, end):  # iterate from table data start to end row
        sheet[f"{column}{i}"].value = datetime.strptime(sheet[f"{column}{i}"].value, "%d/%m/%Y").date()  # converting to standard date formate
    return wb


def deleteFooter(wb, end):
    """
        Delete all rows below the specified end (last) row in the active sheet of an Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the worksheet to be processed.
        - end (int): The row index until which rows below will be deleted (exclusive).

        Returns:
        - openpyxl.Workbook: The modified workbook after deleting the footer rows.

        Notes:
        - This function removes rows below the specified 'end' index in the active sheet.
        - The 'end' parameter denotes the row index until which rows below will be deleted.
        - If 'end' is the last row index, it will effectively clear the sheet below that row.
        - The original workbook is modified in place.
    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through max row in the sheet to last row of table data
        sheet.delete_rows(x)  # delete row
    return wb


def removingNoneRows(wb, start, end, refColumn):
    """
        Remove rows with empty values (None) in a specified reference column within the given Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the worksheet to be processed.
        - start (int): The starting row index (inclusive) from which removal should begin.
        - end (int): The ending row index (exclusive) until which removal should occur.
        - refColumn (str): The column letter (e.g., 'A', 'B') representing the reference column for checking empty values.

        Returns:
        - openpyxl.Workbook: The modified workbook after removing rows with empty values.

        Notes:
        - This function removes rows with empty values (None) in the specified reference column.
        - The 'start' parameter denotes the first row to start checking for empty values.
        - The 'end' parameter denotes the row where checking for empty values stops (exclusive).
        - The 'refColumn' parameter indicates the column to be used as a reference for empty value checks.
        - The original workbook is modified in place.
    """
    sheet = wb.active
    for x in range(end, start, -1):  # iterating through table data end row to table data start row
        a_cell = f"{refColumn}{x}"  # get reference column cell address
        if sheet[a_cell].value is None:  # if reference column cell value is none
            sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merge scattered rows in a specified merging column based on a reference column within the given Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the worksheet to be processed.
        - start (int): The starting row index (inclusive) from which merging should begin.
        - end (int): The ending row index (exclusive) until which merging should occur.
        - refColumn (str): The column letter (e.g., 'A', 'B') representing the reference column for row identification.
        - mergingColumn (str): The column letter (e.g., 'C', 'D') representing the column where merging should be applied.

        Returns:
        - openpyxl.Workbook: The modified workbook after merging scattered rows.

        Notes:
        - This function merges rows in the specified merging column based on the values in the reference column.
        - The 'start' parameter denotes the first row to start merging.
        - The 'end' parameter denotes the row where merging stops (exclusive).
        - The 'refColumn' parameter indicates the column used for identifying the starting row of a group to be merged.
        - The 'mergingColumn' parameter indicates the column where merging should be applied.
        - The original workbook is modified in place.
    """
    sheet = wb.active
    dataToMerge = []    # array to store row data which was scattered
    for i in range(start, end):  # iterate through start and end row
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1st index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def dbs1_validation(wb):  # the conversion of pdf to excel may cause this statement to have both 8 or 9 columns, this core logic can handle both
    """
        Validate the number of columns in the active sheet of an Excel workbook for a specific core logic.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the worksheet to be validated.

        Returns:
        - bool: True if the number of columns is neither 8 nor 9, indicating an invalid format.
                False if the number of columns is either 8 or 9, indicating a valid format.

        Notes:
        - This function checks whether the number of columns in the active sheet is compatible with the expected format.
        - The validation is based on a specific core logic that can handle either 8 or 9 columns.
        - If the number of columns is 8 or 9, the format is considered valid; otherwise, it is considered invalid.
        - The original workbook is not modified by this function.
    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn1 = 8
    countOfColumn2 = 9
    if max_column == countOfColumn1:
        return False
    if max_column == countOfColumn2:
        return False
    else:
        return True


def dbs1_main(wb):
    """
        Main processing logic for standardizing and transforming the content of a DBS Bank India Ltd. Excel file.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the worksheet to be processed.

        Returns:
        - dict: A dictionary containing the processed workbook and a message.
                If the workbook format is invalid, returns an error message and None.

        Notes:
        - This function performs a series of transformations on the provided Excel workbook to standardize its content.
        - It checks the validity of the workbook format using the 'dbs1_validation' function.
        - If the format is invalid, it prints an error message and returns a dictionary with a message and None.
        - If the format is valid, it proceeds with transforming the data based on specified operations.
        - The original workbook is modified in place.
    """
    sheet = wb.active
    if dbs1_validation(wb):  # validating the column count for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # returning response with error msg
    else:
        startText = "Transaction date"  # header text in column A
        endText = "Summary"  # text define the end of table data
        startEndRefColumn = "A"  # reference column contains the start and end text
        deleteFlagStartText = "DBS Bank India Ltd."  # starting row reference text to delete the rows by range
        deleteFlagStopText = "Transaction date"  # ending row reference text to delete the rows by range
        deleteFlagRefColumn = "A"  # column containing starting row reference text and ending row reference text to delete the rows by range
        columnToMerg1 = "D"  # column to merge the row data
        refColumnToMerg = "A"  # reference column (date column) to merge the rows of desired column
        refTextToRemoveRows1 = "Account statement"  # reference text to remove unwanted rows
        refTextToRemoveRows2 = "Transaction date"  # reference text to remove unwanted rows
        dateConversionColumn1 = "A"  # column to convert date to standard formate
        dateConversionColumn2 = "B"  # column to convert date to standard formate
        stringAlignColumn1 = "D"  # column to align string by removing "\n"
        stringAlignColumn2 = "E"  # column to align string by removing "\n"
        deleteColumnRefText = "Branch code"  # header text to delete column
        refHeaderText1 = "Transaction date"  # header text to replace with standardised column name
        refHeaderText2 = "Value date"  # header text to replace with standardised column name
        refHeaderText3 = "Description"  # header text to replace with standardised column name
        refHeaderText4 = "Cheque/Reference number"  # header text to replace with standardised column name
        refHeaderText5 = "Debit"  # header text to replace with standardised column name
        refHeaderText6 = "Credit"  # header text to replace with standardised column name
        refHeaderText7 = "Balance"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Value_Date"  # standard column name
        headerText3 = "Narration"  # standard column name
        headerText4 = "ChequeNo_RefNo"  # standard column name
        headerText5 = "Withdrawal"  # standard column name
        headerText6 = "Deposit"  # standard column name
        headerText7 = "Balance"  # standard column name
        negativeValueColumnRefText1 = "Withdrawal"  # no need of removing negative value to positive
        headerTextToReplaceEmptyCellToNone1 = "ChequeNo_RefNo"  # header text to make empty cells to none in a column
        headerTextToReplaceEmptyCellToNone2 = "Withdrawal"  # header text to make empty cells to none in a column
        headerTextToReplaceEmptyCellToNone3 = "Deposit"  # header text to make empty cells to none in a column
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        duplicateHeaderRemoved = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagStopText, deleteFlagRefColumn)  # deleting the rows by range
        start, end = Excel.get_start_end_row_index(duplicateHeaderRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        rowsMergedD = mergingRows(duplicateHeaderRemoved, start, end, refColumnToMerg, columnToMerg1)  # merging rows of desired column
        noneRowsRemoved = removingNoneRows(rowsMergedD, start, end, refColumnToMerg)  # removing none rows (empty rows) in date column
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        Excel.remove_rows(wb, start, end, refTextToRemoveRows1, refColumnToMerg)  # remove multiple rows by reference text in reference column
        Excel.remove_rows(wb, start, end, refTextToRemoveRows2, refColumnToMerg)  # remove multiple rows by reference text in reference column
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        dateConvertedA = dateConversion(wb, start + 1, end, dateConversionColumn1)  # converting date to standard date formate, start-1 to Skip Header
        dateConvertedB = dateConversion(dateConvertedA, start + 1, end, dateConversionColumn2)  # converting date to standard date formate, start-1 to Skip Header
        footerDeleted = deleteFooter(dateConvertedB, end - 1)  # delete all the rows below the end(last) row, end-1 to Include End Footer
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # delete rows above the start index row, start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        alignedStringD = Excel.string_align(headerDeleted, start, end + 1, stringAlignColumn1)  # aligning string in column by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        alignedStringE = Excel.string_align(alignedStringD, start, end + 1, stringAlignColumn2)  # aligning string in column by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        removedNoneD = removeNone(alignedStringE, start, end + 1, stringAlignColumn1)  # removing "None" string from a column cells end+1 to Include Last Row
        removedNoneE = removeNone(removedNoneD, start, end + 1, stringAlignColumn2)  # removing "None" string from a column cells end+1 to Include Last Row
        branchCodeDeleted = Excel.delete_column(removedNoneE, deleteColumnRefText)  # deleting desired column by header text
        lastCol = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        transdate = Excel.alter_header_name(branchCodeDeleted, refHeaderText1, headerText1, lastCol)  # alter header name by standard column name
        valuedate = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name by standard column name
        naration = Excel.alter_header_name(valuedate, refHeaderText3, headerText3, lastCol)  # alter header name by standard column name
        chqno = Excel.alter_header_name(naration, refHeaderText4, headerText4, lastCol)  # alter header name by standard column name
        debit = Excel.alter_header_name(chqno, refHeaderText5, headerText5, lastCol)  # alter header name by standard column name
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)  # alter header name by standard column name
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)  # alter header name by standard column name
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slnoCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))  # creted new column slno
        negativeValueChecked = Excel.check_neagativeValue_by_column(slnoCreated, negativeValueColumnRefText1)  # no need of converting negative value to positive
        replacedNoneCHQNO = Excel.empty_cell_to_none(negativeValueChecked, start, end + 1, headerTextToReplaceEmptyCellToNone1)  # making empty cells in a column to none by using the header text as reference
        replacedNoneWITHDRAWAL = Excel.empty_cell_to_none(replacedNoneCHQNO, start, end + 1, headerTextToReplaceEmptyCellToNone2)  # making empty cells in a column to none by using the header text as reference
        replacedNoneDEPOSIT = Excel.empty_cell_to_none(replacedNoneWITHDRAWAL, start, end + 1, headerTextToReplaceEmptyCellToNone3)  # making empty cells in a column to none by using the header text as reference
        columnFinalised = Excel.finalise_column(wb, columns)  # standardizing count of column
        createdTransTypeColumn = Excel.transaction_type_column(wb)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/LVB-0697__05-12-2023-19-02-44.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/LVB_-_0145P.W_-_1L1675876_unlocked__12-09-2023-15-56-14.xlsx"
    path = "C:/Users/Admin/Desktop/0ba801750000001800837_ESTATEMENT_022023_0ba8017500000018__27-12-2023-23-53-03.xlsx"
    wb = openpyxl.load_workbook(path)
    result = dbs1_main(wb)
    result["data"].save('C:/Users/Admin/Desktop/DBS1output.xlsx')
