from datetime import datetime

import openpyxl

from CommonClass import Excel


def seperate_debit_credit_column(wb, sourceColumn, withdrawal, deposit):
    """
        Separate debit and credit data from a source column and store them in separate columns.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - sourceColumn (str): The column containing combined debit/credit data.
        - withdrawal (str): The column where the separated debit data will be stored.
        - deposit (str): The column where the separated credit data will be stored.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object.

        Note:
        This function iterates through the specified source column in the Workbook and separates debit and credit data.
        If the string "Dr" is found in the source column, it extracts the debit amount and stores it in the specified
        withdrawal column. If the string "Cr" is found, it extracts the credit amount and stores it in the specified
        deposit column.

    """
    sheet = wb.active
    for i in range(2, sheet.max_row + 1):  # iterating through table data 1st row (2-> skip header row) to sheet max row
        if "Dr" in str(sheet[f"{sourceColumn}{i}"].value):  # if "Dr" in source column
            temp = str(sheet[f"{sourceColumn}{i}"].value).split("(")  # split cell value with "(" to get debit amount
            sheet[f"{withdrawal}{i}"].value = temp[0].replace(",", "")  # store debit amount in withdrawal column
        if "Cr" in str(sheet[f"{sourceColumn}{i}"].value):  # if string "Cr" in source column
            temp = str(sheet[f"{sourceColumn}{i}"].value).split("(")  # split cell value with "(" to get credit amount
            sheet[f"{deposit}{i}"].value = temp[0].replace(",", "")  # store credit amount in deposit column
    return wb


def createColumn(wb, columnName):
    """
        Create a new column in the specified openpyxl Workbook with the desired header name.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object where the new column will be added.
        - columnName (str): The desired header name for the new column.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object.

        Note:
        This function adds a new column to the active sheet of the Workbook with the specified columnName as the header.
        The new column is added to the right of the existing columns.

    """
    sheet = wb.active
    sheet[f"{chr(65+sheet.max_column)}1"].value = columnName  # adding max column count with 65 (ascii value) to get next column of last column
    return wb


def dateConvertion(wb, start, end, column):
    """
        Convert date values in a specified column of an openpyxl Workbook to the standard date format.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index where the conversion should begin (inclusive).
        - end (int): The row index where the conversion should end (exclusive).
        - column (str): The column letter (e.g., 'A', 'B', 'C') representing the column containing date values.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object.

        Note:
        This function iterates through the specified range of rows and converts the date values in the given column
        to the standard date format ('%d-%m-%Y'). The modified Workbook is then returned.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%m-%Y").date()  # converting to standard date formate
    return wb


def deleteHeader(wb, start):
    """
        Delete header rows from an openpyxl Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index specifying where the header rows end (exclusive).

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object.

        Note:
        This function iterates through the range of rows from the specified 'start' index to the first row (exclusive)
        and deletes each row, effectively removing the header rows. The modified Workbook is then returned.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating though table data start row to 1st column of sheet
        sheet.delete_rows(x)  # delete rows
    return wb


def deleteFooter(wb, end):
    """
        Delete footer rows from an openpyxl Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - end (int): The row index specifying where the footer rows end (exclusive).

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object.

        Note:
        This function iterates through the range of rows from the maximum row in the sheet to the specified 'end' index (exclusive),
        and deletes each row, effectively removing the footer rows. The modified Workbook is then returned.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through max row of sheet to table data end row
        sheet.delete_rows(x)  # deleting rows
    return wb


def removeNoneRows(wb, start, end, column):
    """
        Remove rows from an openpyxl Workbook where the cell in the specified reference column is None.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index specifying where the data starts (inclusive).
        - end (int): The row index specifying where the data ends (exclusive).
        - column (str): The column letter representing the reference column to check for None values.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object.

        Note:
        This function iterates through the range of rows from 'end' to 'start' (exclusive), checks if the cell in the specified
        reference column is None, and removes the rows where the condition is met. The modified Workbook is then returned.

    """
    sheet = wb.active
    for x in range(end - 1, start, -1):  # iterating through table data end row to table data start row
        if sheet[f"{column}{x}"].value is None:  # if cell value is none
            sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merge consecutive rows in the specified column of an openpyxl Workbook based on a reference column.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index specifying where the merging operation starts (inclusive).
        - end (int): The row index specifying where the merging operation ends (exclusive).
        - refColumn (str): The column letter representing the reference column to determine row grouping.
        - mergingColumn (str): The column letter representing the column to merge the data.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object.

        Note:
        This function iterates through the specified range of rows from 'start' to 'end' (exclusive) in the Workbook.
        It merges consecutive rows in the 'mergingColumn' based on the presence of data in the 'refColumn'.
        The reference column 'refColumn' is used to determine the starting row for merging.
        The modified Workbook is then returned.

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through table data start row to table data end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def kotak2_validation(wb):
    """
        Validate the column count for the core logic in a Kotak2 Bank statement Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.

        Returns:
        - bool: True if the column count does not match the expected count, False otherwise.

        Note:
        This function checks whether the number of columns in the active sheet of the provided Workbook 'wb'
        matches the expected column count for the designed core logic (countOfColumn). If the column count is
        different from the expected count, it returns True indicating a validation failure. Otherwise, it returns
        False indicating a successful validation.

    """
    sheet = wb.active
    max_column = sheet.max_column  # get sheet max column
    countOfColumn = 5  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def kotak2_main(wb):
    """
        Process and standardize a Kotak Bank statement Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.

        Returns:
        - dict: A dictionary containing the processed workbook and an optional message.
        Example: {"data": wb, "msg": None}

        Note:
        This function performs several data processing steps to standardize the format of a Kotak Bank statement
        stored in an Excel workbook ('wb'). It includes operations such as merging rows, removing unnecessary rows,
        deleting headers and footers, creating new columns, and converting dates to a standard format.

        If the column count validation using 'kotak2_validation' fails, it prints an error message and returns a
        response dictionary with an error message.

    """
    if kotak2_validation(wb):  # validating column count for core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return response with error msg
    else:
        sheet = wb.active
        startText = "Date"  # header text to define table data start row
        stopText = "Statement  Summary"   # text define table data end row
        startEndRefColumn = "A"  # column containing start end text
        deleteFlagStartText = "Period"  # starting row reference text to delete rows by range
        deleteFlagEndText = "Narration"  # ending row reference text to delete rows by range
        deleteFlagRefColumn = "B"  # column contains starting row reference text and ending row reference text to delete the rows by range
        columnToMerg1 = "B"  # column to merge misaligned rows
        columnToMerg2 = "C"  # column to merge misaligned rows
        refColumnToMerg = "A"  # reference column to merge other column rows
        refTextToRemove = "None"  # reference text to remove from column
        dateConversionColumn1 = "A"  # column to convert date to standard date formate
        newColumnName1 = "Withdrawal"  # header text to create new column
        newColumnName2 = "Deposit"  # hedaer text to create new column
        sourceDataColumn = "D"  # column containing credit and debit data
        withdrawalColumn = "F"  # column to store debit amount
        depositColumn = "G"  # column to store credit amount
        refTextToDeleteColumn = "Withdrawal (Dr)"  # header text to delete column
        refStringFromBalanceColumn = "(Cr)"  # reference string to define amount type(debit or credit)
        removeStringRefColumn = "D"  # column to remove reference string
        refHeaderText1 = "Date"  # header text to replace with standardised column name
        refHeaderText2 = "Narration"  # header text to replace with standardised column name
        refHeaderText3 = "Chq/Ref No"  # header text to replace with standardised column name
        refHeaderText4 = "Balance"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Narration"  # standard column name
        headerText3 = "ChequeNo_RefNo"  # standard column name
        headerText4 = "Balance"  # standard column name
        columns = ["Sl.No.", "Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        refStringToRemove = ","  # reference string to remove from column
        columnToRemoveString1 = "D"  # column to remove reference text from column
        columnToAlignString1 = "B"  # column to align string by removing "\n"
        headerToReplaceEmptyCellToNone1 = "ChequeNo_RefNo"  # header text to replace empty cells to none in column
        start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)  # get start and end row index to specify table data with in
        dupHeaderRemoved = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagEndText, deleteFlagRefColumn)  # deleting the rows by range
        start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)  # get start and end row index to specify table data with in
        mergedColumnB = mergingRows(dupHeaderRemoved, start, end, refColumnToMerg, columnToMerg1)  # merging the rows of desired column
        mergedColumnC = mergingRows(mergedColumnB, start, end, refColumnToMerg, columnToMerg2)  # merging the rows of desired column
        noneRowsRemoved = removeNoneRows(mergedColumnC, start, end, refColumnToMerg)  # removing rows if cell in reference column is none
        start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)  # get start and end row index to specify table data with in
        removedNoneB = Excel.remove_string(noneRowsRemoved, start, end, refTextToRemove, columnToMerg1)  # remove desired string from a column
        removedNoneC = Excel.remove_string(removedNoneB, start, end, refTextToRemove, columnToMerg2)  # remove desired string from a column
        footerDeleted = deleteFooter(removedNoneC, end - 1)  # deleting footer rows
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # deleting header rows
        start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)  # get start and end row index to specify table data with in
        convertedDateA = dateConvertion(headerDeleted, start + 1, end + 1, dateConversionColumn1)  # converting date to standard date formate
        debitCreated = createColumn(convertedDateA, newColumnName1)  # creating new withdrawal column
        creditCreated = createColumn(convertedDateA, newColumnName2)  # creating new new deposit column
        dataSeperated = seperate_debit_credit_column(wb, sourceDataColumn, withdrawalColumn, depositColumn)  # seperating debit and credit data to diffrent columns
        sourceDataColumnDeleted = Excel.delete_column(dataSeperated, refTextToDeleteColumn)  # deleting "Withdrawal (Dr)" column
        CRremoved = Excel.remove_string(sourceDataColumnDeleted, start + 1, end + 1, refStringFromBalanceColumn, removeStringRefColumn)  # remove desired string from a column
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 -> ASCII value
        slnoCreated = Excel.create_slno_column(CRremoved, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
        lastCol = 65 + sheet.max_column  # 65 => ASCII value "A"
        transdate = Excel.alter_header_name(convertedDateA, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
        narration = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        chqno = Excel.alter_header_name(narration, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        finalisedColumns = Excel.finalise_column(balance, columns)  # standardizing count of column
        comaRemoved = Excel.remove_string(finalisedColumns, start, end + 1, refStringToRemove, columnToRemoveString1)  # remove desired string from a column
        stringAlignedB = Excel.string_align(comaRemoved, start, end + 1, columnToAlignString1)  # aligning string in column by removing the \n from the string
        replacedToNoneCHQNO = Excel.empty_cell_to_none(stringAlignedB, start, end + 1, headerToReplaceEmptyCellToNone1)  # making empty cells in a column to none by using the header text as reference
        createdTransTypeColumn = Excel.transaction_type_column(replacedToNoneCHQNO)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Desktop/Kotak-9006_unlocked__25-12-2023-16-59-51.xlsx"
    # path = "C:/Users/Admin/Downloads/Kotak_-_5887 (1)_unlocked__28-12-2023-11-17-40.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    result = kotak2_main(wb)
    # result["data"].save("C:/Users/Admin/Desktop/Kotak2output.xlsx")