from datetime import datetime

import openpyxl
from openpyxl.utils import column_index_from_string

from CommonClass import Excel


def removeString(wb, start, end, refText, column):
    """
        Remove a particular string from a specified column in the given range.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The starting row index.
        - end (int): The ending row index.
        - refText (str): The string to be removed from the specified column.
        - column (str): The column letter (e.g., "A", "B") from which the string should be removed.

        Returns:
        - Workbook: The modified Openpyxl Workbook object.

        Note:
        This function iterates through the specified range and checks if the reference text
        is present in the cell value of the specified column. If found, it replaces the reference
        text with an empty string.

    """
    sheet = wb.active
    for x in range(start, end):  # iterating through table data start row to table data end row
        if refText in str(sheet[f"{column}{x}"].value):  # if reference text in cell value
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace(refText, "")  # replace reference text with empty string
    return wb


def deleteColumn(wb, column):
    """
        Delete a specified column from the given Openpyxl Workbook.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - column (str): The column letter (e.g., "A", "B") to be deleted.

        Returns:
        - Workbook: The modified Openpyxl Workbook object.

        Note:
        This function uses the Openpyxl library to delete the specified column from the active sheet of the Workbook.

    """
    sheet = wb.active
    column_index = openpyxl.utils.column_index_from_string(column)  # getting column index
    sheet.delete_cols(column_index)  # deleting column
    return wb


def makeNone(wb, start, end, column):
    """
        Replace empty cells with None in the specified column of the given Openpyxl Workbook.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The starting row index for the operation.
        - end (int): The ending row index for the operation.
        - column (str): The column letter (e.g., "A", "B") to process.

        Returns:
        - Workbook: The modified Openpyxl Workbook object.

        Note:
        This function iterates through the specified column in the active sheet of the Workbook,
        replaces empty cells with None, and returns the modified Workbook.

    """
    sheet = wb.active
    for x in range(start, end):  # iterating through table data start row to table data end row
        if len(sheet[f"{column}{x}"].value) < 1:  # if length of cell value < 1, it's empty cell
            sheet[f"{column}{x}"].value = None  # assign none to cell value
    return wb


def removeNone(wb, start, end, column):
    """
        Remove the string "None" from cells in the specified column of the given Openpyxl Workbook.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The starting row index for the operation.
        - end (int): The ending row index for the operation.
        - column (str): The column letter (e.g., "A", "B") to process.

        Returns:
        - Workbook: The modified Openpyxl Workbook object.

        Note:
        This function iterates through the specified column in the active sheet of the Workbook,
        removes the string "None" from non-empty cells, and returns the modified Workbook.

    """
    sheet = wb.active
    for x in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):  # if cell value is not none and string "None" in cell value
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")  # replace it with empty string
    return wb


def dateConversion(wb, start, end, column):
    """
        Convert the date values in the specified column of the given Openpyxl Workbook to a standard date format.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The starting row index for the operation.
        - end (int): The ending row index for the operation.
        - column (str): The column letter (e.g., "A", "B") containing the date values.

        Returns:
        - Workbook: The modified Openpyxl Workbook object.

        Note:
        This function iterates through the specified column in the active sheet of the Workbook,
        converts date values to the standard date format ("%d/%m/%Y"), and returns the modified Workbook.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d/%m/%Y").date()  # convret date to standard date formate
    return wb


def removeHeader(wb, start):
    """
        Remove rows above the specified start index row in the active sheet of the given Openpyxl Workbook.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The index of the row where the table data starts.

        Returns:
        - Workbook: The modified Openpyxl Workbook object.

        Note:
        This function iterates through rows from the specified start index to the first row (1),
        deleting each row and returns the modified Workbook.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to 1st row of sheet
        sheet.delete_rows(x)  # delete row
    return wb


def removeNoneRows(wb, start, end, column):
    """
        Remove rows in the active sheet of the given Openpyxl Workbook where the specified reference column cell value is None.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The index of the row where the table data starts.
        - end (int): The index of the row where the table data ends (exclusive).
        - column (str): The reference column letter (e.g., 'A').

        Returns:
        - Workbook: The modified Openpyxl Workbook object.

        Note:
        This function iterates through rows from the end row to the start row (exclusive),
        deleting each row where the reference column cell value is None, and returns the modified Workbook.

    """
    sheet = wb.active
    for x in range(end - 1, start, -1):  # iterating through table data end row to table data start row
        if sheet[f"{column}{x}"].value is None:  # if reference column cell value is none
            sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    def mergingRows(wb, start, end, refColumn, mergingColumn):
        """
        Merge the rows of a desired column based on a reference column in the active sheet of the given Openpyxl Workbook.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The index of the row where the table data starts.
        - end (int): The index of the row where the table data ends (exclusive).
        - refColumn (str): The reference column letter (e.g., 'A') to identify starting rows.
        - mergingColumn (str): The column letter (e.g., 'B') whose rows need to be merged.

        Returns:
        - Workbook: The modified Openpyxl Workbook object.

        Note:
        This function iterates through rows in the specified range and merges rows in the desired column ('mergingColumn')
        based on the presence of values in the reference column ('refColumn'). It concatenates the values of consecutive rows
        in 'mergingColumn' until a new starting row in 'refColumn' is encountered.

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through table data start row to table data end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def indian_bank1_validation(wb):
    """
        Validate columns for the core logic in the active sheet of the given Openpyxl Workbook.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.

        Returns:
        - bool: True if the number of columns is not equal to the expected count, False otherwise.

        Note:
        This function checks if the number of columns in the active sheet of the workbook matches the expected count.
        If the count is not equal to the expected value, it returns True, indicating a validation failure. Otherwise,
        it returns False, indicating successful validation.

    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 8  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def indian_bank1_main(wb):
    """
        Process and standardize the data in the active sheet of the given Openpyxl Workbook based on the requirements
        specific to Indian Bank1.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.

        Returns:
        - dict: A dictionary containing the processed workbook and a message.

        Note:
        This function performs various operations on the input workbook to standardize the data according to
        Indian Bank1 specifications. It validates the columns, removes unwanted rows, deletes specific columns,
        merges rows, converts date formats, alters header names, creates additional columns, and more.

    """
    sheet = wb.active
    if indian_bank1_validation(wb):  # validate columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return response with error msg
    else:
        startText = "ValueDate"  # header text to define table data start row
        endText = " Statement Downloaded By"  # reference text to define table data end row
        startEndRefColumn = "A"  # column contains the start and end text
        deleteFlagStartText1 = "Page No"  # starting row reference text to delete rows by range
        deleteFlagEndText1 = "ValueDate"  # ending row reference text to delete rows by range
        deleteFlagStartText2 = "Statement Downloaded By"  # starting row reference text to delete rows by range
        deleteFlagEndText2 = "END OF STATEMENT"  # ending row reference text to delete rows by range
        deleteFlagRefColumn = "A"  # column contains starting row reference text and ending row reference text to delete the rows by range
        removeRowRefText1 = "BALANCE B/F"  # reference string to remove row
        removeRowRefColumn1 = "D"  # column to remove row with reference string
        removeRowRefText2 = "Download Limit"  # reference string to remove row
        removeRowRefColumn2 = "A"  # column to remove row with reference string
        removeRowRefText3 = "Page No"  # reference string to remove row
        removeRowRefColumn3 = "A"  # column to remove row with reference string
        columnToMerg = "D"  # column to merge misaligned rows
        refColumnToMerg = "A"  # reference column to merge rows of other columns
        removeNoneRowRefColumn = "A"  # reference column to remove none rows (unwanted rows)
        dateConersionColumn1 = "A"  # column to convert date to standard date formate
        dateConersionColumn2 = "B"  # column to convert date to standard date formate
        stringAlignColumn1 = "A"  # column to aligning string by removing the \n from string
        stringAlignColumn2 = "B"  # column to aligning string by removing the \n from string
        stringAlignColumn3 = "C"  # column to aligning string by removing the \n from string
        stringAlignColumn4 = "D"  # column to aligning string by removing the \n from string
        stringAlignColumn5 = "E"  # column to aligning string by removing the \n from string
        stringAlignColumn6 = "F"  # column to aligning string by removing the \n from string
        stringAlignColumn7 = "G"  # column to aligning string by removing the \n from string
        stringAlignColumn8 = "H"  # column to aligning string by removing the \n from string
        makeNoneColumn1 = "A"  # column to make empty cells to none
        deleteColumn1 = "C"  # column to delete
        refHaederText1 = "ValueDate"  # header text to replace with standardised column name
        refHaederText2 = "PostDate"  # header text to replace with standardised column name
        refHaederText3 = "Description"  # header text to replace with standardised column name
        refHaederText4 = "Cheque No"  # header text to replace with standardised column name
        refHaederText5 = "DR"  # header text to replace with standardised column name
        refHaederText6 = "CR"  # header text to replace with standardised column name  # header text to replace with standardised column name
        refHaederText7 = "Balance"  # header text to replace with standardised column name
        headerText1 = "Value_Date"  # standard column name
        headerText2 = "Transaction_Date"  # standard column name
        headerText3 = "Narration"  # standard column name
        headerText4 = "ChequeNo_RefNo"  # standard column name
        headerText5 = "Withdrawal"  # standard column name
        headerText6 = "Deposit"  # standard column name
        headerText7 = "Balance"  # standard column name
        UnWantedRefText1 = "CR"  # reference string to remove from column
        columnToRemoveUnWantedText1 = "G"  # column to remove unwanted string
        start = 1  # assigning table data start row as 1
        end = sheet.max_row  # assigning table data end row as max row of sheet
        columnTextToMakeEmptyCellToNone1 = "ChequeNo_RefNo"  # column header text to make empty cells to none
        columnTextToMakeEmptyCellToNone2 = "Withdrawal"  # column header text to make empty cells to none
        columnTextToMakeEmptyCellToNone3 = "Deposit"  # column header text to make empty cells to none
        alignedStringA = Excel.string_align(wb, start, end + 1, stringAlignColumn1)  # aligning string in column by removing the \n from the string -> \n -> next line
        alignedStringB = Excel.string_align(alignedStringA, start, end + 1, stringAlignColumn2)  # aligning string in column by removing the \n from the string -> \n -> next line
        alignedStringC = Excel.string_align(alignedStringB, start, end + 1, stringAlignColumn3)  # aligning string in column by removing the \n from the string -> \n -> next line
        alignedStringD = Excel.string_align(alignedStringC, start, end + 1, stringAlignColumn4)  # aligning string in column by removing the \n from the string -> \n -> next line
        alignedStringE = Excel.string_align(alignedStringD, start, end + 1, stringAlignColumn5)  # aligning string in column by removing the \n from the string -> \n -> next line
        alignedStringF = Excel.string_align(alignedStringE, start, end + 1, stringAlignColumn6)  # aligning string in column by removing the \n from the string -> \n -> next line
        alignedStringG = Excel.string_align(alignedStringF, start, end + 1, stringAlignColumn7)  # aligning string in column by removing the \n from the string -> \n -> next line
        alignedStringH = Excel.string_align(alignedStringG, start, end + 1, stringAlignColumn8)  # aligning string in column by removing the \n from the string -> \n -> next line
        noneRemovedA = removeNone(alignedStringH, start, end + 1, stringAlignColumn1)  # removing "None" string from A colum
        noneRemovedB = removeNone(noneRemovedA, start, end + 1, stringAlignColumn2)  # removing "None" string from B colum
        noneRemovedC = removeNone(noneRemovedB, start, end + 1, stringAlignColumn3)  # removing "None" string from C colum
        noneRemovedD = removeNone(noneRemovedC, start, end + 1, stringAlignColumn4)  # removing "None" string from D colum
        noneRemovedE = removeNone(noneRemovedD, start, end + 1, stringAlignColumn5)  # removing "None" string from E colum
        noneRemovedF = removeNone(noneRemovedE, start, end + 1, stringAlignColumn6)  # removing "None" string from F colum
        noneRemovedG = removeNone(noneRemovedF, start, end + 1, stringAlignColumn7)  # removing "None" string from G colum
        noneRemovedH = removeNone(noneRemovedG, start, end + 1, stringAlignColumn8)  # removing "None" string from H colum
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # assign sheet max row to end
        dupHeader1Removed = Excel.delete_rows_by_range(wb, start, end + 1, deleteFlagStartText1, deleteFlagEndText1, deleteFlagRefColumn)  # deleting the rows by range
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # assign sheet max row to end
        dupHeader2Removed = Excel.delete_rows_by_range(dupHeader1Removed, start, end + 1, deleteFlagStartText2, deleteFlagEndText2, deleteFlagRefColumn)  # deleting the rows by range
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # assign sheet max row to end
        openBalRowsRemoved = Excel.remove_rows(dupHeader2Removed, start, end + 1, removeRowRefText1, removeRowRefColumn1)  # remove multiple rows by reference text in reference column
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # assign sheet max row to end
        unWantedRowRemoved = Excel.remove_rows(openBalRowsRemoved, start, end + 1, removeRowRefText2, removeRowRefColumn2)  # remove multiple rows by reference text in reference column
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        pageNoRowRemoved = Excel.remove_rows(unWantedRowRemoved, start, end, removeRowRefText3, removeRowRefColumn3)  # remove multiple rows by reference text in reference column
        headerRemoved = removeHeader(pageNoRowRemoved, start - 1)  # start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        makedNoneA = makeNone(wb, start, end + 1, makeNoneColumn1)  # making empty cells to none
        mergedRowsD = mergingRows(makedNoneA, start, end + 1, refColumnToMerg, columnToMerg)  # merging the rows of desired column
        noneRowsRemoved = removeNoneRows(pageNoRowRemoved, start, end + 1, removeNoneRowRefColumn)  # removing none rows by reference column
        start, end = Excel.get_start_end_row_index(headerRemoved, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        convertedDateA = dateConversion(headerRemoved, start + 1, end + 1, dateConersionColumn1)  # converting date to standard date formate, start+1 to Skip Header, end+1 to Include Last Row
        convertedDateB = dateConversion(convertedDateA, start + 1, end + 1, dateConersionColumn2)  # converting date to standard date formate, start+1 to Skip Header, end+1 to Include Last Row
        deletedBRANCHcolumn = deleteColumn(convertedDateB, deleteColumn1)  # deleting column C
        lastCol = 65 + Excel.column_count(wb)  # 65 ASCII value
        valuedate = Excel.alter_header_name(deletedBRANCHcolumn, refHaederText1, headerText1, lastCol)  # altering header to standard header name
        transdate = Excel.alter_header_name(valuedate, refHaederText2, headerText2, lastCol)  # altering header to standard header name
        narration = Excel.alter_header_name(transdate, refHaederText3, headerText3, lastCol)  # altering header to standard header name
        chqno = Excel.alter_header_name(narration, refHaederText4, headerText4, lastCol)  # altering header name to standard header name
        debit = Excel.alter_header_name(chqno, refHaederText5, headerText5, lastCol)  # altering header name to standard header name
        credit = Excel.alter_header_name(debit, refHaederText6, headerText6, lastCol)  # altering header name to standard header name
        balance = Excel.alter_header_name(credit, refHaederText7, headerText7, lastCol)  # altering header name to standard header name
        removedCR = removeString(balance, start, end + 1, UnWantedRefText1, columnToRemoveUnWantedText1)  # removing string "CR" from column G end+1 to include last row
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 ASCII value
        slCreated = Excel.create_slno_column(removedCR, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column end+1 to include last row
        replacedNoneCHQNO = Excel.empty_cell_to_none(slCreated, start, end + 1, columnTextToMakeEmptyCellToNone1)  # making empty cells to none, end+1 to include last row
        replacedNoneWITHDRAWAL = Excel.empty_cell_to_none(replacedNoneCHQNO, start, end + 1, columnTextToMakeEmptyCellToNone2)  # making empty cells to none, end+1 to include last row
        replacedNonedeposit = Excel.empty_cell_to_none(replacedNoneWITHDRAWAL, start, end + 1, columnTextToMakeEmptyCellToNone3)  # making empty cells to none end+1 to include last row
        createdTransTypeColumn = Excel.transaction_type_column(replacedNonedeposit)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/SRT_-_INDIAN_BANK_-_6096825697___23-09-2023-10-24-18.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    result = indian_bank1_main(wb)
    # result.save("C:/Users/Admin/Desktop/FinalOutput/INDIAN_BANK1output.xlsx")
