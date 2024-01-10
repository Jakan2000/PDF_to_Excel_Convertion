from datetime import datetime

import openpyxl
from openpyxl.utils import column_index_from_string

from CommonClass import Excel


def aligningColumn(wb, start, end, mergingColumn, refColumn):
    """
        Overview:
        Aligns misaligned column data by copying values from the reference column to the merging column when the reference column cell value is not None.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - start (int): The starting row index for processing.
        - end (int): The ending row index for processing.
        - mergingColumn (str): The column where misaligned data will be aligned.
        - refColumn (str): The reference column used to align data.

        Returns:
        - wb (Workbook): The modified Excel workbook object after aligning the specified column data.

        Note:
        The function assumes that the workbook contains an active sheet. It aligns the data in the specified merging column based on the
        values present in the reference column. The alignment is performed only when the reference column cell value is not None.
    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row and table data end row
        if sheet[f"{refColumn}{i}"].value is not None:  # if reference column cell value is none
            sheet[f"{mergingColumn}{i}"].value = sheet[f"{refColumn}{i}"].value  # assign reference column cell value to merging column cell value
            sheet[f"{refColumn}{i}"].value = None  # assign reference column cell value to None
    return wb


def dateConvertion(wb, start, end, column):
    """
        Converts dates in the specified column to a standard date format.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - start (int): The starting row index for processing.
        - end (int): The ending row index for processing.
        - column (str): The column containing date values to be converted.

        Returns:
        - wb (Workbook): The modified Excel workbook object after converting the specified column's date values.

        Note:
        The function assumes that the workbook contains an active sheet. It iterates through the specified column's data
        and converts the date values to a standard date format ('%d-%b-%Y'). The input dates are assumed to be in the 'dd-MMM-yyyy' format.
    """
    sheet = wb.active
    for i in range(start, end):  # iterate through table data start row and table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), '%d-%b-%Y').date()  # converting date to standard date formate
    return wb


def remove_space(wb, start, end, column):
    """
        Removes spaces from strings in the specified column and replaces them with an empty string.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - start (int): The starting row index for processing.
        - end (int): The ending row index for processing.
        - column (str): The column containing strings with spaces to be removed.

        Returns:
        - wb (Workbook): The modified Excel workbook object after removing spaces from the specified column's strings.

        Note:
        The function assumes that the workbook contains an active sheet. It iterates through the specified column's data
        and replaces any spaces in the strings with an empty string.
    """
    sheet = wb.active
    for x in range(start, end):  # iterating through table data start row to table data end row
        sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace(" ", "")  # replacing space with empty string
    return wb


def removeHeader(wb, start):
    """
        Deletes rows above the specified start index row.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - start (int): The index of the row from which to start removing rows.

        Returns:
        - wb (Workbook): The modified Excel workbook object after removing rows above the specified start index.

        Note:
        The function assumes that the workbook contains an active sheet. It iterates through rows from the specified start
        index to the first row of the sheet, deleting each row.
    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to 1st row of the sheet
        sheet.delete_rows(x) # delete row
    return wb


def deleteColumn(wb, column):
    """
        Deletes the specified column in the Excel workbook.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - column (str): The header name of the column to be deleted.

        Returns:
        - wb (Workbook): The modified Excel workbook object after deleting the specified column.

        Note:
        The function assumes that the workbook contains an active sheet. It uses the openpyxl.utils.column_index_from_string
        function to obtain the column index from the header name and then deletes the entire column.
    """
    sheet = wb.active
    column_index = openpyxl.utils.column_index_from_string(column)  # getting column index by header name using inbuild function
    sheet.delete_cols(column_index)  # delete column
    return wb


def removeRowsByDateLength(wb, start, end, column):
    """
        Removes rows from an Excel workbook based on the length of the date string in a specified column.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - start (int): The starting row index of the table data.
        - end (int): The ending row index of the table data.
        - column (str): The header name of the column containing date values.

        Returns:
        - wb (Workbook): The modified Excel workbook object after removing rows based on date length.

        Note:
        The function iterates through the specified range of rows (from end to start) and checks the length of the date
        string in the specified column. If the length is less than a predefined year length, the row is deleted.
    """
    sheet = wb.active
    yearLength = 6  # length of the string to remove
    for x in range(end, start, -1):  # iterating through table data start row and table data end row
        if len(str(sheet[f"{column}{x}"].value)) < yearLength:  # if cell value length is < year length
            sheet.delete_rows(x)  # delete row
    return wb


def removeNone(wb, start, end, column):
    """
        Removes occurrences of the string "None" from cells in a specified column within the specified range.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - start (int): The starting row index of the table data.
        - end (int): The ending row index of the table data.
        - column (str): The header name of the column containing string values.

        Returns:
        - wb (Workbook): The modified Excel workbook object after removing "None" from the specified column.

        Note:
        The function iterates through the specified range of rows and checks if the cell value in the specified column
        is not None and if "None" is present in the cell value string. If the conditions are met, "None" is replaced
        with an empty string.
    """
    sheet = wb.active
    for x in range(start, end):  # iterate through table data start column to table data end column
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):  # if cell value is not none and "None" in cell value string
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")  # replace "None" with empty string
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merges rows in a specified column based on a reference column in the given range.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - start (int): The starting row index of the table data.
        - end (int): The ending row index of the table data.
        - refColumn (str): The header name of the reference column.
        - mergingColumn (str): The header name of the column to be merged.

        Returns:
        - wb (Workbook): The modified Excel workbook object after merging rows in the specified column.

        Note:
        The function iterates through the specified range of rows and identifies starting rows based on the
        non-None values in the reference column. It then concatenates the values in the merging column for
        consecutive starting rows and updates the corresponding cell with the merged data. The last row is also
        merged after the loop.
    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through table data start row to table data end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def mergingDateColumn(wb, start, end, column):
    """
        Merges date column cell values for incomplete date strings in the given range.

        Parameters:
        - wb (Workbook): The input Excel workbook object.
        - start (int): The starting row index of the table data.
        - end (int): The ending row index of the table data.
        - column (str): The header name of the date column to be merged.

        Returns:
        - wb (Workbook): The modified Excel workbook object after merging date column cell values.

        Note:
        The function iterates through the specified range of rows and checks for incomplete date strings in the
        date column. If an incomplete date is identified (length less than inCompleteDateLen) and the next row's
        date has a complete year (length equal to yearLength), it concatenates them to form a complete date.
    """
    sheet = wb.active
    inCompleteDateLen = 10  # incomplete date length (string)
    yearLength = 4  # year length (string)
    for i in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{column}{i}"].value is not None:  # if cell value in date column is not none
            if len(str(sheet[f"{column}{i}"].value)) < inCompleteDateLen and len(str(sheet[f"{column}{i + 1}"].value)) == yearLength:  # if length of date in cell is less than incomplete date length and length of date in next row cell is equal to year length
                s = str(sheet[f"{column}{i}"].value) + " " + str(sheet[f"{column}{i + 1}"].value)  # concat the date with year
                sheet[f"{column}{i}"].value = s  # assign it to date cell
    return wb


def icici3_validation(wb):
    """
        Validates columns for the core logic.

        Parameters:
        - wb (Workbook): The input Excel workbook object.

        Returns:
        - bool: True if the column count does not match the designed core logic, False otherwise.

        Note:
        The function checks if the number of columns in the active sheet matches the expected count defined
        by the designed core logic (countOfColumn). If the count does not match, it returns True; otherwise, False.
    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 9  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def icici3_main(wb):
    """
        Main function for processing ICICI3 Bank statements with specific operations.

        Parameters:
        - wb (Workbook): The input Excel workbook object.

        Returns:
        - dict: A dictionary containing processed workbook data and a message.

        Note:
        The function performs a series of operations on the input workbook based on specific requirements for ICICI Bank statements.
        If the column count validation fails using icici3_validation function, an error message is returned.
        Otherwise, it executes a sequence of operations, such as merging date columns, aligning misaligned rows, removing "None" values,
        deleting columns, altering header names, creating new columns, and replacing empty cells with "None". The processed workbook
        is returned in the 'data' key of the response dictionary.
    """
    sheet = wb.active
    if icici3_validation(wb):  # validate columns for core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return response with error msg
    else:
        startText = "Sr No"  # header text to define table data start row
        endText = ""  # reference text to define table data end column -> hear if empty string means max row of sheet
        startEndDefColumn = "A"  # reference column contains the start and end text
        deleteFlagStartText = "DETAILED STATEMENT"  # starting row reference text to delete rows by range
        deleteFlagStopText = "Sr No"  # ending row reference text to delete rows by range
        DeleteFlagRefColumn = "A"  # column contains starting row reference text and ending row reference text to delete the rows by range
        dateMergColumn1 = "B"  # column to merge incomplete date in date column cell
        dateMergColumn2 = "C"  # column to merge incomplete date in date column cell
        columnToMerg1 = "E"  # column to merge misaligned row by reference column to merge
        refColumnToMerg = "A"  # reference column to merge misaligned rows of a column
        removeNoneColumn1 = "B"  # column remove "None" from string by column
        removeNoneColumn2 = "C"  # column remove "None" from string by column
        removeNoneColumn3 = "E"  # column remove "None" from string by column
        refColumnToDeleteDateRow = "B"  # reference column to removing rows by date length
        columnToDelete = "A"  # column to delete
        stringAlignColumn1 = "A"  # column to aligning string by removing the \n from the string -> \n -> next line
        stringAlignColumn2 = "B"  # column to aligning string by removing the \n from the string -> \n -> next line
        stringAlignColumn3 = "C"  # column to aligning string by removing the \n from the string -> \n -> next line
        stringAlignColumn4 = "D"  # column to aligning string by removing the \n from the string -> \n -> next line
        stringAlignColumn5 = "E"  # column to aligning string by removing the \n from the string -> \n -> next line
        stringAlignColumn6 = "F"  # column to aligning string by removing the \n from the string -> \n -> next line
        stringAlignColumn7 = "G"  # column to aligning string by removing the \n from the string -> \n -> next line
        dateConversionColumn1 = "A"  # column to convert date to standard date formate
        dateConversionColumn2 = "B"  # column to convert date to standard date formate
        refHeaderText1 = "ValueDate"  # header text to replace with standardised column name
        refHeaderText2 = "TransactionDate"  # header text to replace with standardised column name
        refHeaderText3 = "ChequeNumber"  # header text to replace with standardised column name
        refHeaderText4 = "Transaction Remarks"  # header text to replace with standardised column name
        refHeaderText5 = "DebitAmount"  # header text to replace with standardised column name
        refHeaderText6 = "CreditAmount"  # header text to replace with standardised column name
        refHeaderText7 = "Balance(INR)"  # header text to replace with standardised column name
        headerText1 = "Value_Date"  # standard column name
        headerText2 = "Transaction_Date"  # standard column name
        headerText3 = "ChequeNo_RefNo"  # standard column name
        headerText4 = "Narration"  # standard column name
        headerText5 = "Withdrawal"  # standard column name
        headerText6 = "Deposit"  # standard column name
        headerText7 = "Balance"  # standard column name
        columnToAlign = "G"   # column to aligning misaligned data if reference column cell value is none
        refColumnToAlign = "H"  # reference column to align misaligned column data
        stringToRemove2 = "NA"  # reference string to remove from column
        columnToRemoveString4 = "F"  # column to remove reference string
        columnToRemoveString5 = "E"  # column to remove reference string
        replaceEmptyColumeByNone1 = "ChequeNo_RefNo"  # column header text to replace empty cells with none
        replaceEmptyColumeByNone2 = "Withdrawal"  # column header text to replace empty cells with none
        replaceEmptyColumeByNone3 = "Deposit"  # column header text to replace empty cells with none
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # taking max row in sheet as end row
        dupHeaderRemoved = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagStopText, DeleteFlagRefColumn)  # deleting the rows by range
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # taking max row in sheet as end row
        mergedDateB = mergingDateColumn(dupHeaderRemoved, start, end, dateMergColumn1)  # merging date column cell values -> incomplete date string
        mergedDateC = mergingDateColumn(mergedDateB, start, end, dateMergColumn2)  # merging date column cell values -> incomplete date string
        mergedColumnE = mergingRows(mergedDateC, start, end, refColumnToMerg, columnToMerg1)  # merging rows of E column
        noneRemovedB = removeNone(mergedColumnE, start, end, removeNoneColumn1)  # remove "None" from string in E column
        noneRemovedC = removeNone(noneRemovedB, start, end, removeNoneColumn2)  # remove "None" from string in C column
        noneRemovedE = removeNone(noneRemovedC, start, end, removeNoneColumn3)  # remove "None" from string in E column
        dateRowRemoved = removeRowsByDateLength(noneRemovedE, start, end, refColumnToDeleteDateRow)  # removing rows by date length
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # taking max row in sheet as end row
        headerRemoved = removeHeader(dateRowRemoved, start - 1)  # delete rows above the start index row, start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # taking max row in sheet as end row
        deletedA = deleteColumn(dateRowRemoved, columnToDelete)  # deleting column A
        alignedA = Excel.string_align(deletedA, start, end + 1, stringAlignColumn1)  # aligning string by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        alignedB = Excel.string_align(alignedA, start, end + 1, stringAlignColumn2)  # aligning string by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        alignedC = Excel.string_align(alignedB, start, end + 1, stringAlignColumn3)  # aligning string by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        alignedD = Excel.string_align(alignedC, start, end + 1, stringAlignColumn4)  # aligning string by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        alignedE = Excel.string_align(alignedD, start, end + 1, stringAlignColumn5)  # aligning string by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        alignedF = Excel.string_align(alignedE, start, end + 1, stringAlignColumn6)  # aligning string by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        alignedG = Excel.string_align(alignedF, start, end + 1, stringAlignColumn7)  # aligning string by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        noneRemovedA = removeNone(alignedG, start, end + 1, stringAlignColumn1)  # remove "None" from string by column, end+1 to Include Last Row
        noneRemovedB = removeNone(noneRemovedA, start, end + 1, stringAlignColumn2)  # remove "None" from string by column, end+1 to Include Last Row
        noneRemovedC = removeNone(noneRemovedB, start, end + 1, stringAlignColumn3)  # remove "None" from string by column, end+1 to Include Last Row
        noneRemovedD = removeNone(noneRemovedC, start, end + 1, stringAlignColumn4)  # remove "None" from string by column, end+1 to Include Last Row
        noneRemovedE = removeNone(noneRemovedD, start, end + 1, stringAlignColumn5)  # remove "None" from string by column, end+1 to Include Last Row
        noneRemovedF = removeNone(noneRemovedE, start, end + 1, stringAlignColumn6)  # remove "None" from string by column, end+1 to Include Last Row
        noneRemovedG = removeNone(noneRemovedF, start, end + 1, stringAlignColumn7)  # remove "None" from string by column, end+1 to Include Last Row
        removedSpaceA = remove_space(noneRemovedG, start, end + 1, dateConversionColumn1)  # end+1 to Include Last Row
        removedSpaceB = remove_space(noneRemovedG, start, end + 1, dateConversionColumn2)  # end+1 to Include Last Row
        dateConvertedA = dateConvertion(removedSpaceB, start + 1, end + 1, dateConversionColumn1)  # start+1 to Skip Header, end+1 to Include Last Row
        dateConvertedB = dateConvertion(dateConvertedA, start + 1, end + 1, dateConversionColumn2)  # start+1 to Skip Header, end+1 to Include Last Row
        lastCol = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        valuedate = Excel.alter_header_name(dateConvertedB, refHeaderText1, headerText1, lastCol)  # altering header name of Value_Date column to standard column nam
        transdate = Excel.alter_header_name(valuedate, refHeaderText2, headerText2, lastCol)  # altering header name of Transaction_Date column to standard column name
        refno = Excel.alter_header_name(transdate, refHeaderText3, headerText3, lastCol)  # altering header name of ChequeNo_RefNo to standard column name
        naration = Excel.alter_header_name(refno, refHeaderText4, headerText4, lastCol)  # altering header name of Narration to standard column name
        debit = Excel.alter_header_name(naration, refHeaderText5, headerText5, lastCol)  # altering header name of Withdrawal to standard column name
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)  # altering header name of Deposit to standard column name
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)  # altering header name of Balance to standard column name
        alignedColumnH = aligningColumn(balance, start, end, columnToAlign, refColumnToAlign)  # aligning misaligned column data if reference column cell value is none
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slCreated = Excel.create_slno_column(alignedColumnH, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
        noneReplacedF = Excel.replace_to_none(slCreated, start, end + 1, stringToRemove2, columnToRemoveString4)  # replace to None when stringToRemove2 is in cell of columnToRemoveString4
        noneReplacedE = Excel.replace_to_none(noneReplacedF, start, end + 1, stringToRemove2, columnToRemoveString5)  # replace to None when stringToRemove2 is in cell of columnToRemoveString5
        replacednonechqno = Excel.empty_cell_to_none(noneReplacedE, start, end + 1, replaceEmptyColumeByNone1)  # making empty cells in a column to none by using the header text as reference
        replacednonewithdrawal = Excel.empty_cell_to_none(replacednonechqno, start, end + 1, replaceEmptyColumeByNone2)  # making empty cells in a column to none by using the header text as reference
        replacednoneDeposit = Excel.empty_cell_to_none(replacednonewithdrawal, start, end + 1, replaceEmptyColumeByNone3)  # making empty cells in a column to none by using the header text as reference
        createdTransTypeColumn = Excel.transaction_type_column(replacednoneDeposit)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/ilovepdf_merged_7__26-09-2023-10-37-33.xlsx"
    wb = openpyxl.load_workbook(path)
    result = icici3_main(wb)
    result.save('C:/Users/Admin/Desktop/FinalOutput/ICICI3output.xlsx')
