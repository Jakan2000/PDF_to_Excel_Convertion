from datetime import datetime

import openpyxl

from AlignmentData import addAlignmentData
from CommonClass import Excel


def removeString(wb, start, end, refText, column):
    """
        Replaces a reference string with an empty string in a specified column of an openpyxl Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The starting row index for the operation.
        - end (int): The ending row index (exclusive) for the operation.
        - refText (str): The reference string to be replaced in the specified column.
        - column (str): The column letter where the replacement should occur.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object.

        Note:
        The function iterates through the specified range of rows in the specified column.
        If the reference string is found in a cell value, it replaces the reference string with an empty string.
        The modified Workbook is then returned.

    """
    sheet = wb.active
    for x in range(start, end):  # iterating through table data start row to table data end row
        if refText in str(sheet[f"{column}{x}"].value):  # if reference string in cell value
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace(refText, "")  # replace it with empty string
    return wb


def dateConvertion(wb, start, end, column):
    """
        Converts dates in a specified column of an openpyxl Workbook to a standard date format.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The starting row index for the operation.
        - end (int): The ending row index (exclusive) for the operation.
        - column (str): The column letter where date conversion should occur.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object.

        Note:
        The function iterates through the specified range of rows in the specified column.
        It uses the datetime.strptime method to parse the existing date format ("%d-%b-%y")
        and assigns the converted date back to the cell value.
        The modified Workbook is then returned.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%b-%y").date()  # convert date to standard date formate
    return wb


def deleteHeader(wb, start):
    """
        Deletes header rows from an openpyxl Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index from which header rows should be deleted.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object.

        Note:
        The function iterates through the specified range of rows (from 'start' to the 1st row) in reverse order,
        using openpyxl's delete_rows method to remove each row.
        The modified Workbook is then returned.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to 1st row of sheet
        sheet.delete_rows(x)  # delete row
    return wb


def alignColumns(wb, start, end, headData, refColumnToAlign):
    """
        Aligns columns manually in an openpyxl Workbook based on specified criteria.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The starting row index for iterating through the table data.
        - end (int): The ending row index for iterating through the table data.
        - headData (list): A list containing data for alignment criteria, where:
        - headData[0] (str): Account number.
        - headData[1] (str): Name.
        - headData[2] (str): Period.
        - refColumnToAlign (str): The reference column (e.g., 'G') used for alignment.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object.

        Note:
        The function iterates through the specified range of rows (from 'start' to 'end').
        If a cell in the specified reference column is not None, it extracts data from relevant columns (A to G),
        calls addAlignmentData() function to add the data to the Excel file, and appends the row index to error_records.
        Finally, for each error record, the function updates the 'Narration' column and makes other columns None.

    """
    sheet = wb.active
    error_records = []
    for i in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"G{i}"].value is not None:  # if cell in column G is not none
            data = [headData[0], headData[1], headData[2]]  # headData[0] -> account number, headData[1] -> name, headData[2] -> period
            data.append(i - 1)  # appending row
            data.append(str(sheet[f"A{i}"].value))  # append A column cell value
            data.append(str(sheet[f"B{i}"].value))  # append B column cell value
            data.append(str(sheet[f"C{i}"].value))  # append C column cell value
            data.append(str(sheet[f"D{i}"].value))  # append D column cell value
            data.append(str(sheet[f"E{i}"].value))  # append E column cell value
            data.append(str(sheet[f"F{i}"].value))  # append F column cell value
            data.append(str(sheet[f"G{i}"].value))  # append G column cell value
            addAlignmentData(data)  # calling addAlignmentData() function to add the data to excel file
            error_records.append(i)  # appending the error records row index
    for x in error_records:  # iterating through error_records array
        sheet[f"B{x}"].value = "Error Record"  # assign "error record" in narration column
        sheet[f"C{x}"].value = None  # making other column to none
        sheet[f"D{x}"].value = None  # making other columns to none
        sheet[f"E{x}"].value = None  # making other columns to none
        sheet[f"F{x}"].value = None  # making other columns to none
        sheet[f"G{x}"].value = None  # making other columns to none
    return wb


# TODO: get name , acnum - use num vs char instead of digit count
def headerData(wb, start, end):
    """
        Extracts header data from header rows in an openpyxl Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The starting row index for iterating through the table data.
        - end (int): The ending row index for iterating through the table data.

        Returns:
        - list: A list containing header data, where:
        - list[0] (str): Account number.
        - list[1] (str): Name.
        - list[2] (str): Period.

        Note:
        The function iterates through the specified range of rows (from 'start' to 'end') in reverse order.
        It searches for specific strings ('Period' in D column and 'Account No' in A column) to extract relevant information.
        The account number is extracted by removing unnecessary characters and spaces.

    """
    sheet = wb.active
    acnum = "Undefined"
    name = "Undefined"
    period = "Undefined"
    for i in range(start, 0, -1):  # iterating through table data start row to 1st row of sheet
        if "Period" in str(sheet[f"D{i}"].value):   # if "period" in D column cell value
            period = str(sheet[f"E{i}"].value)  # assigning E column cell value
        if "Account No" in str(sheet[f"A{i}"].value):  # if "account number" in A column cell value
            spl = str(sheet[f"A{i}"].value).split("Account No.")  # splitting account number from junk data
            a = spl[1].strip().replace('\n', '').split(" ")  # removing spaces from start and end string and removing
            acnum = ""
            for char in a[0]:  # iterating through characters in a[0]
                if char.isdigit():  # if character is digit
                    acnum += char  # storing it in acnum variable
    headData = [acnum, name, period]  # appending account number, name, period to headData array
    return headData


def deleteRowByDateLen(wb, start, dateLen, refColumn):
    """
        Deletes rows from an openpyxl Workbook where the length of the date in a specific column is less than a given value.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The starting row index for iterating through the table data.
        - dateLen (int): The minimum length of the date in characters to compare against.
        - refColumn (str): The reference column containing the date data.

        Returns:
        - openpyxl.Workbook: The updated Workbook object after deleting rows.

        Note:
        The function iterates through the specified range of rows (from 'sheet.max_row' to 'start') in reverse order.
        It checks the length of the date in the specified column ('refColumn') and deletes rows where the length is less than 'dateLen'.

    """
    sheet = wb.active
    for x in range(sheet.max_row, start, -1):  # iterating through max row of sheet to table data start tow
        if len(str(sheet[f"{refColumn}{x}"].value)) < dateLen:  # if length of date in column cell < dateLen
            sheet.delete_rows(x)  # delete rows
    return wb


def removeFooter(wb, end):
    """
        Deletes footer rows from an openpyxl Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - end (int): The row index indicating the end of the table data.

        Returns:
        - openpyxl.Workbook: The updated Workbook object after deleting footer rows.

        Note:
        The function iterates through the specified range of rows (from 'sheet.max_row' to 'end') in reverse order.
        It deletes rows to clean up the data, assuming that rows after the 'end' index are part of the footer.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through sheet max row to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def removeHeader(wb, start):
    """
        Removes header rows from an openpyxl Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index indicating the start of the table data.

        Returns:
        - openpyxl.Workbook: The updated Workbook object after removing header rows.

        Note:
        The function iterates through the specified range of rows (from 'start' to 1) in reverse order.
        It deletes rows to clean up the data, assuming that rows before the 'start' index are part of the header.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through
        sheet.delete_rows(x)  # delete row
    return wb


def removeNoneRows(wb, start, end, column):
    """
        Deletes rows in an openpyxl Workbook where the specified reference column cell is None.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data.
        - column (str): The column letter or number indicating the reference column.

        Returns:
        - openpyxl.Workbook: The updated Workbook object after removing rows with None values.

        Note:
        The function iterates through the specified range of rows (from 'end - 1' to 'start' in reverse order).
        It deletes rows where the cell value in the specified reference column is None.

    """
    sheet = wb.active
    for x in range(end - 1, start, -1):  # iterating through table data end row to table data start row
        if sheet[f"{column}{x}"].value is None:  # if cell value is none
            sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merges rows in an openpyxl Workbook based on the values in a specified reference column.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data.
        - refColumn (str): The column letter or number containing the reference values.
        - mergingColumn (str): The column letter or number to be merged based on the reference values.

        Returns:
        - openpyxl.Workbook: The updated Workbook object after merging rows based on the specified columns.

        Note:
        The function iterates through the specified range of rows (from 'start' to 'end').
        It merges rows in the specified merging column based on the values in the reference column.
        The last row in the range will be merged outside the loop to ensure it is not skipped.

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through table data start row to table data end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def kotak1_validation(wb):
    """
        Validates the column count for the core logic in a Kotak Mahindra Bank statement.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.

        Returns:
        - bool: True if the column count is not equal to the expected count, False otherwise.

        Note:
        The function checks the number of columns in the active sheet of the Workbook.
        If the column count is not equal to the expected count (12), it returns True, indicating an invalid format.
        Otherwise, it returns False, indicating a valid format.

    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 12  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def kotak1_main(wb):
    """
        Process and standardize a Kotak Mahindra Bank statement.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.

        Returns:
        - dict: A dictionary containing the processed Workbook and an optional message.
        Example: {"data": wb, "msg": None}

        Note:
        This function processes the provided Workbook to standardize the format of a Kotak Mahindra Bank statement.
        It performs various operations such as removing unwanted rows, merging misaligned rows, deleting specific rows,
        aligning columns manually, converting date formats, altering header names, and creating new columns.

        If the Workbook fails the validation (kotak1_validation), it prints an error message and returns a response
        dictionary indicating the validation failure.

    """
    if kotak1_validation(wb):  # validating column count for core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return response with error msg
    else:
        sheet = wb.active
        startText = "Chq/Ref No."  # header text to define table data start row
        endText = "Statement Summary"   # text define table data end row
        startEndRefColumn = "C"  # column containing start end text
        dupHeaderStartText = "Contd."  # starting row reference text to delete rows by range
        dupHeaderEndText = "Page"  # ending row reference text to delete rows by range
        dupHeaderRefColumn = "A"  # column contains starting row reference text to delete the rows by range
        columnToMerg1 = "B"  # column to merge misaligned row
        refColumnToMerg = "A"  # reference colum to merge row of other columns
        dateLength = 5  # date string length
        deleteRowRefText1 = "OPENING BALANCE"  # reference text to delete
        deleteRowRefColumn = "B"  # reference column to delete row with reference text
        refColumnToAlign = "G"  # reference column to align column string
        dateConversionColumn1 = "A"  # column to convert date to standard date formate
        refHeaderText1 = "Date"  # header text to replace with standardised column name
        refHeaderText2 = "Narration"  # header text to replace with standardised column name
        refHeaderText3 = "Chq/Ref No."  # header text to replace with standardised column name
        refHeaderText4 = "Withdrawal (Dr)"  # header text to replace with standardised column name
        refHeaderText5 = "Deposit (Cr)"  # header text to replace with standardised column name
        refHeaderText6 = "Balance"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Narration"  # standard column name
        headerText3 = "ChequeNo_RefNo"  # standard column name
        headerText4 = "Withdrawal"  # standard column name
        headerText5 = "Deposit"  # standard column name
        headerText6 = "Balance"  # standard column name
        refTextToRemove1 = "(Cr)"  # reference text to remove from column
        stringRemoveColumn1 = "F"  # column to remove reference string
        refStringToRemove = ","  # reference string to remove from column
        columnToRemoveString1 = "F"  # column to remove reference string
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify table data with in
        dupHeaderRemoved = Excel.delete_rows_by_range(wb, start, end, dupHeaderStartText, dupHeaderEndText, dupHeaderRefColumn)  # deleting unwanted rows by range
        start, end = Excel.get_start_end_row_index(dupHeaderRemoved, startText, endText, startEndRefColumn)  # get start and end row index to specify table data with in
        columnMergedB = mergingRows(dupHeaderRemoved, start, end, refColumnToMerg, columnToMerg1)  # merging rows of colum B
        noneRowsRemoved = removeNoneRows(columnMergedB, start, end, refColumnToMerg)  # deleting row if reference column cell is none
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)  # get start and end row index to specify table data with in
        footerRemoved = removeFooter(noneRowsRemoved, end - 1)  # end-1 to Include End Footer
        start, end = Excel.get_start_end_row_index(footerRemoved, startText, endText, startEndRefColumn)  # get start and end row index to specify table data with in
        nullRowDeleted = deleteRowByDateLen(footerRemoved, start, dateLength, refColumnToMerg)  # deleting rows by date length
        start, end = Excel.get_start_end_row_index(footerRemoved, startText, endText, startEndRefColumn)  # get start and end row index to specify table data with in
        headData = headerData(nullRowDeleted, start, end)  # deleting header rows
        headerRemoved = removeHeader(nullRowDeleted, start - 1)  # removing header rows, start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerRemoved, startText, endText, startEndRefColumn)  # get start and end row index to specify table data with in
        openbalDeleted = Excel.remove_row(headerRemoved, start, end, deleteRowRefText1, deleteRowRefColumn)  # remove a single row by checking the referance text is in the column cell
        alignColumns(openbalDeleted, start + 1, end, headData, refColumnToAlign)  # column data to align manually, start+1 to Skip Header
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 -> ASCII value
        slCreated = Excel.create_slno_column(openbalDeleted, start, end, chr(columnToCreateSlNo))  # creating new slno column
        start, end = Excel.get_start_end_row_index(slCreated, startText, endText, startEndRefColumn)  # get start and end row index to specify table data with in
        convertedDateA = dateConvertion(slCreated, start + 1, end + 1, dateConversionColumn1)  # converting date to standard date formate, start+1 to Skip Header, end+1 to Include Last Row
        lastCol = 65 + sheet.max_column  # 65 => ASCII value "A"
        transdate = Excel.alter_header_name(convertedDateA, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
        naration = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        chqno = Excel.alter_header_name(naration, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        debit = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        credit = Excel.alter_header_name(debit, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(credit, refHeaderText6, headerText6, lastCol)  # alter header name from the excel file to the standard column name
        stringRemovedCR = removeString(balance, start, end + 1, refTextToRemove1, stringRemoveColumn1)  # replacing reference string with empty string in column
        columnFinalised = Excel.finalise_column(stringRemovedCR, columns)  # standardizing count of column
        comaRemoved = Excel.remove_string(columnFinalised, start, end + 1, refStringToRemove, columnToRemoveString1)  # replacing reference string with empty string in column
        createdTransTypeColumn = Excel.transaction_type_column(comaRemoved)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/Kotak1._Apr-22_637102__06-09-2023-14-01-34.xlsx"
    # path = "C:/Users/Admin/Desktop/1._Kotak_-_7212_-_01-Apr-20_to_31-Mar-21__27-12-2023-11-20-55.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    result = kotak1_main(wb)
    # result.save("C:/Users/Admin/Desktop/FinalOutput/Kotak1output.xlsx")
