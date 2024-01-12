from datetime import datetime

import openpyxl

from CommonClass import Excel


def dateConvertion(wb, start, end, column):
    """
        Convert dates in a specified column to the standard date format.

        Parameters:
        - wb (openpyxl.workbook.Workbook): Excel workbook containing the data.
        - start (int): Start row index for the data.
        - end (int): End row index for the data (inclusive).
        - column (str): Column letter containing the date values.

        Returns:
        openpyxl.workbook.Workbook: Modified workbook with dates converted to the standard format.

        Overview:
        - Iterates through the rows in the specified range.
        - Converts each date in the specified column to the standard date format "%d/%b/%Y".

        Note:
        - Ensure that the 'openpyxl' library is installed.
        - 'start' and 'end' parameters define the range of rows to process.
        - 'column' parameter specifies the column letter containing the date values.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d/%b/%Y").date()  # converting to standard date formate
    return wb


def delete_header(wb, start):
    """
        Delete rows above the specified start index row in the workbook.

        Parameters:
        - wb (openpyxl.workbook.Workbook): Excel workbook containing the data.
        - start (int): Start row index from which rows above will be deleted.

        Returns:
        openpyxl.workbook.Workbook: Workbook with rows deleted above the specified start index.

        Overview:
        - Iterates through the rows from the specified start index to the 1st row.
        - Deletes each row in the specified range.

        Note:
        - Ensure that the 'openpyxl' library is installed.
        - 'start' parameter defines the start index from which rows above will be deleted.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to sheet 1st row
        sheet.delete_rows(x)  # delete row
    return wb


def delete_footer(wb, end):
    """
        Delete all rows below the specified end (last) row in the workbook.

        Parameters:
        - wb (openpyxl.workbook.Workbook): Excel workbook containing the data.
        - end (int): End row index, indicating the last row to retain in the workbook.

        Returns:
        openpyxl.workbook.Workbook: Workbook with rows deleted below the specified end (last) row.

        Overview:
        - Iterates through the rows from the sheet's max row to the specified end row.
        - Deletes each row in the specified range.

        Note:
        - Ensure that the 'openpyxl' library is installed.
        - 'end' parameter defines the end row index, indicating the last row to retain.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through sheet's max row to table data end row
        sheet.delete_rows(x)  # delete column
    return wb


def delete_nonerows(wb, start, end, column):
    """
        Remove unwanted rows where the reference column cell value is None.

        Parameters:
        - wb (openpyxl.workbook.Workbook): Excel workbook containing the data.
        - start (int): Start row index, indicating the beginning of the data range.
        - end (int): End row index, indicating the last row in the data range.
        - column (str): Reference column letter (e.g., 'A').

        Returns:
        openpyxl.workbook.Workbook: Workbook with unwanted rows removed.

        Overview:
        - Iterates through the rows from the end row to the start row.
        - Deletes rows where the reference column cell value is None.

        Note:
        - Ensure that the 'openpyxl' library is installed.
        - 'start' parameter defines the start row index, indicating the beginning of the data range.
        - 'end' parameter defines the end row index, indicating the last row in the data range.
        - 'column' parameter specifies the reference column letter.

    """
    sheet = wb.active
    for x in range(end, start, -1):  # iterating through table data end row to table data start row
        if sheet[f"{column}{x}"].value is None:  # if reference column cell value is none
            sheet.delete_rows(x)  # delete row
    return wb


def remove_none(wb, start, end, Column):
    """
        Remove occurrences of "None" string from a specific column.

        Parameters:
        - wb (openpyxl.workbook.Workbook): Excel workbook containing the data.
        - start (int): Start row index, indicating the beginning of the data range.
        - end (int): End row index, indicating the last row in the data range.
        - column (str): Column letter (e.g., 'A').

        Returns:
        openpyxl.workbook.Workbook: Workbook with "None" occurrences removed from the specified column.

        Overview:
        - Iterates through the rows from the start row to the end row.
        - Removes occurrences of the string "None" from the specified column.

        Note:
        - Ensure that the 'openpyxl' library is installed.
        - 'start' parameter defines the start row index, indicating the beginning of the data range.
        - 'end' parameter defines the end row index, indicating the last row in the data range.
        - 'column' parameter specifies the column letter.

    """
    sheet = wb.active
    for x in range(start, end):  # iterating through table data start row to table data end row
        if "None" in str(sheet[f"{Column}{x}"].value):  # if string "None" in cell value
            sheet[f"{Column}{x}"].value = str(sheet[f"{Column}{x}"].value).replace("None", "")  # replace it with empty string
    return wb


def merging_rows(wb, start, end, refColumn, mergingColumn):
    """
        Merge consecutive rows of a specified column based on the non-None values in a reference column.

        Parameters:
        - wb (openpyxl.workbook.Workbook): Excel workbook containing the data.
        - start (int): Start row index, indicating the beginning of the data range.
        - end (int): End row index, indicating the last row in the data range.
        - refColumn (str): Reference column letter (e.g., 'A') for identifying starting rows.
        - mergingColumn (str): Column letter (e.g., 'B') to merge consecutive rows.

        Returns:
        openpyxl.workbook.Workbook: Workbook with merged rows in the specified column.

        Overview:
        - Iterates through the rows from the start row to the end row.
        - Merges consecutive rows of the specified column based on the non-None values in the reference column.

        Note:
        - Ensure that the 'openpyxl' library is installed.
        - 'start' parameter defines the start row index, indicating the beginning of the data range.
        - 'end' parameter defines the end row index, indicating the last row in the data range.
        - 'refColumn' parameter specifies the reference column letter.
        - 'mergingColumn' parameter specifies the column letter to merge consecutive rows.

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through table data start row to table data end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def delete_pagenorow(wb, start, end, Column, refText):
    """
        Delete rows containing a specified reference text in a given column within a specified range.

        Parameters:
        - wb (openpyxl.workbook.Workbook): Excel workbook containing the data.
        - start (int): Start row index, indicating the beginning of the data range.
        - end (int): End row index, indicating the last row in the data range.
        - Column (str): Column letter (e.g., 'A') to search for the reference text.
        - refText (str): Reference text to identify rows for deletion.

        Returns:
        openpyxl.workbook.Workbook: Workbook with specified rows deleted.

        Overview:
        - Iterates through the rows from the end row to the start row.
        - Deletes rows containing the specified reference text in the specified column within the given range.

        Note:
        - Ensure that the 'openpyxl' library is installed.
        - 'start' parameter defines the start row index, indicating the beginning of the data range.
        - 'end' parameter defines the end row index, indicating the last row in the data range.
        - 'Column' parameter specifies the column letter to search for the reference text.
        - 'refText' parameter specifies the reference text to identify rows for deletion.

    """
    sheet = wb.active
    for x in range(end, start, -1):  # iterating through table data end row to table data start row
        if refText in str(sheet[f"{Column}{x}"].value):  # if reference text in cell value
            end -= 1
            sheet.delete_rows(x)  # delete row
    return wb


def icici1_validation(wb):
    """
        Validate the columns of an ICICI bank statement for the core logic.

        Parameters:
        - wb (openpyxl.workbook.Workbook): Excel workbook containing the data.

        Returns:
        bool: True if the column count is not equal to the designed core logic, False otherwise.

        Overview:
        - Checks if the number of columns in the active sheet matches the designed core logic.
        - The designed core logic specifies the expected number of columns in the ICICI bank statement.

        Note:
        - Ensure that the 'openpyxl' library is installed.
        - The function returns True if the column count does not match the designed core logic.
        - The function returns False if the column count matches the designed core logic.

    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 10  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def icici1_main(wb):
    """
        Process and standardize an ICICI1 bank statement in the provided Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.Workbook): Excel workbook containing the ICICI bank statement.

        Returns:
        dict: A dictionary containing the processed workbook and a message.

        Overview:
        - Validates the columns of the ICICI bank statement using the `icici1_validation` function.
        - If the validation fails, returns an error message.
        - Processes and standardizes the data by merging, aligning, deleting, and converting columns.
        - Returns a dictionary containing the processed workbook and a success message.

        Note:
        - Ensure that the 'openpyxl' library is installed.
        - The function returns a dictionary with keys 'data' (processed workbook) and 'msg' (message).

    """
    sheet = wb.active
    if icici1_validation(wb):  # validate columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return the response with error msg
    else:
        startText = "Sl"  # header text in column A
        endText = "Opening Bal:"  # text define the end of the table data row
        startEndDefColumn = "A"  # reference column contains the start and end text
        deletePageNoRowFromColumn = "A"  # reference column to delete page no rows
        referanceTextToDelete = "Page"  # reference text to delete page no rows
        refColumn = "A"  # reference column to merge other columns
        column1 = "B"  # column to merge
        column2 = "C"  # column to merge
        column3 = "D"  # column to merge
        column4 = "E"  # column to merge
        column5 = "G"  # column to merge
        refTextToDeleteColumn1 = "SlNo"  # reference column header text to delete column
        refTextToDeleteColumn2 = "TranId"  # reference column header text to delete column
        refTextToDeleteColumn3 = "TransactionPosted Date"  # reference column header text to delete column
        refHeaderText1 = "ValueDate"  # header text to replace with standardised column name
        refHeaderText2 = "TransactionDate"  # header text to replace with standardised column name
        refHeaderText3 = "Cheque no /Ref No"  # header text to replace with standardised column name
        refHeaderText4 = "TransactionRemarks"  # header text to replace with standardised column name
        refHeaderText5 = "Withdrawal (Dr)"  # header text to replace with standardised column name
        refHeaderText6 = "Deposit(Cr)"  # header text to replace with standardised column name
        refHeaderText7 = "Balance"  # header text to replace with standardised column name
        headerText1 = "Value_Date"  # standard column name
        headerText2 = "Transaction_Date"  # standard column name
        headerText3 = "ChequeNo_RefNo"  # standard column name
        headerText4 = "Narration"  # standard column name
        headerText5 = "Withdrawal"  # standard column name
        headerText6 = "Deposit"  # standard column name
        headerText7 = "Balance"  # standard column name
        stringAlignColumn4 = "A"  # column to align string -> removing "\n"(next line) from string
        stringAlignColumn5 = "B"  # column to align string -> removing "\n"(next line) from string
        dateStringAlignColumn1 = "C"  # column to align string -> removing "\n"(next line) from string
        dateStringAlignColumn2 = "D"  # column to align string -> removing "\n"(next line) from string
        stringAlignColumn6 = "E"  # column to align string -> removing "\n"(next line) from string
        stringAlignColumn7 = "F"  # column to align string -> removing "\n"(next line) from string
        stringAlignColumn3 = "G"  # column to align string -> removing "\n"(next line) from string
        stringAlignColumn8 = "H"  # column to align string -> removing "\n"(next line) from string
        stringAlignColumn9 = "I"  # column to align string -> removing "\n"(next line) from string
        stringAlignColumn10 = "J"  # column to align string -> removing "\n"(next line) from string
        dateConversionColumn1 = "A"  # column to convert date to standard date formate
        dateConversionColumn2 = "B"  # column to convert date to standard date formate
        noneRemoveColumn1 = "A"  # column to remove "None" string -> replace string "None" with empty string
        noneRemoveColumn2 = "B"  # column to remove "None" string -> replace string "None" with empty string
        noneRemoveColumn3 = "C"  # column to remove "None" string -> replace string "None" with empty string
        noneRemoveColumn4 = "D"  # column to remove "None" string -> replace string "None" with empty string
        noneRemoveColumn5 = "E"  # column to remove "None" string -> replace string "None" with empty string
        noneRemoveColumn6 = "F"  # column to remove "None" string -> replace string "None" with empty string
        noneRemoveColumn7 = "G"  # column to remove "None" string -> replace string "None" with empty string
        negativeValueColumnRefText1 = "Withdrawal"  # no need of converting negative value to positive
        headerTextToMakeEmptyCellsToNone1 = "Withdrawal"  # header text to make empty cells to none
        headerTextToMakeEmptyCellsToNone2 = "Deposit"  # header text to make empty cells to none
        headerTextToMakeEmptyCellsToNone3 = "ChequeNo_RefNo"  # header text to make empty cells to none
        columnToRemoveString1 = "E"  # column to remove desired string
        columnToRemoveString2 = "F"  # column to remove desired string
        columnToRemoveString3 = "G"  # column to remove desired string
        stringToRemove1 = ","  # reference string to remove from columns
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify the data with in
        pageNoRowDeletedwb = delete_pagenorow(wb, start, end, deletePageNoRowFromColumn, referanceTextToDelete)  # deleting rows containing  page no in it
        start, end = Excel.get_start_end_row_index(pageNoRowDeletedwb, startText, endText, startEndDefColumn)   # get start and end row index to specify the data with in
        mergcolumnB = merging_rows(pageNoRowDeletedwb, start, end, refColumn, column1)  # merging the rows of desired column
        mergColumnC = merging_rows(mergcolumnB, start, end, refColumn, column2)  # merging the rows of desired column
        mergColumnD = merging_rows(mergColumnC, start, end, refColumn, column3)  # merging the rows of desired column
        mergColumnE = merging_rows(mergColumnD, start, end, refColumn, column4)  # merging the rows of desired column
        mergColumnG = merging_rows(mergColumnD, start, end, refColumn, column5)  # merging the rows of desired column
        removeNoneB = remove_none(mergColumnG, start, end, column1)  # removing "None" string from columns
        removeNoneC = remove_none(removeNoneB, start, end, column2)  # removing "None" string from columns
        removeNoneD = remove_none(removeNoneC, start, end, column3)  # removing "None" string from columns
        removeNoneE = remove_none(removeNoneD, start, end, column4)  # removing "None" string from columns
        deletedNoneRows = delete_nonerows(removeNoneE, start, end, refColumn)  # removing the unwanted rows, when the reference column cell value is none
        start, end = Excel.get_start_end_row_index(deletedNoneRows, startText, endText, startEndDefColumn)   # get start and end row index to specify the data with in
        footerDeleted = delete_footer(deletedNoneRows, end - 1)  # delete all the rows below the end(last) row, end-1 to Include End Footer
        headerDeleted = delete_header(footerDeleted, start - 1)  # delete rows above the start index row, start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndDefColumn)   # get start and end row index to specify the data with in
        stringAlignedA = Excel.string_align(headerDeleted, start, end + 1, stringAlignColumn4)  # end+1 to Include Last Row
        stringAlignedB = Excel.string_align(headerDeleted, start, end + 1, stringAlignColumn5)  # end+1 to Include Last Row
        stringAlignedC = Excel.string_align(stringAlignedB, start, end + 1, dateStringAlignColumn1)  # end+1 to Include Last Row
        stringAlignedD = Excel.string_align(stringAlignedC, start, end + 1, dateStringAlignColumn2)  # end+1 to Include Last Row
        stringAlignedE = Excel.string_align(stringAlignedD, start, end + 1, stringAlignColumn6)  # end+1 to Include Last Row
        stringAlignedF = Excel.string_align(stringAlignedE, start, end + 1, stringAlignColumn7)  # end+1 to Include Last Row
        stringAlignedG = Excel.string_align(stringAlignedF, start, end + 1, stringAlignColumn3)  # end+1 to Include Last Row
        stringAlignedH = Excel.string_align(stringAlignedG, start, end + 1, stringAlignColumn8)  # end+1 to Include Last Row
        stringAlignedI = Excel.string_align(stringAlignedH, start, end + 1, stringAlignColumn9)  # end+1 to Include Last Row
        stringAlignedJ = Excel.string_align(stringAlignedI, start, end + 1, stringAlignColumn10)  # end+1 to Include Last Row
        deletedColumn1 = Excel.delete_column(stringAlignedD, refTextToDeleteColumn1)  # deleting an existing column
        deletedColumn2 = Excel.delete_column(deletedColumn1, refTextToDeleteColumn2)  # deleting an existing column
        deletedColumn3 = Excel.delete_column(deletedColumn2, refTextToDeleteColumn3)  # deleting an existing column
        lastCol = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        valueDate = Excel.alter_header_name(deletedColumn2, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
        transDate = Excel.alter_header_name(valueDate, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        chequeNo = Excel.alter_header_name(transDate, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        naration = Excel.alter_header_name(chequeNo, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        debit = Excel.alter_header_name(naration, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)  # alter header name from the excel file to the standard column name
        dateConvertedA = dateConvertion(balance, start + 1, end + 1, dateConversionColumn1)  # converting date in a column to standard date formate, start+1 to Skip Header, end-1 to Include Last Row
        dateConvertedB = dateConvertion(dateConvertedA, start + 1, end + 1, dateConversionColumn2)  # converting date in a column to standard date formate, start+1 to Skip Header, end-1 to Include Last Row
        noneRemovedA = remove_none(dateConvertedB, start, end + 1, noneRemoveColumn1)  # removing "None" string from columns, end+1 to Include Last Row
        noneRemovedB = remove_none(noneRemovedA, start, end + 1, noneRemoveColumn2)  # removing "None" string from columns, end+1 to Include Last Row
        noneRemovedC = remove_none(noneRemovedB, start, end + 1, noneRemoveColumn3)  # removing "None" string from columns, end+1 to Include Last Row
        noneRemovedD = remove_none(noneRemovedC, start, end + 1, noneRemoveColumn4)  # removing "None" string from columns, end+1 to Include Last Row
        noneRemovedE = remove_none(noneRemovedD, start, end + 1, noneRemoveColumn5)  # removing "None" string from columns, end+1 to Include Last Row
        noneRemovedF = remove_none(noneRemovedE, start, end + 1, noneRemoveColumn6)  # removing "None" string from columns, end+1 to Include Last Row
        noneRemovedG = remove_none(noneRemovedF, start, end + 1, noneRemoveColumn7)  # removing "None" string from columns, end+1 to Include Last RowcolumnToCreateSlNo = 65 + column_count(wb)
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slCreated = Excel.create_slno_column(noneRemovedG, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
        negativeValueChecked = Excel.check_neagativeValue_by_column(slCreated, negativeValueColumnRefText1)  # no need to convert negative value to positive
        withdrawalNoneReplaced = Excel.empty_cell_to_none(negativeValueChecked, start, end + 1, headerTextToMakeEmptyCellsToNone1)  # making empty cells in a column to none by using the header text as reference
        depositNoneReplaced = Excel.empty_cell_to_none(withdrawalNoneReplaced, start, end + 1, headerTextToMakeEmptyCellsToNone2)  # making empty cells in a column to none by using the header text as reference
        chqnoNoneReplaced = Excel.empty_cell_to_none(depositNoneReplaced, start, end + 1, headerTextToMakeEmptyCellsToNone3)  # making empty cells in a column to none by using the header text as reference
        stringRemovedE = Excel.remove_string(chqnoNoneReplaced, start, end + 1, stringToRemove1, columnToRemoveString1)  # removing particular column from desired string
        stringRemovedF = Excel.remove_string(stringRemovedE, start, end + 1, stringToRemove1, columnToRemoveString2)  # removing particular column from desired string
        stringRemovedG = Excel.remove_string(stringRemovedF, start, end + 1, stringToRemove1, columnToRemoveString3)  # removing particular column from desired string
        createdTransTypeColumn = Excel.transaction_type_column(stringRemovedG)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/ICICI_-_3281__05-09-2023-15-28-06.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    result = icici1_main(wb)
    # result.save('C:/Users/Admin/Desktop/FinalOutput/ICICI1output.xlsx')
