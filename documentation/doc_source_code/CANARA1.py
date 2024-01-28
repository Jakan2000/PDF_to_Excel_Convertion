from datetime import datetime

import openpyxl

from CommonClass import Excel


def dateConvertion(wb, start, end, column):
    """
        Convert date values in a specified column to a standard date format in the given Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index for date conversion (inclusive).
        - end (int): The ending row index for date conversion (exclusive).
        - column (str): The column letter (e.g., 'A', 'B') containing date values to be converted.

        Returns:
        openpyxl.Workbook: The modified Workbook object after date conversion.

         Notes:
        - This function assumes the date values in the specified column are in the format '%d-%b-%y'.
        - The converted dates are stored in the same column in the standard date format.
        - The 'start' parameter denotes the first row to start converting dates.
        - The 'end' parameter denotes the row where date conversion stops (exclusive).
        - The original Workbook is modified in place.

    """
    sheet = wb.active
    for i in range(start, end):  # iterate through start and end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%b-%y").date()  # convert to standard date formate
    return wb


def delete_header(wb, start):
    """
        Delete rows above the specified starting row index, effectively removing the header from the Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index (inclusive) to begin deletion of rows above.

        Returns:
        openpyxl.Workbook: The modified Workbook object after removing rows above the specified starting index.

        Notes:
        - This function removes rows from the top of the sheet up to the specified starting index.
        - The 'start' parameter denotes the first row that should be retained; rows above this index are deleted.
        - The original Workbook is modified in place.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to sheet's 1st row
        sheet.delete_rows(x)  # delete row
    return wb


def alignColumn(wb, start, end, fromColumn, toColumn):
    """
        Align the data in the specified column by copying values from another column.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index (inclusive) for alignment.
        - end (int): The ending row index (exclusive) for alignment.
        - fromColumn (str): The source column letter to copy values from.
        - toColumn (str): The destination column letter to align by copying values.

        Returns:
        The modified Workbook object after aligning the specified column.

         Notes:
        - This function aligns the data in 'toColumn' by copying values from 'fromColumn'.
        - Rows within the specified range [start, end) are processed.
        - If the destination cell in 'toColumn' is None, the value from the corresponding cell in 'fromColumn' is copied.
        - The original value in 'fromColumn' is set to None after copying.
        - The function modifies the provided Workbook object in place.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through start and end rows
        h_cell = f"{toColumn}{i}"  # get cell address
        if sheet[h_cell].value is None:  # if H column cell is None
            sheet[h_cell].value = sheet[f"I{i}"].value  # assign I column cell value to the H column cell
            sheet[f"{fromColumn}{i}"].value = None  # then make I column cell value to None
    return wb


def deleteNoneRows(wb, start, end, refColumn):
    """
        Delete rows where the specified date column cell is None.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index (inclusive) for row deletion.
        - end (int): The ending row index (exclusive) for row deletion.
        - refColumn (str): The letter of the date column for checking None values.

        Returns:
        openpyxl.Workbook: The modified Workbook object after deleting rows with None values in the specified column.

        Notes:
        - This function deletes rows within the specified range where the date column cell is None.
        - Rows with a None value in the 'refColumn' are removed.
        - The function modifies the provided Workbook object in place.

    """
    sheet = wb.active
    for x in range(end, start, -1):  # iterate through end to start row
        a_cell = f"A{x}"  # get date column cell address
        if sheet[a_cell].value is None:  # if date column cell is None
            end -= 1
            sheet.delete_rows(x)  # delete the row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merge consecutive rows in the specified column based on a reference column.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index (inclusive) for merging.
        - end (int): The ending row index (exclusive) for merging.
        - refColumn (str): The letter of the reference column to determine row merging.
        - mergingColumn (str): The letter of the column whose rows need to be merged.

         Notes:
        - This function is designed to merge rows in the specified 'mergingColumn' based on the non-empty values in the 'refColumn'.
        - Consecutive rows with non-empty values in the 'refColumn' are merged into a single row in the 'mergingColumn'.
        - The last row in the specified range (end - 1) is also considered for merging.
        - The function modifies the provided Workbook object in place.

        Returns:
        openpyxl.Workbook: The modified Workbook object after merging consecutive rows in the specified column.

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through start and end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def deleteRowByLength(wb, start, end, refColumn, pgNoLength):
    """
        Delete rows in the specified range based on the string length in the reference column.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index (inclusive) for deletion.
        - end (int): The ending row index (exclusive) for deletion.
        - refColumn (str): The letter of the reference column containing string lengths for deletion.
        - pgNoLength (int): The maximum length of the string in the reference column to retain a row.

        Returns:
        openpyxl.Workbook: The modified Workbook object after deleting rows based on string length.

        Notes:
        - This function deletes rows within the specified range where the string length in the 'refColumn' is less than 'pgNoLength'.
        - Rows with non-empty values in the 'refColumn' and a string length less than 'pgNoLength' are removed.
        - The function modifies the provided Workbook object in place.

    """
    sheet = wb.active
    for x in range(end, start, -1):  # iterating through end and start row
        if sheet[f"{refColumn}{x}"].value is not None and (len(str(sheet[f"{refColumn}{x}"].value))) < 5:  # if reference column cell value is not none and reference column cell value length is < 5
            sheet.delete_rows(x)  # delete row
    return wb


def deleteFooter(wb, end):
    """
        Delete all rows below the specified ending row index, effectively removing the footer from the Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - end (int): The ending row index (inclusive) to retain in the Workbook.

        Returns:
        openpyxl.Workbook: The modified Workbook object after removing rows below the specified ending index.

        Notes:
        - This function deletes all rows below the provided 'end' index, essentially removing footer information.
        - Ensure that 'end' corresponds to the last row of your desired data to avoid unintentional data loss.
        - The modified Workbook is returned, and the original Workbook is updated in-place.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):   # iterating from data ending row to the last row
        sheet.delete_rows(x)  # delete row
    return wb


def mergHeaderText(wb, start, column):
    """
        Merge header text in the specified column from the next row and update the header.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The row index where the header text merging starts.
        - column (str): The column letter (e.g., 'A', 'B') containing the header text.

        Returns:
        openpyxl.Workbook: The modified Workbook object after merging the header text.

        Notes:
        - This function merges the header text in the specified column from the next row and updates the header.
        - The header text from the next row is concatenated with the current header text.
        - The original Workbook is updated in-place, and the modified Workbook is returned.

    """
    sheet = wb.active
    txt = sheet[f"{column}{start}"].value + sheet[f"{column}{start + 1}"].value  # merging header text from next row
    sheet[f"{column}{start}"].value = txt  # assigning conceited string to the main header
    return wb


def canara1_validation(wb):
    """
        Validate columns for the core logic implemented in the Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.

        Returns:
        bool: True if the number of columns in the Workbook is not equal to the expected count; False otherwise.

        Notes:
        - This function checks if the number of columns in the Workbook matches the expected count for the core logic.
        - If the count of columns is not equal to the expected count, the function returns True, indicating validation failure.
        - If the count of columns matches the expected count, the function returns False, indicating successful validation.

    """
    sheet = wb.active
    max_column = sheet.max_column   # get max column in the sheet, using predefined function
    countOfColumn = 9   # need to change the logic for 8 column -> previous logic 9 columns, the column count of our core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def canara1_main(wb):
    """
        Perform data processing and formatting for a specific logic (canara1) on the given Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.

        Returns:
        dict: A dictionary containing processed Workbook data and a message.
        - 'data' (openpyxl.Workbook): Processed Workbook data.
        - 'msg' (str): A message describing the operation performed.

        Notes:
        - This function applies a series of data processing steps and formatting to the Workbook based on the canara1 logic.
        - The processing steps include removing unwanted rows and headers, merging columns, aligning data, converting date formats, deleting columns, and more.
        - The function returns a dictionary with the processed Workbook data and a message indicating the success of the operation.

    """
    sheet = wb.active  # get active sheet
    if canara1_validation(wb):  # validate columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response
    else:
        startText = "TRANS"  # header text in column A
        endText = "UNLESS THE CONSTITUENT BRINGS TO THE NOTICE OF THE BANK"  # text define the end of the table data
        startEndRefColumn = "A"  # column containing the text to define start and end -> data within
        mergHeaderColumn1 = "A"  # column to merge the row data
        mergHeaderColumn2 = "B"  # column to merge the row data
        duplicateHeaderTxt1 = "TRANS"  # reference header text to remove duplicate header
        duplicateHeaderTxt2 = "DATE"  # reference header text to remove duplicate header
        duplicateHeaderRefColumn = "A"  # reference column to remove duplicate header text
        startText2 = "TRANS DATE"  # header text in column A
        endText2 = "Statement Summary :"  # text define the end of the table data
        pgNoLength = 5  # to remove the page number by using the length
        refColumnToDeletePgNoRow = "A"  # reference column to delete the page number row
        refColumnToMerg = "A"  # reference column (date column) to merge the rows of other column
        columnToMerg1 = "E"  # column to merge the scattered rows
        refColumnToDeleteNoneRows = "A"  # reference column to merge delete none rows
        fromColumn = "I"  # aligning column data -> data should take from column
        toColumn = "H"  # aligning column data -> data should assign to column
        stringAlignColumn = "E"  # column to align string -> remove "\n" from the string
        refTextToDeleteRow = "B/F"  # reference text to delete the row
        refColumnToDeleteRow = "E"  # reference column to delete the unwanted rows
        dateConversionColumn1 = "A"  # column to convert date to the standard date formate
        dateConversionColumn2 = "B"  # column to convert date to the standard date formate
        deleteColumnRefText1 = "BRANCH"  # reference text to delete a column
        refHeaderText1 = "TRANS DATE"  # header text to replace with standardised column name
        refHeaderText2 = "VALUE DATE"  # header text to replace with standardised column name
        refHeaderText3 = "REF/CHQ.NO"  # header text to replace with standardised column name
        refHeaderText4 = "DESCRIPTION"  # header text to replace with standardised column name
        refHeaderText5 = "WITHDRAWS"  # header text to replace with standardised column name
        refHeaderText6 = "DEPOSIT"  # header text to replace with standardised column name
        refHeaderText7 = "BALANCE"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Value_Date"  # standard column name
        headerText3 = "ChequeNo_RefNo"  # standard column name
        headerText4 = "Narration"  # standard column name
        headerText5 = "Withdrawal"  # standard column name
        headerText6 = "Deposit"  # standard column name
        headerText7 = "Balance"  # standard column name
        negativeValueColumnRefText1 = "Withdrawal"   # no need to convert the negative value to positive
        columns = ["Sl.No.", "Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        AheaderTextMerged = mergHeaderText(wb, start, mergHeaderColumn1)  # merging splited header text
        BheaderTextMerged = mergHeaderText(AheaderTextMerged, start, mergHeaderColumn2)  # merging splited header text
        duplicateHeaderRemoved1 = Excel.remove_rows(BheaderTextMerged, start, end, duplicateHeaderTxt1, duplicateHeaderRefColumn)  # removing duplicate header row
        start, end = Excel.get_start_end_row_index(duplicateHeaderRemoved1, startText, endText, startEndRefColumn)  # get start and end row index to specify the data with in
        duplicateHeaderRemoved2 = Excel.remove_rows(duplicateHeaderRemoved1, start, end, duplicateHeaderTxt2, duplicateHeaderRefColumn)  # removing duplicate header row
        start, end = Excel.get_start_end_row_index(duplicateHeaderRemoved2, startText, endText, startEndRefColumn)  # get start and end row index to specify the data with in
        footerDeleted = deleteFooter(duplicateHeaderRemoved2, end)   # delete all rows below the end(last) row
        start, end = Excel.get_start_end_row_index(footerDeleted, startText2, endText2, startEndRefColumn)  # get start and end row index to specify the data with in
        pgNoRowDeleted = deleteRowByLength(footerDeleted, start, end, refColumnToDeletePgNoRow, pgNoLength)   # deleting rows by the string length in reference column
        start, end = Excel.get_start_end_row_index(pgNoRowDeleted, startText2, endText2, startEndRefColumn)  # get start and end row index to specify the data with in
        mergColumnE = mergingRows(pgNoRowDeleted, start, end, refColumnToMerg, columnToMerg1)  # merging rows of desired column
        noneRowsDeleted = deleteNoneRows(mergColumnE, start, end, refColumnToDeleteNoneRows)  # delete row's if date column row is none
        start, end = Excel.get_start_end_row_index(noneRowsDeleted, startText2, endText2, startEndRefColumn)  # get start and end row index to specify the data with in
        alignedColumnH = alignColumn(noneRowsDeleted, start, end, fromColumn, toColumn)  # aligning the misaligned column data
        start, end = Excel.get_start_end_row_index(alignedColumnH, startText2, endText2, startEndRefColumn)  # get start and end row index to specify the data with in
        footerDeleted = deleteFooter(alignedColumnH, end - 1)  # end-1 to Include Last Footer Row, delete all the rows below the end(last) row
        headerDeleted = delete_header(footerDeleted, start - 1)  # start-1 to Skip Header, delete rows above the start index row
        start, end = Excel.get_start_end_row_index(alignedColumnH, startText2, endText2, startEndRefColumn)  # get start and end row index to specify the data with in
        alignedE = Excel.string_align(headerDeleted, start, end, stringAlignColumn)  # aligning string in column by removing the \n from the string -> \n -> next line
        removedOpeningBalance = Excel.remove_row(alignedE, start, end, refTextToDeleteRow, refColumnToDeleteRow)  # remove a single row by checking the referance text is in the column cell
        start, end = Excel.get_start_end_row_index(alignedColumnH, startText2, endText2, startEndRefColumn)  # get start and end row index to specify the data with in
        dateConvertedA = dateConvertion(removedOpeningBalance, start + 1, end + 1, dateConversionColumn1)  # start+1 to Skip Header, end+1 to Include Last Row, converting date to standard date formate
        dateConvertedB = dateConvertion(dateConvertedA, start + 1, end + 1, dateConversionColumn2)  # start+1 to Skip Header, end+1 to Include Last Row, converting date to standard date formate
        dalatedColumnBRANCH = Excel.delete_column(dateConvertedB, deleteColumnRefText1)  # deleting desired column
        lastCol = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        transdate = Excel.alter_header_name(dalatedColumnBRANCH, refHeaderText1, headerText1, lastCol)  # alter header name by standard column name
        valuedate = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name by standard column name
        chqno = Excel.alter_header_name(valuedate, refHeaderText3, headerText3, lastCol)  # alter header name by standard column name
        naration = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)  # alter header name by standard column name
        debit = Excel.alter_header_name(naration, refHeaderText5, headerText5, lastCol)  # alter header name by standard column name
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)  # alter header name by standard column name
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)  # alter header name by standard column name
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slnoCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))  # creating new column - slno
        negativeColumnChecked = Excel.check_neagativeValue_by_column(slnoCreated, negativeValueColumnRefText1)  # no need to convert the negative value to positive
        columnFinalised = Excel.finalise_column(wb, columns)  # standardizing count of column
        createdTransTypeColumn = Excel.transaction_type_column(negativeColumnChecked)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/1.Canara_-_6183__11-09-2023-18-32-43.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    result = canara1_main(wb)
    # result.save('C:/Users/Admin/Desktop/FinalOutput/CANARA1output.xlsx')
