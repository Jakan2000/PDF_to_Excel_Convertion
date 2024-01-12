from datetime import datetime

import openpyxl

from CommonClass import Excel


def makeNone(wb, start, end, column):
    """
        Replace empty cells in a specific column with 'None' in a given Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook in which empty cells will be replaced with 'None'.
        - start (int): The starting row index for the table data in the specified column.
        - end (int): The ending row index (inclusive) for the table data in the specified column.
        - column (str): The column letter (e.g., 'A', 'B', 'C') in which empty cells will be replaced with 'None'.

        Returns:
        - openpyxl.Workbook: The modified Excel workbook with empty cells in the specified column replaced by 'None'.

        Note:
        This function iterates through the specified range of rows in the given column and replaces the values of
        empty cells with 'None'. It is particularly useful when standardizing and cleaning data in Excel workbooks.

    """
    sheet = wb.active
    for x in range(start, end):  # iterating through table data start row and table data end row
        if len(sheet[f"{column}{x}"].value) < 1:  # if length of cell value < 1
            sheet[f"{column}{x}"].value = None  # assign cell value to None
    return wb


def removeNone(wb, start, end, column):
    """
        Remove the string "None" from cells in a specific column within a given range in an Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook from which the string "None" will be removed.
        - start (int): The starting row index for the table data in the specified column.
        - end (int): The ending row index (exclusive) for the table data in the specified column.
        - column (str): The column letter (e.g., 'A', 'B', 'C') from which the string "None" will be removed.

        Returns:
        - openpyxl.Workbook: The modified Excel workbook with the string "None" removed from the specified column.

        Note:
        This function iterates through the specified range of rows in the given column and removes occurrences of
        the string "None" from non-empty cells. It is useful for cleaning and standardizing data in Excel workbooks.

    """
    sheet = wb.active
    for x in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):  # if cell value is not None and string "None" in cell value
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")  # replace string "None" with empty string
    return wb


def dateConversion(wb, start, end, column):
    """
        Convert date values in a specific column within a given range to a standard date format in an Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the date values to be converted.
        - start (int): The starting row index for the table data in the specified column.
        - end (int): The ending row index (exclusive) for the table data in the specified column.
        - column (str): The column letter (e.g., 'A', 'B', 'C') containing the date values to be converted.

        Returns:
        - openpyxl.Workbook: The modified Excel workbook with date values in the specified column converted to a standard format.

        Note:
        This function iterates through the specified range of rows in the given column and converts date values to
        a standard date format ("%d %b %Y"). It is useful for standardizing date formats in Excel workbooks.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d %b %Y").date()  # converting date to standard date formate
    return wb


def deleteHeader(wb, start):
    """
       Delete header rows from an Excel workbook, up to the specified starting row index.

       Parameters:
       - wb (openpyxl.Workbook): The Excel workbook containing header rows to be deleted.
       - start (int): The starting row index (inclusive) up to which header rows will be deleted.

       Returns:
       - openpyxl.Workbook: The modified Excel workbook with header rows deleted.

       Note:
       This function iterates through the specified range of rows, starting from the bottom,
       and deletes rows up to the specified starting row index. It is useful for removing header
       information from Excel workbooks.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to 1st row of sheet
        sheet.delete_rows(x)  # delete rows
    return wb


def deleteFooter(wb, end):
    """
       Delete footer rows from an Excel workbook, starting from the end and up to the specified ending row index.

       Parameters:
       - wb (openpyxl.Workbook): The Excel workbook containing footer rows to be deleted.
       - end (int): The ending row index (inclusive) up to which footer rows will be deleted.

       Returns:
       - openpyxl.Workbook: The modified Excel workbook with footer rows deleted.

       Note:
       This function iterates through the specified range of rows, starting from the bottom
       of the Excel sheet, and deletes rows up to the specified ending row index. It is useful
       for removing footer information from Excel workbooks.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through sheet max row to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def removeNoneRows(wb, start, end, column):
    """
        Remove rows from an Excel workbook where the value in the specified reference column is None.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the data to be processed.
        - start (int): The starting row index from which to begin checking for None values.
        - end (int): The ending row index up to which to check for None values.
        - column (str): The reference column letter (e.g., 'A', 'B') containing the values to be checked.

        Returns:
        - openpyxl.Workbook: The modified Excel workbook with rows removed where the specified column value is None.

        Note:
        This function iterates through the specified range of rows, checking the value in the specified reference column.
        Rows with a None value in the reference column are removed from the workbook. It is useful for cleaning up data
        by removing rows with missing or undefined values in a particular column.

    """
    sheet = wb.active
    for x in range(end - 1, start, -1):  # iterating through table data end row to table data start row
        if sheet[f"{column}{x}"].value is None:  # if cell value is None
            sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merge consecutive rows in an Excel workbook based on a reference column value.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook containing the data to be processed.
        - start (int): The starting row index from which to begin merging rows.
        - end (int): The ending row index up to which to merge rows.
        - refColumn (str): The reference column letter (e.g., 'A', 'B') used for identifying consecutive rows.
        - mergingColumn (str): The column letter (e.g., 'B', 'C') whose values will be merged for consecutive rows.

        Returns:
        - openpyxl.Workbook: The modified Excel workbook with consecutive rows merged in the specified column.

        Note:
        This function iterates through the specified range of rows, merging consecutive rows in the specified column
        based on the value in the reference column. The merged data is updated in the workbook, and the modified workbook
        is then returned.

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through start and end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def yes1_validation(wb):
    """
        Validate the number of columns in the given Excel workbook for a specific core logic.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook to be validated.

        Returns:
        - bool: True if the workbook does not have the expected number of columns, False otherwise.

        Note:
        This function checks whether the provided workbook has the expected number of columns required for a specific
        core logic. The expected column count is defined by the variable 'countOfColumn'. If the workbook has a different
        number of columns, the function returns True, indicating a validation failure. Otherwise, it returns False,
        indicating that the workbook is valid for the specified core logic.

    """
    sheet = wb.active
    max_column = sheet.max_column   # get max column in the sheet, using predefined function
    countOfColumn = 7  # the column count of our core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


#todo chequeno_refno column is Encoded eg. [1.387420221031e+24]
def yes1_main(wb):
    """
        Perform a series of data processing tasks on the provided Excel workbook based on a specific core logic.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook to be processed.

        Returns:
        - dict: A dictionary containing the processed workbook and a message.

        Note:
        This function executes a sequence of operations on the provided workbook to process and standardize the data.
        If the workbook does not meet the validation criteria defined by the `yes1_validation` function, an error message
        is returned. Otherwise, the function performs tasks such as aligning strings, removing "None" values, merging
        misaligned rows, converting date formats, and standardizing column names. The resulting workbook is then returned
        along with a success message.

    """
    sheet = wb.active
    if yes1_validation(wb):  # validating columns for the core logic written
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # returning response with error msg
    else:
        startText = "TransactionDate"  # header text in column A
        endText = "Opening Balance"  # text define the end of the table data
        startEndRefColumn = "A"  # column containing the text to define start and end -> data within
        deleteFlagStartText = "Customer Id"  # starting row reference text to delete the rows by range
        deleteFlagStopText = "TransactionDate"  # ending row reference text to delete the rows by range
        deleteFlagRefColumn = "A"  # column containing starting row reference text and ending row reference text to delete the rows by range
        columnToMerg = "D"  # column to merge misaligned rows
        refColumnToMerg = "A"  # reference column to merge misaligned rows of other columns
        refColumnToRemoveNoneRows = "A"  # reference column to remove none rows
        dateConversionColumn1 = "A"  # column to convert date to standard date formate
        dateConversionColumn2 = "B"  # column to convert date to standard date formate
        stringAlignColumn1 = "A"  # column to align string by removing the \n from string
        stringAlignColumn2 = "B"  # column to align string by removing the \n from string
        stringAlignColumn3 = "C"  # column to align string by removing the \n from string
        stringAlignColumn4 = "D"  # column to align string by removing the \n from string
        stringAlignColumn5 = "E"  # column to align string by removing the \n from string
        stringAlignColumn6 = "F"  # column to align string by removing the \n from string
        stringAlignColumn7 = "G"  # column to align string by removing the \n from string
        makeNoneRefColumn1 = "A"  # column to make empty cells to none
        refHeaderText1 = "TransactionDate"  # header text to replace with standardised column name
        refHeaderText2 = "Value Date"  # header text to replace with standardised column name
        refHeaderText3 = "Cheque No/Reference No"  # header text to replace with standardised column name
        refHeaderText4 = "Description"  # header text to replace with standardised column name
        refHeaderText5 = "Withdrawals"  # header text to replace with standardised column name
        refHeaderText6 = "Deposits"  # header text to replace with standardised column name
        refHeaderText7 = "Running Balance"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Value_Date"  # standard column name
        headerText3 = "ChequeNo_RefNo"  # standard column name
        headerText4 = "Narration"  # standard column name
        headerText5 = "Withdrawal"  # standard column name
        headerText6 = "Deposit"  # standard column name
        headerText7 = "Balance"  # standard column name
        headerTextToReplaceToNone1 = "Withdrawal"  # reference header text to replace empty cells to None
        headerTextToReplaceToNone2 = "Deposit"  # reference header text to replace empty cells to None
        start = 1  # table data start row
        end = sheet.max_row  # table data end row
        alignedStringA = Excel.string_align(wb, start, end, stringAlignColumn1)  # aligning string by removing \n from column string
        alignedStringB = Excel.string_align(alignedStringA, start, end, stringAlignColumn2)  # aligning string by removing \n from column string
        alignedStringC = Excel.string_align(alignedStringB, start, end, stringAlignColumn3)  # aligning string by removing \n from column string
        alignedStringD = Excel.string_align(alignedStringC, start, end, stringAlignColumn4)  # aligning string by removing \n from column string
        alignedStringE = Excel.string_align(alignedStringD, start, end, stringAlignColumn5)  # aligning string by removing \n from column string
        alignedStringF = Excel.string_align(alignedStringE, start, end, stringAlignColumn6)  # aligning string by removing \n from column string
        alignedStringG = Excel.string_align(alignedStringF, start, end, stringAlignColumn7)  # aligning string by removing \n from column string
        noneRemovedA = removeNone(alignedStringG, start, end, stringAlignColumn1)  # removing string "None" from column A
        noneRemovedB = removeNone(noneRemovedA, start, end, stringAlignColumn2)  # removing string "None" from column B
        noneRemovedC = removeNone(noneRemovedB, start, end, stringAlignColumn3)  # removing string "None" from column C
        noneRemovedD = removeNone(noneRemovedC, start, end, stringAlignColumn4)  # removing string "None" from column D
        noneRemovedE = removeNone(noneRemovedD, start, end, stringAlignColumn5)  # removing string "None" from column E
        noneRemovedF = removeNone(noneRemovedE, start, end, stringAlignColumn6)  # removing string "None" from column F
        noneRemovedG = removeNone(noneRemovedF, start, end, stringAlignColumn7)  # removing string "None" from column G
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        dupHeaderRemoved = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagStopText, deleteFlagRefColumn)  # deleting rows by range
        start, end = Excel.get_start_end_row_index(dupHeaderRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        madeNoneA = makeNone(dupHeaderRemoved, start, end, makeNoneRefColumn1)  # making empty cells of column A to none
        mergedColumnD = mergingRows(madeNoneA, start, end, refColumnToMerg, columnToMerg)  # merging misaligned rows of column D
        noneStringRemovedD = removeNone(mergedColumnD, start, end, columnToMerg)  # removing string "None" from column D
        noneRowsRemoved = removeNoneRows(noneStringRemovedD, start, end, refColumnToRemoveNoneRows)  # removing rows if reference column cell is None
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        footerDeleted = deleteFooter(noneRowsRemoved, end - 1)  # deleting footer rows end-1 to Inclide End Footer
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # deleting header rows start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        end = sheet.max_row  # table data end row
        convertedDateA = dateConversion(headerDeleted, start + 1, end + 1, dateConversionColumn1)  # converting date to standard date formate, start+1 to Skip Header, end+1 to IncludeLast Row
        convertedDateB = dateConversion(convertedDateA, start + 1, end + 1, dateConversionColumn2)  # converting date to standard date formate, start+1 to Skip Header, end+1 to IncludeLast Row
        lastCol = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        transdate = Excel.alter_header_name(convertedDateB, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
        valuedate = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        chqno = Excel.alter_header_name(valuedate, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        narration = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        debit = Excel.alter_header_name(narration, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)  # alter header name from the excel file to the standard column name
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
        replacedNoneWITHDRAWAL = Excel.empty_cell_to_none(slCreated, start, end + 1, headerTextToReplaceToNone1)  # making empty cells of "Withdrawal" column to None
        replacedNoneDEPOSIT = Excel.empty_cell_to_none(replacedNoneWITHDRAWAL, start, end + 1, headerTextToReplaceToNone2)  # making empty cells of "Deposit" to None
        createdTransTypeColumn = Excel.transaction_type_column(replacedNoneDEPOSIT)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    path = ""
    wb = openpyxl.load_workbook(path)
    result = yes1_main(wb)
    # result["data"].save("C:/Users/Admin/Desktop/YES1output.xlsx")
