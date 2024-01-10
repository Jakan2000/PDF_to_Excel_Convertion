from datetime import datetime

import openpyxl
from CommonClass import Excel


def dateConvertion(wb, start, end, column):
    """
        Converts date values in a specified column of an openpyxl Workbook object to a standard date format.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index where the conversion should start (inclusive).
        - end (int): The row index where the conversion should end (exclusive).
        - column (str): The column letter (e.g., 'A', 'B', 'C') containing the date values.

        Returns:
        - openpyxl.Workbook: The modified Workbook object with date values in the specified column converted to the standard format.

        Note:
        The function iterates through the specified column and converts each date value to the standard date format '%d-%b-%Y'.
        The modified Workbook is then returned.
    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        original_date = sheet[f"{column}{i}"].value  # getting date from sheet cell value
        new_date = datetime.strptime(original_date, '%d-%b-%Y').date()  # converting to standard date formate
        sheet[f"{column}{i}"].value = new_date  # assigning converted date to sheet cell value
    return wb


def deleteHeader(wb, start):
    """
        Deletes header data from the active sheet of an openpyxl Workbook object.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index where the deletion should start (inclusive).

        Returns:
        - openpyxl.Workbook: The modified Workbook object with header rows deleted.

        Note:
        The function iterates through the specified rows, starting from the specified row index 'start'
        and deletes each row until it reaches the 1st row of the sheet, effectively removing header data.
    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to 1st row of sheet
        sheet.delete_rows(x)  # delete row
    return (wb)


def deleteFooter(wb, end):
    """
        Deletes footer data from the active sheet of an openpyxl Workbook object.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - end (int): The row index where the deletion should end (exclusive).

        Returns:
        - openpyxl.Workbook: The modified Workbook object with footer rows deleted.

        Note:
        The function iterates through the specified rows, starting from the last row of the sheet,
        and deletes each row until it reaches the specified 'end' row, effectively removing footer data.
    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through sheet max row to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def iob2_validation(wb):
    """
        Validates the column count in the active sheet of an openpyxl Workbook object for a specific core logic.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object to be validated.

        Returns:
        - bool: True if the number of columns in the active sheet is not equal to the expected core logic column count (7),
                False otherwise.

        Note:
        The function checks if the maximum number of columns in the active sheet is not equal to the expected core logic column count (7).
        If the condition is met, it returns True, indicating a validation failure. Otherwise, it returns False, indicating a successful validation.
    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 7  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def iob2_main(wb):
    """
        Processes and standardizes an openpyxl Workbook object according to IOB2 specifications.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.

        Returns:
        - dict: A dictionary containing the modified Workbook object and a message.

        Note:
        The function performs the following operations:
        1. Validates the column count for core logic using iob2_validation.
        2. If validation fails, it prints an error message and returns a response dictionary with an error message.
        3. Otherwise, it processes the Workbook:
           - Deletes header and footer rows based on specified start and end text.
           - Deletes the 'COD' column.
           - Creates a new 'Sl. No.' column.
           - Aligns strings in multiple columns.
           - Alters header names to match the standard column names.
           - Converts dates to a standard date format.
           - Removes specified strings from columns.
           - Standardizes column count and creates a new 'Transaction Type' column.

        The modified Workbook and a success message are included in the response dictionary.
    """
    sheet = wb.active()
    if iob2_validation(wb):  # validating column count for core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return response with error msg
    else:
        startText = "DATE"  # header text to define table data start row
        endText = "* denotes cancelled transaction"   # text define table data end row
        startEndRefColumn = "A"  # column containing start end text
        refTextToDeleteColumn = "COD"  # reference header text to delete column
        stringAlignColumn1 = "A"  # aligning string in column by removing the \n from the string
        stringAlignColumn2 = "B"  # aligning string in column by removing the \n from the string
        stringAlignColumn3 = "C"  # aligning string in column by removing the \n from the string
        stringAlignColumn4 = "D"  # aligning string in column by removing the \n from the string
        stringAlignColumn5 = "E"  # aligning string in column by removing the \n from the string
        stringAlignColumn6 = "F"  # aligning string in column by removing the \n from the string
        refHeaderText1 = "DATE"  # header text to replace with standardised column name
        refHeaderText2 = "CHQNO"  # header text to replace with standardised column name
        refHeaderText3 = "NARATION"  # header text to replace with standardised column name
        refHeaderText4 = "DEBIT"  # header text to replace with standardised column name
        refHeaderText5 = "CREDIT"  # header text to replace with standardised column name
        refHeaderText6 = "BALANCE"  # header text to replace with standardised column name
        headerText1 = "Value_Date"  # standard column name
        headerText2 = "ChequeNo_RefNo"  # standard column name
        headerText3 = "Narration"  # standard column name
        headerText4 = "Withdrawal"  # standard column name
        headerText5 = "Deposit"  # standard column name
        headerText6 = "Balance"  # standard column name
        dateConversionColumn = "A"  # column to convert date to standard date formate
        refStringToRemove = "None"  # remove desired string from a column
        refColumnToRemoveString1 = "B"  # column to remove reference string
        refColumnToRemoveString2 = "D"  # column to remove reference string
        refColumnToRemoveString3 = "E"  # column to remove reference string
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        deleteHeader(wb, start - 1)  # deleting header text
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        deleteFooter(wb, end - 1)  # deleting footer text
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        Excel.delete_column(wb, refTextToDeleteColumn)  # deleting COD column
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 -> ASCII value
        Excel.create_slno_column(wb, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
        lastCol = 65 + Excel.column_count(wb)  # 65 -> ASCII value
        columnA = Excel.string_align(wb, start, end + 1, stringAlignColumn1)  # aligning string in column by removing the \n from the string -> \n -> next line
        columnB = Excel.string_align(wb, start, end + 1, stringAlignColumn2)  # aligning string in column by removing the \n from the string -> \n -> next line
        columnC = Excel.string_align(wb, start, end + 1, stringAlignColumn3)  # aligning string in column by removing the \n from the string -> \n -> next line
        columnD = Excel.string_align(wb, start, end + 1, stringAlignColumn4)  # aligning string in column by removing the \n from the string -> \n -> next line
        columnE = Excel.string_align(wb, start, end + 1, stringAlignColumn5)  # aligning string in column by removing the \n from the string -> \n -> next line
        columnF = Excel.string_align(wb, start, end + 1, stringAlignColumn6)  # aligning string in column by removing the \n from the string -> \n -> next line
        valuedate = Excel.alter_header_name(wb, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
        chqno = Excel.alter_header_name(wb, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        narration = Excel.alter_header_name(wb, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        debit = Excel.alter_header_name(wb, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        credit = Excel.alter_header_name(wb, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(wb, refHeaderText6, headerText6, lastCol)  # alter header name from the excel file to the standard column name
        dateConvertion(wb, start + 1, end + 1, dateConversionColumn)  # converting date to standard date formate
        Excel.remove_string(wb, start, end + 1, refStringToRemove, refColumnToRemoveString1)  # remove desired string from a column
        Excel.remove_string(wb, start, end + 1, refStringToRemove, refColumnToRemoveString2)  # remove desired string from a column
        Excel.remove_string(wb, start, end + 1, refStringToRemove, refColumnToRemoveString3)  # remove desired string from a column
        Excel.finalise_column(wb, columns)  # standardizing column count
        Excel.transaction_type_column(wb)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/last_year_statement_july_1to_october_31.xlsx"
    wb = openpyxl.load_workbook(path)
    result = iob2_main(wb)
    result["data"].save("C:/Users/Admin/Desktop/last_year_statement_july_1to_october_31.xlsx")