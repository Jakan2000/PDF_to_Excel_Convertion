from datetime import datetime

import openpyxl

from CommonClass import Excel


def deleteHeader(wb, start):
    """
        Delete header rows from the specified Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be modified.
        - start (int): The row index indicating the start of the table data.

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook after deleting header rows.

        Note:
        This function deletes rows in reverse order starting from the specified 'start' row index.
        The 'wb' parameter is modified in-place, and the modified workbook is returned.
    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row
        sheet.delete_rows(x)  # delete row
    return wb


def removeFooter(wb, end):
    """
        Remove footer rows from the specified Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be modified.
        - end (int): The row index indicating the end of the table data.

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook after removing footer rows.

        Note:
        This function deletes rows in reverse order starting from the last row of the sheet
        up to the specified 'end' row index. The 'wb' parameter is modified in-place,
        and the modified workbook is returned.
    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through max row of sheet to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def dateConvertion(wb, start, end, column):
    """
        Convert date values in the specified column to a standard date format in the given Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be modified.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data.
        - column (str): The column letter representing the column with date values.

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook after converting date values.

        Note:
        This function iterates through the rows in the specified range (from 'start' to 'end')
        and converts the date values in the specified 'column' to a standard date format.
        The 'wb' parameter is modified in-place, and the modified workbook is returned.
    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(sheet[f"{column}{i}"].value, "%d %b %Y").date()  # converting to standard date formate
    return wb


def removeRowsByDateLength(wb, start, end, column):
    """
        Delete rows in the specified Excel workbook where the date values in the specified column have insufficient length.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be modified.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data.
        - column (str): The column letter representing the column with date values.

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook after removing rows with insufficient date length.

        Note:
        This function iterates through the rows in the specified range (from 'end' to 'start')
        and deletes rows where the length of the date values in the specified 'column'
        is less than a predefined 'yearLength'. The 'wb' parameter is modified in-place,
        and the modified workbook is returned.
    """
    sheet = wb.active
    yearLength = 6  # year string length in date column cell
    for x in range(end, start, -1):  # iterating through table data end row to table data start row
        if len(str(sheet[f"{column}{x}"].value)) < yearLength:  # if length of cell value < year length
            sheet.delete_rows(x)  # delete row
    return wb


def mergingRowsByDateLength(wb, start, end, refColumn, mergingColumn):
    """
        Merge misaligned rows of the specified column by date length as reference in an Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be modified.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data.
        - refColumn (str): The column letter representing the reference column for determining row merging.
        - mergingColumn (str): The column letter representing the column to be merged.

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook after merging misaligned rows.

        Note:
        This function iterates through the rows in the specified range (from 'start' to 'end')
        and merges rows in the 'mergingColumn' based on the date length in the 'refColumn'.
        Rows are merged if the 'refColumn' cell value is not None and has a length greater than or equal to 'dateWithFullLen'.
        The 'wb' parameter is modified in-place, and the modified workbook is returned.
    """
    sheet = wb.active
    dateWithFullLen = 10  # string length of complete date
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through table data start row to table data end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None and len(str(slno)) >= dateWithFullLen:  # if cell value is not none and length of cell value >= dateWithFullLen
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                cell_address = f"{mergingColumn}{i}"
                dataToMerge.append(cell_address)  # store cell address in the 0 index
                dataToMerge.append(sheet[cell_address].value)  # store data from the 1 index
            else:  # if dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[cell_address].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                cell_address = f"{mergingColumn}{i}"
                dataToMerge.append(cell_address)  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[cell_address].value)  # appending row data to the corresponding index
        else:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[cell_address].value = st1  # assign conceited data to the cell
    return wb  # return work book by merging the corresponding rows in the column


def removeNone(wb, start, end, column):
    """
        Remove the string "None" from the specified column in an Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be modified.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data.
        - column (str): The column letter representing the column from which "None" strings will be removed.

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook after removing "None" strings.

        Note:
        This function iterates through the rows in the specified range (from 'start' to 'end')
        and removes occurrences of the string "None" from the specified 'column'.
        The 'wb' parameter is modified in-place, and the modified workbook is returned.
    """
    sheet = wb.active
    for x in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):  # if cell value is not none and string "None" in cell value
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")  # replace it with empty string
    return wb


def mergingDateColumn(wb, start, end, column):
    """
        Merge data in an Excel workbook's date column by combining incomplete date strings with corresponding year strings.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be modified.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data.
        - column (str): The column letter representing the date column to be merged.

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook after merging data in the specified date column.

        Note:
        This function iterates through the rows in the specified range (from 'start' to 'end')
        and merges incomplete date strings with corresponding year strings in the specified 'column'.
        The 'wb' parameter is modified in-place, and the modified workbook is returned.
    """
    sheet = wb.active
    inCompleteDateLen = 10  # incomplete date string length
    yearLength = 4  # year string length length
    for i in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{column}{i}"].value is not None:  # if cell value is not none
            if (len(str(sheet[f"{column}{i}"].value)) < inCompleteDateLen and len(str(sheet[f"{column}{i + 1}"].value)) == yearLength):  #if cell value string cell length < incomplete date length and string length of next cell value == year length
                s = str(sheet[f"{column}{i}"].value) + " " + str(sheet[f"{column}{i + 1}"].value)  # concat date with year
                sheet[f"{column}{i}"].value = s  # assign concated string to cell value
    return wb


def sbi1_validation(wb):
    """
        Validate the column count in an Excel workbook for the SBI Type 1 format core logic.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be validated.

        Returns:
        - bool: True if the column count is not equal to the designed core logic count, False otherwise.

        Note:
        This function checks whether the maximum column count in the active sheet of the provided workbook ('wb')
        is equal to the designed core logic count (7 columns for SBI Type 1 format).
        If the column count is not equal, it returns True, indicating a validation failure.
        Otherwise, it returns False, indicating that the column count is valid.
    """
    sheet = wb.active
    max_column = sheet.max_column  # getting sheet max column
    countOfColumn = 7  # count of column for designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def sbi1_main(wb):
    """
        Process an Excel workbook containing SBI Type 1 format data.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be processed.

        Returns:
        - dict: A dictionary containing processed data and an optional message.
            - "data": The modified Excel workbook.
            - "msg": An optional message. If None, the processing was successful.

        Note:
        This function performs a series of data processing steps on the provided Excel workbook ('wb') to standardize the format
        and structure of SBI Type 1 statements. It includes tasks such as header and footer removal, column merging,
        string alignment, date conversion, and header name alteration.
    """
    sheet = wb.active
    if sbi1_validation(wb):  # validating column count for core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return response with error msg
    else:
        startText = "Txn Date"  # header text to define table data start row
        endText = "Please do not share"  # text define table data end row
        startEndDefColumn = "A"  # column contains start and end reference text
        dupHeaderText = "Txn Date"  # reference duplicate header text to remove rows
        dupHeaderRefColumn = "A"  # reference column to remove duplicate header text
        columnToMerg1 = "A"  # column to merge misaligned date rows
        columnToMerg2 = "B"  # column to merge misaligned date rows
        refColumnToMerg = "A"  # reference column to merge other columns
        columnToMerg3 = "C"  # merging rows by reference column date length
        columnToMerg4 = "D"  # column to merge misaligned row
        removeNoneColumn1 = "A"  # column to remove string "None"
        removeNoneColumn2 = "B"  # column to remove string "None"
        removeNoneColumn3 = "C"  # column to remove string "None"
        removeNoneColumn4 = "D"  # column to remove string "None"
        columnToConvertDate1 = "A"  # column to convert date to standard date formate
        columnToConvertDate2 = "B"  # column to convert date to standard date formate
        stringAlignColumn1 = "B"  # column to align string by removing \n from it
        stringAlignColumn2 = "C"  # column to align string by removing \n from it
        stringAlignColumn3 = "D"  # column to align string by removing \n from it
        refHeaderText1 = "Txn Date"  # header text to replace with standardised column name
        refHeaderText2 = "ValueDate"  # header text to replace with standardised column name
        refHeaderText3 = "Description"  # header text to replace with standardised column name
        refHeaderText4 = "Ref No./ChequeNo."  # header text to replace with standardised column name
        refHeaderText5 = "Debit"  # header text to replace with standardised column name
        refHeaderText6 = "Credit"  # header text to replace with standardised column name
        refHeaderText7 = "Balance"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Value_Date"  # standard column name
        headerText3 = "Narration"  # standard column name
        headerText4 = "ChequeNo_RefNo"  # standard column name
        headerText5 = "Withdrawal"  # standard column name
        headerText6 = "Deposit"  # standard column name
        headerText7 = "Balance"  # standard column name
        headerToReplaceEmptyCellToNone1 = "ChequeNo_RefNo"  # column header text to replace empty cells to none
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify table data with in
        dupHeaderRemoved = Excel.remove_rows(wb, start, end + 1, dupHeaderText, dupHeaderRefColumn)  # remove multiple rows by reference text in reference column
        start, end = Excel.get_start_end_row_index(dupHeaderRemoved, startText, endText, startEndDefColumn)  # get start and end row index to specify table data with in
        columnMergA = mergingDateColumn(dupHeaderRemoved, start, end + 1, columnToMerg1)  # merging data in date column
        columnMergB = mergingDateColumn(columnMergA, start, end + 1, columnToMerg2)  # merging data in date column
        noneRemovedFromA = removeNone(columnMergB, start, end + 1, removeNoneColumn1)  # removing string "None" from column A
        noneRemovedFromB = removeNone(noneRemovedFromA, start, end + 1, removeNoneColumn2)  # removing string "None" from column "B"
        columnMergC = mergingRowsByDateLength(noneRemovedFromB, start + 1, end + 1, refColumnToMerg, columnToMerg3)  # merging rows of column "C" with date length as reference, start+1 to Skip Header
        columnMergD = mergingRowsByDateLength(columnMergC, start + 1, end + 1, refColumnToMerg, columnToMerg4)  # start+1 to Skip Header
        unWantedRowsRemoved = removeRowsByDateLength(columnMergD, start, end + 1, refColumnToMerg)  # removing rows by date length (rows having only years)
        start, end = Excel.get_start_end_row_index(unWantedRowsRemoved, startText, endText, startEndDefColumn)  # get start and end row index to specify table data with in
        if "The count of transactions for the selected date range exceeds" in sheet[f"A{end-1}"].value:  # in some statements this line may appear, so if appear
            end = end-1  # decrement the end row index to -1
            footerRemoved = removeFooter(wb, end - 1)  # removing footer rows, end-1 to Include End row
        footerRemoved = removeFooter(wb, end - 1)  # removing footer rows, end-1 to Include End row
        dateConvertedA = dateConvertion(unWantedRowsRemoved, start + 1, end, columnToConvertDate1)  # converting date to standard date formate, start+1 to Skip Header
        dateConvertedB = dateConvertion(dateConvertedA, start + 1, end, columnToConvertDate2)  # convert date to standard date formate, start+1 to Skip Header
        start, end = Excel.get_start_end_row_index(dateConvertedB, startText, endText, startEndDefColumn)  # get start and end row index to specify table data with in
        noneRemovedFromC = removeNone(dateConvertedB, start, end + 1, removeNoneColumn3)  # removing string "None" from column "C"
        noneRemovedFromD = removeNone(noneRemovedFromC, start, end + 1, removeNoneColumn4)  # removing string "None" from column "D"
        headerDeleted = deleteHeader(footerRemoved, start - 1)  # deleting header rows, start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(dateConvertedB, startText, endText, startEndDefColumn)  # get start and end row index to specify table data with in
        alignedColumnStringB = Excel.string_align(headerDeleted, start, end + 1, stringAlignColumn1)  # aligning string in colum rows by removing \n, end+1 to Include Last Row
        alignedColumnStringC = Excel.string_align(alignedColumnStringB, start, end + 1, stringAlignColumn2)  # aligning string in colum rows by removing \n, end+1 to Include Last Row
        alignedColumnStringD = Excel.string_align(alignedColumnStringC, start, end + 1, stringAlignColumn3)  # aligning string in colum rows by removing \n, end+1 to Include Last Row
        lastCol = 65 + Excel.column_count(wb)  # 65 => ASCII value "A", column_count() function returns count of column in excel sheet
        transdate = Excel.alter_header_name(alignedColumnStringD, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
        valuedate = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        naration = Excel.alter_header_name(valuedate, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        chqno = Excel.alter_header_name(naration, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        debit = Excel.alter_header_name(chqno, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)  # alter header name from the excel file to the standard column name
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A", column_count() function returns count of column in excel sheet
        slCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
        replacedToNoneCHQNO = Excel.empty_cell_to_none(slCreated, start, end + 1, headerToReplaceEmptyCellToNone1)  # making empty cells to none in column "ChequeNo_RefNo"
        createdTransTypeColumn = Excel.transaction_type_column(replacedToNoneCHQNO)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Desktop/SBI_2__25-12-2023-12-05-02.xlsx"
    # path = "C:/Users/Admin/Downloads/NKP SBI STATEMENT__28-12-2023-16-47-30.xlsx"
    # path = "C:/Users/Admin/Downloads/50100546053022_1695798594660_HDFC__28-12-2023-17-02-53.xlsx"
    # path = "C:/Users/Admin/Desktop/New folder/SBI/2. May - 2022__29-12-2023-01-52-42.xlsx"
    # path = "C:/Users/Admin/Desktop/New folder/SBI/3. June - 2022__29-12-2023-02-01-21.xlsx"
    # path = "C:/Users/Admin/Desktop/New folder/SBI/4. July - 2022__29-12-2023-02-12-20.xlsx"
    # path = "C:/Users/Admin/Desktop/5. August - 2022__29-12-2023-02-18-56.xlsx"
    # path = "C:/Users/Admin/Desktop/6. September - 2022__29-12-2023-02-36-26.xlsx"
    # path = "C:/Users/Admin/Desktop/7. October - 2022__29-12-2023-02-55-50.xlsx"
    # path = "C:/Users/Admin/Desktop/8. November - 2022__29-12-2023-02-51-10.xlsx"
    # path = "C:/Users/Admin/Desktop/9. December - 2022__29-12-2023-02-57-57.xlsx"
    # path = "C:/Users/Admin/Desktop/10. January - 2023__29-12-2023-02-59-45.xlsx"
    # path = "C:/Users/Admin/Desktop/11. February - 2023__29-12-2023-03-02-23.xlsx"
    # path = "C:/Users/Admin/Desktop/12. March - 2023__29-12-2023-03-04-04.xlsx"
    path = "C:/Users/Admin/Downloads/40499XXXXXX_HEJ2C9U8_unlocked__29-12-2023-14-16-03.xlsx"
    wb = openpyxl.load_workbook(path)
    result = sbi1_main(wb)
    result["data"].save('C:/Users/Admin/Desktop/SBI1output.xlsx')
