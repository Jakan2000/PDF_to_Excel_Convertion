import openpyxl


# 9
def aligningHeader(wb):
    sheet = wb.active
    column_a = sheet['A']
    start = 0

    end = 0
    for cell in column_a:
        start += 1
        if "Date" in str(cell.value):
            print(start)
            break
    for cell in column_a:
        end += 1
        if "Statement Summary" in str(cell.value):
            print(end)
            break

    wb.save('C:/Users/Admin/Desktop/KSV/Python/ExcelOutput/KotakYearlyoutput.xlsx')

# 8
def removingNoneFromColumn(wb):
    sheet = wb.active
    column_c = sheet['C']
    for cell in column_c:
        if "None" in str(cell.value):
            cell.value = str(cell.value).replace("None", "")
    column_b = sheet['B']
    for cell in column_b:
        if "None" in str(cell.value):
            cell.value = str(cell.value).replace("None", "")
    column_a = sheet['A']
    for cell in column_a:
        if "None" in str(cell.value):
            cell.value = str(cell.value).replace("None", "")
    aligningHeader(wb)

# 7
def mergingrows2(wb):
    sheet = wb.active
    column_a = sheet['A']
    start = 0
    end = 0
    for cell in column_a:
        start += 1
        if "Date" in str(cell.value):
            # print(start)
            break
    for cell in column_a:
        end += 1
        if "Statement Summary" in str(cell.value):
            # print(end)
            break
    for i in range(start, end):
        if len(str(sheet[f"A{i}"].value)) > 10:
            # print(sheet[f"A{i}"].value)
            sheet[f"B{i-1}"].value = sheet[f"B{i-1}"].value + sheet[f"B{i}"].value
            sheet[f"C{i-1}"].value = sheet[f"C{i-1}"].value + sheet[f"C{i}"].value
    for x in range(end-1, start, -1):
        if len(str(sheet[f"A{x}"].value)) > 10:
            print(sheet[f"A{x}"].value)
            sheet.delete_rows(x)
    removingNoneFromColumn(wb)

# 6
def removingNullRows(wb, start, end):
    sheet = wb.active
    for x in range(end, start, -1):
        if sheet[f"A{x}"].value is None:
            sheet.delete_rows(x)
    mergingrows2(wb)

# 5
def mergingColumnC(wb, start, end):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        a_cell = "A" + f"{i}"
        c_cell = "C" + f"{i}"
        date = sheet[a_cell].value
        if date is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append("C" + f"{i}")
                dataToMerge.append(sheet[c_cell].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append("C" + f"{i}")
                dataToMerge.append(sheet[c_cell].value)
        if date is None:
            dataToMerge.append(sheet[c_cell].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    removingNullRows(wb, start, end)

# 4
def mergingColumnB(wb, start, end):
    sheet = wb.active
    # print(f"{start}     {end}")
    dataToMerge = []
    for i in range(start, end):
        a_cell = "A" + f"{i}"
        b_cell = "B" + f"{i}"
        date = sheet[a_cell].value
        if date is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append("B" + f"{i}")
                dataToMerge.append(sheet[b_cell].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append("B" + f"{i}")
                dataToMerge.append(sheet[b_cell].value)
        if date is None:
            dataToMerge.append(sheet[b_cell].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    mergingColumnC(wb, start, end)

# 3
def mergingColumnA(wb):
    sheet = wb.active
    column_a = sheet['A']
    start = 0
    end = 0
    for cell in column_a:
        start += 1
        if "Date" in str(cell.value):
            # print(start)
            break
    for cell in column_a:
        end += 1
        if "Statement Summary" in str(cell.value):
            # print(end)
            break
    for i in range(start, end - 1):
        if sheet[f"A{i}"].value is not None and len(str(sheet[f"A{i}"].value)) < 8:
            sheet[f"A{i}"].value = str(sheet[f"A{i}"].value) + str(sheet[f"A{i + 1}"].value)
            # print(sheet[f"A{i}"].value)
    mergingColumnB(wb, start, end)

# 2
def deletingDuplicateHeader2(wb):
    sheet = wb.active
    column_a = sheet['A']
    start = 0
    end = 0
    for cell in column_a:
        start += 1
        if "Date" in str(cell.value):
            # print(start)
            break
    for cell in column_a:
        end += 1
        if "Statement Summary" in str(cell.value):
            # print(end)
            break
    for x in range(end, start, -1):
        if "Date" in str(sheet[f"A{x}"].value):
            sheet.delete_rows(x)
        elif "Period" in str(sheet[f"B{x}"].value):
            sheet.delete_rows(x)
    mergingColumnA(wb)

# 1
def deletingDuplicateHeader1(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb.active
    column_a = sheet['A']
    start = 0
    end = 0
    for cell in column_a:
        start += 1
        if "Date" in str(cell.value):
            # print(start)
            break
    for cell in column_a:
        end += 1
        if "Statement Summary" in str(cell.value):
            break
    # Aligning Column Header
    for x in range(end, start - 1, -1):
        if sheet[f"B{x}"].value is None and "Narration" in str(sheet[f"C{x}"].value):
            for column in range(2, 8):  # Columns B to G
                current_cell = sheet.cell(row=x, column=column)
                next_cell = sheet.cell(row=x, column=column + 1)
                if current_cell.value is None:
                    # Copy the value from the next cell to the current cell
                    current_cell.value = next_cell.value
                    # Clear the value from the next cell
                    next_cell.value = None
    delete_next = False
    first_set_encountered = False
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        cell_value = row[0].value
        if cell_value == "Period":
            if first_set_encountered:
                start_row = row[0].row + 1  # Start one row below "Period" for subsequent sets
                delete_next = True
            else:
                first_set_encountered = True
        elif cell_value == "Narration" and delete_next:
            end_row = row[0].row - 1  # End one row above "Narration"
            for row_num in range(start_row, end_row + 1):
                sheet.delete_rows(start_row)
            delete_next = False
    deletingDuplicateHeader2(wb)
# 1
deletingDuplicateHeader1(path="C:/Users/Admin/Downloads/7111741150__13-09-2023-12-07-32.xlsx")