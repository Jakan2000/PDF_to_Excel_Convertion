from datetime import datetime

import openpyxl

from CommonClass import Excel


def dateHeaderAlter(wb, refText, actualText, lastCol):
    """
        Alter header names in a Workbook based on a reference text.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - refText (str): The reference text to identify the headers that need alteration.
        - actualText (str): The actual text to replace the identified headers.
        - lastCol (int): The ASCII value representing the last column to iterate through.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object with altered headers.

        Notes:
        - This function iterates through the header row of the Workbook and alters the headers
          based on the provided reference text and actual replacement text.
        - Headers with a length less than 5 and containing the reference text will be replaced
          with the specified actual text.
    """
    sheet = wb.active
    column = 65  # ascii value of "A"
    row = 1  # header row index
    while column < lastCol:  # iterating through all columns
        if refText in str(sheet[f"{chr(column)}{row}"].value) and len(str(sheet[f"{chr(column)}{row}"].value)) < 5:  # if reference text in cell value and the length is < 5
            sheet[f"{chr(column)}{row}"].value = actualText  # assign the actual text in the cell
        column += 1
    return wb


def dateConvertion(wb, start, end, column):
    """
        Convert the date in a specified column to a standard date format in a Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index in the specified column to begin the conversion.
        - end (int): The ending row index in the specified column to end the conversion (exclusive).
        - column (str): The column letter where date conversion is to be performed.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object with date conversion applied.

        Notes:
        - This function iterates through the specified column in the Workbook, converts
          each cell's date value to a standard date format (%d-%m-%Y), and updates the cell value.
    """
    sheet = wb.active
    for i in range(start, end):  # iterate through table data start row and table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), '%d-%m-%Y').date()  # convert the date to standard date formate
    return wb


def removeNone(wb, start, end, column):
    """
        Replace "None" with an empty string in the specified column of a Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index in the specified column to begin the replacement.
        - end (int): The ending row index in the specified column to end the replacement (exclusive).
        - column (str): The column letter where the replacement is to be performed.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object with "None" replaced by an empty string.

        Notes:
        - This function iterates through the specified column in the Workbook, checks each cell's value,
          and replaces the value with an empty string if the value is not None and contains the string "None".
    """
    sheet = wb.active
    for x in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):  # if cell value in column is not none and cell value contains "None" in it
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")  # replace string "None" with empty string
    return wb


def deleteHeader(wb, start):
    """
       Delete rows above the specified start index row in the Workbook.

       Parameters:
       - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
       - start (int): The index of the row from which deletion starts (inclusive).

       Returns:
       - openpyxl.Workbook: The modified openpyxl Workbook object with rows deleted above the specified start index.

       Notes:
       - This function removes rows from the active sheet of the Workbook, starting from the specified index 'start'
         and moving upward towards the first row (header skipped).
       - It is particularly useful to remove unnecessary header information or other unwanted rows above the data.
    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterate through table data header row (header skipped) to 1st row of the excel sheet
        sheet.delete_rows(x)  # delete row
    return wb


def deleteRow(wb, start, refText, refColumn):
    """
        Delete multiple rows in the Workbook using a reference text in a specific column.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The index of the row from which deletion starts (inclusive).
        - refText (str): The reference text used to identify rows for deletion.
        - refColumn (str): The column letter containing the reference text.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object with rows deleted based on the reference text.

        Notes:
        - This function removes rows from the active sheet of the Workbook based on the presence of the specified
          reference text in the given column.
        - It is useful for removing rows containing specific information or unwanted data based on a reference text.
    """

    sheet = wb.active
    for x in range(sheet.max_row, start, -1):  # iterate through max row in sheet to table data start row
        if refText in str(sheet[f"{refColumn}{x}"].value):  # if reference text in cell value
            sheet.delete_rows(x)  # delete row
    return wb


def deleteFooter(wb, end):
    """
        Delete all rows below the specified end index (last row) in the Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - end (int): The index of the last row (inclusive) up to which rows will be retained.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object with rows deleted below the specified end index.

        Notes:
        - This function removes all rows below the specified end index in the active sheet of the Workbook.
        - It is useful for deleting footer rows or unnecessary information present after the actual data in an Excel sheet.
    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterate through max row of sheet to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def removingNullRows(wb, start, end, column):
    """
        Remove rows with null values in the specified column within a given range in the Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The index of the starting row (inclusive) within which null rows will be removed.
        - end (int): The index of the last row (inclusive) up to which rows with null values will be removed.
        - column (str): The column letter or label (e.g., 'A', 'B') containing the date column used as a reference.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object with null rows removed based on the specified column.

        Notes:
        - This function iterates through the specified range and removes rows where the specified column has a null value.
        - It is useful for cleaning up data by removing rows with missing or null values in a particular column.
    """
    sheet = wb.active
    for x in range(end, start, -1):  # iterating through table data end row to table data start row
        if sheet[f"{column}{x}"].value is None:  # if cell value is none in desired column
            sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merge consecutive rows with the same reference value in the specified columns within a given range in the Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The index of the starting row (inclusive) within which rows will be merged.
        - end (int): The index of the last row (inclusive) up to which rows will be merged.
        - refColumn (str): The column letter or label (e.g., 'A', 'B') containing the reference values for row merging.
        - mergingColumn (str): The column letter or label (e.g., 'C', 'D') where consecutive rows' data will be merged.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object with consecutive rows merged based on the specified columns.

        Notes:
        - This function iterates through the specified range and merges consecutive rows in the specified column based on the reference column.
        - Rows are merged by concatenating the values in the specified merging column.
        - The merged value is placed in the first cell of the consecutive rows, and subsequent cells are cleared.
    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through start and end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    st = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def removeRowsOnRange(wb, start, end, startText, endText, column):
    """
        Delete rows within a specified range based on the presence of start and end text in a given column of a Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The index of the starting row (inclusive) within which rows may be deleted.
        - end (int): The index of the last row (inclusive) up to which rows may be deleted.
        - startText (str): The text indicating the start of the range to be deleted when found in the specified column.
        - endText (str): The text indicating the end of the range to be deleted when found in the specified column.
        - column (str): The column letter or label (e.g., 'A', 'B') where the start and end texts are checked for row deletion.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object after deleting rows within the specified range.

        Notes:
        - This function iterates through the specified range and deletes rows based on the presence of start and end text in the specified column.
        - Rows are deleted from the first occurrence of the start text until the last occurrence of the end text (inclusive).
    """
    sheet = wb.active
    delete_flag = False
    rows_to_delete = []
    for i in range(start, end):  # iterate from start to end row
        if startText in str(sheet[f"{column}{i}"].value):  # if start text is in the reference column cell it's the starting row
            delete_flag = True  # make delete flag true
        if delete_flag:  # if delete flag is true append the row to rows_to_delete array
            rows_to_delete.append(i)  # append row to rows_to_delete array
        if endText in str(sheet[f"{column}{i}"].value):  # if stop text is in the reference column cell it's the last row
            delete_flag = False  # make delete flag false
    for idx in reversed(rows_to_delete):  # iterate array in reversed order to avoid the index problem while deleting the rows
        sheet.delete_rows(idx)  # delete row
    return wb


def federal1_validation(wb):
    """
        Validate the columns in the given Workbook for compliance with the designed core logic.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.

        Returns:
        - bool: True if the number of columns in the Workbook is not equal to the designed core logic count; False otherwise.

        Notes:
        - This function checks if the number of columns in the active sheet of the Workbook matches the expected count according to the designed core logic.
        - The designed core logic specifies a predefined count of columns (countOfColumn), and this function returns True if the actual number of columns is different, indicating a validation failure.
    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 10  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def federal1_main(wb):
    """
        Process and standardize the data in the given Workbook based on the designed core logic for Federal format.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.

        Returns:
        - dict: A dictionary containing the processed Workbook and an optional message.
            - 'data' (openpyxl.Workbook): The processed Workbook.
            - 'msg' (str or None): A message indicating the result of the processing. If there is no error, msg is set to None.

        Notes:
        - This function performs a series of operations on the Workbook to standardize the data according to the designed core logic for the Federal format.
        - It includes operations such as removing unnecessary rows, merging rows, aligning strings, deleting columns, converting date formats, and standardizing column names.
        - The function returns a dictionary containing the processed Workbook and an optional message indicating the result of the processing.
    """
    sheet = wb.active
    if federal1_validation(wb):  # validate columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return the response with error msg
    else:
        startText = "Particulars"  # header text in column C
        endText = "GRAND TOTAL"  # text define the end of the table data row
        startEndRefColumn = "C"  # reference column contains the start and end text
        dupHeaderStartText = "Page"  # reference start text to delete rows by range
        dupHeaderEndText = "Deposits"  # reference end text to delete rows by range
        dupHeaderRefColumn = "H"  # reference column to delete rows by range
        columnToMerg1 = "C"  # column to merge the miss aligned row data
        refColumnToMerg = "A"  # reference column to merge the row data (date column)
        deleteOpenBalText = "Opening Balance"  # reference text to delete row
        deleteOpenBalRefColumn = "C"  # column which contains the reference text to remove row
        stringAlignColumn1 = "A"  # column to align string by removing "\n"
        stringAlignColumn2 = "B"  # column to align string by removing "\n"
        stringAlignColumn3 = "C"  # column to align string by removing "\n"
        stringAlignColumn4 = "D"  # column to align string by removing "\n"
        stringAlignColumn5 = "E"  # column to align string by removing "\n"
        stringAlignColumn6 = "F"  # column to align string by removing "\n"
        stringAlignColumn7 = "G"  # column to align string by removing "\n"
        stringAlignColumn8 = "H"  # column to align string by removing "\n"
        stringAlignColumn9 = "I"  # column to align string by removing "\n"
        deleteColumnRefText1 = "TranType"  # reference header text to delete column
        deleteColumnRefText2 = "Tran Id"  # reference header text to delete column
        deleteColumnRefText3 = "Cr/Dr"  # reference header text to delete column
        dateConversionColumn1 = "A"  # column to convert date to standard date formate
        dateConversionColumn2 = "B"  # column to convert date to standard date formate
        refHeaderText1 = "Date"  # header text to replace with standardised column name
        refHeaderText2 = "Value Date"  # header text to replace with standardised column name
        refHeaderText3 = "Particulars"  # header text to replace with standardised column name
        refHeaderText4 = "ChequeDetails"  # header text to replace with standardised column name
        refHeaderText5 = "Withdrawals"  # header text to replace with standardised column name
        refHeaderText6 = "Deposits"  # header text to replace with standardised column name
        refHeaderText7 = "Balance"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Value_Date"  # standard column name
        headerText3 = "Narration"  # standard column name
        headerText4 = "ChequeNo_RefNo"  # standard column name
        headerText5 = "Withdrawal"  # standard column name
        headerText6 = "Deposit"  # standard column name
        headerText7 = "Balance"  # standard column name
        negativeValueColumnRefText1 = "Withdrawal"  # no need of converting negative value to positive
        headerTextToMakeEmptyCellsToNone1 = "Withdrawal"  # header text to make empty cells to none
        headerTextToMakeEmptyCellsToNone2 = "Deposit"  # header text to make empty cells to none
        headerTextToMakeEmptyCellsToNone3 = "ChequeNo_RefNo"  # header text to make empty cells to none
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        dupHeaderRemoved = removeRowsOnRange(wb, start, end, dupHeaderStartText, dupHeaderEndText, dupHeaderRefColumn)  # removing rows by range
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        columnMergedC = mergingRows(dupHeaderRemoved, start + 2, end, refColumnToMerg, columnToMerg1)  # merging the rows of desired column, start+2 to Skip Opening Balance Row
        noneRowsRemoved = removingNullRows(columnMergedC, start + 1, end - 1, refColumnToMerg)  #removing null rows by date column as reference, end-1 to Skip End Footer
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        footerDeleted = deleteFooter(noneRowsRemoved, end - 1)  # delete all the rows below the end(last) row, end-1 to Include
        openBalRowDeleted = deleteRow(wb, start, deleteOpenBalText, deleteOpenBalRefColumn)  # deleting unwanted rows
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        headerDeleted = deleteHeader(openBalRowDeleted, start - 1)  # delete rows above the start index row, start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        stringAlignedA = Excel.string_align(headerDeleted, start, end + 1, stringAlignColumn1)  # aligning string in column by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        stringAlignedB = Excel.string_align(stringAlignedA, start, end + 1, stringAlignColumn2)  # aligning string in column by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        stringAlignedC = Excel.string_align(stringAlignedB, start, end + 1, stringAlignColumn3)  # aligning string in column by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        stringAlignedD = Excel.string_align(stringAlignedC, start, end + 1, stringAlignColumn4)  # aligning string in column by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        stringAlignedE = Excel.string_align(stringAlignedD, start, end + 1, stringAlignColumn5)  # aligning string in column by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        stringAlignedF = Excel.string_align(stringAlignedE, start, end + 1, stringAlignColumn6)  # aligning string in column by removing the \n from the string -> \n -> next linen end+1 to Include Last Row
        stringAlignedG = Excel.string_align(stringAlignedF, start, end + 1, stringAlignColumn7)  # aligning string in column by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        stringAlignedH = Excel.string_align(stringAlignedG, start, end + 1, stringAlignColumn8)  # aligning string in column by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        stringAlignedI = Excel.string_align(stringAlignedH, start, end + 1, stringAlignColumn9)  # aligning string in column by removing the \n from the string -> \n -> next line, end+1 to Include Last Row
        removedNoneA = removeNone(stringAlignedI, start, end + 1, stringAlignColumn1)  # replace "None" with empty string i column cells, end+1 to Include Last Row
        removedNoneB = removeNone(removedNoneA, start, end + 1, stringAlignColumn2)  # replace "None" with empty string i column cells, end+1 to Include Last Row
        removedNoneC = removeNone(removedNoneB, start, end + 1, stringAlignColumn3)  # replace "None" with empty string i column cells, end+1 to Include Last Row
        removedNoneF = removeNone(removedNoneC, start, end + 1, stringAlignColumn6)  # replace "None" with empty string i column cells, end+1 to Include Last Row
        removedNoneG = removeNone(removedNoneF, start, end + 1, stringAlignColumn7)  # replace "None" with empty string i column cells, end+1 to Include Last Row
        removedNoneH = removeNone(removedNoneG, start, end + 1, stringAlignColumn8)  # replace "None" with empty string i column cells, end+1 to Include Last Row
        removedNoneI = removeNone(removedNoneH, start, end + 1, stringAlignColumn9)  # replace "None" with empty string i column cells, end+1 to Include Last Row
        trantypeDeleted = Excel.delete_column(wb, deleteColumnRefText1)  # deleting column using header text
        tranIdDeleted = Excel.delete_column(wb, deleteColumnRefText2)  # deleting column using header text
        crdrDeleted = Excel.delete_column(wb, deleteColumnRefText3)  # deleting column using header text
        convertedDateA = dateConvertion(crdrDeleted, start + 1, end + 1, dateConversionColumn1)  # convert the date to standard date formate, start+1 to Skip Header, end+1 to Include Last Row
        convertedDateB = dateConvertion(convertedDateA, start + 1, end + 1, dateConversionColumn2)  # convert the date to standard date formate, start+1 to Skip Header, end+1 to Include Last Row
        lastCol = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        transdate = dateHeaderAlter(wb, refHeaderText1, headerText1, lastCol)  # altering header name with actual text
        valuedate = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        naration = Excel.alter_header_name(valuedate, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        chqno = Excel.alter_header_name(naration, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        debit = Excel.alter_header_name(chqno, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)  # alter header name from the excel file to the standard column name
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slnoCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))  # created new column-slno
        neagativeValueChecked = Excel.check_neagativeValue_by_column(slnoCreated, negativeValueColumnRefText1)  # no need to convert negative value to positive
        withdrawalNoneReplaced = Excel.empty_cell_to_none(neagativeValueChecked, start + 1, end + 1, headerTextToMakeEmptyCellsToNone1)  # making empty cells in a column to none by using the header text as reference
        depositNoneReplaced = Excel.empty_cell_to_none(withdrawalNoneReplaced, start + 1, end + 1, headerTextToMakeEmptyCellsToNone2)  # making empty cells in a column to none by using the header text as reference
        chqnoNoneReplaced = Excel.empty_cell_to_none(depositNoneReplaced, start + 1, end + 1, headerTextToMakeEmptyCellsToNone3)  # making empty cells in a column to none by using the header text as reference
        Excel.finalise_column(wb, columns)  # standardizing count of column
        createdTransTypeColumn = Excel.transaction_type_column(wb)  # created new column transaction type
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/2. R RAVICHANDRAN - Federal - 2416 Pass - RAVI016 __11-09-2023-15-59-25.XLSX"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/1. Federal - 3448__21-12-2023-18-35-39.xlsx"
    path = "C:/Users/Admin/Downloads/Account Statement 2214XXXXXX3265_unlocked__30-12-2023-18-38-58.xlsx"
    wb = openpyxl.load_workbook(path)
    result = federal1_main(wb)
    result["data"].save('C:/Users/Admin/Desktop/FEDERAL1output.xlsx')
