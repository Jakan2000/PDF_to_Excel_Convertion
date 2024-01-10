from datetime import datetime

import openpyxl

from CommonClass import Excel


def aligningAllColumns(wb, start, end, refColumn):
    """
        Aligns the column data in the specified range when the value in the reference column is None.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook object.
        - start (int): The starting row index in the table data.
        - end (int): The ending row index in the table data.
        - refColumn (str): The reference column letter.

        Returns:
        - openpyxl.Workbook: The modified Excel workbook.

        Note:
        - This function is designed for aligning columns in a specified range when the value in the reference column is None.
        - Columns from F to B are shifted to the right, and the value in column E is moved to column F, and so on.
        - The specified range is inclusive of both the start and end indices.
    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start column to table data end column
        if sheet[f"{refColumn}{i}"].value is None:  # if cell value in reference column is none
            sheet[f'F{i}'].value = sheet[f'E{i}'].value  # assign E column cell value to F column cell value
            sheet[f'E{i}'].value = sheet[f'D{i}'].value  # assign D column cell value to E column cell value
            sheet[f'D{i}'].value = sheet[f'C{i}'].value  # assign C column cell value to D column
            sheet[f'C{i}'].value = sheet[f'B{i}'].value  # assign B column cell value to C column
            sheet[f'B{i}'].value = None  # assign none to B column cell value
    return wb


def dateConvertion(wb, start, end, column):
    """
        Converts the date values in the specified column to a standard date format in the given Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook object.
        - start (int): The starting row index in the table data.
        - end (int): The ending row index in the table data.
        - column (str): The column letter containing date values.

        Returns:
        - openpyxl.Workbook: The modified Excel workbook.

        Note:
        - This function converts the date values in the specified column to the standard date format "%d-%m-%Y".
        - The specified range is inclusive of both the start and end indices.
    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%m-%Y").date()  # converting to standard date formate
    return wb


def deleteHeader(wb, start):
    """
        Deletes rows above the specified start index row in the given Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook object.
        - start (int): The index of the row from which deletion starts.

        Returns:
        - openpyxl.Workbook: The modified Excel workbook.

        Note:
        - This function deletes rows from the top of the sheet (above the start index) down to the 1st row of the sheet.
        - The specified start index is inclusive, and rows above it will be removed.
    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to 1st row of sheet
        sheet.delete_rows(x)  # delete row
    return (wb)


def deleteFooter(wb, end):
    """
        Deletes all rows below the specified end index row in the given Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook object.
        - end (int): The index of the last row. Rows below this index will be deleted.

        Returns:
        - openpyxl.Workbook: The modified Excel workbook.

        Note:
        - This function deletes rows from the bottom of the sheet (below the end index) up to the last row of the sheet.
        - The specified end index is inclusive, and rows below it will be removed.
    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through max row of sheet to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def removeNoneRows(wb, start, end, column):
    """
        Removes rows with None values in the specified column within the given range.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook object.
        - start (int): The starting index of the range.
        - end (int): The ending index of the range.
        - column (str): The column letter or name where None values will be checked.

        Returns:
        - openpyxl.Workbook: The modified Excel workbook.

        Note:
        - This function iterates through the specified range (from end to start) and removes rows
          where the cell value in the specified column is None.
        - The specified range is inclusive (both start and end indices are included).
    """
    sheet = wb.active
    for x in range(end - 1, start, -1):  # iterating through table data end row to table data start row
        if sheet[f"{column}{x}"].value is None:  # if cell value is none
            sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merges consecutive rows with the same value in the reference column for a specified merging column.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook object.
        - start (int): The starting index of the range.
        - end (int): The ending index of the range.
        - refColumn (str): The column letter or name used as a reference for merging rows.
        - mergingColumn (str): The column letter or name where rows will be merged based on the reference column.

        Returns:
        - openpyxl.Workbook: The modified Excel workbook.

        Note:
        - This function iterates through the specified range and merges consecutive rows with the same value
          in the reference column for the specified merging column.
        - The merged value is a concatenation of the original values in the merging column for each consecutive row.
        - The last row in the specified range will be merged outside the loop.
    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through table data start row to table data end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def icici2_validation(wb):
    """
        Validates the number of columns in the Excel workbook for the ICICI2 core logic.

        Parameters:
        - wb (openpyxl.Workbook): The Excel workbook object.

        Returns:
        - bool: True if the number of columns is not equal to the expected count, otherwise False.

        Note:
        - This function checks whether the number of columns in the active sheet of the workbook is equal to
          the expected count for the ICICI2 core logic.
        - If the count of columns is not equal to the expected count, the validation fails (returns True).
        - If the count of columns is equal to the expected count, the validation passes (returns False).
    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 6  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def icici2_main(wb):
    """
        Overview:
        Processes an Excel workbook containing financial data, performing various data cleaning and restructuring operations.

        Parameters:
        - wb (Workbook): The input Excel workbook object.

        Returns:
        - response (dict): A dictionary containing the processed workbook (`data`) and an optional error message (`msg`).
                          If there is a column count mismatch, an error message is included; otherwise, the message is set to None.

        Operations:
        1. Column Validation:
           - Checks if the workbook has the expected number of columns using the `icici2_validation` function.
           - If the validation fails, it prints an error message, creates a response dictionary with a corresponding error message, and returns the response.

        2. Data Processing Steps:
           - Defines various constants such as start and end text, reference columns, header texts, and column names.
           - Utilizes helper functions from an `Excel` module (assuming it's defined elsewhere) for operations like deleting rows,
             merging columns, aligning columns, converting dates, deleting columns, altering header names, string alignment,
             creating serial numbers, standardizing column count, removing specific strings, making empty cells to None, and creating a new transaction type column.

        3. Workbook Processing Flow:
           - Retrieves the start and end row indices based on specific header texts to define the data range.
           - Deletes unnecessary rows using the `Excel.delete_rows_by_range` function.

        Note:
        The function follows a structured approach to process financial data in an Excel workbook. It ensures that the data is cleaned, aligned, and formatted
        according to specified standards. The `Excel` module is assumed to contain helper functions for common Excel operations.
        """
    sheet = wb.active
    if icici2_validation(wb):  # validate columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # returning response with error msg
    else:
        startText = "DATE"  # header text in column A
        # endText = "Account Related Other Information"  # text defines end of table data row
        endText = "Statement of Fixed Deposit Linked to Account Number"
        startEndDefColumn = "A"  # reference column contains the start and end text
        deleteFlagStartText = "Page"  # starting row reference text to delete the rows by range
        deleteFlagEndText = "MODE"  # ending row reference text to delete the rows by range
        refColumn1 = "B"  # column contains starting row reference text and ending row reference text to delete the rows by range
        refColumnToMerg = "A"  # reference column to merge th misaligned rows
        columnToMerg1 = "C"  # column to merge misaligned data by reference column to merge
        openBalRefText = "B/F"  # reference column to remove row
        openBalRefColumn = "C"  # reference column to remove open balance row
        refColumnToAlignAllColumns = "F"  # reference column to align misaligned column data
        dateConversionColumn = "A"  # column to convert date to standard date formate
        refTextDeleteColumn = "MODE**"  # reference header text to delete column
        refHeaderText1 = "DATE"  # header text to replace with standardised column name
        refHeaderText2 = "PARTICULARS"  # header text to replace with standardised column name
        refHeaderText3 = "DEPOSITS"  # header text to replace with standardised column name
        refHeaderText4 = "WITHDRAWALS"  # header text to replace with standardised column name
        refHeaderText5 = "BALANCE"  # header text to replace with standardised column name
        lastCol = 65 + sheet.max_column  # 65 => ASCII value "A"
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Narration"  # standard column name
        headerText3 = "Deposit"  # standard column name
        headerText4 = "Withdrawal"  # standard column name
        headerText5 = "Balance"  # standard column name
        stringAlignColumn1 = "A"  # column to align string -> removing "\n"(next line) from string
        stringAlignColumn2 = "B"  # column to align string -> removing "\n"(next line) from string
        stringAlignColumn3 = "C"  # column to align string -> removing "\n"(next line) from string
        stringAlignColumn4 = "D"  # column to align string -> removing "\n"(next line) from string
        stringAlignColumn5 = "E"  # column to align string -> removing "\n"(next line) from string
        headerToMakeEmptyCellToNOne1 = "Deposit"  # reference header text tomake empty cells to none in column
        headerToMakeEmptyCellToNOne2 = "Withdrawal"  # reference header text tomake empty cells to none in column
        refStringToRemoveFromColumn1 = "None"  # reference string to remove from desired column
        columnToRemoveNone = "C"  # column to remove reference string
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify the data with in
        rowsRemoveD = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagEndText, refColumn1)  # deleting the rows by range
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify the data with in
        mergColumnC = mergingRows(rowsRemoveD, start, end, refColumnToMerg, columnToMerg1)  # merging the rows of desired column
        removeNull = removeNoneRows(mergColumnC, start, end, refColumnToMerg)  # removing the unwanted rows, when the reference column cell value is none
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify the data with in
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify the data with in
        footerDeleted = deleteFooter(removeNull, end - 1)  # delete all the rows below the end(last) row, end-1 to Include Last Row
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # delete rows above the start index row, start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndDefColumn)  # get start and end row index to specify the data with in
        removeOpenBal = Excel.remove_row(headerDeleted, start, end, openBalRefText, openBalRefColumn)  # remove a single row by checking the referance text is in the column cell
        start, end = Excel.get_start_end_row_index(removeOpenBal, startText, endText, startEndDefColumn)  # get start and end row index to specify the data with in
        alignedAllColumns = aligningAllColumns(removeOpenBal, start, end+1, refColumnToAlignAllColumns)
        dateConvertedA = dateConvertion(alignedAllColumns, start + 1, end + 1, dateConversionColumn)  # start+1 to Skip Header, end+1 to Include Last Row
        deletedModeColumn = Excel.delete_column(wb, refTextDeleteColumn)  # aligning the column data when the reference column is None
        date = Excel.alter_header_name(deletedModeColumn, refHeaderText1, headerText1, lastCol - 1)  # alter header name from excel file to the standard column name
        naration = Excel.alter_header_name(date, refHeaderText2, headerText2, lastCol - 1)  # alter header name from excel file to the standard column name
        deposits = Excel.alter_header_name(naration, refHeaderText3, headerText3, lastCol - 1)  # alter header name from excel file to the standard column name
        withdrawal = Excel.alter_header_name(deposits, refHeaderText4, headerText4, lastCol - 1)  # alter header name from excel file to the standard column name
        balance = Excel.alter_header_name(withdrawal, refHeaderText5, headerText5, lastCol - 1)  # alter header name from excel file to the standard column name
        alignedA = Excel.string_align(balance, start, end, stringAlignColumn1)  # aligning string in column by removing the \n from the string -> \n -> next line
        alignedB = Excel.string_align(alignedA, start, end, stringAlignColumn2)  # aligning string in column by removing the \n from the string -> \n -> next line
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slCreated = Excel.create_slno_column(alignedB, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
        columnFinalised = Excel.finalise_column(slCreated, columns)  # standardizing count of column
        noneRemovedDeposit = Excel.remove_string(columnFinalised, start, end+1, refStringToRemoveFromColumn1, columnToRemoveNone)  # remove desired string from a column
        depositMadeNone = Excel.empty_cell_to_none(noneRemovedDeposit, start, end+1, headerToMakeEmptyCellToNOne1)  # making empty cells in a column to none by using the header text as reference
        withdrawalMadeNone = Excel.empty_cell_to_none(depositMadeNone, start, end+1, headerToMakeEmptyCellToNOne2)  # making empty cells in a column to none by using the header text as reference
        createdTransTypeColumn = Excel.transaction_type_column(withdrawalMadeNone)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/ICICI_-_2207PW-088601502207_unlocked__15-09-2023-12-58-00.xlsx"
    # path = "C:/Users/Admin/Downloads/2._Rajamani_-_ICICI_8226 (1)__23-11-2023-13-21-31.xlsx"
    path = "C:/Users/Admin/Downloads/GOKUL_ICICI_-_PASS_GOKU2210_unlocked__28-12-2023-17-36-15.xlsx"
    wb = openpyxl.load_workbook(path)
    result = icici2_main(wb)
    # result.save('C:/Users/Admin/Desktop/FinalOutput/ICICI2output.xlsx')
    result["data"].save('C:/Users/Admin/Desktop/ICICI2output.xlsx')
