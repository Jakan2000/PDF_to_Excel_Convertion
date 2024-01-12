from datetime import datetime

import openpyxl

from CommonClass import Excel


def dateConversion(wb, start, end, column):
    """
        Convert date values in a specified column of the active sheet in an Openpyxl Workbook to a standard date format.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The starting row index for the date conversion operation.
        - end (int): The ending row index for the date conversion operation.
        - column (str): The column letter where date values are present.

        Returns:
        - Workbook: The Openpyxl Workbook object after performing the date conversion.

        Note:
        This function iterates through the specified range of rows in the specified column of the active sheet and converts
        the date values to a standard date format ("%d-%b-%Y").

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%b-%Y").date()  # converting to standard date formate
    return wb


def deleteHeader(wb, start):
    """
        Delete header rows above the specified starting row index in the active sheet of an Openpyxl Workbook.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The row index from where the header rows will be deleted.

        Returns:
        - Workbook: The Openpyxl Workbook object after deleting the header rows.

        Note:
        This function iterates through the specified range of rows (from start to the 1st row) in the active sheet and deletes
        each row.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to 1st row of sheet
        sheet.delete_rows(x)  # dalete row
    return wb


def deleteFooter(wb, start, refColumn):
    """
       Delete footer rows below the specified starting row index in the active sheet of an Openpyxl Workbook.

       Parameters:
       - wb (Workbook): The Openpyxl Workbook object.
       - start (int): The row index from where the footer rows will be deleted.
       - refColumn (str): The reference column letter (e.g., 'A') to check for None or "None" in cell values.

       Returns:
       - Workbook: The Openpyxl Workbook object after deleting the footer rows.

       Note:
       This function iterates through the specified range of rows (from end to the starting row) in the active sheet.
       It checks if the cell value in the specified reference column is None or "None" and deletes the row accordingly.

    """
    sheet = wb.active
    for x in range(sheet.max_row, start, -1):  # iterating through table data end row to max row of sheet
        if sheet[f"{refColumn}{x}"].value is None or "None" in str(sheet[f"{refColumn}{x}"].value):  # if cell value is None or string "None"
            sheet.delete_rows(x)  # delete row
        else:
            break
    return wb


def removeNoneRows(wb, start, end, column):
    """
       Remove rows in an Openpyxl Workbook where the cell value is None in the specified reference column.

       Parameters:
       - wb (Workbook): The Openpyxl Workbook object.
       - start (int): The starting row index for iteration.
       - end (int): The ending row index (exclusive) for iteration.
       - column (str): The reference column letter (e.g., 'A') to check for None in cell values.

       Returns:
       - Workbook: The Openpyxl Workbook object after removing rows with None in the specified column.

       Note:
       This function iterates through the specified range of rows (from end to start) in the active sheet.
       It checks if the cell value in the specified reference column is None and deletes the row accordingly.

    """
    sheet = wb.active
    for x in range(end - 1, start, -1):  # iterating through table data end row to table data start row
        if sheet[f"{column}{x}"].value is None:  # if cell value is none
            sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merge consecutive rows in the specified column based on a reference column.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The starting row index for iteration.
        - end (int): The ending row index (exclusive) for iteration.
        - refColumn (str): The reference column letter (e.g., 'A') to check for starting rows.
        - mergingColumn (str): The column letter (e.g., 'B') in which rows will be merged.

        Returns:
        - Workbook: The Openpyxl Workbook object after merging consecutive rows in the specified column.

        Note:
        This function iterates through the specified range of rows in the active sheet.
        It merges consecutive rows in the specified column ('mergingColumn') based on the presence of data in the reference column ('refColumn').

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through table data start row to table data end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def deleteRowsByRange(wb, start, end, startText, stopText, startRefcolumn, stopRefColumn):
    """
        Delete rows in a specified range based on start and stop conditions.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The starting row index for iteration.
        - end (int): The ending row index (exclusive) for iteration.
        - startText (str): The text to identify the starting row for deletion.
        - stopText (str): The text to identify the stopping row for deletion.
        - startRefColumn (str): The reference column letter (e.g., 'A') to check for starting text.
        - stopRefColumn (str): The reference column letter (e.g., 'B') to check for stopping text.

        Returns:
        - Workbook: The Openpyxl Workbook object after deleting rows within the specified range.

        Note:
        This function iterates through the specified range of rows in the active sheet.
        It deletes rows based on the conditions specified by the start and stop texts in the respective reference columns.

    """
    sheet = wb.active
    delete_flag = False
    rows_to_delete = []
    for i in range(start, end):  # iterating through table data start row to table data end row
        if startText in str(sheet[f"{startRefcolumn}{i}"].value):  # if start text is in the reference column cell it's the starting row
            delete_flag = True  # make delete flag true
        if delete_flag:  # if delete flag is true append the row to rows_to_delete array
            rows_to_delete.append(i)  # append row to rows_to_delete array
        if stopText in str(sheet[f"{stopRefColumn}{i}"].value):  # if stop text is in the reference column cell it's the last row
            delete_flag = False  # make delete flag false
    for x in reversed(rows_to_delete):  # iterate array in reversed order to avoid the index problem while deleting the rows
        sheet.delete_rows(x)  # delete row
    return wb


def indusind1_validation(wb):
    """
        Validate columns for the core logic.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.

        Returns:
        - bool: True if the column count doesn't match the expected count, False otherwise.

        Note:
        This function checks if the number of columns in the active sheet matches the expected count for the designed core logic.
        It returns True if the count doesn't match, indicating an invalid format, and False otherwise.

    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 6  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def indusind1_main(wb):
    """
        Perform data cleaning and transformation for IndusInd Bank statement.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.

        Returns:
        - dict: A dictionary containing cleaned workbook and message.

        Note:
        This function expects an Openpyxl Workbook object (`wb`) containing a statement from IndusInd Bank.
        It performs various data cleaning and transformation operations, including deleting unwanted rows,
        merging misaligned rows, converting date formats, standardizing column names, and creating additional columns.
        The result is stored in a new workbook, and a response dictionary is returned with the cleaned workbook (`data`)
        and a message (`msg`) indicating the success of the process.

    """
    sheet = wb.active
    if indusind1_validation(wb):  # validate columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return response with error msg
    else:
        startText = "Date"  # header text to define table data start row
        endText = ""  # reference text to define table data end row -> hear if empty string means max row of sheet
        startEndRefColumn = "A"  # column contains the start and end text
        deleteFlagStartText1 = "Page"  # starting row reference text to delete rows by range
        deleteFlagStopText1 = "Account No"  # ending row reference text to delete rows by range
        deleteFlagStartTextRefColumn1 = "A"  # column contains starting row reference text to delete the rows by range
        deleteFlagStopTextRefColumn1 = "B"  # column contains ending row reference text to delete the rows by range
        columnToMerg = "B"  # column to merge misaligned rows
        refColumnToMerg = "A"  # reference column to merge rows of other columns
        deleteFooterRefColumn = "B"  # reference column to delete footer data
        dateConversionColumn1 = "A"  # column to convert date to standard date formate
        refHeaderText1 = "Date"  # header text to replace with standardised column name
        refHeaderText2 = "Particulars"  # header text to replace with standardised column name
        refHeaderText3 = "Chq./Ref. No"  # header text to replace with standardised column name
        refHeaderText4 = "WithDrawal"  # header text to replace with standardised column name
        refHeaderText5 = "Deposit"  # header text to replace with standardised column name
        refHeaderText6 = "Balance"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Narration"  # standard column name
        headerText3 = "ChequeNo_RefNo"  # standard column name
        headerText4 = "Withdrawal"  # standard column name
        headerText5 = "Deposit"  # standard column name
        headerText6 = "Balance"  # standard column name
        stringAlignColumn1 = "B"  # column to align string
        refTextToRemoveRow1 = "Brought Forward"  # reference text to remove row
        columnToRemoveRow = "B"  # column containing reference text to remove row
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # assign sheet max row to end
        dupHeaderDeleted1 = deleteRowsByRange(wb, start, end, deleteFlagStartText1, deleteFlagStopText1, deleteFlagStartTextRefColumn1, deleteFlagStopTextRefColumn1)  # deleting rows on range
        start, end = Excel.get_start_end_row_index(dupHeaderDeleted1, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # assign sheet max row to end
        dupHeaderDeleted2 = deleteRowsByRange(dupHeaderDeleted1, start, end, deleteFlagStartText1, deleteFlagStopText1, deleteFlagStopTextRefColumn1, deleteFlagStopTextRefColumn1)  # deleting rows on range
        start, end = Excel.get_start_end_row_index(dupHeaderDeleted2, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # assign sheet max row to end
        mergedColumnB = mergingRows(dupHeaderDeleted2, start, end, refColumnToMerg, columnToMerg)  # merging the rows of desired column
        noneRowsRemoved = removeNoneRows(mergedColumnB, start, end, refColumnToMerg)  # removing row if cell value is None in reference column
        footerDeleted = deleteFooter(noneRowsRemoved, start, deleteFooterRefColumn)  # deleting footer rows
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # deleting header rows, start-1 to Skip The Header
        start, end = Excel.get_start_end_row_index(dupHeaderDeleted2, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        end = sheet.max_row  # assign sheet max row to end
        convertedDateA = dateConversion(headerDeleted, start + 1, end + 1, dateConversionColumn1)  # converting date to standard date formate by columns
        lastCol = 65 + Excel.column_count(wb)  # 65 -> ASCII value
        transdate = Excel.alter_header_name(convertedDateA, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
        narration = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        chqno = Excel.alter_header_name(narration, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        debit = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        credit = Excel.alter_header_name(debit, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(credit, refHeaderText6, headerText6, lastCol)  # alter header name from the excel file to the standard column name
        alignedStringB = Excel.string_align(balance, start, end + 1, stringAlignColumn1)  # aligning string in column by removing the \n from the string -> \n -> next line
        broughtForwardRowRemoved = Excel.remove_row(alignedStringB, start, end + 1, refTextToRemoveRow1, columnToRemoveRow)  # removing row by reference text
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 ASCII value
        slCreated = Excel.create_slno_column(broughtForwardRowRemoved, start, end + 1, chr(columnToCreateSlNo))
        columnFinalised = Excel.finalise_column(slCreated, columns)  # standardizing count of column
        createdTransTypeColumn = Excel.transaction_type_column(columnFinalised)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/Senthil_indusind_pdf.io__23-09-2023-14-19-31.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    result = indusind1_main(wb)
    # result.save("C:/Users/Admin/Desktop/FinalOutput/INDUSIND1output.xlsx")
