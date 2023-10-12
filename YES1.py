import os
from datetime import datetime

import openpyxl

from KSV.FormatingExcelFiles.CommonClass import Excel


def makeNone(wb, start, end, column):
    sheet = wb.active
    for x in range(start, end):
        if len(sheet[f"{column}{x}"].value) < 1:
            sheet[f"{column}{x}"].value = None
    return wb


def removeNone(wb, start, end, column):
    sheet = wb.active
    for x in range(start, end):
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")
    return wb


def dateConversion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d %b %Y").date()
    return wb


def deleteHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def deleteFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


def removeNoneRows(wb, start, end, column):
    sheet = wb.active
    for x in range(end - 1, start, -1):
        if sheet[f"{column}{x}"].value is None:
            sheet.delete_rows(x)
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def yes1_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 7
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False

#todo chequeno_refno column is Encoded eg. [1.387420221031e+24]
def yes1_main(wb):
    sheet = wb.active
    if yes1_validation(wb):
        raise Exception(f"<= INVALID FORMATE =>  <Count Of Column Mismatch>")
    else:
        startText = """TransactionDate"""
        endText = "Opening Balance"
        startEndRefColumn = "A"
        deleteFlagStartText = "Customer Id"
        deleteFlagStopText = """TransactionDate"""
        deleteFlagRefColumn = "A"
        columnToMerg = "D"
        refColumnToMerg = "A"
        refColumnToRemoveNoneRows = "A"
        dateConversionColumn1 = "A"
        dateConversionColumn2 = "B"
        stringAlignColumn1 = "A"
        stringAlignColumn2 = "B"
        stringAlignColumn3 = "C"
        stringAlignColumn4 = "D"
        stringAlignColumn5 = "E"
        stringAlignColumn6 = "F"
        stringAlignColumn7 = "G"
        makeNoneRefColumn1 = "A"
        refHeaderText1 = "TransactionDate"
        refHeaderText2 = "Value Date"
        refHeaderText3 = "Cheque No/Reference No"
        refHeaderText4 = "Description"
        refHeaderText5 = "Withdrawals"
        refHeaderText6 = "Deposits"
        refHeaderText7 = "Running Balance"
        headerText1 = "Transaction_Date"
        headerText2 = "Value_Date"
        headerText3 = "ChequeNo_RefNo"
        headerText4 = "Narration"
        headerText5 = "Withdrawal"
        headerText6 = "Deposit"
        headerText7 = "Balance"
        headerTextToReplaceToNone1 = "Withdrawal"
        headerTextToReplaceToNone2 = "Deposit"
        start = 1
        end = sheet.max_row
        alignedStringA = Excel.string_align(wb, start, end, stringAlignColumn1)
        alignedStringB = Excel.string_align(alignedStringA, start, end, stringAlignColumn2)
        alignedStringC = Excel.string_align(alignedStringB, start, end, stringAlignColumn3)
        alignedStringD = Excel.string_align(alignedStringC, start, end, stringAlignColumn4)
        alignedStringE = Excel.string_align(alignedStringD, start, end, stringAlignColumn5)
        alignedStringF = Excel.string_align(alignedStringE, start, end, stringAlignColumn6)
        alignedStringG = Excel.string_align(alignedStringF, start, end, stringAlignColumn7)
        noneRemovedA = removeNone(alignedStringG, start, end, stringAlignColumn1)
        noneRemovedB = removeNone(noneRemovedA, start, end, stringAlignColumn2)
        noneRemovedC = removeNone(noneRemovedB, start, end, stringAlignColumn3)
        noneRemovedD = removeNone(noneRemovedC, start, end, stringAlignColumn4)
        noneRemovedE = removeNone(noneRemovedD, start, end, stringAlignColumn5)
        noneRemovedF = removeNone(noneRemovedE, start, end, stringAlignColumn6)
        noneRemovedG = removeNone(noneRemovedF, start, end, stringAlignColumn7)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        dupHeaderRemoved = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagStopText, deleteFlagRefColumn)
        start, end = Excel.get_start_end_row_index(dupHeaderRemoved, startText, endText, startEndRefColumn)
        madeNoneA = makeNone(dupHeaderRemoved, start, end, makeNoneRefColumn1)
        mergedColumnD = mergingRows(madeNoneA, start, end, refColumnToMerg, columnToMerg)
        noneStringRemovedD = removeNone(mergedColumnD, start, end, columnToMerg)
        noneRowsRemoved = removeNoneRows(noneStringRemovedD, start, end, refColumnToRemoveNoneRows)
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)
        footerDeleted = deleteFooter(noneRowsRemoved, end - 1)  # end-1 to Inclide End Footer
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndRefColumn)
        end = sheet.max_row
        convertedDateA = dateConversion(headerDeleted, start + 1, end + 1, dateConversionColumn1)  # start+1 to Skip Header, end+1 to IncludeLast Row
        convertedDateB = dateConversion(convertedDateA, start + 1, end + 1, dateConversionColumn2)  # start+1 to Skip Header, end+1 to IncludeLast Row
        lastCol = 65 + Excel.column_count(wb)
        transdate = Excel.alter_header_name(convertedDateB, refHeaderText1, headerText1, lastCol)
        valuedate = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)
        chqno = Excel.alter_header_name(valuedate, refHeaderText3, headerText3, lastCol)
        narration = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)
        debit = Excel.alter_header_name(narration, refHeaderText5, headerText5, lastCol)
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        slCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))
        replacedNoneWITHDRAWAL = Excel.empty_cell_to_none(slCreated, start, end + 1, headerTextToReplaceToNone1)
        replacedNoneDEPOSIT = Excel.empty_cell_to_none(replacedNoneWITHDRAWAL, start, end + 1, headerTextToReplaceToNone2)
        return wb


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/8._YES_bank_-_8241_Aug-Oct__23-09-2023-16-16-58.xlsx"
    wb = openpyxl.load_workbook(path)
    result = yes1_main(wb)
    result.save("C:/Users/Admin/Desktop/FinalOutput/YES1output.xlsx")
