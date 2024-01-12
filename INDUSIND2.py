from datetime import datetime

import openpyxl

from CommonClass import Excel


def replace_to_none(wb, start, end, refText, column):
    """
        Replace cells with a specific reference string to None.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The starting row index.
        - end (int): The ending row index.
        - refText (str): The reference string to search for in the specified column.
        - column (str): The column letter (e.g., "A", "B", etc.) where replacement will be performed.

        Returns:
        - Workbook: The Openpyxl Workbook object after performing the replacements.

        Notes:
        - This function iterates through the specified range of rows in the given column and replaces occurrences of the
          specified reference string with None.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        if refText in str(sheet[f"{column}{i}"].value):  # if reference text in cell value
            sheet[f"{column}{i}"].value = None  # assign None to cell value
    return wb


def deleteHeader(wb, start):
    """
        Delete rows from the specified start row to the 1st row of the sheet.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - start (int): The starting row index.

        Returns:
        - Workbook: The Openpyxl Workbook object after deleting the specified rows.

        Notes:
        - This function iterates through the rows from the specified start index to the 1st row of the sheet and deletes each row.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to 1st row off sheet
        sheet.delete_rows(x)  # delete row
    return wb


def deleteFooter(wb, end):
    """
        Delete rows from the specified end row to the max row of the sheet.

        Parameters:
        - wb (Workbook): The Openpyxl Workbook object.
        - end (int): The ending row index.

        Returns:
        - Workbook: The Openpyxl Workbook object after deleting the specified rows.

        Notes:
        - This function iterates through the rows from the max row of the sheet to the specified end index and deletes each row.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through max row of sheet to table data end row
        sheet.delete_rows(x)  # deleting row
    return wb


def dateConversion(wb, start, end, column):
    """
        Converts date values in a specified column of an openpyxl Workbook object to a standard date format.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index where the conversion should start (inclusive).
        - end (int): The row index where the conversion should end (exclusive).
        - column (str): The column letter (e.g., 'A', 'B', 'C') containing the date values.

        Returns:
        - openpyxl.Workbook: The modified Workbook object with the date values converted.

        Note:
        The function uses the datetime.strptime method to parse the date values in the specified column
        and then updates the Workbook with the parsed date values in a standard date format ("%b %d, %Y").

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%b %d, %Y").date()  # converting to standard date formate
    return wb


def indusind2_validation(wb):
    """
        Validates the column count in the active sheet of an openpyxl Workbook object for a specific core logic.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object to be validated.

        Returns:
        - bool: True if the number of columns in the active sheet is not equal to the expected core logic column count (6),
                False otherwise.

        Note:
        The function checks if the maximum number of columns in the active sheet is not equal to the expected core logic column count (6).
        If the condition is met, it returns True, indicating a validation failure. Otherwise, it returns False, indicating a successful validation.

    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 6  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def indusind2_main(wb):
    """
        Performs data processing and validation for a specific format in the active sheet of an openpyxl Workbook object.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data to be processed.

        Returns:
        - dict: A dictionary containing processed data and a message.
            - "data": The modified Workbook object.
            - "msg": A message indicating the result of the processing.

        Note:
        The function performs various operations on the provided Workbook object to standardize and process the data.
        It involves validation, header renaming, column deletion, date conversion, string alignment, and other operations.

        The processed data is returned along with a message indicating the success or failure of the processing.

    """
    sheet = wb.active
    if indusind2_validation(wb):  # validate columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return response with error msg
    else:
        startText = "Date"  # header text to define table data start row
        endText = "This is a computer generated statement"  # text defining table data end row
        startEndDefColumn = "A"  # column containing start and end text
        deleteFlagStartText1 = "Page"  # starting row reference text to delete rows by range
        deleteFlagStopText1 = "Balance"  # ending row reference text to delete rows by range
        deleteFlagRefColumn1 = "F"  # column contains starting row reference text and ending row reference text to delete the rows by range
        deleteFlagStartText2 = "Page"  # starting row reference text to delete rows by range
        deleteFlagStopText2 = "Credit"  # ending row reference text to delete rows by range
        deleteFlagRefColumn2 = "E"  # column contains starting row reference text and ending row reference text to delete the rows by range
        dateConversionColumn1 = "A"  # column to convert date to standard date formate
        stringAlignColumn1 = "C"  # column to aligning string by removing the \n from string
        negativeValueColumnRefText1 = "Withdrawal"  # no need to convert negative value to positive
        refTextToReplace = "-"  # reference text to replace with none
        refColumnToReplaceText1 = "C"  # reference column to replace reference text with none
        refColumnToReplaceText2 = "D"  # reference column to replace reference text with none
        refTextToDeleteColumn1 = "Type"  # reference header text to delete column
        refHeaderText1 = "Date"  # header text to replace with standardised column name
        refHeaderText2 = "Description"  # header text to replace with standardised column name
        refHeaderText3 = "Debit"  # header text to replace with standardised column name
        refHeaderText4 = "Credit"  # header text to replace with standardised column name
        refHeaderText5 = "Balance"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Narration"  # standard column name
        headerText3 = "Withdrawal"  # standard column name
        headerText4 = "Deposit"  # standard column name
        headerText5 = "Balance"  # standard column name
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify data with in
        dupHeaderRemoved1 = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText1, deleteFlagStopText1, deleteFlagRefColumn1)  # deleting rows on range
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify data with in
        dupHeaderRemoved2 = Excel.delete_rows_by_range(dupHeaderRemoved1, start, end, deleteFlagStartText2, deleteFlagStopText2, deleteFlagRefColumn2)  # deleting rows on range
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify data with in
        convertedDateA = dateConversion(dupHeaderRemoved2, start + 1, end, dateConversionColumn1)  # converting date to standard date formate
        alignedStringC = Excel.string_align(convertedDateA, start, end, stringAlignColumn1)  # aligning string in column by removing the \n from the string -> \n -> next line
        footerDeleted = deleteFooter(alignedStringC, end - 1)  # deleting footer from table data end row to max row of sheet
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # deleting rows from table data start row to 1st row of sheet
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify data with in
        deletedColumnTYPE = Excel.delete_column(headerDeleted, refTextToDeleteColumn1)  # deleting header by reference header name
        lastCol = 65 + Excel.column_count(wb)  # 65 -> ASCII value
        transdate = Excel.alter_header_name(deletedColumnTYPE, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
        narration = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        debit = Excel.alter_header_name(narration, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        credit = Excel.alter_header_name(debit, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(credit, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
        replacedNoneD = replace_to_none(balance, start, end + 1, refTextToReplace, refColumnToReplaceText1)  # replacing cell with reference string to none
        replacedNoneE = replace_to_none(replacedNoneD, start, end + 1, refTextToReplace, refColumnToReplaceText2)  # replacing cell with reference string to none
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 -> ASCII value
        slCreated = Excel.create_slno_column(replacedNoneE, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
        columnFinalised = Excel.finalise_column(slCreated, columns)  # standardizing column count
        createdTransTypeColumn = Excel.transaction_type_column(columnFinalised)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/1._Indusind_-_2673__07-10-2023-12-03-48.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    wb = indusind2_main(wb)
    # wb["data"].save("C:/Users/Admin/Desktop/FinalOutput/INDUSIND2output.xlsx")
