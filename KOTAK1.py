import os
from datetime import datetime
from openpyxl.utils import column_index_from_string
import openpyxl
from FormatingExcelFiles.AlignmentData import addAlignmentData
from FormatingExcelFiles.CommonClass import Excel


def removeString(wb, start, end, refText, column):
    sheet = wb.active
    for x in range(start, end):
        if refText in str(sheet[f"{column}{x}"].value):
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace(refText, "")
    return wb


def dateConvertion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%b-%y").date()
    return wb


def deleteHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def alignColumns(wb, start, end, headData, refColumnToAlign):
    sheet = wb.active
    error_records = []
    for i in range(start, end):
        if sheet[f"G{i}"].value is not None:
            data = [headData[0], headData[1], headData[2]]
            data.append(i - 1)
            data.append(str(sheet[f"A{i}"].value))
            data.append(str(sheet[f"B{i}"].value))
            data.append(str(sheet[f"C{i}"].value))
            data.append(str(sheet[f"D{i}"].value))
            data.append(str(sheet[f"E{i}"].value))
            data.append(str(sheet[f"F{i}"].value))
            data.append(str(sheet[f"G{i}"].value))
            addAlignmentData(data)
            error_records.append(i)
    for x in error_records:
        sheet[f"B{x}"].value = "Error Record"
        sheet[f"C{x}"].value = None
        sheet[f"D{x}"].value = None
        sheet[f"E{x}"].value = None
        sheet[f"F{x}"].value = None
        sheet[f"G{x}"].value = None


# TODO: get name , acnum - use num vs char instead of digit count
def headerData(wb, start, end):
    sheet = wb.active
    acnum = "Undefined"
    name = "Undefined"
    period = "Undefined"
    for i in range(start, 0, -1):
        if "Period" in str(sheet[f"D{i}"].value):
            period = str(sheet[f"E{i}"].value)
        if "Account No" in str(sheet[f"A{i}"].value):
            spl = str(sheet[f"A{i}"].value).split("Account No.")
            a = spl[1].strip().replace('\n', '').split(" ")
            acnum = ""
            for char in a[0]:
                if char.isdigit():
                    acnum += char
    headData = [acnum, name, period]
    return headData


def deleteRowByDateLen(wb, start, dateLen, refColumn):
    sheet = wb.active
    for x in range(sheet.max_row, start, -1):
        if len(str(sheet[f"{refColumn}{x}"].value)) < dateLen:
            sheet.delete_rows(x)
    return wb


def removeFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


def removeHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def removeNoneRows(wb, start, end, column):
    sheet = wb.active
    for x in range(end - 1, start, -1):
        if sheet[f"{column}{x}"].value is None:
            sheet.delete_rows(x)
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def kotak1_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 12
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def kotak1_main(wb):
    if kotak1_validation(wb):
        raise Exception(f"<= INVALID FORMATE =>  <Count Of Column Mismatch>")
    else:
        sheet = wb.active
        startText = "Chq/Ref No."
        endText = "Statement Summary"
        startEndRefColumn = "C"
        dupHeaderStartText = "Contd."
        dupHeaderEndText = "Page"
        dupHeaderRefColumn = "A"
        columnToMerg1 = "B"
        refColumnToMerg = "A"
        dateLength = 5
        deleteRowRefText1 = "OPENING BALANCE"
        deleteRowRefColumn = "B"
        refColumnToAlign = "G"
        dateConversionColumn1 = "A"
        refHeaderText1 = "Date"
        refHeaderText2 = "Narration"
        refHeaderText3 = "Chq/Ref No."
        refHeaderText4 = "Withdrawal (Dr)"
        refHeaderText5 = "Deposit (Cr)"
        refHeaderText6 = "Balance"
        headerText1 = "Transaction_Date"
        headerText2 = "Narration"
        headerText3 = "ChequeNo_RefNo"
        headerText4 = "Withdrawal"
        headerText5 = "Deposit"
        headerText6 = "Balance"
        refTextToRemove1 = "(Cr)"
        stringRemoveColumn1 = "F"
        refStringToRemove = ","
        columnToRemoveString1 = "F"
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        dupHeaderRemoved = Excel.delete_rows_by_range(wb, start, end, dupHeaderStartText, dupHeaderEndText, dupHeaderRefColumn)
        start, end = Excel.get_start_end_row_index(dupHeaderRemoved, startText, endText, startEndRefColumn)
        columnMergedB = mergingRows(dupHeaderRemoved, start, end, refColumnToMerg, columnToMerg1)
        noneRowsRemoved = removeNoneRows(columnMergedB, start, end, refColumnToMerg)
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)
        footerRemoved = removeFooter(noneRowsRemoved, end - 1)  # end-1 to Include End Footer
        start, end = Excel.get_start_end_row_index(footerRemoved, startText, endText, startEndRefColumn)
        nullRowDeleted = deleteRowByDateLen(footerRemoved, start, dateLength, refColumnToMerg)
        start, end = Excel.get_start_end_row_index(footerRemoved, startText, endText, startEndRefColumn)
        headData = headerData(nullRowDeleted, start, end)
        headerRemoved = removeHeader(nullRowDeleted, start - 1)  # start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerRemoved, startText, endText, startEndRefColumn)
        openbalDeleted = Excel.remove_row(headerRemoved, start, end, deleteRowRefText1, deleteRowRefColumn)
        alignColumns(openbalDeleted, start + 1, end, headData, refColumnToAlign)  # start+1 to Skip Header
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        slCreated = Excel.create_slno_column(openbalDeleted, start, end, chr(columnToCreateSlNo))
        start, end = Excel.get_start_end_row_index(slCreated, startText, endText, startEndRefColumn)
        convertedDateA = dateConvertion(slCreated, start + 1, end + 1, dateConversionColumn1)  # start+1 to Skip Header, end+1 to Include Last Row
        lastCol = 65 + sheet.max_column  # 65 => ASCII value "A"
        transdate = Excel.alter_header_name(convertedDateA, refHeaderText1, headerText1, lastCol)
        naration = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)
        chqno = Excel.alter_header_name(naration, refHeaderText3, headerText3, lastCol)
        debit = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)
        credit = Excel.alter_header_name(debit, refHeaderText5, headerText5, lastCol)
        balance = Excel.alter_header_name(credit, refHeaderText6, headerText6, lastCol)
        stringRemovedCR = removeString(balance, start, end + 1, refTextToRemove1, stringRemoveColumn1)
        columnFinalised = Excel.finalise_column(stringRemovedCR, columns)
        comaRemoved = Excel.remove_string(columnFinalised, start, end + 1, refStringToRemove, columnToRemoveString1)
        createdTransTypeColumn = Excel.transaction_type_column(comaRemoved)
        return wb


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/Kotak1._Apr-22_637102__06-09-2023-14-01-34.xlsx"
    wb = openpyxl.load_workbook(path)
    result = kotak1_main(wb)
    result.save("C:/Users/Admin/Desktop/FinalOutput/Kotak1output.xlsx")
