from datetime import datetime

import openpyxl

from KSV.FormatingExcelFiles.CommonClass import Excel


def seperate_debit_credit_column(wb, sourceColumn, withdrawal, deposit):
    sheet = wb.active
    for i in range(2, sheet.max_row + 1):
        if "Dr" in str(sheet[f"{sourceColumn}{i}"].value):
            temp = str(sheet[f"{sourceColumn}{i}"].value).split("(")
            sheet[f"{withdrawal}{i}"].value = temp[0].replace(",", "")
        if "Cr" in str(sheet[f"{sourceColumn}{i}"].value):
            temp = str(sheet[f"{sourceColumn}{i}"].value).split("(")
            sheet[f"{deposit}{i}"].value = temp[0].replace(",", "")
    return wb


def createColumn(wb, columnName):
    sheet = wb.active
    sheet[f"{chr(65+sheet.max_column)}1"].value = columnName
    return wb


def dateConvertion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%m-%Y").date()
    return wb


def deleteHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def deleteFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


def removeNoneRows(wb, start, end, column):
    sheet = wb.active
    for x in range(end - 1, start, -1):
        if sheet[f"{column}{x}"].value is None:
            sheet.delete_rows(x)
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def kotak2_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 5
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def kotak2_main(wb):
    if kotak2_validation(wb):
        raise Exception(f"<= INVALID FORMATE =>  <Count Of Column Mismatch>")
    else:
        sheet = wb.active
        startText = "Date"
        stopText = "Statement  Summary"
        startEndRefColumn = "A"
        deleteFlagStartText = "Period"
        deleteFlagEndText = "Narration"
        deleteFlagRefColumn = "B"
        columnToMerg1 = "B"
        columnToMerg2 = "C"
        refColumnToMerg = "A"
        refTextToRemove = "None"
        dateConversionColumn1 = "A"
        newColumnName1 = "Withdrawal"
        newColumnName2 = "Deposit"
        sourceDataColumn = "D"
        withdrawalColumn = "F"
        depositColumn = "G"
        refTextToDeleteColumn = "Withdrawal (Dr)"
        refStringFromBalanceColumn = "(Cr)"
        removeStringRefColumn = "D"
        refHeaderText1 = "Date"
        refHeaderText2 = "Narration"
        refHeaderText3 = "Chq/Ref No"
        refHeaderText4 = "Balance"
        headerText1 = "Transaction_Date"
        headerText2 = "Narration"
        headerText3 = "ChequeNo_RefNo"
        headerText4 = "Balance"
        columns = ["Sl.No.", "Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]
        refStringToRemove = ","
        columnToRemoveString1 = "D"
        columnToAlignString1 = "B"
        headerToReplaceEmptyCellToNone1 = "ChequeNo_RefNo"
        start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)
        dupHeaderRemoved = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagEndText, deleteFlagRefColumn)
        start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)
        mergedColumnB = mergingRows(dupHeaderRemoved, start, end, refColumnToMerg, columnToMerg1)
        mergedColumnC = mergingRows(mergedColumnB, start, end, refColumnToMerg, columnToMerg2)
        noneRowsRemoved = removeNoneRows(mergedColumnC, start, end, refColumnToMerg)
        start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)
        removedNoneB = Excel.remove_string(noneRowsRemoved, start, end, refTextToRemove, columnToMerg1)
        removedNoneC = Excel.remove_string(removedNoneB, start, end, refTextToRemove, columnToMerg2)
        footerDeleted = deleteFooter(removedNoneC, end - 1)
        headerDeleted = deleteHeader(footerDeleted, start - 1)
        start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)
        convertedDateA = dateConvertion(headerDeleted, start + 1, end + 1, dateConversionColumn1)
        debitCreated = createColumn(convertedDateA, newColumnName1)
        creditCreated = createColumn(convertedDateA, newColumnName2)
        dataSeperated = seperate_debit_credit_column(wb, sourceDataColumn, withdrawalColumn, depositColumn)
        sourceDataColumnDeleted = Excel.delete_column(dataSeperated, refTextToDeleteColumn)
        CRremoved = Excel.remove_string(sourceDataColumnDeleted, start + 1, end + 1, refStringFromBalanceColumn, removeStringRefColumn)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        slnoCreated = Excel.create_slno_column(CRremoved, start, end + 1, chr(columnToCreateSlNo))
        lastCol = 65 + sheet.max_column  # 65 => ASCII value "A"
        transdate = Excel.alter_header_name(convertedDateA, refHeaderText1, headerText1, lastCol)
        narration = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)
        chqno = Excel.alter_header_name(narration, refHeaderText3, headerText3, lastCol)
        balance = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)
        finalisedColumns = Excel.finalise_column(balance, columns)
        comaRemoved = Excel.remove_string(finalisedColumns, start, end + 1, refStringToRemove, columnToRemoveString1)
        stringAlignedB = Excel.string_align(comaRemoved, start, end + 1, columnToAlignString1)
        replacedToNoneCHQNO = Excel.empty_cell_to_none(stringAlignedB, start, end + 1, headerToReplaceEmptyCellToNone1)
        return wb


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/Kotak_-_5887.xlsx"
    wb = openpyxl.load_workbook(path)
    result = kotak2_main(wb)
    result.save("C:/Users/Admin/Desktop/FinalOutput/Kotak2output.xlsx")