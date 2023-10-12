import os
from datetime import datetime

import openpyxl

from KSV.FormatingExcelFiles.CommonClass import Excel


def removeNone(wb, start, end, column):
    sheet = wb.active
    for x in range(start, end):
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")
    return wb


def deleteHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def dateConversion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(sheet[f"{column}{i}"].value, "%d/%m/%Y").date()
    return wb


def deleteFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


def removingNoneRows(wb, start, end, refColumn):
    sheet = wb.active
    for x in range(end, start, -1):
        a_cell = f"{refColumn}{x}"
        if sheet[a_cell].value is None:
            sheet.delete_rows(x)
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def dbs1_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 9
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def dbs1_main(wb):
    sheet = wb.active
    if dbs1_validation(wb):
        raise Exception(f"<= INVALID FORMATE =>  <Count Of Column Mismatch>")
    else:
        startText = "Transaction date"
        endText = "Summary"
        startEndRefColumn = "A"
        deleteFlagStartText = "DBS Bank India Ltd."
        deleteFlagStopText = "Transaction date"
        deleteFlagRefColumn = "A"
        columnToMerg1 = "D"
        refColumnToMerg = "A"
        dateConversionColumn1 = "A"
        dateConversionColumn2 = "B"
        stringAlignColumn1 = "D"
        stringAlignColumn2 = "E"
        deleteColumnRefText = "Branch code"
        refHeaderText1 = "Transaction date"
        refHeaderText2 = "Value date"
        refHeaderText3 = "Description"
        refHeaderText4 = "Cheque/Reference number"
        refHeaderText5 = "Debit"
        refHeaderText6 = "Credit"
        refHeaderText7 = "Balance"
        headerText1 = "Transaction_Date"
        headerText2 = "Value_Date"
        headerText3 = "Narration"
        headerText4 = "ChequeNo_RefNo"
        headerText5 = "Withdrawal"
        headerText6 = "Deposit"
        headerText7 = "Balance"
        negativeValueColumnRefText1 = "Withdrawal"
        headerTextToReplaceEmptyCellToNone1 = "ChequeNo_RefNo"
        headerTextToReplaceEmptyCellToNone2 = "Withdrawal"
        headerTextToReplaceEmptyCellToNone3 = "Deposit"
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        duplicateHeaderRemoved = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagStopText, deleteFlagRefColumn)
        start, end = Excel.get_start_end_row_index(duplicateHeaderRemoved, startText, endText, startEndRefColumn)
        rowsMergedD = mergingRows(duplicateHeaderRemoved, start, end, refColumnToMerg, columnToMerg1)
        noneRowsRemoved = removingNoneRows(rowsMergedD, start, end, refColumnToMerg)
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)
        dateConvertedA = dateConversion(noneRowsRemoved, start + 1, end, dateConversionColumn1)  # start-1 to Skip Header
        dateConvertedB = dateConversion(dateConvertedA, start + 1, end, dateConversionColumn2)  # start-1 to Skip Header
        footerDeleted = deleteFooter(dateConvertedB, end - 1)  # end-1 to Include End Footer
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)
        alignedStringD = Excel.string_align(headerDeleted, start, end + 1, stringAlignColumn1)  # end+1 to Include Last Row
        alignedStringE = Excel.string_align(alignedStringD, start, end + 1, stringAlignColumn2)  # end+1 to Include Last Row
        removedNoneD = removeNone(alignedStringE, start, end + 1, stringAlignColumn1)  # end+1 to Include Last Row
        removedNoneE = removeNone(removedNoneD, start, end + 1, stringAlignColumn2)  # end+1 to Include Last Row
        branchCodeDeleted = Excel.delete_column(removedNoneE, deleteColumnRefText)
        lastCol = 65 + Excel.column_count(wb)
        transdate = Excel.alter_header_name(branchCodeDeleted, refHeaderText1, headerText1, lastCol)
        valuedate = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)
        naration = Excel.alter_header_name(valuedate, refHeaderText3, headerText3, lastCol)
        chqno = Excel.alter_header_name(naration, refHeaderText4, headerText4, lastCol)
        debit = Excel.alter_header_name(chqno, refHeaderText5, headerText5, lastCol)
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        slnoCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))
        negativeValueChecked = Excel.check_neagativeValue_by_column(slnoCreated, negativeValueColumnRefText1)
        replacedNoneCHQNO = Excel.empty_cell_to_none(negativeValueChecked, start, end + 1, headerTextToReplaceEmptyCellToNone1)
        replacedNoneWITHDRAWAL = Excel.empty_cell_to_none(replacedNoneCHQNO, start, end + 1, headerTextToReplaceEmptyCellToNone2)
        replacedNoneDEPOSIT = Excel.empty_cell_to_none(replacedNoneWITHDRAWAL, start, end + 1, headerTextToReplaceEmptyCellToNone3)
        return wb


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/LVB_-_0145P.W_-_1L1675876_unlocked__12-09-2023-15-56-14.xlsx"
    wb = openpyxl.load_workbook(path)
    result = dbs1_main(wb)
    result.save('C:/Users/Admin/Desktop/FinalOutput/DBS1output.xlsx')
