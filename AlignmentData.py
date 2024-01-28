import openpyxl




def addAlignmentData(data):
    """
        Adds misaligned data to an existing Excel file.

        Parameters:
        - data (list): A list containing the misaligned data to be added to the Excel file.

        Notes:
        - The function appends the provided misaligned data to an existing Excel file.
        - The Excel file path is hardcoded in the function. Modify the 'path' variable to match your file path.
        - The data is added to the next available row in the active sheet.
        - The function uses ASCII values to determine the column in which the data will be added.
        - After adding the data, the Excel file is saved with the new changes.

    """
    path = "C:/Users/Admin/Desktop/KSV/Python/AlignmentData.xlsx"  # path of the Excel file
    wb = openpyxl.load_workbook(path)  # load the work book
    sheet = wb.active  # get the active sheet
    row = sheet.max_row+1  # get last row new row
    column = 65  # ascii value of column A
    count = 0
    columnCount = len(data)
    for i in range(0, len(data)):  # iterate through data array length
        sheet[f"{chr(column)}{row}"].value = data[i]  # assigning the miss aligned data to Excel sheet
        column += 1  # by incrementing the column by 1
    wb.save("C:/Users/Admin/Desktop/KSV/Python/AlignmentData.xlsx")  # save the Excel file


def column_count(wb):
    sheet = wb.active
    column = 65  # ASCII value of "A"
    count = 0
    for i in range(column, column + sheet.max_column):
        if sheet[f"{chr(i)}1"].value is None and len(str(sheet[f"{chr(i)}1"].value)) > 0:
            break
        count += 1
    return count


def find_column_index_by_header(wb, header):
    sheet = wb.active
    column_index = None
    for column in range(65, 65 + sheet.max_column):  # iterating through all the columns using ascii values
        if header in str(sheet[f"{chr(column)}1"].value):  # if a header text is in the 1st cell of the column
            column_index = column  # store the column index
            break
    return column_index


def find_replace(wb, start, end, find, replace, columns):
    sheet = wb.active
    for column in columns:
        for row in range(start, end):
            if find in str(sheet[f"{column}{row}"].value):
                sheet[f"{column}{row}"].value = str(sheet[f"{column}{row}"].value).replace(find, replace)
    return wb


def negative_value_in_debit_to_credit(wb, start, end, debit_column, credit_column):
    sheet = wb.active
    for row in range(start, end):
        if "-" in str(sheet[f"{debit_column}{row}"].value):
            sheet[f"{credit_column}{row}"].value = str(sheet[f"{debit_column}{row}"].value).replace("-", "")
            sheet[f"{debit_column}{row}"].value = 0
    return wb


def remove_values_from_misaligned_column(wb, start, end, to_column, from_column):
    sheet = wb.active
    for row in range(start, end):
        if sheet[f"{to_column}{row}"].value is not None and len(str(sheet[f"{to_column}{row}"].value)) > 0:
            sheet[f"{from_column}{row}"].value = None
    return wb


def canara1_column_alignment(wb, start, end, data_list):
    sheet = wb.active
    fromcol = set(item["From_Column"] for item in data_list if item["Operation"] == "align_column_data")
    fromcolumn = list(fromcol)[0]
    tocol = set(item["To_Column"] for item in data_list if item["Operation"] == "align_column_data")
    tocolumn = list(tocol)[0]
    debit_column = find_column_index_by_header(wb, "WITHDRAWS")
    credit_column = find_column_index_by_header(wb, "DEPOSIT")
    balance_column = find_column_index_by_header(wb, "BALANCE")
    # print(chr(debit_column))
    # print(chr(credit_column))
    # print(chr(balance_column))
    # print(fromcolumn)
    # print(tocolumn)
    balance = 0
    debit = 0
    credit = 0

    wb = find_replace(wb, start, end, find=",", replace="", columns=[fromcolumn, chr(debit_column), chr(credit_column), chr(balance_column)])
    wb = find_replace(wb, start, end, find="None", replace="", columns=[fromcolumn, chr(debit_column), chr(credit_column), chr(balance_column)])
    wb = negative_value_in_debit_to_credit(wb, start, end, chr(debit_column), chr(credit_column))
    for row in range(start, end):
        # case 1
        if sheet[f"{fromcolumn}{row}"].value is not None and len(str(sheet[f"{fromcolumn}{row}"].value)) > 1 and sheet[f"{tocolumn}{row}"].value is None or len(str(sheet[f"{tocolumn}{row}"].value)) < 1:
            if float(sheet[f"{chr(credit_column)}{row}"].value) > 0:
                balance = float(sheet[f"{fromcolumn}{row}"].value) - float(sheet[f"{chr(credit_column)}{row}"].value)
            if float(sheet[f"{chr(debit_column)}{row}"].value) > 0:
                balance = float(sheet[f"{fromcolumn}{row}"].value) + float(sheet[f"{chr(debit_column)}{row}"].value)
            # print(f"{fromcolumn}{row}")
            # print("balance = ", balance)
            # print("withdrawal", float(sheet[f"{chr(debit_column)}{row}"].value))
            # print("deposit", float(sheet[f"{chr(credit_column)}{row}"].value))
            # if float(sheet[f"{chr(debit_column)}{row}"].value)
            is0 = balance - float(sheet[f"{chr(debit_column)}{row}"].value) + float(sheet[f"{chr(credit_column)}{row}"].value) - float(sheet[f"{fromcolumn}{row}"].value)
            num = format(is0, '.10f')
            is0 = float(num)
            if is0 == 0.0:
                sheet[f"{tocolumn}{row}"].value = sheet[f"{fromcolumn}{row}"].value

        # case 2
        if sheet[f"{fromcolumn}{row}"].value is not None and len(str(sheet[f"{fromcolumn}{row}"].value)) > 1 and sheet[f"{chr(balance_column)}{row}"].value is not None and len(str(sheet[f"{chr(balance_column)}{row}"].value)) > 1 and sheet[f"{chr(credit_column)}{row}"].value is None or len(str(sheet[f"{chr(credit_column)}{row}"].value)) < 1:
            if float(sheet[f"{chr(balance_column)}{row}"].value) > 0:
                balance = float(sheet[f"{fromcolumn}{row}"].value) - float(sheet[f"{chr(balance_column)}{row}"].value)
            if float(sheet[f"{chr(debit_column)}{row}"].value) > 0:
                balance = float(sheet[f"{fromcolumn}{row}"].value) - float(sheet[f"{chr(debit_column)}{row}"].value)
            is0 = balance - float(sheet[f"{chr(debit_column)}{row}"].value) + float(sheet[f"{chr(balance_column)}{row}"].value) - float(sheet[f"{fromcolumn}{row}"].value)
            num = format(is0, '.10f')
            is0 = float(num)
            if is0 == 0.0:
                sheet[f"{chr(credit_column)}{row}"].value = sheet[f"{chr(balance_column)}{row}"].value
                sheet[f"{tocolumn}{row}"].value = sheet[f"{fromcolumn}{row}"].value

        # case 3
        if sheet[f"{fromcolumn}{row}"].value is not None and len(str(sheet[f"{fromcolumn}{row}"].value)) > 1 and sheet[f"{chr(balance_column)}{row}"].value is not None and len(str(sheet[f"{chr(balance_column)}{row}"].value)) > 1 and sheet[f"{chr(credit_column)}{row}"].value is not None and len(str(sheet[f"{chr(credit_column)}{row}"].value)) > 1 and sheet[f"{chr(debit_column)}{row}"].value is None or len(str(sheet[f"{chr(debit_column)}{row}"].value)) < 1:
            if float(sheet[f"{chr(balance_column)}{row}"].value) > 0:
                balance = float(sheet[f"{fromcolumn}{row}"].value) - float(sheet[f"{chr(balance_column)}{row}"].value)
            if float(sheet[f"{chr(credit_column)}{row}"].value) > 0:
                balance = float(sheet[f"{fromcolumn}{row}"].value) - float(sheet[f"{chr(credit_column)}{row}"].value)
            is0 = balance - float(sheet[f"{chr(credit_column)}{row}"].value) + float(sheet[f"{chr(balance_column)}{row}"].value) - float(sheet[f"{fromcolumn}{row}"].value)
            num = format(is0, '.10f')
            is0 = float(num)
            if is0 == 0.0:
                sheet[f"{chr(debit_column)}{row}"].value = sheet[f"{chr(credit_column)}{row}"].value
                sheet[f"{chr(credit_column)}{row}"].value = sheet[f"{chr(balance_column)}{row}"].value
                sheet[f"{tocolumn}{row}"].value = sheet[f"{fromcolumn}{row}"].value

    remove_values_from_misaligned_column(wb, start, end, tocolumn, fromcolumn)
    return wb


def hdfc1_headerData(wb, start, end):
    sheet = wb.active
    s1 = "Undefined"  # variable to store header data
    s2 = "Undefined"  # variable to store header data
    s3 = "Undefined"  # variable to store header data
    s4 = "Undefined"  # variable to store header data
    s5 = "Undefined"  # variable to store header data
    s6 = "Undefined"  # variable to store header data
    s7 = "Undefined"  # variable to store header data
    s8 = "Undefined"  # variable to store header data
    for i in range(start, 0, -1):  # iterating through table data start row to sheet 1st column
        if sheet[f"B{i}"].value is not None and "Account No" in str(sheet[f"B{i}"].value):  # if cell value is not none and "Account No" string in cell value then
            s1 = sheet[f"D{i}"].value  # store value in variable
        if sheet[f"B{i}"].value is not None and "IFSC" in str(sheet[f"B{i}"].value):  # if the cell value is not none and "IFSC" string in cell value then
            s2 = sheet[f"D{i}"].value  # store value in variable
        if sheet[f"A{i}"].value is not None and "Statement From :" in str(sheet[f"A{i}"].value):  # if cell value is not none and "Statement From" string in cell value then
            s3 = f"{sheet[f'A{i}'].value} {sheet[f'B{i}'].value}"  # store value in variable
    spl1 = s1.split(":")  # extracting account number from the string
    a = spl1[3].strip().split(" ")  # extracting account number from the string
    acno = a[0]  # storing account number
    cusid = f"Customer ID : {spl1[2]}"  # storing customer id
    name = "Undefined"  # storing name
    ifsc = f"IFSC : {s2}"  # storing ifsc code
    period = s3  # storing statement period
    openbal = s4  # storing open balance
    closebal = s5  # storing closing balance
    debits = s6  # storing total debit
    credits = s7  # storing total credit
    headData = [acno, name, period]  # header data array
    return headData


def hdfc1_column_error_records(wb, start, end, data_list):
    sheet = wb.active
    refcolumn = set(item["Reference_Column"] for item in data_list if item["Operation"] == "align_error_record")
    refcolumn = list(refcolumn)[0]
    refcolumn = ord(refcolumn)  # convert int to ascii value
    # print(refcolumn)
    start_column = 65
    error_records = []
    for row in range(start, end):
        if sheet[f"{chr(refcolumn)}{row}"].value is not None:
            data = [row - 1]
            for column in range(start_column, refcolumn):
                data.append(str(sheet[f"{chr(column)}{row}"].value))
            addAlignmentData(data)
            error_records.append(row)

    narration_column = find_column_index_by_header(wb, header="Narration")
    date_column = find_column_index_by_header(wb, header="Date")
    for row in error_records:
        for column in range(start_column, refcolumn + 1):
            if column == narration_column:
                sheet[f"{chr(narration_column)}{row}"].value = "error record"
            if column != date_column and column != narration_column:
                sheet[f"{chr(column)}{row}"].value = None
    return wb


def empty_cell_to_0(wb, start, end, columns):
    sheet = wb.active
    # max_column = 65 + sheet.max_column
    # start_column = 65
    for column in columns:
        for row in range(start, end):  # iterating through all the rows from start to end row
            if len(str(sheet[f"{column}{row}"].value)) < 1:  # if cell value length is less than 1
                sheet[f"{column}{row}"].value = "0"  # make it as None
    return wb


def icici3_column_alignment(wb, start, end, data_list):
    fromcolumn = set(item["From_Column"] for item in data_list if item["Operation"] == "align_column_data")
    fromcolumn = list(fromcolumn)[0]
    tocolumn = set(item["To_Column"] for item in data_list if item["Operation"] == "align_column_data")
    tocolumn = list(tocolumn)[0]
    debit_column = find_column_index_by_header(wb, "DebitAmount")
    credit_column = find_column_index_by_header(wb, "CreditAmount")
    balance_column = find_column_index_by_header(wb, "Balance(INR)")
    print("debit column", chr(debit_column))
    print("credit column", chr(credit_column))
    print("balance column", chr(balance_column))
    print("from column", fromcolumn)
    print("to column", tocolumn)
    wb = find_replace(wb, start, end, find="None", replace='', columns=[fromcolumn, chr(debit_column), chr(credit_column), chr(balance_column)])
    wb = find_replace(wb, start, end, find="NA", replace='', columns=[chr(balance_column)])
    wb = empty_cell_to_0(wb, start, end, columns=[fromcolumn, chr(debit_column), chr(credit_column)])
    sheet = wb.active
    balance = 0
    debit = 0
    credit = 0
    for row in range(start, end):
        # case 1
        if sheet[f"{fromcolumn}{row}"].value is not None and len(str(sheet[f"{fromcolumn}{row}"].value)) > 1 and sheet[f"{tocolumn}{row}"].value is None or len(str(sheet[f"{tocolumn}{row}"].value)) < 1:
            if sheet[f"{chr(credit_column)}{row}"].value is not None and len(str(sheet[f"{chr(credit_column)}{row}"].value)) > 0 and float(sheet[f"{chr(credit_column)}{row}"].value) > 0:
                balance = float(sheet[f"{fromcolumn}{row}"].value) - float(sheet[f"{chr(credit_column)}{row}"].value)
            if sheet[f"{chr(debit_column)}{row}"].value is not None and len(str(sheet[f"{chr(debit_column)}{row}"].value)) > 0 and float(sheet[f"{chr(debit_column)}{row}"].value) > 0:
                balance = float(sheet[f"{fromcolumn}{row}"].value) + float(sheet[f"{chr(debit_column)}{row}"].value)
            print(f"{fromcolumn}{row}")
            print("balance = ", balance)
            print("withdrawal", float(sheet[f"{chr(debit_column)}{row}"].value))
            print("deposit", float(sheet[f"{chr(credit_column)}{row}"].value))
            is0 = balance - float(sheet[f"{chr(debit_column)}{row}"].value) + float(sheet[f"{chr(credit_column)}{row}"].value) - float(sheet[f"{fromcolumn}{row}"].value)
            num = format(is0, '.10f')
            is0 = float(num)
            if is0 == 0.0:
                sheet[f"{tocolumn}{row}"].value = sheet[f"{fromcolumn}{row}"].value

        # case 2
        elif sheet[f"{fromcolumn}{row}"].value is not None and len(str(sheet[f"{fromcolumn}{row}"].value)) > 1 and sheet[f"{chr(balance_column)}{row}"].value is not None and len(str(sheet[f"{chr(balance_column)}{row}"].value)) > 1 and sheet[f"{chr(credit_column)}{row}"].value is None or len(str(sheet[f"{chr(credit_column)}{row}"].value)) < 1:
            if sheet[f"{chr(balance_column)}{row}"].value is not None and float(sheet[f"{chr(balance_column)}{row}"].value) > 0:
                balance = float(sheet[f"{fromcolumn}{row}"].value) - float(sheet[f"{chr(balance_column)}{row}"].value)
            if sheet[f"{chr(debit_column)}{row}"].value is not None and float(sheet[f"{chr(debit_column)}{row}"].value) > 0:
                balance = float(sheet[f"{fromcolumn}{row}"].value) - float(sheet[f"{chr(debit_column)}{row}"].value)
            is0 = balance - float(sheet[f"{chr(debit_column)}{row}"].value) + float(sheet[f"{chr(balance_column)}{row}"].value) - float(sheet[f"{fromcolumn}{row}"].value)
            num = format(is0, '.10f')
            is0 = float(num)
            if is0 == 0.0:
                sheet[f"{chr(credit_column)}{row}"].value = sheet[f"{chr(balance_column)}{row}"].value
                sheet[f"{tocolumn}{row}"].value = sheet[f"{fromcolumn}{row}"].value

        # case 3
        elif sheet[f"{fromcolumn}{row}"].value is not None and len(str(sheet[f"{fromcolumn}{row}"].value)) > 1 and sheet[f"{chr(balance_column)}{row}"].value is not None and len(str(sheet[f"{chr(balance_column)}{row}"].value)) > 1 and sheet[f"{chr(credit_column)}{row}"].value is not None and len(str(sheet[f"{chr(credit_column)}{row}"].value)) > 1 and sheet[f"{chr(debit_column)}{row}"].value is None or len(str(sheet[f"{chr(debit_column)}{row}"].value)) < 1:
            if sheet[f"{chr(balance_column)}{row}"].value is not None and float(sheet[f"{chr(balance_column)}{row}"].value) > 0:
                balance = float(sheet[f"{fromcolumn}{row}"].value) - float(sheet[f"{chr(balance_column)}{row}"].value)
            if sheet[f"{chr(credit_column)}{row}"].value is not None and float(sheet[f"{chr(credit_column)}{row}"].value) > 0:
                balance = float(sheet[f"{fromcolumn}{row}"].value) - float(sheet[f"{chr(credit_column)}{row}"].value)
            is0 = balance - float(sheet[f"{chr(credit_column)}{row}"].value) + float(
                sheet[f"{chr(balance_column)}{row}"].value) - float(sheet[f"{fromcolumn}{row}"].value)
            num = format(is0, '.10f')
            is0 = float(num)
            if is0 == 0.0:
                sheet[f"{chr(debit_column)}{row}"].value = sheet[f"{chr(credit_column)}{row}"].value
                sheet[f"{chr(credit_column)}{row}"].value = sheet[f"{chr(balance_column)}{row}"].value
                sheet[f"{tocolumn}{row}"].value = sheet[f"{fromcolumn}{row}"].value

    remove_values_from_misaligned_column(wb, start, end, tocolumn, fromcolumn)
    return wb


def icici3_merging_date_column(wb, start, end, data_list):
    sheet = wb.active
    filtered_data = [entry for entry in data_list if entry["Operation"] == "date_Column"]
    column_json = set(item["Reference_Column"] for item in filtered_data if item["Operation"] == "date_Column")
    columns = list(column_json)
    inCompleteDateLen = 10  # incomplete date length (string)
    yearLength = 4  # year length (string)
    for column in columns:
        for i in range(start, end):  # iterating through table data start row to table data end row
            if sheet[f"{column}{i}"].value is not None:  # if cell value in date column is not none
                if len(str(sheet[f"{column}{i}"].value)) < inCompleteDateLen and len(str(sheet[f"{column}{i + 1}"].value)) == yearLength:  # if length of date in cell is less than incomplete date length and length of date in next row cell is equal to year length
                    s = str(sheet[f"{column}{i}"].value) + "" + str(sheet[f"{column}{i + 1}"].value)  # concat the date with year
                    sheet[f"{column}{i}"].value = s  # assign it to date cell
    return wb, columns


def kotak1_delete_unwanted_rows(wb):
    sheet = wb.active
    for row in range(sheet.max_row, 1, -1):
        if len(str(sheet[f"A{row}"].value)) < 5:
            sheet.delete_rows(row)
    return wb


def kotak1_column_error_records(wb, start, end, data_list):
    sheet = wb.active
    filtered_data = [entry for entry in data_list if entry["Operation"] == "align_error_record"]
    start_column = 65
    error_records = []
    for data in filtered_data:
        refcolumn = data["Reference_Column"]
        for row in range(start, end):
            if sheet[f"{refcolumn}{row}"].value is not None:
                data = [row - 1]
                for column in range(start_column, ord(refcolumn)):
                    data.append(str(sheet[f"{chr(column)}{row}"].value))
                addAlignmentData(data)
                error_records.append(row)

        narration_column = find_column_index_by_header(wb, header="Narration")
        date_column = find_column_index_by_header(wb, header="Date")
        for row in error_records:
            for column in range(start_column, ord(refcolumn) + 1):
                if column == narration_column:
                    sheet[f"{chr(narration_column)}{row}"].value = "error record"
                if column != date_column and column != narration_column:
                    sheet[f"{chr(column)}{row}"].value = None
    wb = kotak1_delete_unwanted_rows(wb)
    return wb


def sbi1_merging_date_column(wb, start, end, data_list):
    sheet = wb.active
    filtered_data = [entry for entry in data_list if entry["Operation"] == "date_Column"]
    column_json = set(item["Reference_Column"] for item in filtered_data if item["Operation"] == "date_Column")
    columns = list(column_json)
    inCompleteDateLen = 10  # incomplete date length (string)
    yearLength = 4  # year length (string)
    for column in columns:
        for i in range(start, end):  # iterating through table data start row to table data end row
            if sheet[f"{column}{i}"].value is not None:  # if cell value in date column is not none
                if len(str(sheet[f"{column}{i}"].value)) < inCompleteDateLen and len(str(sheet[f"{column}{i + 1}"].value)) == yearLength:  # if length of date in cell is less than incomplete date length and length of date in next row cell is equal to year length
                    s = str(sheet[f"{column}{i}"].value) + "" + str(sheet[f"{column}{i + 1}"].value)  # concat the date with year
                    sheet[f"{column}{i}"].value = s  # assign it to date cell
    return wb, columns


def tmb1_column_error_records(wb, start, end, data_list):
    sheet = wb.active
    filtered_data = [entry for entry in data_list if entry["Operation"] == "align_error_record"]
    start_column = 65
    error_records = []
    for data in filtered_data:
        refcolumn = data["Reference_Column"]
        for row in range(start, end):
            if sheet[f"{refcolumn}{row}"].value is None:
                data = [row - 1]
                for column in range(start_column, ord(refcolumn)):
                    data.append(str(sheet[f"{chr(column)}{row}"].value))
                addAlignmentData(data)
                error_records.append(row)

        narration_column = find_column_index_by_header(wb, header="Particulars")
        date_column = find_column_index_by_header(wb, header="Date")
        for row in error_records:
            for column in range(start_column, ord(refcolumn) + 1):
                if column == narration_column:
                    sheet[f"{chr(narration_column)}{row}"].value = "error record"
                if column != date_column and column != narration_column:
                    sheet[f"{chr(column)}{row}"].value = None
    wb = kotak1_delete_unwanted_rows(wb)
    return wb


# if __name__ == '__main__':
    # addAlignmentData(['Kotak1._Apr-22_637102__06-09-2023-14-01-34.xlsx', ' 051112485', 'Undefined', '01-Apr-22 to 30-Apr-22', ['15-Apr-22', 'NEFT FDRLH22105092182 T V THANGARAJFDRL0001789', 'None', 'None', 'NEFTINW-0396757917', '91439', '336,042.86 (Cr)']])