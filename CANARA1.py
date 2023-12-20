from datetime import datetime

import openpyxl

from CommonClass import Excel


def dateConvertion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%b-%y").date()
    return wb


def delete_header(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def alignColumn(wb, start, end, fromColumn, toColumn):
    sheet = wb.active
    for i in range(start, end):
        h_cell = f"{toColumn}{i}"
        if sheet[h_cell].value is None:
            sheet[h_cell].value = sheet[f"I{i}"].value
            sheet[f"{fromColumn}{i}"].value = None
    return wb


def deleteNoneRows(wb, start, end, refColumn):
    sheet = wb.active
    for x in range(end, start, -1):
        a_cell = f"A{x}"
        if sheet[a_cell].value is None:
            end -= 1
            sheet.delete_rows(x)
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def deleteRowByLength(wb, start, end, refColumn, pgNoLength):
    sheet = wb.active
    for x in range(end, start, -1):
        if sheet[f"{refColumn}{x}"].value is not None and (len(str(sheet[f"{refColumn}{x}"].value))) < 5:
            sheet.delete_rows(x)
    return wb


def deleteFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


# TODO: To add validation for header text
def mergHeaderText(wb, start, column):
    sheet = wb.active
    txt = sheet[f"{column}{start}"].value + sheet[f"{column}{start + 1}"].value
    sheet[f"{column}{start}"].value = txt
    return wb


def canara1_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 9   # need to change the logic for 8 column -> previous logic 9 columns
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def canara1_main(wb):
    sheet = wb.active
    # wb.save('C:/Users/Admin/Desktop/CANARA1output.xlsx')
    # exit()
    if canara1_validation(wb):
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response
    else:
        startText = "TRANS"
        endText = "UNLESS THE CONSTITUENT BRINGS TO THE NOTICE OF THE BANK"
        startEndRefColumn = "A"
        mergHeaderColumn1 = "A"
        mergHeaderColumn2 = "B"
        duplicateHeaderTxt1 = "TRANS"
        duplicateHeaderTxt2 = "DATE"
        duplicateHeaderRefColumn = "A"
        startText2 = "TRANS DATE"
        endText2 = "Statement Summary :"
        pgNoLength = 5
        refColumnToDeletePgNoRow = "A"
        refColumnToMerg = "A"
        columnToMerg1 = "E"
        refColumnToDeleteNoneRows = "A"
        fromColumn = "I"
        toColumn = "H"
        stringAlignColumn = "E"
        refTextToDeleteRow = "B/F"
        refColumnToDeleteRow = "E"
        dateConversionColumn1 = "A"
        dateConversionColumn2 = "B"
        deleteColumnRefText1 = "BRANCH"
        refHeaderText1 = "TRANS DATE"
        refHeaderText2 = "VALUE DATE"
        refHeaderText3 = "REF/CHQ.NO"
        refHeaderText4 = "DESCRIPTION"
        refHeaderText5 = "WITHDRAWS"
        refHeaderText6 = "DEPOSIT"
        refHeaderText7 = "BALANCE"
        headerText1 = "Transaction_Date"
        headerText2 = "Value_Date"
        headerText3 = "ChequeNo_RefNo"
        headerText4 = "Narration"
        headerText5 = "Withdrawal"
        headerText6 = "Deposit"
        headerText7 = "Balance"
        negativeValueColumnRefText1 = "Withdrawal"
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        AheaderTextMerged = mergHeaderText(wb, start, mergHeaderColumn1)
        BheaderTextMerged = mergHeaderText(AheaderTextMerged, start, mergHeaderColumn2)
        duplicateHeaderRemoved1 = Excel.remove_rows(BheaderTextMerged, start, end, duplicateHeaderTxt1, duplicateHeaderRefColumn)
        start, end = Excel.get_start_end_row_index(duplicateHeaderRemoved1, startText, endText, startEndRefColumn)
        duplicateHeaderRemoved2 = Excel.remove_rows(duplicateHeaderRemoved1, start, end, duplicateHeaderTxt2, duplicateHeaderRefColumn)
        start, end = Excel.get_start_end_row_index(duplicateHeaderRemoved2, startText, endText, startEndRefColumn)
        footerDeleted = deleteFooter(duplicateHeaderRemoved2, end)
        start, end = Excel.get_start_end_row_index(footerDeleted, startText2, endText2, startEndRefColumn)
        pgNoRowDeleted = deleteRowByLength(footerDeleted, start, end, refColumnToDeletePgNoRow, pgNoLength)
        start, end = Excel.get_start_end_row_index(pgNoRowDeleted, startText2, endText2, startEndRefColumn)
        mergColumnE = mergingRows(pgNoRowDeleted, start, end, refColumnToMerg, columnToMerg1)
        noneRowsDeleted = deleteNoneRows(mergColumnE, start, end, refColumnToDeleteNoneRows)
        start, end = Excel.get_start_end_row_index(noneRowsDeleted, startText2, endText2, startEndRefColumn)
        alignedColumnH = alignColumn(noneRowsDeleted, start, end, fromColumn, toColumn)
        start, end = Excel.get_start_end_row_index(alignedColumnH, startText2, endText2, startEndRefColumn)
        footerDeleted = deleteFooter(alignedColumnH, end - 1)  # end-1 to Include Last Footer Row
        headerDeleted = delete_header(footerDeleted, start - 1)  # start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(alignedColumnH, startText2, endText2, startEndRefColumn)
        alignedE = Excel.string_align(headerDeleted, start, end, stringAlignColumn)
        removedOpeningBalance = Excel.remove_row(alignedE, start, end, refTextToDeleteRow, refColumnToDeleteRow)
        start, end = Excel.get_start_end_row_index(alignedColumnH, startText2, endText2, startEndRefColumn)
        dateConvertedA = dateConvertion(removedOpeningBalance, start + 1, end + 1, dateConversionColumn1)  # start+1 to Skip Header, end+1 to Include Last Row
        dateConvertedB = dateConvertion(dateConvertedA, start + 1, end + 1, dateConversionColumn2)  # start+1 to Skip Header, end+1 to Include Last Row
        dalatedColumnBRANCH = Excel.delete_column(dateConvertedB, deleteColumnRefText1)
        lastCol = 65 + Excel.column_count(wb)
        transdate = Excel.alter_header_name(dalatedColumnBRANCH, refHeaderText1, headerText1, lastCol)
        valuedate = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)
        chqno = Excel.alter_header_name(valuedate, refHeaderText3, headerText3, lastCol)
        naration = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)
        debit = Excel.alter_header_name(naration, refHeaderText5, headerText5, lastCol)
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        slnoCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))
        negativeColumnChecked = Excel.check_neagativeValue_by_column(slnoCreated, negativeValueColumnRefText1)
        createdTransTypeColumn = Excel.transaction_type_column(negativeColumnChecked)
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/1.Canara_-_6183__11-09-2023-18-32-43.xlsx"
    wb = openpyxl.load_workbook(path)
    result = canara1_main(wb)
    result.save('C:/Users/Admin/Desktop/FinalOutput/CANARA1output.xlsx')
