
from datetime import datetime

import openpyxl

from AlignmentData import addAlignmentData
from CommonClass import Excel


def dateConvertion(wb, start, end, column, ref):
    """
        Convert the date values in a specified column to a standard date format.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index to begin processing.
        - end (int): The ending row index to stop processing.
        - column (str): The column letter representing the column containing date values to be converted.
        - ref (str): The reference text to identify the rows with date values for conversion.

        Returns:
        - openpyxl.Workbook: The modified Workbook with date values in the specified column converted to standard date format.

        Notes:
        - This function iterates through the specified column in the given range and converts the date values to a standard date format.
        - The conversion is performed only for the rows where the specified reference text is present in the cell value.
        - The modified Workbook is then returned.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start column to table data end column
        if ref in str(sheet[f"{column}{i}"].value):  # if reference text in cell value
            sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d/%m/%y").date()  # converting date to standard date formate
    return wb


def alignColumns(wb, start, end, headData, refColumnToAlign):
    """
        Manually align miss-aligned data in specified columns.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index to begin processing.
        - end (int): The ending row index to stop processing.
        - headData (list): A list containing statement holder details [account number, name, period].
        - refColumnToAlign (str): The reference column letter to identify rows that require manual alignment.

        Returns:
        - openpyxl.Workbook: The modified Workbook with manually aligned data.

        Notes:
        - This function iterates through the specified range and identifies rows that require manual alignment based on the reference column.
        - For each identified row, it stores the statement holder details, row index, and cell values of specified columns.
        - The data is then passed to the addAlignmentData() function (not provided) to store the misaligned row data to the Excel file.
        - Error records are marked with "Error Record" in the narration column to easily identify them.
        - The modified Workbook is returned.

    """
    sheet = wb.active
    error_records = []  # array to store error records row index
    for i in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{refColumnToAlign}{i}"].value is not None:  # if reference column cell value is not none then the row should be manually aligned
            data = [headData[0], headData[1], headData[2]]  # storing the statement holders details [account number, name, period]
            data.append(i - 1)  # row index (slno)
            data.append(str(sheet[f"A{i}"].value))  # appending cell value of column A
            data.append(str(sheet[f"B{i}"].value))  # appending cell value of column B
            data.append(str(sheet[f"C{i}"].value))  # appending cell value of column C
            data.append(str(sheet[f"D{i}"].value))  # appending cell value of column D
            data.append(str(sheet[f"E{i}"].value))  # appending cell value of column E
            data.append(str(sheet[f"F{i}"].value))  # appending cell value of column F
            data.append(str(sheet[f"G{i}"].value))  # appending cell value of column G
            data.append(str(sheet[f"H{i}"].value))  # appending cell value of column H
            addAlignmentData(data)  # passing the data to the addAlignmentData() function to store the misaligned row data to the excel file
            error_records.append(i)  # storing the error records row index in array
    for x in error_records:  # iterating through error records array
        sheet[f"B{x}"].value = "Error Record"  # assign "Error Record" to the narration column cell to identify that it's a error record
        sheet[f"C{x}"].value = None  # make C column cell to none
        sheet[f"D{x}"].value = None  # make D column cell to none
        sheet[f"E{x}"].value = None  # make E column cell to none
        sheet[f"F{x}"].value = None  # make F column cell to none
        sheet[f"G{x}"].value = None  # make G column cell to none
        sheet[f"H{x}"].value = None  # make H column cell to none
    return wb


def deleteHeader(wb, start, end):
    """
        Delete rows above the specified start index row.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index. Rows above this index will be deleted.
        - end (int): The ending row index (exclusive). Rows above this index will be deleted.

        Returns:
        - openpyxl.Workbook: The modified Workbook with deleted header rows.

        Notes:
        - This function iterates through the specified range of rows in reverse order (from end to start).
        - It deletes each row above the specified starting index.
        - The modified Workbook is returned.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data starting row to 1st row of the sheet
        sheet.delete_rows(x)  # delete row
    return wb


def aligningAllColumns(wb, start, end, refColumn):
    """
        Align column data when the specified reference column is None.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index to begin alignment.
        - end (int): The ending row index (exclusive) to stop alignment.
        - refColumn (str): The reference column letter (e.g., 'A', 'B', 'C').

        Returns:
        - openpyxl.Workbook: The modified Workbook with aligned column data.

        Notes:
        - This function iterates through the specified range of rows.
        - If the value in the specified reference column is None, it aligns the data by shifting values from right to left.
        - The modified Workbook is then returned.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{refColumn}{i}"].value is None:  # if the reference column is None
            sheet[f'C{i}'].value = sheet[f'D{i}'].value  # assigning column D cell value to Column C cell value
            sheet[f'D{i}'].value = sheet[f'E{i}'].value  # assigning column E cell value to column D cell value
            sheet[f'E{i}'].value = sheet[f'F{i}'].value  # assigning column F cell value to column E cell value
            sheet[f'F{i}'].value = sheet[f'G{i}'].value  # assigning column G cell value to column F cell value
            sheet[f'G{i}'].value = sheet[f'H{i}'].value  # assigning column H cell value to column G cell value
            sheet[f'H{i}'].value = None  # assigning column H cell value to None
    return wb


def deleteFooter(wb, start, end):
    """
        Delete all rows below the specified end (last) row.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index to begin deletion (inclusive).
        - end (int): The ending row index to stop deletion (exclusive).

        Returns:
        - openpyxl.Workbook: The modified Workbook with rows deleted below the specified end row.

        Notes:
        - This function iterates through the rows from the last row of the sheet to the specified end row.
        - It deletes each row in the specified range.
        - The modified Workbook is then returned.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterate through max row of sheet to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


# no need of adding header data in the statement
def headerData(wb, start, end):
    """
        Gather statement details [Account number, Customer ID, name, IFSC, period, openbal, closebal, debits, credits].

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index to begin searching for header data.
        - end (int): The ending row index to stop searching for header data.

        Returns:
        - list: A list containing statement details [Account number, Customer ID, name, IFSC, period, openbal, closebal, debits, credits].

        Notes:
        - This function searches for specific strings in the Excel sheet columns to extract relevant header data.
        - It iterates through the rows from the specified start row to the sheet's 1st row.
        - Extracts account number, customer ID, IFSC, and statement period from specific cell values.
        - Returns a list containing statement details.

    """
    sheet = wb.active
    s1 = "Undefined"  # variable to store header data
    s2 = "Undefined"  # variable to store header data
    s3 = "Undefined"  # variable to store header data
    s4 = "Undefined"  # variable to store header data
    s5 = "Undefined"  # variable to store header data
    s6 = "Undefined"  # variable to store header data
    s7 = "Undefined"  # variable to store header data
    s8 = "Undefined"  # variable to store header data
    for i in range(start, 0, -1):  # iterating through table data start row to sheet 1st column
        if sheet[f"B{i}"].value is not None and "Account No" in str(sheet[f"B{i}"].value):  # if cell value is not none and "Account No" string in cell value then
            s1 = sheet[f"D{i}"].value  # store value in variable
        if sheet[f"B{i}"].value is not None and "IFSC" in str(sheet[f"B{i}"].value):  # if the cell value is not none and "IFSC" string in cell value then
            s2 = sheet[f"D{i}"].value  # store value in variable
        if sheet[f"A{i}"].value is not None and "Statement From :" in str(sheet[f"A{i}"].value):  # if cell value is not none and "Statement From" string in cell value then
            s3 = f"{sheet[f'A{i}'].value} {sheet[f'B{i}'].value}"  # store value in variable
    spl1 = s1.split(":")  # extracting account number from the string
    a = spl1[3].strip().split(" ")  # extracting account number from the string
    acno = a[0]  # storing account number
    cusid = f"Customer ID : {spl1[2]}"  # storing customer id
    name = "Undefined"  # storing name
    ifsc = f"IFSC : {s2}"  # storing ifsc code
    period = s3  # storing statement period
    openbal = s4  # storing open balance
    closebal = s5  # storing closing balance
    debits = s6  # storing total debit
    credits = s7  # storing total credit
    headData = [acno, name, period]  # header data array
    return headData


def removeNoneRows(wb, start, end, column):
    """
        Remove unwanted rows when the reference column cell value is None.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index to begin searching for unwanted rows.
        - end (int): The ending row index to stop searching for unwanted rows.
        - column (str): The reference column letter (e.g., 'A', 'B', 'C', etc.).

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object after removing unwanted rows.

        Notes:
        - This function iterates through the rows from the specified end row to the start row.
        - Deletes rows where the reference column cell value is None.
        - Returns the modified Workbook object.

    """
    sheet = wb.active
    for x in range(end, start, -1):  # iterating through table data end row to table data start row
        if sheet[f"{column}{x}"].value is None:  # if reference column cell value is none
            sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merge the rows of a desired column based on a reference column.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index to begin merging rows.
        - end (int): The ending row index to stop merging rows.
        - refColumn (str): The reference column letter (e.g., 'A', 'B', 'C', etc.) to identify starting rows.
        - mergingColumn (str): The column letter (e.g., 'D', 'E', 'F', etc.) whose values will be merged.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object after merging rows.

        Notes:
        - This function iterates through rows from the specified start to end.
        - Merges rows in the specified mergingColumn based on the presence of a value in the reference column.
        - The reference column value indicates the starting row of the merging operation.
        - The merged result is stored in the corresponding cell of the reference column.
        - The last row may not be merged during iteration, so an additional merge is performed for the last row.

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through start and end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def hdfc1_validation(wb):
    """
        Validate columns for the core logic in an HDFC workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the HDFC Excel file.

        Returns:
        - bool: True if the number of columns doesn't match the expected count, False otherwise.

        Notes:
        - This function checks whether the number of columns in the workbook matches the expected count.
        - The expected count is set to 7 columns as per the designed core logic.

    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 7  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def hdfc1_main(wb):
    """
        Process HDFC1 bank statement data in an Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.Workbook): Excel workbook containing HDFC bank statement data.

        Returns:
        dict: A dictionary containing the modified workbook ('data') and a message ('msg').

        Overview:
        - Validates the column count in the workbook using hdfc1_validation function.
        - Performs various data cleaning and standardization tasks:
            - Deletes unnecessary rows based on specified start and end texts.
            - Merges misaligned rows in a specified column.
            - Removes rows where a reference column cell value is None.
            - Deletes footer rows.
            - Aligns column data using reference columns and manual alignment.
            - Converts date columns to a standard date format.
            - Renames headers to standardize them.
            - Adds a serial number column.
            - Checks for negative values in specific columns.
            - Standardizes the overall column count.
            - Adds a transaction type column.

        Note:
        - Ensure that the hdfc1_validation, mergingRows, removeNoneRows, deleteFooter,
          headerData, deleteHeader, aligningAllColumns, alignColumns, dateConvertion,
          and other related functions are defined and imported before using this function.

    """
    sheet = wb.active
    if hdfc1_validation(wb):  # validate columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return the response with error msg
    else:
        startText = "Narration"  # header text in column B
        endText = "STATEMENT SUMMARY :-"  # text define the end of the table data row
        startEndRefColumn = "B"  # reference column contains the start and end text
        deleteFlagStartText = "HDFC BANK LIMITED"  # starting row reference text to delete the rows by range
        deleteFlagStopText = "Statement From :"  # ending row reference text to delete the rows by range
        deleteFlagRefColumn = "A"  # column containing starting row reference text and ending row reference text to delete the rows by range
        dateConversionColumn = "A"  # column to convert date to standard date formate
        columnToMerg1 = "B"  # column to merge misaligned row data
        refColumnToMerg = "A"  # reference column (date column) to merge the rows of other column
        refColumnToAlignAllColumn = "C"  # reference column to align column data using aligningAllColumns() function
        refColumnToAlignColumn = "H"  # reference column to align column data manually -> stored in excel file
        dateConversionColumn1 = "A"  # column to convert date to standard date formate
        dateConversionColumn2 = "D"  # column to convert date to standard date formate
        dateRefText = "/"  # reference text to identify the date -> which splits the date, month, year
        refHeaderText1 = "Date"  # header text to replace with standardised column name
        refHeaderText2 = "Narration"  # header text to replace with standardised column name
        refHeaderText3 = "Chq./Ref.No."  # header text to replace with standardised column name
        refHeaderText4 = "Value Dt"  # header text to replace with standardised column name
        refHeaderText5 = "Withdrawal Amt."  # header text to replace with standardised column name
        refHeaderText6 = "Deposit Amt."  # header text to replace with standardised column name
        refHeaderText7 = "Closing Balance"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Narration"  # standard column name
        headerText3 = "ChequeNo_RefNo"  # standard column name
        headerText4 = "Value_Date"  # standard column name
        headerText5 = "Withdrawal"  # standard column name
        headerText6 = "Deposit"  # standard column name
        headerText7 = "Balance"  # standard column name
        negativeValueColumnRefText1 = "Withdrawal"  # no need of changing negative value to positive in columns
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        dupRowsDeleted = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagStopText, deleteFlagRefColumn)  # deleting the rows by range
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        mergedColumnB = mergingRows(dupRowsDeleted, start, end, refColumnToMerg, columnToMerg1)  # merging misaligned rows of desired column
        noneRemoved = removeNoneRows(mergedColumnB, start, end - 1, refColumnToMerg)  # removing none rows by date column as referance
        start, end = Excel.get_start_end_row_index(noneRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        footerDeleted = deleteFooter(noneRemoved, start, end - 1)  # delete all the rows below the end(last) row end-1 to IncludeEnd Footer
        start, end = Excel.get_start_end_row_index(footerDeleted, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        headData = headerData(footerDeleted, start, end)  # no need of adding header data to the statement
        start, end = Excel.get_start_end_row_index(footerDeleted, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        headerDeleted = deleteHeader(footerDeleted, start - 1, end)  # delete rows above the start index row start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        columnAligned = aligningAllColumns(headerDeleted, start, end + 1, refColumnToAlignAllColumn)  # aligning the column data when the reference column is None, end+1 to Include Last Row
        alignColumns(wb, start, end, headData, refColumnToAlignColumn)  # data to be align manually -> miss aligned while converting pdf to excel
        convertedDateA = dateConvertion(wb, start + 1, end + 1, dateConversionColumn1, dateRefText)  # converting the date in a column to standard date formate, start+1 to Skip Header, end+1 to Include Last Row
        convertedDateD = dateConvertion(convertedDateA, start + 1, end + 1, dateConversionColumn2, dateRefText)  # converting the date in a column to standard date formate,start+1 to Skip Header, end+1 to Include Last Row
        lastCol = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        transdate = Excel.alter_header_name(convertedDateD, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
        naration = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        chqno = Excel.alter_header_name(naration, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        valuedate = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        debit = Excel.alter_header_name(valuedate, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)  # alter header name from the excel file to the standard column name
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slnoCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
        negativeValueChecked = Excel.check_neagativeValue_by_column(slnoCreated, negativeValueColumnRefText1)  # no need of converting negative value to positive
        Excel.finalise_column(wb, columns)  # standardizing count of column
        createdTransTypeColumn = Excel.transaction_type_column(negativeValueChecked)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/hdfc.xlsx"
    # path = "C:/Users/Admin/Downloads/50100443515738_1703914785718__30-12-2023-13-53-08.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    result = hdfc1_main(wb)
    # result["data"].save('C:/Users/Admin/Desktop/HDFC1output.xlsx')
