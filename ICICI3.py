import os
from datetime import datetime
import openpyxl
from openpyxl.utils import column_index_from_string

from FormatingExcelFiles.CommonClass import Excel


def aligningColumn(wb, start, end, mergingColumn, refColumn):
    sheet = wb.active
    for i in range(start, end):
        if sheet[f"{refColumn}{i}"].value is not None:
            sheet[f"{mergingColumn}{i}"].value = sheet[f"{refColumn}{i}"].value
            sheet[f"{refColumn}{i}"].value = None
    return wb


def dateConvertion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), '%d-%b-%Y').date()
    return wb


def remove_space(wb, start, end, column):
    sheet = wb.active
    for x in range(start, end):
        sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace(" ", "")
    return wb


def removeHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def deleteColumn(wb, column):
    sheet = wb.active
    column_index = openpyxl.utils.column_index_from_string(column)
    sheet.delete_cols(column_index)
    return wb


def removeRowsByDateLength(wb, start, end, column):
    sheet = wb.active
    yearLength = 6
    for x in range(end, start, -1):
        if len(str(sheet[f"{column}{x}"].value)) < yearLength:
            sheet.delete_rows(x)
    return wb


def removeNone(wb, start, end, column):
    sheet = wb.active
    for x in range(start, end):
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def mergingDateColumn(wb, start, end, column):
    sheet = wb.active
    inCompleteDateLen = 10
    yearLength = 4
    for i in range(start, end):
        if sheet[f"{column}{i}"].value is not None:
            if len(str(sheet[f"{column}{i}"].value)) < inCompleteDateLen and len(
                    str(sheet[f"{column}{i + 1}"].value)) == yearLength:
                s = str(sheet[f"{column}{i}"].value) + " " + str(sheet[f"{column}{i + 1}"].value)
                sheet[f"{column}{i}"].value = s
    return wb


def icici3_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 9
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def icici3_main(wb):
    sheet = wb.active
    if icici3_validation(wb):
        raise Exception(f"<= INVALID FORMATE =>  <Count Of Column Mismatch>")
    else:
        startText = "Sr No"
        endText = ""
        startEndDefColumn = "A"
        deleteFlagStartText = "DETAILED STATEMENT"
        deleteFlagStopText = "Sr No"
        DeleteFlagRefColumn = "A"
        dateMergColumn1 = "B"
        dateMergColumn2 = "C"
        columnToMerg1 = "E"
        refColumnToMerg = "A"
        removeNoneColumn1 = "B"
        removeNoneColumn2 = "C"
        removeNoneColumn3 = "E"
        refColumnToDeleteDateRow = "B"
        columnToDelete = "A"
        stringAlignColumn1 = "A"
        stringAlignColumn2 = "B"
        stringAlignColumn3 = "C"
        stringAlignColumn4 = "D"
        stringAlignColumn5 = "E"
        stringAlignColumn6 = "F"
        stringAlignColumn7 = "G"
        dateConversionColumn1 = "A"
        dateConversionColumn2 = "B"
        refHeaderText1 = "ValueDate"
        refHeaderText2 = "TransactionDate"
        refHeaderText3 = "ChequeNumber"
        refHeaderText4 = "Transaction Remarks"
        refHeaderText5 = "DebitAmount"
        refHeaderText6 = "CreditAmount"
        refHeaderText7 = "Balance(INR)"
        headerText1 = "Value_Date"
        headerText2 = "Transaction_Date"
        headerText3 = "ChequeNo_RefNo"
        headerText4 = "Narration"
        headerText5 = "Withdrawal"
        headerText6 = "Deposit"
        headerText7 = "Balance"
        columnToAlign = "G"
        refColumnToAlign = "H"
        stringToRemove2 = "NA"
        columnToRemoveString4 = "F"
        columnToRemoveString5 = "E"
        replaceEmptyColumeByNone1 = "ChequeNo_RefNo"
        replaceEmptyColumeByNone2 = "Withdrawal"
        replaceEmptyColumeByNone3 = "Deposit"
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)
        end = sheet.max_row
        dupHeaderRemoved = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagStopText, DeleteFlagRefColumn)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)
        end = sheet.max_row
        mergedDateB = mergingDateColumn(dupHeaderRemoved, start, end, dateMergColumn1)
        mergedDateC = mergingDateColumn(mergedDateB, start, end, dateMergColumn2)
        mergedColumnE = mergingRows(mergedDateC, start, end, refColumnToMerg, columnToMerg1)
        noneRemovedB = removeNone(mergedColumnE, start, end, removeNoneColumn1)
        noneRemovedC = removeNone(noneRemovedB, start, end, removeNoneColumn2)
        noneRemovedE = removeNone(noneRemovedC, start, end, removeNoneColumn3)
        dateRowRemoved = removeRowsByDateLength(noneRemovedE, start, end, refColumnToDeleteDateRow)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)
        end = sheet.max_row
        headerRemoved = removeHeader(dateRowRemoved, start - 1)  # start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)
        end = sheet.max_row
        deletedA = deleteColumn(dateRowRemoved, columnToDelete)
        alignedA = Excel.string_align(deletedA, start, end + 1, stringAlignColumn1)  # end+1 to Include Last Row
        alignedB = Excel.string_align(alignedA, start, end + 1, stringAlignColumn2)  # end+1 to Include Last Row
        alignedC = Excel.string_align(alignedB, start, end + 1, stringAlignColumn3)  # end+1 to Include Last Row
        alignedD = Excel.string_align(alignedC, start, end + 1, stringAlignColumn4)  # end+1 to Include Last Row
        alignedE = Excel.string_align(alignedD, start, end + 1, stringAlignColumn5)  # end+1 to Include Last Row
        alignedF = Excel.string_align(alignedE, start, end + 1, stringAlignColumn6)  # end+1 to Include Last Row
        alignedG = Excel.string_align(alignedF, start, end + 1, stringAlignColumn7)  # end+1 to Include Last Row
        noneRemovedA = removeNone(alignedG, start, end + 1, stringAlignColumn1)  # end+1 to Include Last Row
        noneRemovedB = removeNone(noneRemovedA, start, end + 1, stringAlignColumn2)  # end+1 to Include Last Row
        noneRemovedC = removeNone(noneRemovedB, start, end + 1, stringAlignColumn3)  # end+1 to Include Last Row
        noneRemovedD = removeNone(noneRemovedC, start, end + 1, stringAlignColumn4)  # end+1 to Include Last Row
        noneRemovedE = removeNone(noneRemovedD, start, end + 1, stringAlignColumn5)  # end+1 to Include Last Row
        noneRemovedF = removeNone(noneRemovedE, start, end + 1, stringAlignColumn6)  # end+1 to Include Last Row
        noneRemovedG = removeNone(noneRemovedF, start, end + 1, stringAlignColumn7)  # end+1 to Include Last Row
        removedSpaceA = remove_space(noneRemovedG, start, end + 1, dateConversionColumn1)  # end+1 to Include Last Row
        removedSpaceB = remove_space(noneRemovedG, start, end + 1, dateConversionColumn2)  # end+1 to Include Last Row
        dateConvertedA = dateConvertion(removedSpaceB, start + 1, end + 1, dateConversionColumn1)  # start+1 to Skip Header, end+1 to Include Last Row
        dateConvertedB = dateConvertion(dateConvertedA, start + 1, end + 1, dateConversionColumn2)  # start+1 to Skip Header, end+1 to Include Last Row
        lastCol = 65 + Excel.column_count(wb)
        valuedate = Excel.alter_header_name(dateConvertedB, refHeaderText1, headerText1, lastCol)
        transdate = Excel.alter_header_name(valuedate, refHeaderText2, headerText2, lastCol)
        refno = Excel.alter_header_name(transdate, refHeaderText3, headerText3, lastCol)
        naration = Excel.alter_header_name(refno, refHeaderText4, headerText4, lastCol)
        debit = Excel.alter_header_name(naration, refHeaderText5, headerText5, lastCol)
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)
        alignedColumnH = aligningColumn(balance, start, end, columnToAlign, refColumnToAlign)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        slCreated = Excel.create_slno_column(alignedColumnH, start, end + 1, chr(columnToCreateSlNo))
        noneReplacedF = Excel.replace_to_none(slCreated, start, end + 1, stringToRemove2, columnToRemoveString4)
        noneReplacedE = Excel.replace_to_none(noneReplacedF, start, end + 1, stringToRemove2, columnToRemoveString5)
        replacednonechqno = Excel.empty_cell_to_none(noneReplacedE, start, end + 1, replaceEmptyColumeByNone1)
        replacednonewithdrawal = Excel.empty_cell_to_none(replacednonechqno, start, end + 1, replaceEmptyColumeByNone2)
        replacednoneDeposit = Excel.empty_cell_to_none(replacednonewithdrawal, start, end + 1, replaceEmptyColumeByNone3)
        createdTransTypeColumn = Excel.transaction_type_column(replacednoneDeposit)
        return wb


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/ilovepdf_merged_7__26-09-2023-10-37-33.xlsx"
    wb = openpyxl.load_workbook(path)
    result = icici3_main(wb)
    result.save('C:/Users/Admin/Desktop/FinalOutput/ICICI3output.xlsx')
