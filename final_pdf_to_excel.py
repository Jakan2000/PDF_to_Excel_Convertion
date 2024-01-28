import configparser
import json
import os
from datetime import datetime
from io import BytesIO

import openpyxl
import psycopg2
import requests
from PyPDF2 import PdfWriter, PdfReader
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from psycopg2 import extras

from AlignmentData import canara1_column_alignment, hdfc1_column_error_records, hdfc1_headerData, find_replace
from AlignmentData import sbi1_merging_date_column, tmb1_column_error_records
from AlignmentData import icici3_column_alignment, icici3_merging_date_column, column_count, kotak1_column_error_records


def get_unwanted_row_index_merged_row(wb, start, end, data_list):
    sheet = wb.active
    filtered_data = [entry for entry in data_list if entry["Operation"] == "date_Column"]
    refcolumn = filtered_data[0]["Reference_Column"]
    indxes = []
    for row in range(start, end):
        if sheet[f"{refcolumn}{row}"].value is None:
            indxes.append(row)
    return indxes


def remove_string_none(wb, start, end):
    sheet = wb.active
    max_column = 65 + sheet.max_column
    start_column = 65
    for column in range(start_column, max_column):
        for row in range(start, end):
            if "None" in str(sheet[f"{chr(column)}{row}"].value):
                sheet[f"{chr(column)}{row}"].value = str(sheet[f"{chr(column)}{row}"].value).replace("None", "")
    return wb


def delete_unwanted_rows(wb, start, end, bank, type):
    def delete_footer(wb, end):
        sheet = wb.active
        for row in range(sheet.max_row, end, -1):  # iterating from data ending row to the last row
            sheet.delete_rows(row)  # delete row
        return wb

    def delete_header(wb, start):
        sheet = wb.active
        for row in range(start, 0, -1):  # iterating from header row to the 0 index row
            sheet.delete_rows(row)  # delete row
        return wb

    def delete_row_by_string_length(wb, start, end):
        sheet = wb.active
        for row in range(start, end):
            if len(str(sheet[f"A{row}"].value)) < 4:
                sheet.delete_rows(row)  # delete row
        return wb

    delete_footer(wb, end)
    delete_header(wb, start)
    if bank == "canara" and type == "type1":
        delete_row_by_string_length(wb, start, end)
    return wb


def string_align(wb, start, end):
    sheet = wb.active
    max_column = 65 + sheet.max_column
    start_column = 65
    for column in range(start_column, max_column):
        for row in range(start, end):  # iterate through start row and end row
            if sheet[f"{chr(column)}{row}"].value is not None:
                sheet[f"{chr(column)}{row}"].value = str(sheet[f"{chr(column)}{row}"].value).replace('\n', '')
    return wb


def empty_cell_to_none(wb, start, end):
    sheet = wb.active
    max_column = 65 + sheet.max_column
    start_column = 65
    for column in range(start_column, max_column):
        for row in range(start, end):  # iterating through all the rows from start to end row
            if len(str(sheet[f"{chr(column)}{row}"].value)) < 1:  # if cell value length is less than 1
                sheet[f"{chr(column)}{row}"].value = None  # make it as None
    return wb


def get_start_end_row_index(wb, data_list):
    start_row_text_json = set(item["Keywords"] for item in data_list if item["Operation"] == "start_row")
    startText = list(start_row_text_json)[0]
    # print(startText)

    end_row_text_json = set(item["Keywords"] for item in data_list if item["Operation"] == "end_row")
    endText = list(end_row_text_json)[0]
    # print(endText)

    start_refcolumn_json = set(item["Reference_Column"] for item in data_list if item["Operation"] == "start_row")
    startRefColumn = list(start_refcolumn_json)[0]
    # print(startRefColumn)

    end_refcolumn_json = set(item["Reference_Column"] for item in data_list if item["Operation"] == "end_row")
    endRefColumn = list(end_refcolumn_json)[0]
    # print(endRefColumn)

    sheet = wb.active
    start = 0
    end = 0
    for cell in sheet[startRefColumn]:  # iterating through all the cells in startEndDefColumn
        start += 1  # increment start value
        if startText in str(cell.value):  # if start text in the cell break the loop, and the cell index is stored in start variable
            break
    for cell in sheet[endRefColumn]:  # iterating through all the cells in startEndDefColumn
        end += 1  # increment end value
        if endText is None:
            pass
        elif endText in str(cell.value):  # if end text in the cell, break the loop, and the cell index is stored in end variable
            break
    return start, end


def delete_row(wb, index):
    sheet = wb.active
    # sorted_index = sorted(index, reverse=True)
    sorted_indices = sorted(set(index), reverse=True)
    for row in sorted_indices:  # iterate through a sorted array
        if isinstance(row, int):
            sheet.delete_rows(row)  # delete row
    return wb


def delete_rows_by_range(wb, start, end, data_list):
    sheet = wb.active
    delete_flag = False
    rows_to_delete = []
    filtered_data = [entry for entry in data_list if entry["Operation"] == "delete_row_by_range"]
    for data in filtered_data:
        for row in range(start, end):  # iterate from start to end row
            from_text = data["From_Text"]
            to_text = data["To_Text"]
            from_reference_column = data["From_Text_Reference_Column"]
            to_reference_column = data["To_Text_Reference_Column"]
            if from_text in str(sheet[f"{from_reference_column}{row}"].value):  # if start text is in the reference column cell it's the starting row
                delete_flag = True  # make delete flag true
            if delete_flag:  # if delete flag is true append the row to rows_to_delete array
                rows_to_delete.append(row)  # append row to rows_to_delete array
            if to_text in str(sheet[f"{to_reference_column}{row}"].value):  # if stop text is in the reference column cell it's the last row
                delete_flag = False  # make delete flag false
    return rows_to_delete


def merging_rows(wb, start, end, data_list):
    sheet = wb.active
    max_column = 65 + sheet.max_column
    start_column = 65
    dataToMerge = []
    filtered_data = [entry for entry in data_list if entry["Operation"] == "merging_rows_reference_column"]
    refcolumn = filtered_data[0]["Reference_Column"]
    for column in range(start_column, max_column):
        for row in range(start, end):
            date = sheet[f"{refcolumn}{row}"].value  # getting reference column cell value
            if date is not None:  # if reference column cell is not none then it's the starting row
                if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                    dataToMerge.append(f"{chr(column)}{row}")  # store cell address in the 0 index
                    dataToMerge.append(sheet[f"{chr(column)}{row}"].value)  # store data from the 1 index
                else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                    s = ""  # empty string to merge the row data
                    for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                        s += str(dataToMerge[j])  # concat the row data
                    cell_address = dataToMerge[0]  # take current cell address from 0 index
                    sheet[str(cell_address)].value = s  # assign conceited data to the cell
                    dataToMerge = []  # emptying dataToMerge ot find the next row starting
                    dataToMerge.append(f"{chr(column)}{row}")  # appending next starting row address in the 0 index
                    dataToMerge.append(sheet[f"{chr(column)}{row}"].value)  # appending row data to the corresponding index
            if date is None:  # if date is none this is not the starting row
                dataToMerge.append(sheet[f"{chr(column)}{row}"].value)  # append data to the corresponding index
        # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
        st1 = ""  # empty string to merge the row data
        for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
            st1 += str(dataToMerge[m])  # concat the row data
        cell_address = dataToMerge[0]  # take current cell address from 0 index
        sheet[str(cell_address)].value = st1  # assign conceited data to the cell
        dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def convert_to_date_formate(wb, start, end, data_list):
    sheet = wb.active
    filtered_data = [entry for entry in data_list if entry["Operation"] == "date_Column"]
    # print(filtered_data)
    for data in filtered_data:
        for row in range(start, end):
            try:
                sheet[f"{data["Reference_Column"]}{row}"].value = datetime.strptime(sheet[f"{data["Reference_Column"]}{row}"].value, "%d-%b-%y").date()
            except Exception:
                try:
                    sheet[f"{data["Reference_Column"]}{row}"].value = datetime.strptime(str(sheet[f"{data["Reference_Column"]}{row}"].value), "%d-%m-%Y").date()
                except Exception:
                    try:
                        sheet[f"{data["Reference_Column"]}{row}"].value = datetime.strptime(str(sheet[f"{data["Reference_Column"]}{row}"].value),"%d/%m/%Y").date()
                    except Exception:
                        try:
                            sheet[f"{data["Reference_Column"]}{row}"].value = datetime.strptime(str(sheet[f"{data["Reference_Column"]}{row}"].value),"%d-%b-%Y").date()
                        except Exception:
                            try:
                                sheet[f"{data["Reference_Column"]}{row}"].value = datetime.strptime(str(sheet[f"{data["Reference_Column"]}{row}"].value),"%d/%m/%y").date()
                            except Exception:
                                try:
                                    sheet[f"{data["Reference_Column"]}{row}"].value = datetime.strptime(str(sheet[f"{data["Reference_Column"]}{row}"].value),"%d/%b/%Y").date()
                                except Exception:
                                    try:
                                        sheet[f"{data["Reference_Column"]}{row}"].value = datetime.strptime(str(sheet[f"{data["Reference_Column"]}{row}"].value), "%d-%b-%Y").date()
                                    except Exception:
                                        try:
                                            sheet[f"{data["Reference_Column"]}{row}"].value = datetime.strptime(str(sheet[f"{data["Reference_Column"]}{row}"].value), "%b %d, %Y").date()
                                        except Exception:
                                            try:
                                                sheet[f"{data["Reference_Column"]}{row}"].value = datetime.strptime(str(sheet[f"{data["Reference_Column"]}{row}"].value), "%d %b %Y").date()
                                            except Exception:
                                                pass
    return wb


def fetch_data_from_db(bank_value, type_value, table_name):
    config = configparser.ConfigParser()
    config.read(".env")
    postgres_credentials = {
        "user": config.get("DEFAULT", "USER"),
        "password": config.get("DEFAULT", "PASSWORD"),
        "host": config.get("DEFAULT", "HOST"),
        "port": config.get("DEFAULT", "PORT"),
        "database": config.get("DEFAULT", "DATABASE"),
    }
    # Construct the connection string
    database_url = f"postgresql://{postgres_credentials['user']}:{postgres_credentials['password']}@{postgres_credentials['host']}:{postgres_credentials['port']}/{postgres_credentials['database']}"
    # Connect to the PostgreSQL database
    conn = psycopg2.connect(database_url)
    # Create a cursor object to execute SQL queries
    cursor = conn.cursor(cursor_factory=extras.RealDictCursor)
    try:
        # Construct the SELECT query with specific columns and WHERE conditions using f-string
        select_query = f"SELECT * FROM {table_name} WHERE \"Bank\" = %s AND \"Type\" = %s"
        # Execute the SELECT query with the provided values for "Bank" and "Type"
        cursor.execute(select_query, (bank_value, type_value))
        # Fetch all the rows as a list of dictionaries
        result = cursor.fetchall()
        # Convert the result to JSON format
        json_data = json.dumps(result, indent=2)
        return json_data
    finally:
        # Close the cursor and connection
        cursor.close()
        conn.close()


def get_unwanted_row_index(wb, start, end, data_list):
    sheet = wb.active
    row_index = []
    filtered_data = [entry for entry in data_list if entry["Operation"] == "delete_row"]
    for data in filtered_data:
        reftext = data["Keywords"]
        refcolumn = data["Reference_Column"]
        for row in range(start, end):
            if sheet[f"{refcolumn}{row}"].value is not None:
                if reftext in str(sheet[f"{refcolumn}{row}"].value):
                    row_index.append(row)
    return row_index


def get_row_lessthan_string_length(wb, start, end, columns, length):
    sheet = wb.active
    cells_to_none = []
    for column in columns:
        for row in range(start, end):
            if len(str(sheet[f"{column}{row}"].value)) <= length:
                cells_to_none.append(row)  # delete row
    return cells_to_none


def cell_to_none(wb, start, end, columns, rows):
    sheet = wb.active
    for column in columns:
        for row in rows:
            sheet[f"{column}{row}"].value = None
    return wb


def align_string_in_all_cells(wb, bank, type):
    sheet = wb.active
    start_column = 65
    max_column = sheet.max_column
    start_row = 1
    max_row = sheet.max_row
    for column in range(start_column, 65+max_column):
        for row in range(start_row, max_row):
            sheet[f"{chr(column)}{row}"].value = str(sheet[f"{chr(column)}{row}"].value).replace('\n', '')
    return wb


def find_column_index_by_header(wb, header):
    sheet = wb.active
    column_index = None
    for column in range(65, 65 + sheet.max_column):  # iterating through all the columns using ascii values
        if header in str(sheet[f"{chr(column)}1"].value):  # if a header text is in the 1st cell of the column
            column_index = column  # store the column index
            break
    return column_index


def delete_closing_balance_row(wb, column_mapping_data):
    sheet = wb.active
    column_mapping_data = json.loads(column_mapping_data)  # got a problem while reading json without loads
    narration = set(item["Source_Column"] for item in column_mapping_data if item["Target_Column"] == "Narration")
    narration = list(narration)[0]
    chequeno_refno = set(item["Source_Column"] for item in column_mapping_data if item["Target_Column"] == "ChequeNo_RefNo")
    chequeno_refno = list(chequeno_refno)[0]
    withdrawal = set(item["Source_Column"] for item in column_mapping_data if item["Target_Column"] == "Withdrawal")
    withdrawal = list(withdrawal)[0]
    deposit = set(item["Source_Column"] for item in column_mapping_data if item["Target_Column"] == "Deposit")
    deposit = list(deposit)[0]
    balance = set(item["Source_Column"] for item in column_mapping_data if item["Target_Column"] == "Balance")
    balance = list(balance)[0]
    # print(narration)
    # print(chequeno_refno)
    # print(withdrawal)
    # print(deposit)
    # print(balance)
    narration = find_column_index_by_header(wb, narration)
    chequeno_refno = find_column_index_by_header(wb, chequeno_refno)
    withdrawal = find_column_index_by_header(wb, withdrawal)
    deposit = find_column_index_by_header(wb, deposit)
    balance = find_column_index_by_header(wb, balance)
    # print(chr(narration))
    # print(chr(chequeno_refno))
    # print(chr(withdrawal))
    # print(chr(deposit))
    # print(chr(balance))
    if (sheet[f"{chr(narration)}{sheet.max_row}"].value is None and
        sheet[f"{chr(chequeno_refno)}{sheet.max_row}"].value is None and
        sheet[f"{chr(withdrawal)}{sheet.max_row}"].value is None and
        sheet[f"{chr(deposit)}{sheet.max_row}"].value is None and
        sheet[f"{chr(balance)}{sheet.max_row}"].value == sheet[f"{chr(balance)}{sheet.max_row-1}"].value):
        wb = delete_row(wb, [sheet.max_row])
    return wb


def create_column(wb, header):
    sheet = wb.active
    last_column = 65 + column_count(wb)
    sheet[f"{chr(last_column)}1"].value = header
    return wb


def split_transaction_value_date(wb, start, end, data_list):
    sheet = wb.active
    filtered_data = [entry for entry in data_list if entry["Operation"] == "split_transaction_value_date"]
    # wb = create_column(wb, "Value_Date")
    # value_date = find_column_index_by_header(wb, "Value_Date")
    last_column = 65 + column_count(wb)
    for data in filtered_data:
        refcolumn = data["Reference_Column"]
        refstring = data["Keywords"]
        for row in range(start, end):
            if row == start:
                temp = str(sheet[f"{refcolumn}{row}"].value).split(refstring)
                sheet[f"{chr(last_column)}{row}"].value = temp[1].replace(")", "")
            elif refstring in str(sheet[f"{refcolumn}{row}"].value):
                temp = str(sheet[f"{refcolumn}{row}"].value).split(refstring)
                sheet[f"{refcolumn}{row}"].value = temp[0]
                sheet[f"{chr(last_column)}{row}"].value = temp[1].replace(")", "")

    return wb


def main_fun(wb, bank, type, caller):
    sheet = wb.active
    bank_value = bank
    type_value = type
    json_result = fetch_data_from_db(bank_value, type_value, table_name="ksv.bank_statement_operations")
    column_mapping = fetch_data_from_db(bank_value, type_value, table_name="ksv.bank_statement_column_mapping")
    # print(json_result)
    # print(column_mapping)
    data_list = json.loads(json_result)

    distinct_operations_json = set(record["Operation"] for record in data_list)
    distinct_operations = list(distinct_operations_json)
    print(distinct_operations)

    wb = align_string_in_all_cells(wb, bank, type)

    start, end = get_start_end_row_index(wb, data_list)

    wb = remove_string_none(wb, start, end + 1)

    # wb = string_align(wb, start, end + 1)
    wb = empty_cell_to_none(wb, start + 1, end + 1)

    if "delete_row_by_range" in distinct_operations:
        rows_to_delete = delete_rows_by_range(wb, start + 1, sheet.max_row, data_list)
        wb = delete_row(wb, rows_to_delete)

    start, end = get_start_end_row_index(wb, data_list)

    wb = delete_unwanted_rows(wb, start - 1, end - 1, bank, type)
    start, end = get_start_end_row_index(wb, data_list)

    start, end = get_start_end_row_index(wb, data_list)

    if "delete_row" in distinct_operations:
        rows_to_delete = get_unwanted_row_index(wb, start + 1, end + 1, data_list)
        wb = delete_row(wb, rows_to_delete)

    start, end = get_start_end_row_index(wb, data_list)
    end = sheet.max_row

    if "split_transaction_value_date" in distinct_operations:
        wb = split_transaction_value_date(wb, start, end + 1, data_list)

    if "date_Column" in distinct_operations:
        wb = convert_to_date_formate(wb, start + 1, end + 1, data_list)

    if bank == "icici" and type == "type3":
        wb, columns = icici3_merging_date_column(wb, start + 1, end + 1, data_list)
        wb = find_replace(wb, start + 1, end + 1, find="None", replace="", columns=columns)
        cells_to_none = get_row_lessthan_string_length(wb, start + 1, end + 1, columns, length=5)
        wb = cell_to_none(wb, start, end, columns, cells_to_none)

    if bank == "sbi" and type == "type1":
        wb, columns = sbi1_merging_date_column(wb, start + 1, end + 1, data_list)
        wb = find_replace(wb, start + 1, end + 1, find="None", replace="", columns=columns)
        cells_to_none = get_row_lessthan_string_length(wb, start + 1, end + 1, columns, length=5)
        wb = cell_to_none(wb, start, end, columns, cells_to_none)

    wb = merging_rows(wb, start + 1, end + 1, data_list)

    wb = remove_string_none(wb, start, end + 1)

    wb = empty_cell_to_none(wb, start + 1, end + 1)

    rows_to_delete_merged_rows = get_unwanted_row_index_merged_row(wb, start + 1, end + 1, data_list)
    wb = delete_row(wb, rows_to_delete_merged_rows)
    start, end = get_start_end_row_index(wb, data_list)

    if "date_Column" in distinct_operations:
        wb = convert_to_date_formate(wb, start + 1, end + 1, data_list)

    if "remove_string" in distinct_operations:
        filtered_data = [entry for entry in data_list if entry["Operation"] == "remove_string"]
        column_json = set(item["Reference_Column"] for item in filtered_data if item["Operation"] == "remove_string")
        columns = list(column_json)
        find_json = set(item["Keywords"] for item in filtered_data if item["Operation"] == "remove_string")
        find = list(find_json)
        for value in find:
            wb = find_replace(wb, start, end + 1, value, "", columns)

    if "delete_closing_balance" in distinct_operations:
        wb = delete_closing_balance_row(wb, column_mapping)

    if "align_column_data" or "align_error_record" in distinct_operations:
        if bank == "canara" and type == "type1":
           wb = canara1_column_alignment(wb, start + 1, end + 1, data_list)
        if bank == "hdfc" and type == "type1":
           wb = hdfc1_column_error_records(wb, start + 1, end + 1, data_list)
        # if bank == "icici" and type == "type3":
        #     wb = icici3_column_alignment(wb, start + 1, end + 1, data_list)
        if bank == "kotak" and type == "type1":
            wb = kotak1_column_error_records(wb, start + 1, end + 1, data_list)
        if bank == "tmb" and type == "type1":
            wb = tmb1_column_error_records(wb, start + 1, end + 1, data_list)
    return wb


def delete_files_with_criteria(folder_path, keyword, extension):
    # Get a list of all files in the folder
    files = os.listdir(folder_path)
    # Iterate through the files and delete those with the specified extension and keyword
    for file_name in files:
        if extension in file_name and keyword in file_name:
            file_path = os.path.join(folder_path, file_name)
            os.remove(file_path)


def convert_url_to_bytes(pdf_url):
    bytes_list = []
    response = requests.get(pdf_url)
    response.raise_for_status()  # Check for any request errors
    bytes_stream = BytesIO(response.content)
    reader = PdfReader(bytes_stream)

    for page in reader.pages:
        writer = PdfWriter()
        writer.add_page(page)
        with BytesIO() as bytes_stream:
            writer.write(bytes_stream)
            bytes_stream.seek(0)
            bytes_list.append(bytes_stream.getvalue())
    return bytes_list


def convert_bytes_to_excel(pdf_bytes):
    def create_output_excel(output_xlsx):
        # Create a new Excel workbook
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = ''
        workbook.save(output_xlsx)

    def join_data(temp_xlsx, output_xlsx):
        # Open the source and destination workbooks
        source_workbook_1 = load_workbook(output_xlsx)
        sheet1 = source_workbook_1.active
        source_workbook_2 = load_workbook(temp_xlsx)
        sheet2 = source_workbook_2['Sheet1']
        # Append data from the source to the destination workbook
        for row in sheet2.iter_rows(min_row=2, values_only=True):
            sheet1.append(row)
        # Save the changes to the destination workbook
        source_workbook_1.save(output_xlsx)

    def convert_excel(pdf, temp_xlsx, output_xlsx):
        # Convert PDF bytes to Excel and join data
        with open(temp_xlsx, 'wb') as f:
            document = ap.Document(BytesIO(pdf))
            save_option = ap.ExcelSaveOptions()
            save_option.minimize_the_number_of_worksheets = True
            document.save(f, options=save_option)
        join_data(temp_xlsx, output_xlsx)

    # Generate timestamp for unique filenames
    now = datetime.now()
    t = now.strftime("__%d-%m-%Y-%H-%M-%S")
    # Define filenames
    output_xlsx = 'output' + t + '.xlsx'
    temp_xlsx = 'temp' + t + '.xlsx'
    # Create an initial Excel workbook
    create_output_excel(output_xlsx)
    # Convert PDF bytes to Excel and join data for each page
    for page_bytes in pdf_bytes:
        convert_excel(page_bytes, temp_xlsx, output_xlsx)
    # Remove temporary file
    os.remove(temp_xlsx)

    # Load and return the final workbook
    return load_workbook(output_xlsx), output_xlsx


def pdf_to_excel_main(pdf_url, bank, type, caller):
    pdf_bytes = convert_url_to_bytes(pdf_url)  # converting the pdf URL to bytes
    output_workbook, output_workbook_xlsx = convert_bytes_to_excel(pdf_bytes)  # converting bytes to excel
    sheet = output_workbook.active  # getting the first active sheet
    max_column = sheet.max_column  # getting the max column
    if max_column < 2:  # if max column is < 2 its Insufficient Data
        response = {"data": None,
                    "file_name": None,
                    "msg": "Insufficient Data to convert_bytes_to_excel To Process Driver Dictionary"}
        return response
    else:
        response = main_fun(wb, bank, type, caller)  # receiving the response as json

    # Deleting the temp files created by the Aspose library in the project folder
    delete_files_with_criteria(folder_path="C:/Users/Admin/PycharmProjects/pythonProject1/KSV/FormatingExcelFiles", keyword="output", extension='.xlsx')  # deleting the temp file create in the project folder, created by aspose library
    delete_files_with_criteria(folder_path="C:/Users/Admin/PycharmProjects/pythonProject1/KSV/FormatingExcelFiles", keyword="temp", extension='.xlsx')  # deleting the temp file create in the project folder, created by aspose library
    print("pdf_to_excel_main : ", response)
    return response



if __name__ == "__main__":
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/axis_type1__17-01-2024-17-17-05.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/canara_type1__19-01-2024-17-28-41.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/city_union_type1__22-01-2024-17-15-42.xlsx"  # 0 not coming
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/dbs_type1__22-01-2024-19-37-31.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/equitas_type1__22-01-2024-20-05-47.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/federal_type1__22-01-2024-20-28-05.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/hdfc_type1__23-01-2024-10-46-01.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/ICICI_type1__23-01-2024-13-31-13.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/ICICI_type2__23-01-2024-14-19-57.xlsx"  # align_column_data not required
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/icici_type3__26-01-2024-16-14-58.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/indian_bank_type1__23-01-2024-22-29-53.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/indusind_type1__24-01-2024-21-59-57.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/Indusind_type2__28-01-2024-12-20-58.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/IOB_type1__28-01-2024-13-23-48.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/IOB_type2__28-01-2024-16-20-25.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/Kotak_type1__28-01-2024-18-06-39.xlsx"
    # path = "kotak type 2 not converted by aspose"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/SBI_type1__28-01-2024-19-01-41.xlsx"
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/TMB_type1__28-01-2024-19-18-55.xlsx"
    path = "C:/Users/Admin/Desktop/KSV/source_excel_files/yes_bank_type1__28-01-2024-19-44-49.xlsx"
    wb = openpyxl.load_workbook(path)
    # result = main_fun(wb=wb, bank="yes", type="type1", caller="appsmith")
    pdf_to_excel_main(pdf_url=, bank=, type, caller)
    result.save('C:/Users/Admin/Desktop/temp_output.xlsx')
