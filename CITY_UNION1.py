import os
from datetime import datetime
from openpyxl.utils import column_index_from_string
import openpyxl

from FormatingExcelFiles.CommonClass import Excel


def dateConversion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d/%m/%Y").date()
    return wb


def deleteHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def deleteFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


def cityunion1_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 6
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def cityunion1_main(wb):
    sheet = wb.active
    if cityunion1_validation(wb):
        raise Exception(f"<=INVALID FORMATE =>  <Count Of Column Mismatch>")
    else:
        startText = "DATE"
        endText = "TOTAL"
        startEndRefColumn = "A"
        deleteFlagStartText = "Regd. Office"
        deleteFlagStopText = "DATE"
        deleteFlagRefColumn = "A"
        dateConversionColumn = "A"
        refHeaderText1 = "DATE"
        refHeaderText2 = "DESCRIPTION"
        refHeaderText3 = "CHEQUE NO"
        refHeaderText4 = "DEBIT"
        refHeaderText5 = "CREDIT"
        refHeaderText6 = "BALANCE"
        headerText1 = "Transaction_Date"
        headerText2 = "Narration"
        headerText3 = "ChequeNo_RefNo"
        headerText4 = "Withdrawal"
        headerText5 = "Deposit"
        headerText6 = "Balance"
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]
        negativeValueColumnRefText1 = "Withdrawal"
        headerTextToReplaceNone1 = "ChequeNo_RefNo"
        headerTextToReplaceNone2 = "Withdrawal"
        headerTextToReplaceNone3 = "Value_Date"
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        dupHeadersRemoved = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagStopText,
                                              deleteFlagRefColumn)
        start, end = Excel.get_start_end_row_index(dupHeadersRemoved, startText, endText, startEndRefColumn)
        footerDeleted = deleteFooter(dupHeadersRemoved, end - 1)  # end-1 to Include Last Row
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndRefColumn)
        convertedToDateA = dateConversion(headerDeleted, start + 1, end + 1,
                                          dateConversionColumn)  # start+1 to Skip Header, end+1 to include last row
        lastCol = 65 + Excel.column_count(wb)
        transdate = Excel.alter_header_name(convertedToDateA, refHeaderText1, headerText1, lastCol)
        narration = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)
        chqno = Excel.alter_header_name(narration, refHeaderText3, headerText3, lastCol)
        debit = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)
        credit = Excel.alter_header_name(debit, refHeaderText5, headerText5, lastCol)
        balance = Excel.alter_header_name(credit, refHeaderText6, headerText6, lastCol)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        slnoCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))
        columnFinalised = Excel.finalise_column(balance, columns)
        negativeColumnChecked = Excel.check_neagativeValue_by_column(columnFinalised, negativeValueColumnRefText1)
        replacedNoneCHQNO = Excel.empty_cell_to_none(negativeColumnChecked, start, end + 1, headerTextToReplaceNone1)
        replacedNoneWITHDRAWAL = Excel.empty_cell_to_none(replacedNoneCHQNO, start, end + 1, headerTextToReplaceNone2)
        replacedNoneVALUEDATE = Excel.empty_cell_to_none(replacedNoneWITHDRAWAL, start, end + 1, headerTextToReplaceNone3)
        createdTransTypeColumn = Excel.transaction_type_column(replacedNoneVALUEDATE)
        return wb


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/CITY_UNION_BANK_-_SB-500101012199098__23-09-2023-18-18-25.xlsx"
    wb = openpyxl.load_workbook(path)
    result = cityunion1_main(wb)
    result.save("C:/Users/Admin/Desktop/FinalOutput/CITY_UNION1output.xlsx")
