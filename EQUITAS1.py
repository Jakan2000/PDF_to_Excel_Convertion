from datetime import datetime

import openpyxl

from CommonClass import Excel


def dateConversion(wb, start, end, column):
    """
        Convert date values in a specified column to a standard date format in the given Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index for date conversion (inclusive).
        - end (int): The ending row index for date conversion (exclusive).
        - column (str): The column letter (e.g., 'A', 'B') containing date values to be converted.

        Returns:
        - openpyxl.Workbook: The modified Workbook object after date conversion.

        Notes:
        - This function assumes the date values in the specified column are in the format '%d-%b-%Y'.
        - The converted dates are stored in the same column in the standard date format.
        - The 'start' parameter denotes the first row to start converting dates.
        - The 'end' parameter denotes the row where date conversion stops (exclusive).
        - The original Workbook is modified in place.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating from table data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%b-%Y").date()  # converting date to standard formate
    return wb


def deleteHeader(wb, start):
    """
        Delete rows above the specified start index row in the given Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The row index indicating the starting point for deletion (inclusive).

        Returns:
        - openpyxl.Workbook: The modified Workbook object after deleting the specified rows.

        Notes:
        - This function deletes rows from the active sheet of the Workbook.
        - The 'start' parameter denotes the row index from which rows will be deleted and onwards.
        - Rows are deleted in-place, and the original Workbook is modified.
        - If 'start' is 1, the entire sheet will be cleared, as it includes the header row.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterate from table data start row to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def deleteFooter(wb, end):
    """
        Delete all rows below the specified end index row in the given Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - end (int): The row index indicating the ending point for deletion (exclusive).

        Returns:
        - openpyxl.Workbook: The modified Workbook object after deleting the specified rows.

        Notes:
        - This function deletes rows from the active sheet of the Workbook.
        - The 'end' parameter denotes the row index until which rows will be deleted (exclusive).
        - Rows are deleted in-place, and the original Workbook is modified.
        - If 'end' is equal to the maximum row count, it will result in an empty sheet.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterate through max row in sheet to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def removeNoneRows(wb, start, end, column):
    """
        Remove rows with 'None' values in the specified column within the given Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The row index indicating the starting point for removal (inclusive).
        - end (int): The row index indicating the ending point for removal (exclusive).
        - column (str): The column letter (e.g., 'A', 'B') containing the 'None' values to be removed.

        Returns:
        - openpyxl.Workbook: The modified Workbook object after removing rows with 'None' values.

        Notes:
        - This function removes rows in-place from the active sheet of the Workbook.
        - The 'start' parameter denotes the row index from which the removal starts (inclusive).
        - The 'end' parameter denotes the row index until which the removal occurs (exclusive).
        - Rows are removed if the specified column's cell value is 'None'.
        - If 'start' and 'end' are equal, no rows will be removed.

    """
    sheet = wb.active
    for x in range(end - 1, start, -1):  # iterating through table data end row and table data start row
        if sheet[f"{column}{x}"].value is None:  # if column's cell value is none
            sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merge consecutive rows in a specified column within the given Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The row index indicating the starting point for row merging (inclusive).
        - end (int): The row index indicating the ending point for row merging (exclusive).
        - refColumn (str): The column letter (e.g., 'A', 'B') containing the reference values for row merging.
        - mergingColumn (str): The column letter (e.g., 'C', 'D') whose rows will be merged based on the reference column.

        Returns:
        - openpyxl.Workbook: The modified Workbook object after merging rows in the specified column.

        Notes:
        - This function modifies the Workbook in-place by merging rows in the specified column.
        - Rows are merged based on the consecutive presence of reference values in the 'refColumn'.
        - The 'start' parameter denotes the row index from which the merging starts (inclusive).
        - The 'end' parameter denotes the row index until which the merging occurs (exclusive).
        - If 'start' and 'end' are equal, no rows will be merged.
        - The merged data is concatenated and stored in the row where the first occurrence of the reference value is found.

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through start and end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def equitas1_validation(wb):
    """
        Validate the number of columns in the Workbook to ensure compatibility with the core logic.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.

        Returns:
        - bool: True if the number of columns in the Workbook is not equal to the expected count, False otherwise.

        Notes:
        - This function checks whether the Workbook has the expected number of columns required by the core logic.
        - The core logic is designed for a specific count of columns, and this function validates if the Workbook adheres to that count.
        - If the Workbook has a column count different from the expected count, it may indicate an incompatible format.

    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 6  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def equitas1_main(wb):
    """
        Execute core logic for processing Equitas Bank statements in Excel format.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.

        Returns:
        - dict: A dictionary containing processed Workbook and status message.
                The dictionary has two keys:
                - "data": The processed Workbook with modifications.
                - "msg": A status message indicating the success or failure of the operation.

        Notes:
        - This function serves as the main entry point for processing Equitas Bank statements in Excel format.
        - It validates the format of the Workbook using the `equitas1_validation` function.
        - If the validation is successful, it performs a series of operations to clean, merge, and standardize the data.
        - The processed Workbook is returned along with a status message in the dictionary.

    """
    sheet = wb.active
    if equitas1_validation(wb):  # validate columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return response with error msg
    else:
        startText = "Date"  # header text in column A
        endText = "*** End of the Statement ***"  # text define the end of the table data row
        startEndRefColumn = "A"  # reference column contains the start and end text
        deleteFlagStartText = "Page"    # starting row reference text to delete the rows by range
        deleteFlagStopText = "Deposit"  # ending row reference text to delete the rows by range
        deleteFlagRefColumn = "E"  # column containing starting row reference text and ending row reference text to delete the rows by range
        removeRowRefText = "INR"  # reference text to remove a row
        removeRowRefColumn = "D"  # reference column to remove the row
        columnToMerg = "C"  # column to merge the row data
        refColumnToMerg = "A"  # reference column (date column) to merge the rows of desired column
        dateConversionColumn1 = "A"  # convert date to standard date formate
        refHeaderText1 = "Date"  # header text to replace with standardised column name
        refHeaderText2 = "Reference No. / Cheque No."  # header text to replace with standardised column name
        refHeaderText3 = "Narration"  # header text to replace with standardised column name
        refHeaderText4 = "Withdrawal"  # header text to replace with standardised column name
        refHeaderText5 = "Deposit"  # header text to replace with standardised column name
        refHeaderText6 = "ClosingBalance"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "ChequeNo_RefNo"  # standard column name
        headerText3 = "Narration"  # standard column name
        headerText4 = "Withdrawal"  # standard column name
        headerText5 = "Deposit"  # standard column name
        headerText6 = "Balance"  # standard column name
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        negativeValueColumnRefText1 = "Withdrawal"  # no need of removing the negative value to positive
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        end = sheet.max_row  # assigning end row as max row
        dupHeaderRemoved = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagStopText, deleteFlagRefColumn)  # deleting the rows by range
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        INRrowsRemoved = Excel.remove_rows(dupHeaderRemoved, start, end, removeRowRefText, removeRowRefColumn)  # remove multiple rows by reference text in reference column
        start, end = Excel.get_start_end_row_index(INRrowsRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        mergedColumnC = mergingRows(INRrowsRemoved, start, end, refColumnToMerg, columnToMerg)  # merging the rows of desired column
        noneRowsRemoved = removeNoneRows(mergedColumnC, start, end, refColumnToMerg)  # remove none rows in with respect to desired column (date column)
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        footerDeleted = deleteFooter(noneRowsRemoved, end - 1)  # delete all the rows below the end(last) row end-1 to Include End Footer
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # delete rows above the start index row end-1 to Include End Footer
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        convertedDateA = dateConversion(headerDeleted, start + 1, end + 1, dateConversionColumn1)  # converting date to standard formate in a column, start+1 to Skip Header, end+1 to Include Last Row
        lastCol = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        transdate = Excel.alter_header_name(convertedDateA, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
        chqno = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        narration = Excel.alter_header_name(chqno, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        debit = Excel.alter_header_name(narration, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        credit = Excel.alter_header_name(debit, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(credit, refHeaderText6, headerText6, lastCol)  # alter header name from the excel file to the standard column name
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slnoCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))  # created new column-slno
        columnFinalised = Excel.finalise_column(slnoCreated, columns)  # standardizing count of column
        negativeValueChecked = Excel.check_neagativeValue_by_column(slnoCreated, negativeValueColumnRefText1)  # no need of converting negative value to positive value
        createdTransTypeColumn = Excel.transaction_type_column(negativeValueChecked)  # created new column transaction type
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/Equitas_-_6802_unlocked__23-09-2023-12-01-43.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    result = equitas1_main(wb)
    # result.save("C:/Users/Admin/Desktop/FinalOutput/EQUITAS1output.xlsx")
