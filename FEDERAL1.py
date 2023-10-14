import os
from datetime import datetime

import openpyxl

from KSV.FormatingExcelFiles.CommonClass import Excel


def dateHeaderAlter(wb, refText, actualText, lastCol):
    sheet = wb.active
    column = 65
    row = 1
    while column < lastCol:
        if refText in str(sheet[f"{chr(column)}{row}"].value) and len(str(sheet[f"{chr(column)}{row}"].value)) < 5:
            sheet[f"{chr(column)}{row}"].value = actualText
        column += 1
    return wb


def dateConvertion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), '%d-%m-%Y').date()
    return wb


def removeNone(wb, start, end, column):
    sheet = wb.active
    for x in range(start, end):
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")
    return wb


def deleteHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def deleteRow(wb, start, refText, refColumn):
    sheet = wb.active
    for x in range(sheet.max_row, start, -1):
        if refText in str(sheet[f"{refColumn}{x}"].value):
            sheet.delete_rows(x)
    return wb


def deleteFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


def removingNullRows(wb, start, end, column):
    sheet = wb.active
    for x in range(end, start, -1):
        if sheet[f"{column}{x}"].value is None:
            sheet.delete_rows(x)
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st = ""
    for m in range(1, len(dataToMerge)):
        st += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st
    dataToMerge = []
    return wb


def removeRowsOnRange(wb, start, end, startText, endText, column):
    sheet = wb.active
    delete_flag = False
    rows_to_delete = []
    for i in range(start, end):
        if startText in str(sheet[f"{column}{i}"].value):
            delete_flag = True
        if delete_flag:
            rows_to_delete.append(i)
        if endText in str(sheet[f"{column}{i}"].value):
            delete_flag = False
    for idx in reversed(rows_to_delete):
        sheet.delete_rows(idx)
    return wb


def federal1_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 10
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def federal1_main(wb):
    sheet = wb.active
    if federal1_validation(wb):
        raise Exception(f"<<= {os.path.basename(path)} =>> INVALID FORMATE =>  <Count Of Column Mismatch>")
    else:
        startText = "Particulars"
        endText = "GRAND TOTAL"
        startEndRefColumn = "C"
        dupHeaderStartText = "Page"
        dupHeaderEndText = "Deposits"
        dupHeaderRefColumn = "H"
        columnToMerg1 = "C"
        refColumnToMerg = "A"
        deleteOpenBalText = "Opening Balance"
        deleteOpenBalRefColumn = "C"
        stringAlignColumn1 = "A"
        stringAlignColumn2 = "B"
        stringAlignColumn3 = "C"
        stringAlignColumn4 = "D"
        stringAlignColumn5 = "E"
        stringAlignColumn6 = "F"
        stringAlignColumn7 = "G"
        stringAlignColumn8 = "H"
        stringAlignColumn9 = "I"
        deleteColumnRefText1 = "TranType"
        deleteColumnRefText2 = "Tran Id"
        deleteColumnRefText3 = "Cr/Dr"
        dateConversionColumn1 = "A"
        dateConversionColumn2 = "B"
        refHeaderText1 = "Date"
        refHeaderText2 = "Value Date"
        refHeaderText3 = "Particulars"
        refHeaderText4 = "ChequeDetails"
        refHeaderText5 = "Withdrawals"
        refHeaderText6 = "Deposits"
        refHeaderText7 = "Balance"
        headerText1 = "Transaction_Date"
        headerText2 = "Value_Date"
        headerText3 = "Narration"
        headerText4 = "ChequeNo_RefNo"
        headerText5 = "Withdrawal"
        headerText6 = "Deposit"
        headerText7 = "Balance"
        negativeValueColumnRefText1 = "Withdrawal"
        headerTextToMakeEmptyCellsToNone1 = "Withdrawal"
        headerTextToMakeEmptyCellsToNone2 = "Deposit"
        headerTextToMakeEmptyCellsToNone3 = "ChequeNo_RefNo"
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        dupHeaderRemoved = removeRowsOnRange(wb, start, end, dupHeaderStartText, dupHeaderEndText, dupHeaderRefColumn)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        columnMergedC = mergingRows(dupHeaderRemoved, start + 2, end, refColumnToMerg, columnToMerg1)  # start+2 to Skip Opening Balance Row
        noneRowsRemoved = removingNullRows(columnMergedC, start + 1, end - 1, refColumnToMerg)  # end-1 to Skip End Footer
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)
        footerDeleted = deleteFooter(noneRowsRemoved, end - 1)  # end-1 to Include
        openBalRowDeleted = deleteRow(wb, start, deleteOpenBalText, deleteOpenBalRefColumn)
        headerDeleted = deleteHeader(openBalRowDeleted, start - 1)  # start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndRefColumn)
        stringAlignedA = Excel.string_align(headerDeleted, start, end + 1, stringAlignColumn1)  # end+1 to Include Last Row
        stringAlignedB = Excel.string_align(stringAlignedA, start, end + 1, stringAlignColumn2)  # end+1 to Include Last Row
        stringAlignedC = Excel.string_align(stringAlignedB, start, end + 1, stringAlignColumn3)  # end+1 to Include Last Row
        stringAlignedD = Excel.string_align(stringAlignedC, start, end + 1, stringAlignColumn4)  # end+1 to Include Last Row
        stringAlignedE = Excel.string_align(stringAlignedD, start, end + 1, stringAlignColumn5)  # end+1 to Include Last Row
        stringAlignedF = Excel.string_align(stringAlignedE, start, end + 1, stringAlignColumn6)  # end+1 to Include Last Row
        stringAlignedG = Excel.string_align(stringAlignedF, start, end + 1, stringAlignColumn7)  # end+1 to Include Last Row
        stringAlignedH = Excel.string_align(stringAlignedG, start, end + 1, stringAlignColumn8)  # end+1 to Include Last Row
        stringAlignedI = Excel.string_align(stringAlignedH, start, end + 1, stringAlignColumn9)  # end+1 to Include Last Row
        removedNoneA = removeNone(stringAlignedI, start, end + 1, stringAlignColumn1)  # end+1 to Include Last Row
        removedNoneB = removeNone(removedNoneA, start, end + 1, stringAlignColumn2)  # end+1 to Include Last Row
        removedNoneC = removeNone(removedNoneB, start, end + 1, stringAlignColumn3)  # end+1 to Include Last Row
        removedNoneF = removeNone(removedNoneC, start, end + 1, stringAlignColumn6)  # end+1 to Include Last Row
        removedNoneG = removeNone(removedNoneF, start, end + 1, stringAlignColumn7)  # end+1 to Include Last Row
        removedNoneH = removeNone(removedNoneG, start, end + 1, stringAlignColumn8)  # end+1 to Include Last Row
        removedNoneI = removeNone(removedNoneH, start, end + 1, stringAlignColumn9)  # end+1 to Include Last Row
        trantypeDeleted = Excel.delete_column(removedNoneI, deleteColumnRefText1)
        tranIdDeleted = Excel.delete_column(removedNoneI, deleteColumnRefText2)
        crdrDeleted = Excel.delete_column(removedNoneI, deleteColumnRefText3)
        convertedDateA = dateConvertion(crdrDeleted, start + 1, end + 1, dateConversionColumn1)  # start+1 to Skip Header, end+1 to Include Last Row
        convertedDateB = dateConvertion(convertedDateA, start + 1, end + 1, dateConversionColumn2)  # start+1 to Skip Header, end+1 to Include Last Row
        lastCol = 65 + Excel.column_count(wb)
        transdate = dateHeaderAlter(convertedDateB, refHeaderText1, headerText1, lastCol)
        valuedate = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)
        naration = Excel.alter_header_name(valuedate, refHeaderText3, headerText3, lastCol)
        chqno = Excel.alter_header_name(naration, refHeaderText4, headerText4, lastCol)
        debit = Excel.alter_header_name(chqno, refHeaderText5, headerText5, lastCol)
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        slnoCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))
        neagativeValueChecked = Excel.check_neagativeValue_by_column(slnoCreated, negativeValueColumnRefText1)
        withdrawalNoneReplaced = Excel.empty_cell_to_none(neagativeValueChecked, start + 1, end + 1, headerTextToMakeEmptyCellsToNone1)
        depositNoneReplaced = Excel.empty_cell_to_none(withdrawalNoneReplaced, start + 1, end + 1, headerTextToMakeEmptyCellsToNone2)
        chqnoNoneReplaced = Excel.empty_cell_to_none(depositNoneReplaced, start + 1, end + 1, headerTextToMakeEmptyCellsToNone3)
        createdTransTypeColumn = Excel.transaction_type_column(chqnoNoneReplaced)
        return wb


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/2. R RAVICHANDRAN - Federal - 2416 Pass - RAVI016 __11-09-2023-15-59-25.XLSX"
    wb = openpyxl.load_workbook(path)
    result = federal1_main(wb)
    result.save('C:/Users/Admin/Desktop/FinalOutput/FEDERAL1output.xlsx')
