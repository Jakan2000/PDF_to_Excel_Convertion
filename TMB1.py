from datetime import datetime

import openpyxl

from AlignmentData import addAlignmentData
from CommonClass import Excel


def string_in_column(wb, text):
    """
        Find the column containing the specified reference text in the given Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be searched.
        - text (str): The reference text to be found in the workbook.

        Returns:
        - str: The column letter (in uppercase) containing the reference text. If the text is not found, returns None.

        Note:
        This function iterates through each column and row in the provided Excel workbook ('wb') to find the specified
        reference text ('text'). It returns the column letter (in uppercase) where the text is found. If the text is not
        found in any cell, the function returns None.

    """
    sheet = wb.active
    for column in range(65, sheet.max_column+65):  # iterating through A to sheet max column
        for row in range(1, sheet.max_row):  # iterating through 1st row to sheet max row
            if text in str(sheet[f"{chr(column)}{row}"].value):  # if reference text in cell value
                return chr(column)  # return column ascii value converted to character


def dateConversion(wb, start, end, column):
    """
        Convert the date values in a specified column of an Excel workbook to a standard date format.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be modified.
        - start (int): The starting row index for the table data.
        - end (int): The ending row index (exclusive) for the table data.
        - column (str): The column letter (in uppercase) containing the date values to be converted.

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook with date values converted to the standard format.

        Note:
        This function iterates through the specified column of the provided Excel workbook ('wb') and converts the date
        values to a standard date format ("%d-%m-%Y"). The conversion is applied to rows from the 'start' index to the
        'end-1' index (exclusive).

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%m-%Y").date()  # converting to standard date formate
    return wb


def alignColumns(wb, start, end, headData, refColumnToAlign):
    """
        Align data in specified columns manually and handle misaligned records.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be modified.
        - start (int): The starting row index for the table data.
        - end (int): The ending row index (exclusive) for the table data.
        - headData (list): A list containing header data for the columns to be aligned.
        - refColumnToAlign (str): The reference column letter (in uppercase) to align the specified columns.

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook with aligned columns and handled misaligned records.

        Note:
        This function manually aligns the data in specified columns of the provided Excel workbook ('wb') using a reference
        column ('refColumnToAlign'). It handles misaligned records by updating the 'Narration' column with the string
        "Error Record" and setting corresponding cells in other columns to None. Additionally, the function logs the
        misaligned records in a separate Excel file.

    """
    sheet = wb.active
    error_records = []  # array to store misaligned error records
    for i in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{refColumnToAlign}{i}"].value is None:  # if cell value is none
            data = [headData[0], headData[1], headData[2]]  # [headData[0] -> account number, headData[1] -> name, headData[2] -> period]
            data.append(i - 1)  # appending row index to data array
            data.append(str(sheet[f"A{i}"].value))  # appending A column cell value
            data.append(str(sheet[f"B{i}"].value))  # appending B column cell value
            data.append(str(sheet[f"C{i}"].value))  # appending C column cell value
            data.append(str(sheet[f"D{i}"].value))  # appending D column cell value
            data.append(str(sheet[f"E{i}"].value))  # appending E column cell value
            data.append(str(sheet[f"F{i}"].value))  # appending F column cell value
            addAlignmentData(data)  # adding the miss aligned  data to excel file
            error_records.append(i)  # appending row index of sheet to error records array
    for x in error_records:  # iterating through error records array
        sheet[f"B{x}"].value = "Error Record"  # assigning String "Error Records" to narration column cell value
        sheet[f"C{x}"].value = None  # assign None to C column cell values
        sheet[f"D{x}"].value = None  # assigning None to D column cell value
        sheet[f"E{x}"].value = None  # assigning None to E column cell value
        sheet[f"F{x}"].value = None  # assigning None to F column cell value
    return wb


def aligningAllColumns(wb, start, end, refColumn):
    """
        Align misaligned column data in the specified range based on a reference column.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be modified.
        - start (int): The starting row index for the table data.
        - end (int): The ending row index (exclusive) for the table data.
        - refColumn (str): The reference column letter (in uppercase) based on which misaligned data is aligned.

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook with aligned misaligned column data.

        Note:
        This function aligns misaligned column data in the specified range of the provided Excel workbook ('wb') based on
        the content of a reference column ('refColumn'). It specifically handles cases where the reference column cell value
        is None.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{refColumn}{i}"].value is None:  # if reference column cell value is none
            sheet[f'F{i}'].value = sheet[f'E{i}'].value  # assign E column cell value to F column cell value
            sheet[f'E{i}'].value = sheet[f'D{i}'].value  # assign D column cell value to E column cell value
            sheet[f'D{i}'].value = sheet[f'C{i}'].value  # assign C column cell value to D column cell value
            sheet[f'C{i}'].value = None  # assign None to C column cell value
    return wb


def deleteHeader(wb, start):
    """
        Delete header rows from an Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook to be modified.
        - start (int): The row index indicating the starting point for deleting rows (inclusive).

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook with header rows deleted.

        Note:
        This function deletes rows from the active sheet of the provided Excel workbook ('wb') starting from the specified
        row index ('start') going upwards. It is designed to remove header rows in a table.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def headerData(wb, start, end):
    """
        Get header data from an Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook containing the header data.
        - start (int): The row index indicating the starting point for searching header data (inclusive).
        - end (int): The row index indicating the ending point for searching header data (inclusive).

        Returns:
        - list: An array containing extracted header data [account_number, name, period].

        Note:
        This function searches for relevant header information in the specified range of rows (from 'start' to 'end') of the
        active sheet in the provided Excel workbook ('wb'). It identifies and extracts account number, name, and period
        information based on specific patterns found in the Excel sheet.

    """
    sheet = wb.active
    acnum = "Undefined"  # creating a variable for account number and assigning string "Undefined"
    name = "Undefined"  # creating a variable for name and assigning string "Undefined"
    period = "Undefined"  # creating a variable for period and assigning string "Undefined"
    for i in range(start, 0, -1):  # iterating through table data start row to 1st row of sheet
        if "Name" in str(sheet[f"A{i}"].value):  # if string "Name" in A column cell value
            name = str(sheet[f"B{i}"].value).replace(":", "").strip()  # extract name from junk data
        if "Statement for A/c" in str(sheet[f"A{i}"].value):  # if string "Statement for A/c" in A column cell value
            spl = str(sheet[f"A{i}"].value).split("Between")
            s = spl[0].split("A/c")
            acnum = s[1].strip()  # storing account number
            period = str(spl[1]).replace("and", "to").strip()  # storing period
    headData = [acnum, name, period]  # storing account number, name, period in an array
    return headData  # returning header data array


def deleteFooter(wb, end):
    """
        Delete footer data from an Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook containing the footer data.
        - end (int): The row index indicating the ending point for deleting footer data (exclusive).

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook after deleting footer data.

        Note:
        This function deletes rows in the specified range (from 'end' to the last row) of the active sheet in the provided
        Excel workbook ('wb'). It is intended to remove footer information from the Excel sheet.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through sheet max row to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def removeNoneRows(wb, start, end, column):
    """
        Delete rows from an Excel workbook if a specified reference column cell value is None.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook containing the data.
        - start (int): The row index indicating the starting point for deleting rows.
        - end (int): The row index indicating the ending point for deleting rows (exclusive).
        - column (str): The reference column letter (e.g., 'A', 'B', 'C') to check for None values.

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook after deleting rows.

        Note:
        This function deletes rows in the specified range (from 'start' to 'end' - 1) of the active sheet in the provided
        Excel workbook ('wb') if the specified reference column cell value is None.

    """
    sheet = wb.active
    for x in range(end - 1, start, -1):  # iterating through table data end row to table data start row
        if sheet[f"{column}{x}"].value is None:  # if reference column cell value is none
            sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merge rows of a specified column based on the values in a reference column.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook containing the data.
        - start (int): The row index indicating the starting point for merging rows.
        - end (int): The row index indicating the ending point for merging rows (exclusive).
        - refColumn (str): The reference column letter (e.g., 'A', 'B', 'C') to identify starting rows for merging.
        - mergingColumn (str): The column letter (e.g., 'A', 'B', 'C') whose rows need to be merged.

        Returns:
        - openpyxl.workbook.workbook.Workbook: The modified Excel workbook after merging rows.

        Note:
        This function merges consecutive rows of the specified 'mergingColumn' based on the values in the 'refColumn'.
        Rows are merged if the value in the 'refColumn' is not None, indicating the starting row for merging.
        The last row in the specified range will also be merged.

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through start and end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def tmb1_validation(wb):
    """
        Validate the column count for the core logic of TMB1 format in an Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook containing the data.

        Returns:
        - bool: True if the column count is invalid, False if the column count is valid.

        Note:
        This function compares the maximum column count in the active sheet of the Excel workbook with the expected
        column count for the TMB1 format. If the column count is not equal to the expected count, it is considered
        as an invalid format.

    """
    sheet = wb.active
    max_column = sheet.max_column   # get max column in the sheet, using predefined function
    countOfColumn = 6  # the column count of our core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def tmb1_main(wb):
    """
        Main function to process and standardize data in the TMB1 format in an Excel workbook.

        Parameters:
        - wb (openpyxl.workbook.workbook.Workbook): The Excel workbook containing the TMB1 format data.

        Returns:
        - dict: A dictionary containing the processed workbook and an optional error message.

        Note:
        This function performs various data processing steps to standardize the TMB1 format in the provided Excel workbook.
        If the workbook format is invalid, it prints an error message and returns a dictionary with a None data attribute and
        an error message. Otherwise, it returns a dictionary with the processed workbook and a None message.

    """
    sheet = wb.active
    if tmb1_validation(wb):  # validating columns for the core logic written
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return response with error msg
    else:
        column = string_in_column(wb, text="Closing Balance")  # getting column with reference text (ascii value)
        if column == "D":  # if string_in_column() function returns column D
            startText = "Withdrawals"  # assign string "Withdrawals" to start text variable
        if column == "C":  # if string_in_column() function returns column C
            startText = "Chq. No."  # assign string "Chq. No." to start text variable
        startEndRefColumn = column  # assign string_in_column() function returned column to startEndRefColumn -> column containing start and end reference text
        endText = "Closing Balance"  # text define table data end column
        deleteFlagRefText1 = "Page"  # starting row reference text to delete multiple rows
        deleteFlagRefText2 = "Date"  # ending row reference text to delete multiple rows
        deleteFlagRefColumn = "A"  # column containing reference text to define start row and end row to delete multiple rows
        columnToMerg1 = "B"  # column to merge misaligned rows
        refColumnToMerg = "A"  # reference column to merge misaligned rows of other columns
        refColumnToAlignAllColumn = "F"  # reference column to align misaligned column data
        refColumnToAlign = "F"  # reference column to manually align misaligned column data (error records)
        refHeaderText1 = "Date"  # header text to replace with standardised column name
        refHeaderText2 = "Particulars"  # header text to replace with standardised column name
        refHeaderText3 = "Chq. No."  # header text to replace with standardised column name
        refHeaderText4 = "Withdrawals"  # header text to replace with standardised column name
        refHeaderText5 = "Deposits"  # header text to replace with standardised column name
        refHeaderText6 = "Balance(INR)"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Narration"  # standard column name
        headerText3 = "ChequeNo_RefNo"  # standard column name
        headerText4 = "Withdrawal"  # standard column name
        headerText5 = "Deposit"  # standard column name
        headerText6 = "Balance"  # standard column name
        dateConversionColumn1 = "A"  # column to convert date to standard date formate
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        pageNoRemoved = Excel.remove_rows(wb, start, end, deleteFlagRefText1, deleteFlagRefColumn)  # removing multiple rows with reference text
        start, end = Excel.get_start_end_row_index(pageNoRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        dupHeaderRemoved = Excel.remove_rows(pageNoRemoved, start, end, deleteFlagRefText2, deleteFlagRefColumn)  # removing multiple rows with reference text
        start, end = Excel.get_start_end_row_index(dupHeaderRemoved, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        mergedColumnB = mergingRows(dupHeaderRemoved, start, end, refColumnToMerg, columnToMerg1)  # merging rows of B column
        noneRowsRemoved = removeNoneRows(mergedColumnB, start, end, refColumnToMerg)  # deleting none rows
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)  #  get start and end row index to specify the data with in
        footerDeleted = deleteFooter(noneRowsRemoved, end - 1)  # end-1 to Include End Footer
        start, end = Excel.get_start_end_row_index(footerDeleted, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        headData = headerData(footerDeleted, start, end)  # get header data
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # delete header data, start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndRefColumn)   # get start and end row index to specify the data with in
        allColumnAligned = aligningAllColumns(headerDeleted, start, end + 1, refColumnToAlignAllColumn)  # alligning all columns if F column cell value is none
        alignColumns(allColumnAligned, start, end + 1, headData, refColumnToAlign)  # column data to align manually
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slCreated = Excel.create_slno_column(allColumnAligned, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
        lastCol = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        transdate = Excel.alter_header_name(slCreated, refHeaderText1, headerText1, lastCol)  # alter header name by standard column name
        naration = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name by standard column name
        chqno = Excel.alter_header_name(naration, refHeaderText3, headerText3, lastCol)  # alter header name by standard column name
        debit = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)  # alter header name by standard column name
        credit = Excel.alter_header_name(debit, refHeaderText5, headerText5, lastCol)  # alter header name by standard column name
        balance = Excel.alter_header_name(credit, refHeaderText6, headerText6, lastCol)  # alter header name by standard column name
        convertedDateA = dateConversion(balance, start + 1, end + 1, dateConversionColumn1)  # start+1 to Skip Header, end+1 to Include Last Row
        columnFinalised = Excel.finalise_column(convertedDateA, columns)  # standardizing count of column
        createdTransTypeColumn = Excel.transaction_type_column(columnFinalised)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/TMB_-_2333__23-09-2023-11-52-23.xlsx"
    # path = "C:/Users/Admin/Desktop/TMB_-_0406_Till_march__25-12-2023-11-58-13.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    result = tmb1_main(wb)
    # result["data"].save('C:/Users/Admin/Desktop/TMB1output.xlsx')
