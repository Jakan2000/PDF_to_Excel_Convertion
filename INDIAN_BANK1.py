import os
from datetime import datetime
from openpyxl.utils import column_index_from_string
import openpyxl

from KSV.FormatingExcelFiles.CommonClass import Excel


def removeString(wb, start, end, refText, column):
    sheet = wb.active
    for x in range(start, end):
        if refText in str(sheet[f"{column}{x}"].value):
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace(refText, "")
    return wb


def deleteColumn(wb, column):
    sheet = wb.active
    column_index = openpyxl.utils.column_index_from_string(column)
    sheet.delete_cols(column_index)
    return wb


def makeNone(wb, start, end, column):
    sheet = wb.active
    for x in range(start, end):
        if len(sheet[f"{column}{x}"].value) < 1:
            sheet[f"{column}{x}"].value = None
    return wb


def removeNone(wb, start, end, column):
    sheet = wb.active
    for x in range(start, end):
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")
    return wb


def dateConversion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d/%m/%Y").date()
    return wb


def removeHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def removeNoneRows(wb, start, end, column):
    sheet = wb.active
    for x in range(end - 1, start, -1):
        if sheet[f"{column}{x}"].value is None:
            sheet.delete_rows(x)
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def indian_bank1_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 8
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def indian_bank1_main(wb):
    sheet = wb.active
    if indian_bank1_validation(wb):
        raise Exception(f"<= INVALID FORMATE =>  <Count Of Column Mismatch>")
    else:
        startText = "ValueDate"
        endText = " Statement Downloaded By"
        startEndRefColumn = "A"
        deleteFlagStartText1 = "Page No"
        deleteFlagEndText1 = "ValueDate"
        deleteFlagStartText2 = "Statement Downloaded By"
        deleteFlagEndText2 = "END OF STATEMENT"
        deleteFlagRefColumn = "A"
        removeRowRefText1 = "BALANCE B/F"
        removeRowRefColumn1 = "D"
        removeRowRefText2 = "Download Limit"
        removeRowRefColumn2 = "A"
        removeRowRefText3 = "Page No"
        removeRowRefColumn3 = "A"
        columnToMerg = "D"
        refColumnToMerg = "A"
        removeNoneRowRefColumn = "A"
        dateFormateColumn1 = "A"
        dateFormateColumn2 = "B"
        dateConersionColumn1 = "A"
        dateConersionColumn2 = "B"
        stringAlignColumn1 = "A"
        stringAlignColumn2 = "B"
        stringAlignColumn3 = "C"
        stringAlignColumn4 = "D"
        stringAlignColumn5 = "E"
        stringAlignColumn6 = "F"
        stringAlignColumn7 = "G"
        stringAlignColumn8 = "H"
        makeNoneColumn1 = "A"
        deleteColumn1 = "C"
        refHaederText1 = "ValueDate"
        refHaederText2 = "PostDate"
        refHaederText3 = "Description"
        refHaederText4 = "Cheque No"
        refHaederText5 = "DR"
        refHaederText6 = "CR"
        refHaederText7 = "Balance"
        headerText1 = "Value_Date"
        headerText2 = "Transaction_Date"
        headerText3 = "Narration"
        headerText4 = "ChequeNo_RefNo"
        headerText5 = "Withdrawal"
        headerText6 = "Deposit"
        headerText7 = "Balance"
        UnWantedRefText1 = "CR"
        columnToRemoveUnWantedText1 = "G"
        start = 1
        end = sheet.max_row
        columnTextToMakeEmptyCellToNone1 = "ChequeNo_RefNo"
        columnTextToMakeEmptyCellToNone2 = "Withdrawal"
        columnTextToMakeEmptyCellToNone3 = "Deposit"
        alignedStringA = Excel.string_align(wb, start, end, stringAlignColumn1)
        alignedStringB = Excel.string_align(alignedStringA, start, end, stringAlignColumn2)
        alignedStringC = Excel.string_align(alignedStringB, start, end, stringAlignColumn3)
        alignedStringD = Excel.string_align(alignedStringC, start, end, stringAlignColumn4)
        alignedStringE = Excel.string_align(alignedStringD, start, end, stringAlignColumn5)
        alignedStringF = Excel.string_align(alignedStringE, start, end, stringAlignColumn6)
        alignedStringG = Excel.string_align(alignedStringF, start, end, stringAlignColumn7)
        alignedStringH = Excel.string_align(alignedStringG, start, end, stringAlignColumn8)
        noneRemovedA = removeNone(alignedStringH, start, end, stringAlignColumn1)
        noneRemovedB = removeNone(noneRemovedA, start, end, stringAlignColumn2)
        noneRemovedC = removeNone(noneRemovedB, start, end, stringAlignColumn3)
        noneRemovedD = removeNone(noneRemovedC, start, end, stringAlignColumn4)
        noneRemovedE = removeNone(noneRemovedD, start, end, stringAlignColumn5)
        noneRemovedF = removeNone(noneRemovedE, start, end, stringAlignColumn6)
        noneRemovedG = removeNone(noneRemovedF, start, end, stringAlignColumn7)
        noneRemovedH = removeNone(noneRemovedG, start, end, stringAlignColumn8)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        end = sheet.max_row
        dupHeader1Removed = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText1, deleteFlagEndText1, deleteFlagRefColumn)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        end = sheet.max_row
        dupHeader2Removed = Excel.delete_rows_by_range(dupHeader1Removed, start, end, deleteFlagStartText2, deleteFlagEndText2, deleteFlagRefColumn)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        end = sheet.max_row
        openBalRowsRemoved = Excel.remove_rows(dupHeader2Removed, start, end, removeRowRefText1, removeRowRefColumn1)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        end = sheet.max_row
        unWantedRowRemoved = Excel.remove_rows(openBalRowsRemoved, start, end, removeRowRefText2, removeRowRefColumn2)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        pageNoRowRemoved = Excel.remove_rows(unWantedRowRemoved, start, end, removeRowRefText3, removeRowRefColumn3)
        headerRemoved = removeHeader(pageNoRowRemoved, start - 1)  # start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        makedNoneA = makeNone(wb, start, end, makeNoneColumn1)
        mergedRowsD = mergingRows(makedNoneA, start, end, refColumnToMerg, columnToMerg)
        noneRowsRemoved = removeNoneRows(pageNoRowRemoved, start, end, removeNoneRowRefColumn)
        start, end = Excel.get_start_end_row_index(headerRemoved, startText, endText, startEndRefColumn)
        convertedDateA = dateConversion(headerRemoved, start + 1, end + 1, dateConersionColumn1)  # start+1 to Skip Header, end+1 to Include Last Row
        convertedDateB = dateConversion(convertedDateA, start + 1, end + 1, dateConersionColumn2)  # start+1 to Skip Header, end+1 to Include Last Row
        deletedBRANCHcolumn = deleteColumn(convertedDateB, deleteColumn1)
        lastCol = 65 + Excel.column_count(wb)  # 65 ASCII value
        valuedate = Excel.alter_header_name(deletedBRANCHcolumn, refHaederText1, headerText1, lastCol)
        transdate = Excel.alter_header_name(valuedate, refHaederText2, headerText2, lastCol)
        narration = Excel.alter_header_name(transdate, refHaederText3, headerText3, lastCol)
        chqno = Excel.alter_header_name(narration, refHaederText4, headerText4, lastCol)
        debit = Excel.alter_header_name(chqno, refHaederText5, headerText5, lastCol)
        credit = Excel.alter_header_name(debit, refHaederText6, headerText6, lastCol)
        balance = Excel.alter_header_name(credit, refHaederText7, headerText7, lastCol)
        removedCR = removeString(balance, start, end + 1, UnWantedRefText1, columnToRemoveUnWantedText1)  # end+1 to include last row
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 ASCII value
        slCreated = Excel.create_slno_column(removedCR, start, end + 1, chr(columnToCreateSlNo))  # end+1 to include last row
        replacedNoneCHQNO = Excel.empty_cell_to_none(slCreated, start, end + 1, columnTextToMakeEmptyCellToNone1)  # end+1 to include last row
        replacedNoneWITHDRAWAL = Excel.empty_cell_to_none(replacedNoneCHQNO, start, end + 1, columnTextToMakeEmptyCellToNone2)  # end+1 to include last row
        replacedNonedeposit = Excel.empty_cell_to_none(replacedNoneWITHDRAWAL, start, end + 1, columnTextToMakeEmptyCellToNone3)  # end+1 to include last row
        return wb


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/SRT_-_INDIAN_BANK_-_6096825697___23-09-2023-10-24-18.xlsx"
    wb = openpyxl.load_workbook(path)
    result = indian_bank1_main(wb)
    result.save("C:/Users/Admin/Desktop/FinalOutput/INDIAN_BANK1output.xlsx")
