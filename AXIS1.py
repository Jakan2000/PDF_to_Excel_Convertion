from datetime import datetime

import openpyxl

from CommonClass import Excel


def deleteNoneRows(wb, start, end, refColumn):
    """
        Deletes rows in an Excel file where the specified column has a None value.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index for deletion.
        - end (int): The ending row index for deletion.
        - refColumn (str): The column reference (e.g., 'A', 'B') where None values are checked for deletion.

        Returns:
        openpyxl.Workbook: The updated openpyxl Workbook after deleting the specified rows.

        Notes:
        - The function iterates from the end to the start row and checks if the specified column has a None value.
        - If a row contains a None value in the specified column, that row is deleted.
        - The function modifies the Workbook in-place and returns the updated Workbook.

    """
    sheet = wb.active
    for x in range(end, start, -1):  # iterate from end to start row
        a_cell = f"{refColumn}{x}"  # get cell address
        if sheet[a_cell].value is None:  # if cell value is None
            end -= 1
            sheet.delete_rows(x)  # delete row
    return wb


def dateConvertion(wb, start, end, column):
    """
        Converts date values in a specified column to a standard date format in an Excel file.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index for date conversion.
        - end (int): The ending row index for date conversion.
        - column (str): The column reference (e.g., 'A', 'B') containing date values for conversion.

        Returns:
        openpyxl.Workbook: The updated openpyxl Workbook after converting the specified date values.

        Notes:
        - The function iterates through the specified rows and converts date values in the specified column.
        - The date values are converted to a standard date format ("%d-%m-%Y").
        - The function modifies the Workbook in-place and returns the updated Workbook.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through start and end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%m-%Y").date()  # converting to standard date formate
    return wb


def deleteHeader(wb, start):
    """
        Deletes rows above the specified start index row in an Excel file.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The row index from which rows above will be deleted.

        Returns:
        openpyxl.Workbook: The updated openpyxl Workbook after deleting rows above the specified start index.

        Notes:
        - The function iterates from the specified start index to the header row (row index 0).
        - Rows above the start index (including the start index row) are deleted.
        - The function modifies the Workbook in-place and returns the updated Workbook.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating from header row to the 0 index row
        sheet.delete_rows(x)  # delete row
    return wb


def deleteFooter(wb, end):
    """
       Deletes all rows below the specified end (last) row in an Excel file.

       Parameters:
       - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
       - end (int): The row index until which rows below will be deleted.

       Returns:
       openpyxl.Workbook: The updated openpyxl Workbook after deleting rows below the specified end index.

       Notes:
       - The function iterates from the last row (maximum row index) to the specified end index.
       - Rows below the end index (including the end index row) are deleted.
       - The function modifies the Workbook in-place and returns the updated Workbook.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating from data ending row to the last row
        sheet.delete_rows(x)  # delete row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merges scattered rows in the specified mergingColumn based on a common reference column.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.
        - start (int): The starting row index for merging rows.
        - end (int): The ending row index (exclusive) for merging rows.
        - refColumn (str): The reference column used to identify starting rows for merging.
        - mergingColumn (str): The column in which rows will be merged.

        Returns:
        openpyxl.Workbook: The updated openpyxl Workbook after merging corresponding rows in the specified column.

        Notes:
        - The function iterates through rows in reverse order.
        - Rows are merged in the specified mergingColumn based on a common non-empty cell in the reference column.
        - The merged value is formed by concatenating the data in the mergingColumn for corresponding rows.
        - The last row (end index row) is also considered for merging.
        - The function modifies the Workbook in-place and returns the updated Workbook.

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data which was scattered
    for i in range(end, start, -1):  # iterate through end and start row
        date = sheet[f"{refColumn}{i}"].value  # get last date cell value from refColumn
        if date is not None:  # if date is not none this is the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if date is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(len(dataToMerge) - 1, 0, -1):  # iterate dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying the dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the coresponding index
        if date is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index

    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(len(dataToMerge) - 1, 0, -1):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def axis1_validation(wb):
    """
       Validates the columns in the given Workbook for a specific logic.

       Parameters:
       - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.

       Returns:
       bool: True if the validation is successful, indicating the Workbook has the expected number of columns.
             False if the validation fails, indicating the Workbook does not have the expected number of columns.

       Notes:
       - The function checks if the number of columns in the Workbook matches the expected count.
       - The expected count is set to 7 (countOfColumn), representing the column count for a specific logic.
       - If the count matches, the validation is successful, and the function returns True.
       - If the count does not match, the validation fails, and the function returns False.

    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column in the sheet, using predefined function
    countOfColumn = 7  # column count of our core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def axis1_main(wb):
    """
        Performs data processing and formatting for a specific logic (axis1) on the given Workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel file.

        Returns:
        dict: A dictionary containing processed Workbook data and a message.
              - 'data' (openpyxl.Workbook): Processed Workbook data.
              - 'msg' (str): A message describing the operation performed.

        Notes:
        - This function assumes the input Workbook follows a specific structure for axis1 logic.
        - Data processing includes validation, column alignment, header replacement, and other operations.
        - The modified Workbook is included in the response dictionary under the 'data' key.
        - The 'msg' key contains a message indicating the success or failure of the operation.

    """
    sheet = wb.active  # get active sheet
    if axis1_validation(wb):  # validate columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return response with error msg
    else:
        startText = "Particulars"  # header text in column C
        endText = "CLOSING BALANCE"  # text to define the end of the data
        startEndDefColumn = "C"  # column containing the text to define start and end -> data within
        delRefText1 = "OPENING BALANCE"  # reference text to remove a row
        delRefText2 = "TRANSACTION TOTAL"  # reference text to remove a row
        delRefText3 = "CLOSING BALANCE"  # reference text to remove a row
        deleteFlagRefColumn = "C"  # reference column to delete the row
        stringAlignColumn1 = "C"  # column to align the string in a cell
        stringAlignColumn2 = "G"  # column to align the string in a cell
        dateConversionColumn = "A"  # column to convert the date to -> date formate
        refHeaderText1 = "Tran Date"  # header text to replace with standardised column name
        refHeaderText2 = "Chq No"  # header text to replace with standardised column name
        refHeaderText3 = "Particulars"  # header text to replace with standardised column name
        refHeaderText4 = "Debit"  # header text to replace with standardised column name
        refHeaderText5 = "Credit"  # header text to replace with standardised column name
        refHeaderText6 = "Balance"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # replacement standard column name
        headerText2 = "ChequeNo_RefNo"  # replacement standard column name
        headerText3 = "Narration"  # replacement standard column name
        headerText4 = "Withdrawal"  # replacement standard column name
        headerText5 = "Deposit"  # replacement standard column name
        headerText6 = "Balance"  # replacement standard column name
        deleteColumnRefText = "Init.Br"  # reference text to delete column
        columns = ["Sl.No.", "Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        negativeValueColumnRefText1 = "Withdrawal"  # no need to convert the negative value to positive
        headerTextToEmptyCellToNone1 = "Value_Date"  # header text to make empty cells to none
        headerTextToEmptyCellToNone2 = "ChequeNo_RefNo"  # header text to make empty cells to none
        headerTextToEmptyCellToNone3 = "Withdrawal"  # header text to make empty cells to none
        headerTextToEmptyCellToNone4 = "Deposit"  # header text to make empty cells to none
        refColumnToMerg = "A"  # reference date column to merge rows of other column
        mergingColumn = "C"  # column to merge
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get start and end row index to specify the data with in
        delFooter = deleteFooter(wb, end)  # delete fooder below the end row
        headerDeleted = deleteHeader(delFooter, start - 1)  # delete header above the start row
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndDefColumn)  # after deleting the row index will change
        removed1 = Excel.remove_row(headerDeleted, start, end, delRefText1, deleteFlagRefColumn)  # removing unwanted single row
        start, end = Excel.get_start_end_row_index(removed1, startText, endText, startEndDefColumn)  # get start and end row index to specify the data with in
        removed2 = Excel.remove_row(removed1, start, end, delRefText2, deleteFlagRefColumn)  # removing unwanted single row
        start, end = Excel.get_start_end_row_index(removed2, startText, endText, startEndDefColumn)  # get start and end row index to specify the data with in
        removed3 = Excel.remove_row(removed1, start, end, delRefText3, deleteFlagRefColumn)  # removing unwanted single row
        start, end = Excel.get_start_end_row_index(removed2, startText, endText, startEndDefColumn)  # get start and end row index to specify the data with in
        alignedC = Excel.string_align(removed3, start, end + 1, stringAlignColumn1)  # align the string by column to make the string in a cell to a single row -> end+1 to Include Last Row
        alignedG = Excel.string_align(alignedC, start, end + 1, stringAlignColumn2)  # end+1 to Include Last Row
        # mergedColumnC = mergingRows(alignedG, start, end, refColumnToMerg, mergingColumn)  # merging the splited rows in a column
        noneRowsDeleted = deleteNoneRows(wb, start, end, refColumnToMerg)  # delete empty rows in date column
        start, end = Excel.get_start_end_row_index(removed2, startText, endText, startEndDefColumn)  # get start and end row index to specify the data with in
        convertedDateA = dateConvertion(noneRowsDeleted, start + 1, end + 1, dateConversionColumn)  # convert date to standard date formate -> start+1 to Sip Header, end+1 to Include Last Row
        lastCol = 65 + sheet.max_column  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        trandate = Excel.alter_header_name(convertedDateA, refHeaderText1, headerText1, lastCol)  # alter header name by standard column name
        chqno = Excel.alter_header_name(trandate, refHeaderText2, headerText2, lastCol)  # alter header name by standard column name
        naration = Excel.alter_header_name(chqno, refHeaderText3, headerText3, lastCol)  # alter header name by standard column name
        debit = Excel.alter_header_name(naration, refHeaderText4, headerText4, lastCol)  # alter header name by standard column name
        credit = Excel.alter_header_name(debit, refHeaderText5, headerText5, lastCol)  # alter header name by standard column name
        balance = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)  # alter header name by standard column name
        # deletedColumnG = Excel.delete_column(balance, deleteColumnRefText)  # delete column by column header name
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slnoCreated = Excel.create_slno_column(wb, start, end + 1, chr(columnToCreateSlNo))  # creating new column - slno
        columnFinalised = Excel.finalise_column(slnoCreated, columns)  # standardizing count of column
        # negativeValueChecked = Excel.check_neagativeValue_by_column(columnFinalised, negativeValueColumnRefText1)  # no need to convert the negative value to positive
        valueDateConverted = Excel.empty_cell_to_none(wb, start, end + 1, headerTextToEmptyCellToNone1)  # making the empty cell to none in desired column
        chqnoConverted = Excel.empty_cell_to_none(valueDateConverted, start, end + 1, headerTextToEmptyCellToNone2)  # making empty cell to none in desired column
        withdrawalConverted = Excel.empty_cell_to_none(chqnoConverted, start, end + 1, headerTextToEmptyCellToNone3)  # making empty cell to none in desired column
        depositConverted = Excel.empty_cell_to_none(withdrawalConverted, start, end + 1, headerTextToEmptyCellToNone4)  # making empty cells to none in desired column
        createdTransTypeColumn = Excel.transaction_type_column(depositConverted)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/1.Axis_-_8874-PW_-_GNAN842166790_unlocked__19-09-2023-14-05-39.xlsx"
    # path = "C:/Users/Admin/Downloads/1.SVTTransports-AXIS1437__23-11-2023-17-46-06.xlsx"
    # path = "C:/Users/Admin/Downloads/AXIS_-_5664__31-12-2023-13-17-11.xlsx"
    path = "C:/Users/Admin/Desktop/KSV/source_excel_files/axis_type1__17-01-2024-17-17-05.xlsx"
    wb = openpyxl.load_workbook(path)
    result = axis1_main(wb)
    # result.save('C:/Users/Admin/Desktop/FinalOutput/AXIS1output.xlsx')
    result["data"].save('C:/Users/Admin/Desktop/AXIS1output.xlsx')
