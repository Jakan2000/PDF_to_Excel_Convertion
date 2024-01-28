from datetime import datetime

import openpyxl

from CommonClass import Excel


def dateConversion(wb, start, end, column):
    """
        Converts date values in a specified column of an openpyxl Workbook object to a standard date format.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index where the conversion should start (inclusive).
        - end (int): The row index where the conversion should end (exclusive).
        - column (str): The column letter (e.g., 'A', 'B', 'C') containing the date values.

        Returns:
        - openpyxl.Workbook: The modified Workbook object with the date values converted.

        Note:
        The function uses the datetime.strptime method to parse the date values in the specified column
        and then updates the Workbook with the parsed date values in a standard date format ("%d-%b-%Y").

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through tabe data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%b-%Y").date()  # converting to standard date formate
    return wb


def removeNone(wb, start, end, column):
    """
        Removes occurrences of the string "None" from a specified column in an openpyxl Workbook object.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index where the removal should start (inclusive).
        - end (int): The row index where the removal should end (exclusive).
        - column (str): The column letter (e.g., 'A', 'B', 'C') from which to remove the string "None".

        Returns:
        - openpyxl.Workbook: The modified Workbook object with the specified string removed.

        Note:
        The function iterates through the specified column and checks each cell for the presence of the string "None".
        If the cell value is not None and contains the string "None", it replaces the occurrence with an empty string.

    """
    sheet = wb.active
    for x in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):  # if cell value is not none and string "None" in cell value
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")  # replace with empty string
    return wb


def splitingDate(wb, start, end, column):
    """
        Extracts transaction date from value date in a specified column of an openpyxl Workbook object.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index where the extraction should start (inclusive).
        - end (int): The row index where the extraction should end (exclusive).
        - column (str): The column letter (e.g., 'A', 'B', 'C') containing the date values.

        Returns:
        - openpyxl.Workbook: The modified Workbook object with transaction date and value date columns updated.

        Note:
        The function iterates through the specified column and extracts transaction date and value date
        by splitting the date with "(" symbol. It then updates the Workbook with the extracted values.

    """
    sheet = wb.active
    valueDateColumn = "H"  # column having value date
    for i in range(start, end):  # iterating through table data start row to table data end row
        spl = str(sheet[f"{column}{i}"].value).split("(")  # splitting date with "("
        sheet[f"{column}{i}"].value = spl[0]  # assigning transaction date to transaction date column
        sheet[f"{valueDateColumn}{i}"].value = spl[1].replace(")", "")  # assigning value date to value date column
    return wb


def deleteHeader(wb, start):
    """
        Deletes header data from the active sheet of an openpyxl Workbook object.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - start (int): The row index where the deletion should start (inclusive).

        Returns:
        - openpyxl.Workbook: The modified Workbook object with header rows deleted.

        Note:
        The function iterates through the specified rows, starting from the specified row index 'start'
        and deletes each row until it reaches the 1st row of the sheet, effectively removing header data.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to 1st row of sheet
        sheet.delete_rows(x)  # delete row
    return wb


def deleteFooter(wb, end):
    """
        Deletes footer data from the active sheet of an openpyxl Workbook object.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data.
        - end (int): The row index where the deletion should end (exclusive).

        Returns:
        - openpyxl.Workbook: The modified Workbook object with footer rows deleted.

        Note:
        The function iterates through the specified rows, starting from the last row of the sheet,
        and deletes each row until it reaches the specified 'end' row, effectively removing footer data.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through max row of sheet to table data end row
        sheet.delete_rows(x)  # delete row
    return wb


def iob1_validation(wb):
    """
        Validates the column count in the active sheet of an openpyxl Workbook object for a specific core logic.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object to be validated.

        Returns:
        - bool: True if the number of columns in the active sheet is not equal to the expected core logic column count (7),
                False otherwise.

        Note:
        The function checks if the maximum number of columns in the active sheet is not equal to the expected core logic column count (7).
        If the condition is met, it returns True, indicating a validation failure. Otherwise, it returns False, indicating a successful validation.

    """
    sheet = wb.active
    max_column = sheet.max_column  # get max column using in build keyword(max_column)
    countOfColumn = 7  # column count of designed core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def iob1_main(wb):
    """
        Performs data processing and validation for a specific format in the active sheet of an openpyxl Workbook object.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object containing the data to be processed.

        Returns:
        - dict: A dictionary containing processed data and a message.
            - "data": The modified Workbook object.
            - "msg": A message indicating the result of the processing.

        Note:
        The function performs various operations on the provided Workbook object to standardize and process the data.
        It involves validation, header renaming, column deletion, date conversion, string alignment, and other operations.

        The processed data is returned along with a message indicating the success or failure of the processing.

    """
    sheet = wb.active
    if iob1_validation(wb):  # validating column count for core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # return response with error msg
    else:
        startText = "Date (ValueDate)"  # header text to define table data start row
        endText = "available balance"  # text define table data end row
        startEndRefColumn = "A"  # column containing start end text
        dateSplitColumn = "A"  # column containing transaction date and value date
        stringAlignColumn1 = "A"  # column to aligning string by removing the \n from the string
        stringAlignColumn2 = "B"  # column to aligning string by removing the \n from the string
        stringAlignColumn3 = "C"  # column to aligning string by removing the \n from the string
        stringAlignColumn4 = 'D'  # column to aligning string by removing the \n from the string
        dateConversionColumn1 = "A"  # column to convert date to standard date formate
        dateConversionColumn2 = "H"  # column to convert date to standard date formate
        columnToDeleteRefText1 = "TransactionType"  # reference header text to delete column
        refHeaderText1 = "Date "  # header text to replace with standardised column name
        refHeaderText2 = "Particulars"  # header text to replace with standardised column name
        refHeaderText3 = "RefNo./ChequeNo"  # header text to replace with standardised column name
        refHeaderText4 = "Debit(Rs)"  # header text to replace with standardised column name
        refHeaderText5 = "Credit(Rs)"  # header text to replace with standardised column name
        refHeaderText6 = "Balance(Rs)"  # header text to replace with standardised column name
        refHeaderText7 = "ValueDate"  # header text to replace with standardised column name
        headerText1 = "Transaction_Date"  # standard column name
        headerText2 = "Narration"  # standard column name
        headerText3 = "ChequeNo_RefNo"  # standard column name
        headerText4 = "Withdrawal"  # standard column name
        headerText5 = "Deposit"  # standard column name
        headerText6 = "Balance"  # standard column name
        headerText7 = "Value_Date"  # standard column name
        refTextToReplaceToNone = "-"  # reference text to make cell value to none
        columnToReplaceTextToNone1 = "D"  # column to replace reference text cell to none
        columnToReplaceTextToNone2 = "E"  # column to replace reference text cell to none
        start = 1  # assigning 1st row
        end = sheet.max_row  # assigning sheet's max row
        alignedStringA = Excel.string_align(wb, start, end, stringAlignColumn1)  # column to aligning string by removing the \n from the string
        alignedStringB = Excel.string_align(alignedStringA, start, end, stringAlignColumn2)  # column to aligning string by removing the \n from the string
        alignedStringC = Excel.string_align(alignedStringB, start, end, stringAlignColumn3)  # column to aligning string by removing the \n from the string
        alignedStringD = Excel.string_align(alignedStringC, start, end, stringAlignColumn4)  # column to aligning string by removing the \n from the string
        removedNoneA = removeNone(alignedStringD, start, end, stringAlignColumn1)  # removing string "None" from a column
        removedNoneB = removeNone(removedNoneA, start, end, stringAlignColumn2)  # removing string "None" from a column
        removedNoneC = removeNone(removedNoneB, start, end, stringAlignColumn3)  # removing string "None" from a column
        removedNoneD = removeNone(removedNoneC, start, end, stringAlignColumn4)  # removing string "None" from a column
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        footerDeleted = deleteFooter(wb, end - 1)  # deleting footer rows, end-1 to Include Last Row
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # deleting header rows start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndRefColumn)  # get start and end row index to specify data with in
        dateSplited = splitingDate(headerDeleted, start, end + 1, dateSplitColumn)  # extracting transaction date from value date end+1 to Include Last Row
        convertedDateA = dateConversion(dateSplited, start + 1, end + 1, dateConversionColumn1)  # start+1 to Skip Header, end+1 to include last row
        convertedDateH = dateConversion(convertedDateA, start + 1, end + 1, dateConversionColumn2)  # start+1 to Skip Header, end+1 to include last row
        transTypecolumnDeleted = Excel.delete_column(convertedDateH, columnToDeleteRefText1)  # deleting transaction type column
        lastCol = 65 + Excel.column_count(wb)  # 65 -> ASCII value
        transdate = Excel.alter_header_name(transTypecolumnDeleted, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
        narration = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
        chqno = Excel.alter_header_name(narration, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
        debit = Excel.alter_header_name(chqno, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
        credit = Excel.alter_header_name(debit, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
        balance = Excel.alter_header_name(credit, refHeaderText6, headerText6, lastCol)  # alter header name from the excel file to the standard column name
        valuedate = Excel.alter_header_name(balance, refHeaderText7, headerText7, lastCol)  # alter header name from the excel file to the standard column name
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 -> ASCII value
        slCreated = Excel.create_slno_column(valuedate, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
        replacedToNoneD = Excel.replace_to_none(slCreated, start, end + 1, refTextToReplaceToNone, columnToReplaceTextToNone1)  # replace to None when reference text is in cell of a column
        replacedToNoneE = Excel.replace_to_none(replacedToNoneD, start, end + 1, refTextToReplaceToNone, columnToReplaceTextToNone2)  # replace to None when reference text is in cell of a column
        createdTransTypeColumn = Excel.transaction_type_column(replacedToNoneE)  # creating new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/IOB_-_8713__23-09-2023-17-44-36.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    result = iob1_main(wb)
    # result.save("C:/Users/Admin/Desktop/FinalOutput/IOB1output.xlsx")
