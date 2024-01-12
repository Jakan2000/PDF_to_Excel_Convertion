import configparser

import openpyxl
from minio import Minio
from openpyxl.utils import column_index_from_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook


class Excel:

    def get_start_end_row_index(wb, startText, endText, startEndDefColumn):
        """
            Returns the start and end row index based on specified start and end text in a given column.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook to be analyzed.
            - startText (str): The text marking the start row
            - endText (str): The text marking the end row
            - startEndDefColumn (str): The column containing the start and end text

            Returns:
            tuple: A tuple containing the start and end row indexes (integers).

            Notes:
            - The function iterates through the specified column to find the row indexes of the start and end text.
            - If the start or end text is not found, the corresponding index is set to 0.

        """
        sheet = wb.active
        start = 0
        end = 0
        for cell in sheet[startEndDefColumn]:  # iterating through all the cells in startEndDefColumn
            start += 1  # increment start value
            if startText in str(cell.value):  # if start text in the cell break the loop, and the cell index is stored in start variable
                break
        for cell in sheet[startEndDefColumn]:  # iterating through all the cells in startEndDefColumn
            end += 1  # increment end value
            if endText in str(
                    cell.value):  # if end text in the cell break the loop, and the cell index is stored in end variable
                break
        return start, end

    def create_slno_column(wb, start, end, column):
        """
            Creates a new serial number (Sl.No.) column in the specified range of rows and a given column.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook to be modified.
            - start (int): The starting row index for adding serial numbers.
            - end (int): The ending row index for adding serial numbers.
            - column (str): The column where the serial numbers will be added.

            Returns:
            openpyxl.Workbook: The modified workbook with the new Sl.No. column.

            Notes:
            - The function iterates through the specified range of rows and adds a serial number column.
            - The first row of the specified column is assigned as the header ("Sl.No.").
            - Subsequent rows are assigned consecutive serial numbers.

        """
        sheet = wb.active
        slno = 1  # header row index
        for i in range(start, end):  # iterating through start and end row
            if i == 1:
                sheet[f"{column}{i}"].value = "Sl.No."  # assigning 1st row as header
            else:
                sheet[f"{column}{i}"].value = slno  # assigning next consecutive rows as serial numbers
                slno += 1
        return wb  # return work book with new slno added column

    def column_count(wb):
        """
            Counts the number of columns in an Excel workbook based on the header text in the first row.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook to be analyzed.

            Returns:
            int: The count of columns in the workbook.

            Notes:
            - The function iterates through the columns in the first row and counts until an empty or None cell is encountered.
            - The count represents the number of columns with header text in the first row.

        """
        sheet = wb.active
        column = 65  # ASCII value of "A"
        count = 0
        for i in range(column, column + sheet.max_column):  # iterate through all the column
            if sheet[f"{chr(i)}1"].value is None:  # if in 1st row cell is empty or none that's the column count
                break
            count += 1
        return count  # return count integer

    def creat_column(wb, header):
        """
           Creates a new column with the specified header in an Excel workbook.

           Parameters:
           - wb (openpyxl.Workbook): The Excel workbook where the new column will be added.
           - header (str): The header text for the new column.

           Returns:
           openpyxl.Workbook: The modified workbook with the new column.

           Notes:
           - The function determines the next available column index by using the Excel.column_count function.
           - The header is assigned to the first row of the new column.

        """
        sheet = wb.active
        max_column = Excel.column_count(
            wb) + 1  # getting last column from column_count(wb) and adding +1 to get the next column
        column = openpyxl.utils.get_column_letter(max_column)  # Convert a column index into a column letter (3 -> 'C')
        sheet[f"{column}1"] = header  # assigning the header to the 1st row
        return wb

    def finalise_column(wb, col):
        """
            Standardizes the count of specified columns in an Excel workbook by adding missing columns.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook to be standardized.
            - col (list): A list of standard column headers.

            Returns:
            openpyxl.Workbook: The modified workbook with standardized columns.

            Notes:
            - The function iterates through the specified standard column headers and checks if they exist in the workbook.
            - If a standard column is missing, it is added to the workbook.

        """
        sheet = wb.active
        missing_columns = []  # array to store missing column
        column = 65  # ASCII value of "A"
        for h in range(0, len(col)):  # iterating through the col array
            count = 0
            for i in range(column, column + sheet.max_column):  # iteratig through the columns in Excel file
                if col[h] in str(
                        sheet[f"{chr(i)}1"].value):  # check if current index standard column is present in excel file
                    count += 1  # increment count
            if count == 0:  # if count == 0 then column is not present in Excel file
                missing_columns.append(col[h])  # append the missing colum to the array
        if len(missing_columns) != 0:  # is missing_column array is not empty then there is a missing column
            for i in range(0, len(missing_columns)):  # iterate through missing_column array
                Excel.creat_column(wb, missing_columns[i])  # create the missing columns
        return wb

    def string_align(wb, start, end, column):
        """
           Aligns strings in a specified column by removing newline characters (\n).

           Parameters:
           - wb (openpyxl.Workbook): The Excel workbook containing the strings to be aligned.
           - start (int): The starting row index for aligning strings.
           - end (int): The ending row index for aligning strings.
           - column (str): The column containing the strings to be aligned.

           Returns:
           openpyxl.Workbook: The modified workbook with aligned strings in the specified column.

           Notes:
           - The function iterates through the specified range of rows in the given column.
           - For each cell, it removes newline characters (\n) from the string content.

        """
        sheet = wb.active
        for i in range(start, end):  # iterate through start row and end row
            sheet[f"{column}{i}"].value = str(sheet[f"{column}{i}"].value).replace('\n', '')  # replace \n from the string
        return wb  # return aligned string in the column

    def alter_header_name(wb, refText, actualText, lastCol):
        """
            Alters header names in an Excel workbook from a reference text to a standard column name.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook containing headers to be altered.
            - refText (str): The reference text present in the headers to be replaced.
            - actualText (str): The standard column name to replace the reference text.
            - lastCol (int): The last column index in the Excel file.

            Returns:
            openpyxl.Workbook: The modified workbook with altered header names.

            Notes:
            - The function iterates through the headers in the first row of each column.
            - If the reference text is present in a header, it replaces it with the specified standard column name.
            - The iteration stops when the last column index is reached.

        """
        sheet = wb.active
        column = 65  # ASCII value of "A"
        row = 1  # header will present in ist roe
        while column < lastCol:  # iterate through all the column in Excel file
            if refText in str(sheet[f"{chr(column)}{row}"].value):  # check if reference header text present in the cell
                sheet[f"{chr(column)}{row}"].value = actualText  # replace it with standard header text (actualText)
            column += 1  # increment the column to get next column
        return wb

    def remove_row(wb, start, end, refText, column):
        """
            Removes row in an Excel workbook based on the presence of a reference text in a specified column.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook containing rows to be removed.
            - start (int): The starting row index for removing rows.
            - end (int): The ending row index for removing rows.
            - refText (str): The reference text used to identify rows for removal.
            - column (str): The column where the reference text is checked for removal.

            Returns:
            openpyxl.Workbook: The modified workbook with removed rows.

            Notes:
            - The function iterates through the specified range of rows in the given column.
            - If the reference text is found in a cell, the corresponding row is deleted.
            - The iteration stops when the row is deleted.

        """
        sheet = wb.active
        for x in range(end, start, -1):  # iterate through end row and start row
            if refText in str(sheet[f"{column}{x}"].value):  # if reference text in cell value
                sheet.delete_rows(x)  # delete the row
                break  # stop iteration of the loop
        return wb

    def remove_rows(wb, start, end, refText, column):
        """
            Removes multiple rows in an Excel workbook based on the presence of a reference text in a specified column.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook containing rows to be removed.
            - start (int): The starting row index for removing rows.
            - end (int): The ending row index for removing rows.
            - refText (str): The reference text used to identify rows for removal.
            - column (str): The column where the reference text is checked for removal.

            Returns:
            openpyxl.Workbook: The modified workbook with removed rows.

            Notes:
            - The function iterates through the specified range of rows in the given column.
            - For each row where the reference text is found in the cell, the row is deleted.
            - The iteration stops when the starting row index is reached.

        """
        sheet = wb.active
        for x in range(end, start, -1):  # iterate through end row to start row
            if refText in str(sheet[f"{column}{x}"].value):  # check if reference text in the cell
                sheet.delete_rows(x)  # delete the row
        return wb

    def delete_rows_by_range(wb, start, end, startText, stopText, refcolumn):
        """
            Deletes rows in Excel workbook based on specified range defined by start and stop text in reference column.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook containing rows to be deleted.
            - start (int): The starting row index for identifying the range.
            - end (int): The ending row index for identifying the range.
            - startText (str): The text marking the beginning of the range in the reference column.
            - stopText (str): The text marking the end of the range in the reference column.
            - refcolumn (str): The column where the range is identified using start and stop text.

            Returns:
            openpyxl.Workbook: The modified workbook with deleted rows.

            Notes:
            - The function iterates through the specified range of rows in the given reference column.
            - It identifies the start and stop of the range based on the provided start and stop text.
            - Rows within the identified range are marked for deletion.
            - The deletion is performed in reverse order to avoid index problems.

        """
        sheet = wb.active
        delete_flag = False
        rows_to_delete = []
        for i in range(start, end):  # iterate from start to end row
            if startText in str(sheet[f"{refcolumn}{i}"].value):  # if start text is in the reference column cell it's the starting row
                delete_flag = True  # make delete flag true
            if delete_flag:  # if delete flag is true append the row to rows_to_delete array
                rows_to_delete.append(i)  # append row to rows_to_delete array
            if stopText in str(sheet[f"{refcolumn}{i}"].value):  # if stop text is in the reference column cell it's the last row
                delete_flag = False  # make delete flag false
        for x in reversed(rows_to_delete):  # iterate array in reversed order to avoid the index problem while deleting the rows
            sheet.delete_rows(x)  # delete row
        return wb

    def delete_column(wb, refText):
        """
            Deletes a column in an Excel workbook based on a specified reference text in the column header.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook containing the column to be deleted.
            - refText (str): The reference text used to identify the column for deletion.

            Returns:
            openpyxl.Workbook: The modified workbook with the deleted column.

            Notes:
            - The function iterates through all the columns in the sheet.
            - It searches for the specified reference text in the header of each column.
            - If the reference text is found, the corresponding column is deleted.

        """
        sheet = wb.active
        column_index = None
        for col in range(1, sheet.max_column + 1):  # iterating through all the columns in the sheet
            if refText in sheet.cell(row=1, column=col).value:  # if reference text is in the column header
                column_index = col  # store column index
                break
        if column_index is None:  # if column index is none do nothing
            return wb
        sheet.delete_cols(column_index)  # delete column using column index
        return wb

    def get_header(wb):
        """
            Retrieves the header values from the first row of an Excel workbook.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook from which to retrieve the header.

            Returns:
            list: A list containing the header values from columns A to H.

            Notes:
            - The function directly accesses cells in the first row of columns A to H.
            - It retrieves the values from these cells and returns them as a list.

        """
        sheet = wb.active
        header = [sheet["A1"].value, sheet["B1"].value, sheet["C1"].value, sheet["D1"].value, sheet["E1"].value,
                  sheet["F1"].value, sheet["G1"].value, sheet["H1"].value]
        return header

    def find_column_index_by_header(wb, header):
        """
           Finds the column index in an Excel workbook based on the provided header text.

           Parameters:
           - wb (openpyxl.Workbook): The Excel workbook in which to find the column index.
           - header (str): The header text used to identify the column.

           Returns:
           int or None: The column index if the header is found, or None if the header is not found.

           Notes:
           - The function iterates through all the columns in the sheet using ASCII values.
           - It searches for the specified header text in the first cell of each column.
           - If the header text is found, the corresponding column index is returned.
           - If the header is not found in any column, None is returned.

        """
        sheet = wb.active
        column_index = None
        for column in range(65, 65 + sheet.max_column):  # iterating through all the columns using ascii values
            if header in str(sheet[f"{chr(column)}1"].value):  # if header text is in the 1st cell of the column
                column_index = column  # store the column index
                break
        return column_index

    def check_neagativeValue_by_column(wb, header):  # no need of converting the negative value to positive
        # sheet = wb.active
        # column = Excel.find_column_index_by_header(wb, header)
        # for i in range(2, sheet.max_row + 1):
        #     value = sheet[f"{chr(column)}{i}"].value
        #     if isinstance(value, str) and value.strip() != '' and float(value.replace(',', '')) < 0.0:
        #         temp = str(sheet[f"{chr(column)}{i}"].value).replace(',', '')
        #         sheet[f"{chr(column + 1)}{i}"].value = float(temp.replace("-", ""))
        #         sheet[f"{chr(column)}{i}"].value = None
        return wb

    def empty_cell_to_none(wb, start, end, header):
        """
           Converts empty cells to `None` in a specified column of an Excel workbook.

           Parameters:
           - wb (openpyxl.Workbook): The Excel workbook containing the cells to be modified.
           - start (int): The starting row index for modifying cells.
           - end (int): The ending row index for modifying cells.
           - header (str): The header text used to identify the column for modification.

           Returns:
           openpyxl.Workbook: The modified workbook with empty cells converted to `None`.

           Notes:
           - The function finds the column index using the provided header text.
           - It then iterates through the specified range of rows in that column.
           - If a cell is empty (with a length less than 1), the cell value is set to `None`.

        """
        sheet = wb.active
        column = Excel.find_column_index_by_header(wb, header)  # find column index using header text
        for x in range(start, end):  # iterating through all the rows from start to end row
            if len(str(sheet[f"{chr(column)}{x}"].value)) < 1:  # if cell value length is less than 1, then its a empty cell
                sheet[f"{chr(column)}{x}"].value = None  # make it as None
        return wb

    def remove_string(wb, start, end, refString, column):
        """
            Removes a specified string from cells in a column of an Excel workbook.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook containing the cells to be modified.
            - start (int): The starting row index for modifying cells.
            - end (int): The ending row index for modifying cells.
            - refString (str): The reference string to be removed from the cells.
            - column (str): The column where the reference string is checked for removal.

            Returns:
            openpyxl.Workbook: The modified workbook with the specified string removed from cells.

            Notes:
            - The function iterates through the specified range of cells in the given column.
            - If the reference string is found in a cell, it is replaced with an empty string.

        """
        sheet = wb.active
        for x in range(start, end):  # iterating through all cells from start to end row
            if refString in str(sheet[f"{column}{x}"].value):  # if reference string in cell
                sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace(refString, "")  # replace it with empty string
        return wb

    def replace_to_none(wb, start, end, refText, column):
        """
            Replaces cells containing a specified reference text with `None` in a column range of an Excel workbook.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook containing the cells to be modified.
            - start (int): The starting row index for modifying cells.
            - end (int): The ending row index for modifying cells.
            - refText (str): The reference text to identify cells for replacement.
            - column (str): The column where the reference text is checked for replacement.

            Returns:
            openpyxl.Workbook: The modified workbook with cells containing the reference text replaced with `None`.

            Notes:
            - The function iterates through the specified range of cells in the given column.
            - If the reference text is found in a cell, the cell value is replaced with `None`.

        """
        sheet = wb.active
        for x in range(start, end):  # iterate through start and end row
            if refText in str(sheet[f"{column}{x}"].value):  # if reference text is in the cell of a column
                sheet[f"{column}{x}"].value = None  # assign cell value to None
        return wb

    def transaction_type_column(wb):
        """
            Creates a new "Transaction_Type" column in an Excel workbook based on "Withdrawal" and "Deposit" columns.

            Parameters:
            - wb (openpyxl.Workbook): The Excel workbook to be modified.

            Returns:
            openpyxl.Workbook: The modified workbook with the new "Transaction_Type" column.

            Notes:
            - The function uses existing columns ("Withdrawal" and "Deposit") to determine the transaction type.
            - For each row, it checks the values in the "Withdrawal" and "Deposit" columns to determine if it's a debit or credit.
            - If the value in the "Withdrawal" column is >= 1, the transaction type is set to "Debit."
            - If the value in the "Deposit" column is >= 1, the transaction type is set to "Credit."
            - The function handles errors gracefully, printing a message if there's an issue processing values.

        """
        sheet = wb.active
        Excel.creat_column(wb, header="Transaction_Type")  # creating new transaction type column
        trans_type_column = chr(Excel.find_column_index_by_header(wb, header="Transaction_Type"))  # getting column index of Transaction_Type column
        withdrawal_column = chr(Excel.find_column_index_by_header(wb, header="Withdrawal"))  # getting column index of Withdrawal column
        deposit_column = chr(Excel.find_column_index_by_header(wb, header="Deposit"))  # getting column index of Deposit column

        for i in range(2, sheet.max_row + 1):  # iterating through all rows in sheet by skipping header row (1st)
            withdrawal_value = sheet[f"{withdrawal_column}{i}"].value  # store withdrawal value
            deposit_value = sheet[f"{deposit_column}{i}"].value  # store deposit value
            if withdrawal_value is not None:  # if withdrawal value is not None, the transaction is debited
                try:
                    withdrawal_float = float(str(withdrawal_value))  # get withdrawal value in float (decimal value)
                    if withdrawal_float >= 1:  # if withdrawal value is > 1
                        sheet[
                            f"{trans_type_column}{i}"].value = "Debit"  # set status as debit in transaction type column
                except ValueError as e:
                    print(f"Error processing withdrawal value at row {i}: {e}")  # handling the exception
            if deposit_value is not None:  # if deposit value is not None, the transaction is credited
                try:
                    deposit_float = float(str(deposit_value))  # get the deposit value in float (decimal value)
                    if deposit_float >= 1:  # if deposit value is > 1
                        sheet[
                            f"{trans_type_column}{i}"].value = "Credit"  # set status as credit in transaction type column
                except ValueError as e:
                    print(f"Error processing deposit value at row {i}: {e}")  # handling exception
        return wb

    def minio_upload_pdf(file_path, bucket_name, folder_path):
        """
            Uploads a PDF file to a Minio S3 bucket and returns the presigned URL.

            Parameters:
            - file_path (str): The local path of the PDF file to be uploaded.
            - bucket_name (str): The name of the Minio S3 bucket.
            - folder_path (str): The folder path within the bucket to store the PDF file.

            Returns:
            str: The presigned URL of the uploaded PDF file.

            Notes:
            - The function reads Minio access and secret keys from a configuration file (.env).
            - The Minio client is created with the specified credentials to connect to the Minio server.
            - The function determines the file name from the file path, handling both "/" and "\" separators.
            - The PDF file is uploaded to the specified bucket and folder within Minio.
            - The presigned URL of the uploaded PDF file is generated and returned.

        """
        config = configparser.ConfigParser()
        config.read(".env")  # reading values from environment variable (.env) file
        client = Minio('ksvca-server-01:3502', access_key=config.get("DEFAULT", "MINIO_ACCESS_KEY"),
                       secret_key=config.get("DEFAULT", "MINIO_SECRET_KEY"),secure=False)  # setting credentials to connect with minio
        # file path may have both "/" abd "\" so
        if r"/" in file_path:  # if the path has "/"
            file_name = file_path.split(r'/')[-1]  # split file name from the path
        if R"Downloads\temp_" in file_path:  # if path has "\"
            file_name = file_path.split(r"Downloads\temp_")[-1]  # split file name from the path
        client.fput_object(bucket_name, folder_path + file_name, file_path)  # inserting pdf file into the bucket
        url = client.presigned_get_object(bucket_name, folder_path + file_name, response_headers={'response-content-type': 'application/pdf'})  # getting url of the pdf file
        return url

    def pandas_df_to_openpyxl(df):
        """
           Converts a Pandas DataFrame to an Openpyxl Workbook.

           Parameters:
           - df (pd.DataFrame): The Pandas DataFrame to be converted.

           Returns:
           openpyxl.Workbook: The Openpyxl Workbook containing the DataFrame data.

           Notes:
           - The function creates a new Openpyxl Workbook and a new worksheet.
           - It appends the DataFrame data to the worksheet row by row.
           - The resulting Openpyxl Workbook contains the data from the Pandas DataFrame.

        """
        workbook = Workbook()  # Create a new Openpyxl Workbook
        worksheet = workbook.active  # Create a new worksheet
        # Append the DataFrame data to the worksheet
        for row in dataframe_to_rows(df, index=False, header=False):
            worksheet.append(row)
        return workbook  # returning openpyxl work book
