from datetime import datetime

import openpyxl

from CommonClass import Excel


def dateConversion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%b-%Y").date()
    return wb


def deleteHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def deleteFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


def removeNoneRows(wb, start, end, column):
    sheet = wb.active
    for x in range(end - 1, start, -1):
        if sheet[f"{column}{x}"].value is None:
            sheet.delete_rows(x)
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def equitas1_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 6
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def equitas1_main(wb):
    sheet = wb.active
    if equitas1_validation(wb):
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response
    else:
        startText = "Date"
        endText = "*** End of the Statement ***"
        startEndRefColumn = "A"
        deleteFlagStartText = "Page"
        deleteFlagStopText = "Deposit"
        deleteFlagRefColumn = "E"
        removeRowRefText = "INR"
        removeRowRefColumn = "D"
        columnToMerg = "C"
        refColumnToMerg = "A"
        dateConversionColumn1 = "A"
        refHeaderText1 = "Date"
        refHeaderText2 = "Reference No. / Cheque No."
        refHeaderText3 = "Narration"
        refHeaderText4 = "Withdrawal"
        refHeaderText5 = "Deposit"
        refHeaderText6 = "ClosingBalance"
        headerText1 = "Transaction_Date"
        headerText2 = "ChequeNo_RefNo"
        headerText3 = "Narration"
        headerText4 = "Withdrawal"
        headerText5 = "Deposit"
        headerText6 = "Balance"
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]
        negativeValueColumnRefText1 = "Withdrawal"
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        end = sheet.max_row
        dupHeaderRemoved = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagStopText,
                                             deleteFlagRefColumn)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        INRrowsRemoved = Excel.remove_rows(dupHeaderRemoved, start, end, removeRowRefText, removeRowRefColumn)
        start, end = Excel.get_start_end_row_index(INRrowsRemoved, startText, endText, startEndRefColumn)
        mergedColumnC = mergingRows(INRrowsRemoved, start, end, refColumnToMerg, columnToMerg)
        noneRowsRemoved = removeNoneRows(mergedColumnC, start, end, refColumnToMerg)
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)
        footerDeleted = deleteFooter(noneRowsRemoved, end - 1)  # end-1 to Include End Footer
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # end-1 to Include End Footer
        start, end = Excel.get_start_end_row_index(noneRowsRemoved, startText, endText, startEndRefColumn)
        convertedDateA = dateConversion(headerDeleted, start + 1, end + 1,
                                        dateConversionColumn1)  # start+1 to Skip Header, end+1 to Include Last Row
        lastCol = 65 + Excel.column_count(wb)
        transdate = Excel.alter_header_name(convertedDateA, refHeaderText1, headerText1, lastCol)
        chqno = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)
        narration = Excel.alter_header_name(chqno, refHeaderText3, headerText3, lastCol)
        debit = Excel.alter_header_name(narration, refHeaderText4, headerText4, lastCol)
        credit = Excel.alter_header_name(debit, refHeaderText5, headerText5, lastCol)
        balance = Excel.alter_header_name(credit, refHeaderText6, headerText6, lastCol)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        slnoCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))
        columnFinalised = Excel.finalise_column(slnoCreated, columns)
        negativeValueChecked = Excel.check_neagativeValue_by_column(slnoCreated, negativeValueColumnRefText1)
        createdTransTypeColumn = Excel.transaction_type_column(negativeValueChecked)
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/Equitas_-_6802_unlocked__23-09-2023-12-01-43.xlsx"
    wb = openpyxl.load_workbook(path)
    result = equitas1_main(wb)
    result.save("C:/Users/Admin/Desktop/FinalOutput/EQUITAS1output.xlsx")
