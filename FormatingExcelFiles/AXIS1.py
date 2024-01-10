from datetime import datetime

import openpyxl

from CommonClass import Excel


def deleteNoneRows(wb, start, end, refColumn):  # delete the rows that are none in date column
    sheet = wb.active
    for x in range(end, start, -1):  # iterate from end to start row
        a_cell = f"{refColumn}{x}"  # get the cell address
        if sheet[a_cell].value is None:  # if cell value is None
            end -= 1
            sheet.delete_rows(x)  # delete the row
    return wb


def dateConvertion(wb, start, end, column):  # converting the date to standard date formate
    sheet = wb.active
    for i in range(start, end):  # iterating through start and end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%m-%Y").date()
    return wb


def deleteHeader(wb, start):  # delete the rows above the start index row
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating from header row to the 0 index row
        sheet.delete_rows(x)  # delete the row
    return wb


def deleteFooter(wb, end):  # delete all the rows below the end(last) row
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating from data ending row to the last row
        sheet.delete_rows(x)  # delete the row
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):  # merging the split rows in mergingColumn by
    sheet = wb.active
    dataToMerge = []  # array to store the row data which was scattered
    for i in range(end, start, -1):  # iterate through start and end row
        date = sheet[f"{refColumn}{i}"].value  # get the first date row from refColumn
        if date is not None:  # if date is not none this is the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store the cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store the data from the 1 index
            else:  # if date is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # empty string to merge the row data
                for j in range(len(dataToMerge) - 1, 0, -1):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take the current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign the concated data to the cell
                dataToMerge = []  # emptying the dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending the next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending the row data to the coresponding index
        if date is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append the data to the corresponding index

    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(len(dataToMerge) - 1, 0, -1):  # iterate the dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take the current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign the conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return the work book by merging the corresponding rows in the column


def axis1_validation(wb):  # validate the columns for the logic written
    sheet = wb.active
    max_column = sheet.max_column  # get the max column in the sheet, using predefined function
    countOfColumn = 7  # the column count of our core logic
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def axis1_main(wb):
    sheet = wb.active  # get the active sheet
    if axis1_validation(wb):  # validate the columns for the core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response
    else:
        startText = "Particulars"  # header text in column C
        endText = "CLOSING BALANCE"  # text to define the end of the data
        startEndDefColumn = "C"  # column containing the text to define start and end -> data within
        delRefText1 = "OPENING BALANCE"  # reference text to remove a row
        delRefText2 = "TRANSACTION TOTAL"  # reference text to remove a row
        delRefText3 = "CLOSING BALANCE"  # reference text to remove a row
        deleteFlagRefColumn = "C"  # reference column to delete the row
        stringAlignColumn1 = "C"  # column to align the string in a cell
        stringAlignColumn2 = "G"  # column to align the string in a cell
        dateConversionColumn = "A"  # column to convert the date to -> date formate
        refHeaderText1 = "Tran Date"  # header text to replace with standard column name
        refHeaderText2 = "Chq No"  # header text to replace with standard column name
        refHeaderText3 = "Particulars"  # header text to replace with standard column name
        refHeaderText4 = "Debit"  # header text to replace with standard column name
        refHeaderText5 = "Credit"  # header text to replace with standard column name
        refHeaderText6 = "Balance"  # header text to replace with standard column name
        headerText1 = "Transaction_Date"  # replacement standard column name
        headerText2 = "ChequeNo_RefNo"  # replacement standard column name
        headerText3 = "Narration"  # replacement standard column name
        headerText4 = "Withdrawal"  # replacement standard column name
        headerText5 = "Deposit"  # replacement standard column name
        headerText6 = "Balance"  # replacement standard column name
        deleteColumnRefText = "Init.Br"  # reference text to delete column
        columns = ["Sl.No.", "Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
        negativeValueColumnRefText1 = "Withdrawal"  # no need to convert the negative value to positive
        headerTextToEmptyCellToNone1 = "Value_Date"  # header text to make empty cells to none
        headerTextToEmptyCellToNone2 = "ChequeNo_RefNo"  # header text to make empty cells to none
        headerTextToEmptyCellToNone3 = "Withdrawal"  # header text to make empty cells to none
        headerTextToEmptyCellToNone4 = "Deposit"  # header text to make empty cells to none
        refColumnToMerg = "A"  # reference date column to merge rows of other column
        mergingColumn = "C"  # column to merge
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)  # get the start and end row index to specify the data with in
        delFooter = deleteFooter(wb, end)  # delete the fooder below the end row
        headerDeleted = deleteHeader(delFooter, start - 1)  # delete the header above the start row
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndDefColumn)  # after deleting the row index will change
        removed1 = Excel.remove_row(headerDeleted, start, end, delRefText1, deleteFlagRefColumn)  # removing unwanted single row
        start, end = Excel.get_start_end_row_index(removed1, startText, endText, startEndDefColumn)  # get the start and end row index to specify the data with in
        removed2 = Excel.remove_row(removed1, start, end, delRefText2, deleteFlagRefColumn)  # removing unwanted single row
        start, end = Excel.get_start_end_row_index(removed2, startText, endText, startEndDefColumn)  # get the start and end row index to specify the data with in
        removed3 = Excel.remove_row(removed1, start, end, delRefText3, deleteFlagRefColumn)  # removing unwanted single row
        start, end = Excel.get_start_end_row_index(removed2, startText, endText, startEndDefColumn)  # get the start and end row index to specify the data with in
        alignedC = Excel.string_align(removed3, start, end + 1, stringAlignColumn1)  # align the string by column to make the string in a cell to a single row -> end+1 to Include Last Row
        alignedG = Excel.string_align(alignedC, start, end + 1, stringAlignColumn2)  # end+1 to Include Last Row
        mergedColumnC = mergingRows(alignedG, start, end, refColumnToMerg, mergingColumn)  # merging the splited rows in a column
        noneRowsDeleted = deleteNoneRows(mergedColumnC, start, end, refColumnToMerg)  # delete the empty rows in date column
        start, end = Excel.get_start_end_row_index(removed2, startText, endText, startEndDefColumn)  # get the start and end row index to specify the data with in
        convertedDateA = dateConvertion(noneRowsDeleted, start + 1, end + 1, dateConversionColumn)  # convert the date to standard date formate -> start+1 to Sip Header, end+1 to Include Last Row
        lastCol = 65 + sheet.max_column  # 65 => ASCII value "A" -> by adding 65 + sheet.max_column we get the last column
        trandate = Excel.alter_header_name(convertedDateA, refHeaderText1, headerText1, lastCol)  # alter the header name by standard column name
        chqno = Excel.alter_header_name(trandate, refHeaderText2, headerText2, lastCol)  # alter the header name by standard column name
        naration = Excel.alter_header_name(chqno, refHeaderText3, headerText3, lastCol)  # alter the header name by standard column name
        debit = Excel.alter_header_name(naration, refHeaderText4, headerText4, lastCol)  # alter the header name by standard column name
        credit = Excel.alter_header_name(debit, refHeaderText5, headerText5, lastCol)  # alter the header name by standard column name
        balance = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)  # alter the header name by standard column name
        deletedColumnG = Excel.delete_column(balance, deleteColumnRefText)  # delete the column by column header name
        columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 => ASCII value "A" -> column_count() function return the column count in the sheet
        slnoCreated = Excel.create_slno_column(deletedColumnG, start, end + 1, chr(columnToCreateSlNo))  # creating the new column - slno
        columnFinalised = Excel.finalise_column(slnoCreated, columns)  # standardizing the count of column
        negativeValueChecked = Excel.check_neagativeValue_by_column(columnFinalised, negativeValueColumnRefText1)  # no need to convert the negative value to positive
        valueDateConverted = Excel.empty_cell_to_none(negativeValueChecked, start, end + 1, headerTextToEmptyCellToNone1)  # making the empty cell to none in desired column
        chqnoConverted = Excel.empty_cell_to_none(valueDateConverted, start, end + 1, headerTextToEmptyCellToNone2)  # making the empty cell to none in desired column
        withdrawalConverted = Excel.empty_cell_to_none(chqnoConverted, start, end + 1, headerTextToEmptyCellToNone3)  # making the empty cell to none in desired column
        depositConverted = Excel.empty_cell_to_none(withdrawalConverted, start, end + 1, headerTextToEmptyCellToNone4)  # making the empty cells to none in desired column
        createdTransTypeColumn = Excel.transaction_type_column(depositConverted)  # creating the new transaction type column
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Downloads/1.Axis_-_8874-PW_-_GNAN842166790_unlocked__19-09-2023-14-05-39.xlsx"
    path = "C:/Users/Admin/Downloads/1.SVTTransports-AXIS1437__23-11-2023-17-46-06.xlsx"
    wb = openpyxl.load_workbook(path)
    result = axis1_main(wb)
    # result.save('C:/Users/Admin/Desktop/FinalOutput/AXIS1output.xlsx')
    result.save('C:/Users/Admin/Desktop/AXIS1output.xlsx')
