from datetime import datetime

import openpyxl
from CommonClass import Excel


def dateConvertion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        original_date = sheet[f"{column}{i}"].value
        new_date = datetime.strptime(original_date, '%d-%b-%Y').date()
        sheet[f"{column}{i}"].value = new_date
    return wb


def deleteHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return (wb)


def deleteFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


def iob2_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 7
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def iob2_main(wb):
    # sheet = wb.active()
    if iob2_validation(wb):
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response
    else:
        startText = "DATE"
        endText = "* denotes cancelled transaction"
        startEndRefColumn = "A"
        refTextToDeleteColumn = "COD"
        stringAlignColumn1 = "A"
        stringAlignColumn2 = "B"
        stringAlignColumn3 = "C"
        stringAlignColumn4 = "D"
        stringAlignColumn5 = "E"
        stringAlignColumn6 = "F"
        refHeaderText1 = "DATE"
        refHeaderText2 = "CHQNO"
        refHeaderText3 = "NARATION"
        refHeaderText4 = "DEBIT"
        refHeaderText5 = "CREDIT"
        refHeaderText6 = "BALANCE"
        headerText1 = "Value_Date"
        headerText2 = "ChequeNo_RefNo"
        headerText3 = "Narration"
        headerText4 = "Withdrawal"
        headerText5 = "Deposit"
        headerText6 = "Balance"
        dateConversionColumn = "A"
        refStringToRemove = "None"
        refColumnToRemoveString1 = "B"
        refColumnToRemoveString2 = "D"
        refColumnToRemoveString3 = "E"
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        deleteHeader(wb, start - 1)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        deleteFooter(wb, end - 1)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        Excel.delete_column(wb, refTextToDeleteColumn)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        Excel.create_slno_column(wb, start, end + 1, chr(columnToCreateSlNo))
        lastCol = 65 + Excel.column_count(wb)
        columnA = Excel.string_align(wb, start, end + 1, stringAlignColumn1)
        columnB = Excel.string_align(wb, start, end + 1, stringAlignColumn2)
        columnC = Excel.string_align(wb, start, end + 1, stringAlignColumn3)
        columnD = Excel.string_align(wb, start, end + 1, stringAlignColumn4)
        columnE = Excel.string_align(wb, start, end + 1, stringAlignColumn5)
        columnF = Excel.string_align(wb, start, end + 1, stringAlignColumn6)
        valuedate = Excel.alter_header_name(wb, refHeaderText1, headerText1, lastCol)
        chqno = Excel.alter_header_name(wb, refHeaderText2, headerText2, lastCol)
        narration = Excel.alter_header_name(wb, refHeaderText3, headerText3, lastCol)
        debit = Excel.alter_header_name(wb, refHeaderText4, headerText4, lastCol)
        credit = Excel.alter_header_name(wb, refHeaderText5, headerText5, lastCol)
        balance = Excel.alter_header_name(wb, refHeaderText6, headerText6, lastCol)
        dateConvertion(wb, start + 1, end + 1, dateConversionColumn)
        Excel.remove_string(wb, start, end + 1, refStringToRemove, refColumnToRemoveString1)
        Excel.remove_string(wb, start, end + 1, refStringToRemove, refColumnToRemoveString2)
        Excel.remove_string(wb, start, end + 1, refStringToRemove, refColumnToRemoveString3)
        Excel.finalise_column(wb, columns)
        Excel.transaction_type_column(wb)
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/last_year_statement_july_1to_october_31.xlsx"
    wb = openpyxl.load_workbook(path)
    result = iob2_main(wb)
    result["data"].save("C:/Users/Admin/Desktop/last_year_statement_july_1to_october_31.xlsx")