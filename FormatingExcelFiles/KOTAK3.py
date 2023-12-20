from datetime import datetime

import openpyxl

from CommonClass import Excel


def align_balance_column(wb, start, end, column):
    sheet = wb.active
    for row in range(start, end):
        if sheet[f"{column}{row}"].value is None:
            sheet[f"{column}{row}"].value = sheet[f"{chr(ord(column)-1)}{row}"].value
            sheet[f"{chr(ord(column) - 1)}{row}"].value = None
    return wb


def dateConvertion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%m-%Y").date()
    return wb


# make the cell None if there is only year in the cell
def make_none_date_column(wb, start, end, refColumn):
    sheet = wb.active
    for row in range(start, end):
        if len(str(sheet[f"{refColumn}{row}"].value)) < 5:
            sheet[f"{refColumn}{row}"].value = None
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def align_date_column(wb, start, end, column):
    sheet = wb.active
    for row in range(start, end):
        if sheet[f"{column}{row}"].value is not None:
            if len(str(sheet[f"{column}{row}"].value)) < 9 and len(str(sheet[f"{column}{row}"].value)) > 4:
                sheet[f"{column}{row}"].value = sheet[f"{column}{row + 1}"].value
    return wb


def deleteHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def deleteFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb



def aligningAllColumns(wb, start, end, refColumn):
    sheet = wb.active
    for i in range(start, end):
        if sheet[f"{refColumn}{i}"].value is not None:
            sheet[f'C{i}'].value = sheet[f'D{i}'].value
            sheet[f'D{i}'].value = sheet[f'E{i}'].value
            sheet[f'E{i}'].value = sheet[f'F{i}'].value
            sheet[f'F{i}'].value = sheet[f'G{i}'].value
            sheet[f'G{i}'].value = None
    return wb


def kotak3_validation(wb):
    sheet = wb.active
    max_column = 6
    countOfColumn = Excel.column_count(wb)
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def kotak3_main(wb):
    sheet = wb.active
    countOfColumn = 6
    startText = "Date"
    stopText = "Statement Summary"
    startEndRefColumn = "A"
    deleteFlagStartText = "Period"
    deleteFlagEndText = "Narration"
    deleteFlagRefColumn = "B"
    refColumnToAlignAllColumns = "G"
    dateColumnAlignRefcolumn = "A"
    refTextToRemoveRow1 = "B/F"
    refColumnToRemoveRow1 = "B"
    refColumnToRemoveInvalidDate = "A"
    refColumnToMerg = "A"
    columnToMerg1 = "B"
    columnToMerg2 = "C"
    refStringToRemove1 = "None"
    refStringToRemove2 = "(Cr)"
    refColumnToRemoveString2 = "F"
    dateConversionColumn1 = "A"
    refHeaderText1 = "Date"
    refHeaderText2 = "Narration"
    refHeaderText3 = "Chq/Ref No"
    refHeaderText4 = "Withdrawal (Dr)"
    refHeaderText5 = "Deposit(Cr)"
    refHeaderText6 = "Balance"
    headerText1 = "Transaction_Date"
    headerText2 = "Narration"
    headerText3 = "ChequeNo_RefNo"
    headerText4 = "Withdrawal"
    headerText5 = "Deposit"
    headerText6 = "Balance"
    balance_column = "F"
    columns = ["Sl.No.", "Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]
    start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)
    rows = Excel.find_header(wb, start, end, startText, startEndRefColumn)
    for row in rows:
        Excel.check_header_alignment(wb, row, startEndRefColumn, countOfColumn)
    Excel.delete_rows_by_range(wb, start + 1, end + 1, deleteFlagStartText, deleteFlagEndText, deleteFlagRefColumn)
    start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)
    aligningAllColumns(wb, start, end + 1, refColumnToAlignAllColumns)
    deleteFooter(wb, end - 1)
    start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)
    deleteHeader(wb, start - 1)
    if kotak3_validation(wb):
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response
    start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)
    Excel.remove_row(wb, start, end + 1, refTextToRemoveRow1, refColumnToRemoveRow1)
    start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)
    align_date_column(wb, start, end + 1, dateColumnAlignRefcolumn)
    make_none_date_column(wb, start + 1, end + 1, refColumnToRemoveInvalidDate)
    mergingRows(wb, start, end + 1, refColumnToMerg, columnToMerg1)
    mergingRows(wb, start, end + 1, refColumnToMerg, columnToMerg2)
    Excel.removeNoneRows(wb, start, end + 1, refColumnToMerg)
    start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)
    align_balance_column(wb, start, end, balance_column)
    Excel.remove_string(wb, start, end + 1, refStringToRemove1, columnToMerg1)
    Excel.remove_string(wb, start, end + 1, refStringToRemove1, columnToMerg2)
    Excel.remove_string(wb, start, end + 1, refStringToRemove2, refColumnToRemoveString2)
    dateConvertion(wb, start + 1, end + 1, dateConversionColumn1)
    columnToCreateSlNo = 65 + Excel.column_count(wb)
    Excel.create_slno_column(wb, start, end + 1, chr(columnToCreateSlNo))
    lastCol = 65 + sheet.max_column  # 65 => ASCII value "A"
    transdate = Excel.alter_header_name(wb, refHeaderText1, headerText1, lastCol)
    narration = Excel.alter_header_name(wb, refHeaderText2, headerText2, lastCol)
    chqNo = Excel.alter_header_name(wb, refHeaderText3, headerText3, lastCol)
    debit = Excel.alter_header_name(wb, refHeaderText4, headerText4, lastCol)
    credit = Excel.alter_header_name(wb, refHeaderText5, headerText5, lastCol)
    balance = Excel.alter_header_name(wb, refHeaderText6, headerText6, lastCol)
    Excel.finalise_column(wb, columns)
    Excel.transaction_type_column(wb)
    response = {"data": wb,
                "msg": None}
    return response


if __name__ == "__main__":
    path = "C:/Users/Admin/Desktop/KSV/source_excel_files/sasikala_kotak__14-12-2023-10-37-39.xlsx"
    wb = openpyxl.load_workbook(path)
    result = kotak3_main(wb)
    result["data"].save("C:/Users/Admin/Desktop/Kotak3output.xlsx")