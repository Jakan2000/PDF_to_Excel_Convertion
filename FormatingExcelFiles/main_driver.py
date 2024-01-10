import configparser
import os
import os.path
from datetime import datetime
from io import BytesIO

import aspose.pdf as ap
import pandas
import psycopg2
import requests
from PyPDF2 import PdfReader, PdfWriter
from openpyxl import Workbook, load_workbook

from AXIS1 import axis1_main
from CANARA1 import canara1_main
from CITY_UNION1 import cityunion1_main
from DBS1 import dbs1_main
from EQUITAS1 import equitas1_main
from FEDERAL1 import federal1_main
from HDFC1 import hdfc1_main
from ICICI1 import icici1_main
from ICICI2 import icici2_main
from ICICI3 import icici3_main
from ICICI4 import icici4_main
from INDIAN_BANK1 import indian_bank1_main
from INDUSIND1 import indusind1_main
from INDUSIND2 import indusind2_main
from IOB1 import iob1_main
from IOB2 import iob2_main
from KOTAK1 import kotak1_main
from KOTAK2 import kotak2_main
from KOTAK3 import kotak3_main
from CommonClass import Excel
from SBI1 import sbi1_main
from TMB1 import tmb1_main
from YES1 import yes1_main


def driver(work_book, bank, type, pdf_url, caller):
    banks = {"axis": {"type1": axis1_main},
             "canara": {"type1": canara1_main},
             "city_union": {"type1": cityunion1_main},
             "dbs": {"type1": dbs1_main},
             "equitas": {"type1": equitas1_main},
             "federal": {"type1": federal1_main},
             "hdfc": {"type1": hdfc1_main},
             "icici": {"type1": icici1_main,
                       "type2": icici2_main,
                       "type3": icici3_main,
                       "type4": icici4_main},  # icici type4 will be called explicitly
             "indian_bank": {"type1": indian_bank1_main},
             "indusind": {"type1": indusind1_main,
                          "type2": indusind2_main},
             "iob": {"type1": iob1_main,
                     "type2": iob2_main},
             "kotak": {"type1": kotak1_main,
                       "type2": kotak2_main,
                       "type3": kotak3_main},
             "sbi": {"type1": sbi1_main},
             "tmb": {"type1": tmb1_main},
             "yes_bank": {"type1": yes1_main},
             }
    if bank in banks and type in banks[bank]:  # check is bank and tpe present in the banks json
        result = banks[bank][type](work_book)  # calling the appropriate function -> function call statement
        if result["msg"] is None:  # if msg is none then the data has the excel file
            wb = result["data"]
            response = to_db_and_return_response(wb, pdf_url, caller)  # insert the excel file to DB
            return response
        else:
            return result
    else:
        response = {"data": None,
                    "file_name": None,
                    "msg": f"<Bank or type not found in the dictionary>"}
        return response


def to_db_and_return_response(wb, pdf_url, caller):
    temp = str(os.path.basename(pdf_url)).replace(".pdf", ".xlsx")  # storing the Excel file name in temp variable
    file = temp.replace(".PDF", ".xlsx")  # replace if the .PDF is in Uppercase also
    # creatig a temp file in downloads folder
    download_folder = os.path.expanduser("~\\Downloads")  # get the current download folder path
    download_path = os.path.join(download_folder, f"TEMP_{file}")  # creating the temp file in download folder
    wb.save(download_path)  # save the temp file in download folder
    try:
        df = pandas.read_excel(download_path, na_values=[""])
        os.remove(download_path)  # removing the temp file
        # Convert "Transaction_Date" and "Value_Date" columns to datetime
        df['Transaction_Date'] = pandas.to_datetime(df['Transaction_Date'], format='%Y-%m-%d')
        df['Value_Date'] = pandas.to_datetime(df['Value_Date'], format='%Y-%m-%d')
        # Extract the date component from "Transaction_Date" and "Value_Date" columns
        df['Transaction_Date'] = df['Transaction_Date'].dt.date  # converting the cell to date type
        df['Value_Date'] = df['Value_Date'].dt.date  # converting the cell to date type
        column_name1 = "Sl.No."
        column_name2 = "Transaction_Date"
        column_name3 = "Value_Date"
        column_name4 = "ChequeNo_RefNo"
        column_name5 = "Narration"
        column_name6 = "Transaction_Type"
        column_name7 = "Deposit"
        column_name8 = "Withdrawal"
        column_name9 = "Balance"
        column_data1 = df[column_name1]
        column_data2 = df[column_name2]
        column_data3 = df[column_name3]
        column_data4 = df[column_name4]
        column_data5 = df[column_name5]
        column_data6 = df[column_name6]
        column_data7 = df[column_name7]
        column_data8 = df[column_name8]
        column_data9 = df[column_name9]
        # arranging the column based on DB table
        new_df = pandas.DataFrame(
            {column_name1: column_data1, column_name2: column_data2, column_name3: column_data3,
             column_name4: column_data4, column_name5: column_data5, column_name6: column_data6,
             column_name7: column_data7, column_name8: column_data8, column_name9: column_data9})
        # storing the new data frame in a temp file and removed the slno column / rearranged the columns of deposit and withdrawal
        temp_new_df = pandas.DataFrame(
            {column_name2: column_data2, column_name3: column_data3,
             column_name4: column_data4, column_name5: column_data5, column_name6: column_data6, column_name8: column_data8,
             column_name7: column_data7, column_name9: column_data9})
        # renaming the column of Excel file based on DB table
        df = new_df.rename(
            columns={"Transaction_Type": "trx_type", "Transaction_Date": "trx_date", "Value_Date": "value_date",
                     "ChequeNo_RefNo": "ref_no_org", "Narration": "description", "Deposit": "deposit",
                     "Withdrawal": "withdrawal", "Balance": "balance"})
        column_names = ["trx_type", "trx_date", "value_date", "ref_no_org", "description", "deposit", "withdrawal",
                        "balance"]
        column_data = df[column_names]
        column_data = column_data.applymap(lambda x: None if pandas.isna(x) else x)
        config = configparser.ConfigParser()
        config.read(".env")
        postgres_credentials = {
            "user": config.get("DEFAULT", "USER"),  # reading the data from environment variable (.env) file
            "password": config.get("DEFAULT", "PASSWORD"),  # reading the data from environment variable (.env) file
            "host": config.get("DEFAULT", "HOST"),  # reading the data from environment variable (.env) file
            "port": config.get("DEFAULT", "PORT"),  # reading the data from environment variable (.env) file
            "database": config.get("DEFAULT", "DATABASE"),  # reading the data from environment variable (.env) file
        }
        schema = "ksv"
        table_name = "bank_stmt_lines_t"
        # connecting with DB
        connection = psycopg2.connect(**postgres_credentials)
        cursor = connection.cursor()
        insert_query = f"""
                    INSERT INTO {schema}.{table_name}
                    (trx_type, trx_date, value_date, ref_no_org, description, deposit, withdrawal, balance)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """
        values = [tuple(row) for row in column_data.values]
        cursor.executemany(insert_query, values)  # bulk insert the data
        connection.commit()
        cursor.close()
        connection.close()
        response = "Records inserted successfully."
        if caller == "appsmith":  # if the api calling is from appsmith returning the response as json
            split_file_name = pdf_url.split("/")
            file_name = split_file_name[len(split_file_name) - 1].replace(".pdf", ".xlsx")  # get the file name from the pdf url
            file_name = file_name.replace(".PDF", ".xlsx")   # replacing the Upper case .PDF also
            # uploding to minio to get the url
            download_path = os.path.join(download_folder, f"temp_{file_name}")  # storing the temp excel file, to upload it in minio
            temp_new_df.to_excel(download_path)
            bucket_name = "ksv"
            folder_path = "pdf_to_excel_files/"
            temp_excel_url = Excel.minio_upload_pdf(download_path, bucket_name, folder_path)  # uploading the excel file to minio
            os.remove(download_path)  # remove the temp excel file from download folder
            excel_url = temp_excel_url.split("?")  # spliting the excel file url from expiry link
            excel_url = excel_url[0]
            response = {"data": excel_url,
                        "file_name": file_name,
                        "msg": "Converted PDF to Excel Successfully 😎"}
            return response
        print(response)
        return response
    except Exception as e:
        print(f"An error occurred: {e}")
        response = {"data": None,
                    "file_name": None,
                    "msg": f"An error occurred: {e}"}
        return response


def delete_files_with_criteria(folder_path, keyword, extension):  # delete the temp file created by aspose lib
    # Get a list of all files in the folder
    files = os.listdir(folder_path)
    # Iterate through the files and delete those with the specified extension and keyword
    for file_name in files:
        if extension in file_name and keyword in file_name:
            file_path = os.path.join(folder_path, file_name)
            os.remove(file_path)


def convert_url_to_bytes(pdf_url):
    bytes_list = []
    response = requests.get(pdf_url)
    response.raise_for_status()  # Check for any request errors
    bytes_stream = BytesIO(response.content)
    reader = PdfReader(bytes_stream)

    for page in reader.pages:
        writer = PdfWriter()
        writer.add_page(page)
        with BytesIO() as bytes_stream:
            writer.write(bytes_stream)
            bytes_stream.seek(0)
            bytes_list.append(bytes_stream.getvalue())
    return bytes_list


def convert_bytes_to_excel(pdf_bytes):
    def create_output_excel(output_xlsx):
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = ''
        workbook.save(output_xlsx)

    def join_data(temp_xlsx, output_xlsx):
        source_workbook_1 = load_workbook(output_xlsx)
        sheet1 = source_workbook_1.active
        source_workbook_2 = load_workbook(temp_xlsx)
        sheet2 = source_workbook_2['Sheet1']
        for row in sheet2.iter_rows(min_row=2, values_only=True):
            sheet1.append(row)
        source_workbook_1.save(output_xlsx)

    def convert_excel(pdf, temp_xlsx, output_xlsx):
        with open(temp_xlsx, 'wb') as f:
            document = ap.Document(BytesIO(pdf))
            save_option = ap.ExcelSaveOptions()
            save_option.minimize_the_number_of_worksheets = True
            document.save(f, options=save_option)
        join_data(temp_xlsx, output_xlsx)

    now = datetime.now()
    t = now.strftime("__%d-%m-%Y-%H-%M-%S")
    output_xlsx = 'output' + t + '.xlsx'
    temp_xlsx = 'temp' + t + '.xlsx'
    create_output_excel(output_xlsx)
    for page_bytes in pdf_bytes:
        convert_excel(page_bytes, temp_xlsx, output_xlsx)
    os.remove(temp_xlsx)

    # Load and return the final workbook
    return load_workbook(output_xlsx), output_xlsx


def pdf_to_excel_main(pdf_url, bank, type, caller):
    if bank == "icici" and type == "type4":
        result = icici4_main(pdf_url)  # conversion done using camelot library
        if result["msg"] is None:
            wb = result["data"]
            responce = to_db_and_return_response(wb, pdf_url, caller)  # receive the response as a json
            return responce
        else:
            return result
    else:
        pdf_bytes = convert_url_to_bytes(pdf_url)  # converting the pdf URL to bytes
        output_workbook, output_workbook_xlsx = convert_bytes_to_excel(pdf_bytes)  # converting bytes to excel
        sheet = output_workbook.active  # getting the first active sheet
        max_column = sheet.max_column  # getting the max column
        if max_column < 2:  # if max column is < 2 its Insufficient Data
            response = {"data": None,
                        "file_name": None,
                        "msg": "Insufficient Data to convert_bytes_to_excel To Process Driver Dictionary"}
            return response
        else:
            response = driver(output_workbook, bank, type, pdf_url, caller)  # receiving the response as json

    delete_files_with_criteria(folder_path="C:/Users/Admin/PycharmProjects/pythonProject1/KSV/FormatingExcelFiles", keyword="output", extension='.xlsx')  # deleting the temp file create in the project folder, created by aspose library
    delete_files_with_criteria(folder_path="C:/Users/Admin/PycharmProjects/pythonProject1/KSV/FormatingExcelFiles", keyword="temp", extension='.xlsx')  # deleting the temp file create in the project folder, created by aspose library
    print("pdf_to_excel_main : ", response)
    return response


if __name__ == "__main__":
    response = pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/1.Axis_-_8874-PW_-_GNAN842166790_unlocked.pdf", "axis", "type1", "appsmith")
    # response = pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/1.Canara_-_6183.pdf", "canara", "type1", "appsmith")
    # response = pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/CITY_UNION_BANK_-_SB-500101012199098.pdf", "city_union", "type1", "appsmith")

    # response = pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/LVB_-_0145P.W_-_1L1675876_unlocked.pdf", "dbs", "type1", "appsmith")
    # response = pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/LVB-0697.pdf", "dbs", "type1", "appsmith")


    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/Equitas_-_6802_unlocked.pdf", "equitas", "type1", "appsmith")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/2.%20R%20RAVICHANDRAN%20-%20Federal%20-%202416%20Pass%20-%20RAVI016%20.pdf", "federal", "type1", "appsmith")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/HDFC_-_7768.pdf", "hdfc", "type1", "appsmith")
    # result = pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/HDFC_-_6732.pdf", "hdfc", "type1", "appsmith")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/ICICI_-_3281.pdf", "icici", "type1", "appsmith")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/ICICI_-_2207PW-088601502207_unlocked.pdf", "icici", "type2", "appsmith")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/ilovepdf_merged_7.pdf", "icici", "type3", "appsmith")
    # responce = pdf_to_excel_main("http://ksvca-server-01:3502/ksv/%2Funlock_pdf/2._ICICI_-_4642.pdf", "icici", "type4","appsmith")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/SRT_-_INDIAN_BANK_-_6096825697_.pdf", "indian bank", "type1", "appsmith")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/Senthil_indusind_pdf.io.pdf", "indusind", "type1", "appsmith")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/1._Indusind_-_2673.pdf", "indusind", "type2", "appsmith")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/IOB_-_8713.pdf", "iob", "type1", "appsmith")

    # result = pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/last_year_statement_july_1to_october_31.pdf", "iob", "type2", "appsmith")
    # result = pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/last_year_statement_2022_march_to_june.pdf", "iob", "type2", "appsmith")    #notworking

    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/Kotak1._Apr-22_637102.pdf", "kotak", "type1", "appsmith")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/Kotak_-_5887.PDF", "kotak", "type2", "appsmith")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/SBI12._March_-_2023.pdf", "sbi", "type1", "appsmith")

    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/TMB_-_2333.pdf", "tmb", "type1", "appsmith")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/2._Jan_to_March_-_TMB_-_363100050305246_Pass-1994_unlocked.pdf", "tmb", "type1", "appsmith") # muthu bro statement

    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/8._YES_bank_-_8241_Aug-Oct.pdf", "yes", "type1", "appsmith")

    # response["data"].to_excel(f"C:/Users/Admin/Desktop/{response['file_name']}.xlsx", index=False)