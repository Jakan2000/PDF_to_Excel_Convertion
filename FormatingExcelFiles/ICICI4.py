import tempfile
from datetime import datetime
from io import BytesIO

import camelot
import pandas as pd
import requests
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook import Workbook

from CommonClass import Excel


def dateConversion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        original_date = sheet[f"{column}{i}"].value
        if original_date:  # Check if the cell is not empty
            # Correct the format based on the actual date format in your Excel sheet
            new_date = datetime.strptime(original_date, '%d/%m/%Y').date()
            sheet[f"{column}{i}"].value = new_date
    return wb


def removeNoneRows(wb, start, end, column):
    sheet = wb.active
    for x in range(end - 1, start, -1):
        if sheet[f"{column}{x}"].value is None:
            sheet.delete_rows(x)
    return wb


def remove_header(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def pandas_df_to_openpyxl(df):
    # Create a new Openpyxl Workbook
    workbook = Workbook()
    # Create a new worksheet
    worksheet = workbook.active
    # Append the DataFrame data to the worksheet
    for row in dataframe_to_rows(df, index=False, header=False):
        worksheet.append(row)

    return workbook


def icici4_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 8
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def icici4_main(pdf_url):
    startText = "Value Date"
    endText = ""
    startEndRefColumn = "B"
    headerTextTOMakeEmptyCellTONone = "Value Date"
    refColumnToMerg = "B"
    ColumnToMerg1 = "E"
    refTextToDelteColumn = "S No."
    refColumnToRemoveNoneRows = "A"
    refHeaderText1 = "Value Date"
    refHeaderText2 = "Transaction Date"
    refHeaderText3 = "Cheque Number"
    refHeaderText4 = "Transaction Remarks"
    refHeaderText5 = "Withdrawal Amount"
    refHeaderText6 = "Deposit Amount ( )"
    refHeaderText7 = "Balance ( )"
    headerText1 = "Value_Date"
    headerText2 = "Transaction_Date"
    headerText3 = "ChequeNo_RefNo"
    headerText4 = "Narration"
    headerText5 = "Withdrawal"
    headerText6 = "Deposit"
    headerText7 = "Balance"
    dateConversionColumn1 = "A"
    dateConversionColumn2 = "B"
    refTextToMakeCellNone = "-"
    refColumnToMakeCellNone = "C"
    columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]
    # Download the PDF file from the URL
    response = requests.get(pdf_url)
    pdf_data = BytesIO(response.content)
    # Save the BytesIO content to a temporary PDF file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as temp_pdf:
        temp_pdf.write(pdf_data.getvalue())
        temp_pdf_path = temp_pdf.name
    # Extract tables from PDF using Camelot
    tables = camelot.read_pdf(temp_pdf_path, flavor='stream', pages='all')
    # Concatenate DataFrames for each page into a single DataFrame
    df = pd.concat([table.df for table in tables])
    # Convert DataFrame to Openpyxl Workbook
    wb = pandas_df_to_openpyxl(df)
    # Remove the temporary PDF file
    temp_pdf.close()
    sheet = wb.active
    if icici4_validation(wb):
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        responce = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return responce
    else:
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        end = sheet.max_row
        removedHeader = remove_header(wb, start - 1)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        end = sheet.max_row
        Excel.empty_cell_to_none(wb, start, end + 1, headerTextTOMakeEmptyCellTONone)
        mergedColumnE = mergingRows(wb, start, end+1, refColumnToMerg, ColumnToMerg1)
        Excel.delete_column(wb, refTextToDelteColumn)
        startEndRefColumn = "A"
        removeNoneRows(wb, start, end + 1, refColumnToRemoveNoneRows)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndRefColumn)
        end = sheet.max_row
        lastCol = 65 + Excel.column_count(wb)
        valuedate = Excel.alter_header_name(wb, refHeaderText1, headerText1, lastCol)
        transdate = Excel.alter_header_name(valuedate, refHeaderText2, headerText2, lastCol)
        refno = Excel.alter_header_name(transdate, refHeaderText3, headerText3, lastCol)
        naration = Excel.alter_header_name(refno, refHeaderText4, headerText4, lastCol)
        debit = Excel.alter_header_name(naration, refHeaderText5, headerText5, lastCol)
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        dateConversion(wb, start + 1, end + 1, dateConversionColumn1)
        dateConversion(wb, start + 1, end + 1, dateConversionColumn2)
        Excel.replace_to_none(wb, start, end + 1, refTextToMakeCellNone, refColumnToMakeCellNone)
        slCreated = Excel.create_slno_column(wb, start, end + 1, chr(columnToCreateSlNo))
        Excel.finalise_column(slCreated, columns)
        Excel.transaction_type_column(wb)
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/2._ICICI_-_4642__27-11-2023-17-43-28.xlsx"
    # path = "C:/Users/Admin/Downloads/2._ICICI_-_4642.pdf"
    path = "http://ksvca-server-01:3502/ksv/%2Funlock_pdf/2._ICICI_-_4642.pdf"
    # wb = openpyxl.load_workbook(path)
    result = icici4_main(path)
    if result["data"] is not None:
        result["data"].save('C:/Users/Admin/Desktop/ICICI4output.xlsx')
    else:
        print(result["msg"])
