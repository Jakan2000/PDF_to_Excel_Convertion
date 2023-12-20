from datetime import datetime

import openpyxl

from CommonClass import Excel


def deleteHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def removeFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


def dateConvertion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(sheet[f"{column}{i}"].value, "%d %b %Y").date()
    return wb


def removeRowsByDateLength(wb, start, end, column):
    sheet = wb.active
    yearLength = 6
    for x in range(end, start, -1):
        if len(str(sheet[f"{column}{x}"].value)) < yearLength:
            sheet.delete_rows(x)
    return wb


def mergingRowsByDateLength(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dateWithFullLen = 10
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None and len(str(slno)) >= dateWithFullLen:
            if len(dataToMerge) == 0:
                cell_address = f"{mergingColumn}{i}"
                dataToMerge.append(cell_address)
                dataToMerge.append(sheet[cell_address].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[cell_address].value = s
                dataToMerge = []
                cell_address = f"{mergingColumn}{i}"
                dataToMerge.append(cell_address)
                dataToMerge.append(sheet[cell_address].value)
        else:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[cell_address].value = st1
    return wb


def removeNone(wb, start, end, column):
    sheet = wb.active
    for x in range(start, end):
        if sheet[f"{column}{x}"].value is not None and "None" in str(sheet[f"{column}{x}"].value):
            sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace("None", "")
    return wb


def mergingDateColumn(wb, start, end, column):
    sheet = wb.active
    inCompleteDateLen = 10
    yearLength = 4
    for i in range(start, end):
        if sheet[f"{column}{i}"].value is not None:
            if len(str(sheet[f"{column}{i}"].value)) < inCompleteDateLen and len(
                    str(sheet[f"{column}{i + 1}"].value)) == yearLength:
                s = str(sheet[f"{column}{i}"].value) + " " + str(sheet[f"{column}{i + 1}"].value)
                sheet[f"{column}{i}"].value = s
    return wb


def sbi1_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 7
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def sbi1_main(wb):
    sheet = wb.active
    if sbi1_validation(wb):
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response
    else:
        startText = "Txn Date"
        endText = "Please do not share"
        startEndDefColumn = "A"
        dupHeaderText = "Txn Date"
        dupHeaderRefColumn = "A"
        columnToMerg1 = "A"
        columnToMerg2 = "B"
        refColumnToMerg = "A"
        columnToMerg3 = "C"
        columnToMerg4 = "D"
        removeNoneColumn1 = "A"
        removeNoneColumn2 = "B"
        removeNoneColumn3 = "C"
        removeNoneColumn4 = "D"
        columnToConvertDate1 = "A"
        columnToConvertDate2 = "B"
        stringAlignColumn1 = "B"
        stringAlignColumn2 = "C"
        stringAlignColumn3 = "D"
        refHeaderText1 = "Txn Date"
        refHeaderText2 = "ValueDate"
        refHeaderText3 = "Description"
        refHeaderText4 = "Ref No./ChequeNo."
        refHeaderText5 = "Debit"
        refHeaderText6 = "Credit"
        refHeaderText7 = "Balance"
        headerText1 = "Transaction_Date"
        headerText2 = "Value_Date"
        headerText3 = "Narration"
        headerText4 = "ChequeNo_RefNo"
        headerText5 = "Withdrawal"
        headerText6 = "Deposit"
        headerText7 = "Balance"
        headerToReplaceEmptyCellToNone1 = "ChequeNo_RefNo"
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)
        dupHeaderRemoved = Excel.remove_rows(wb, start, end, dupHeaderText, dupHeaderRefColumn)
        start, end = Excel.get_start_end_row_index(dupHeaderRemoved, startText, endText, startEndDefColumn)
        columnMergA = mergingDateColumn(dupHeaderRemoved, start, end, columnToMerg1)
        columnMergB = mergingDateColumn(columnMergA, start, end, columnToMerg2)
        noneRemovedFromA = removeNone(columnMergB, start, end, removeNoneColumn1)
        noneRemovedFromB = removeNone(noneRemovedFromA, start, end, removeNoneColumn2)
        columnMergC = mergingRowsByDateLength(noneRemovedFromB, start + 1, end, refColumnToMerg, columnToMerg3)  # start+1 to Skip Header
        columnMergD = mergingRowsByDateLength(columnMergC, start + 1, end, refColumnToMerg, columnToMerg4)  # start+1 to Skip Header
        unWantedRowsRemoved = removeRowsByDateLength(columnMergD, start, end, refColumnToMerg)
        start, end = Excel.get_start_end_row_index(unWantedRowsRemoved, startText, endText, startEndDefColumn)
        dateConvertedA = dateConvertion(unWantedRowsRemoved, start + 1, end, columnToConvertDate1)  # start+1 to Skip Header
        dateConvertedB = dateConvertion(dateConvertedA, start + 1, end, columnToConvertDate2)  # start+1 to Skip Header
        start, end = Excel.get_start_end_row_index(dateConvertedB, startText, endText, startEndDefColumn)
        noneRemovedFromC = removeNone(dateConvertedB, start, end, removeNoneColumn3)
        noneRemovedFromD = removeNone(noneRemovedFromC, start, end, removeNoneColumn4)
        footerRemoved = removeFooter(noneRemovedFromD, end - 1)  # end-1 to Include End Footer
        headerDeleted = deleteHeader(footerRemoved, start - 1)  # start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(dateConvertedB, startText, endText, startEndDefColumn)
        alignedColumnStringB = Excel.string_align(headerDeleted, start, end + 1, stringAlignColumn1)  # end+1 to Include Last Row
        alignedColumnStringC = Excel.string_align(alignedColumnStringB, start, end + 1, stringAlignColumn2)  # end+1 to Include Last Row
        alignedColumnStringD = Excel.string_align(alignedColumnStringC, start, end + 1, stringAlignColumn3)  # end+1 to Include Last Row
        lastCol = 65 + Excel.column_count(wb)
        transdate = Excel.alter_header_name(alignedColumnStringD, refHeaderText1, headerText1, lastCol)
        valuedate = Excel.alter_header_name(transdate, refHeaderText2, headerText2, lastCol)
        naration = Excel.alter_header_name(valuedate, refHeaderText3, headerText3, lastCol)
        chqno = Excel.alter_header_name(naration, refHeaderText4, headerText4, lastCol)
        debit = Excel.alter_header_name(chqno, refHeaderText5, headerText5, lastCol)
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)
        balance = Excel.alter_header_name(credit, refHeaderText6, headerText6, lastCol)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        slCreated = Excel.create_slno_column(balance, start, end + 1, chr(columnToCreateSlNo))
        replacedToNoneCHQNO = Excel.empty_cell_to_none(slCreated, start, end + 1, headerToReplaceEmptyCellToNone1)
        createdTransTypeColumn = Excel.transaction_type_column(replacedToNoneCHQNO)
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/SBI12._March_-_2023__06-09-2023-17-54-37.xlsx"
    wb = openpyxl.load_workbook(path)
    result = sbi1_main(wb)
    result.save('C:/Users/Admin/Desktop/FinalOutput/SBI1output.xlsx')
