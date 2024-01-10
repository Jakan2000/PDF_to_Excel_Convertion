import configparser

import openpyxl
from minio import Minio
from openpyxl.utils import column_index_from_string


class Excel:

    def get_start_end_row_index(wb, startText, endText, startEndDefColumn):  # return the start and end row index
        sheet = wb.active
        start = 0
        end = 0
        for cell in sheet[startEndDefColumn]:  # iterating through all the cells in startEndDefColumn
            start += 1  # increment the start value
            if startText in str(cell.value):  # if start text in the cell break the loop, and the cell index is stored in start variable
                break
        for cell in sheet[startEndDefColumn]:  # iterating through all the cells in startEndDefColumn
            end += 1  # increment the end value
            if endText in str(cell.value):   # if end text in the cell break the loop, and the cell index is stored in end variable
                break
        return start, end

    def create_slno_column(wb, start, end, column):  # creating new slno column
        sheet = wb.active
        slno = 1  # header index
        for i in range(start, end):  # iterating through start and end row
            if i == 1:
                sheet[f"{column}{i}"].value = "Sl.No."  # assigning 1st row as header
            else:
                sheet[f"{column}{i}"].value = slno  # assigning next consecutive rows as serial numbers
                slno += 1
        return wb  # return the work book with new slno added column

    def column_count(wb):  # count the column by the header text
        sheet = wb.active
        column = 65  # ASCII value of "A"
        count = 0
        for i in range(column, column + sheet.max_column):  # iterate through all the column
            if sheet[f"{chr(i)}1"].value is None:  # if in 1st row cell is empty or none that's the column count
                break
            count += 1
        return count  # return the count integer

    def creat_column(wb, header):  # creating new column with desired column header
        sheet = wb.active
        max_column = Excel.column_count(wb) + 1  # getting the last column from column_count(wb) and adding +1 to get the next column
        column = openpyxl.utils.get_column_letter(max_column)  # Convert a column index into a column letter (3 -> 'C')
        sheet[f"{column}1"] = header  # assigning the header to the 1st row
        return wb

    def finalise_column(wb, col):  # standardizing the count of column
        sheet = wb.active
        missing_columns = []  # array to store the missing column
        column = 65  # ASCII value of "A"
        for h in range(0, len(col)):  # iterating through the col array
            count = 0
            for i in range(column, column + sheet.max_column):  # iteratig through the columns in Excel file
                if col[h] in str(sheet[f"{chr(i)}1"].value):  # check if current index standard column is present in excel file
                    count += 1  # increment count
            if count == 0:  # if count == 0 then the column is not present in Excel file
                missing_columns.append(col[h])  # append the missing colum to the array
        if len(missing_columns) != 0:  # is missing_column array is not empty then there is a missing column
            for i in range(0, len(missing_columns)):  # iterate through the missing_column array
                Excel.creat_column(wb, missing_columns[i])  # create the missing columns
        return wb

    def string_align(wb, start, end, column):  # aligning the string in column by removing the \n from the string -> \n -> next line
        sheet = wb.active
        for i in range(start, end):  # iterate through start row and end row
            sheet[f"{column}{i}"].value = str(sheet[f"{column}{i}"].value).replace('\n', '')  # replace the \n from the string
        return wb  # return the aligned string in the column

    def alter_header_name(wb, refText, actualText, lastCol):  # alter the header name from the excel file to the standard column name
        sheet = wb.active
        column = 65  # ASCII value of "A"
        row = 1  # header will present in ist roe
        while column < lastCol:  # iterate through all the column in Excel file
            if refText in str(sheet[f"{chr(column)}{row}"].value):  # check if reference header text present in the cell
                sheet[f"{chr(column)}{row}"].value = actualText  # replace it with standard header text (actualText)
            column += 1  # increment the column to get next column
        return wb

    def remove_row(wb, start, end, refText, column):  # remove a single row by checking the referance text is in the column cell
        sheet = wb.active
        for x in range(end, start, -1):  # iterate through end row and start row
            if refText in str(sheet[f"{column}{x}"].value):
                sheet.delete_rows(x)  # delete the row
                break  # stop the iteration of the loop
        return wb

    def remove_rows(wb, start, end, refText, column):  # remove multiple rows
        sheet = wb.active
        for x in range(end, start, -1):  # iterate through end and start row
            if refText in str(sheet[f"{column}{x}"].value):  # check if reference text in the cell
                sheet.delete_rows(x)  # delete the row
        return wb

    def delete_rows_by_range(wb, start, end, startText, stopText, refcolumn):  # deleting the rows by range
        sheet = wb.active
        delete_flag = False
        rows_to_delete = []
        for i in range(start, end):  # iterate from start to end row
            if startText in str(sheet[f"{refcolumn}{i}"].value):  # if start text is in the reference column cell it's the starting row
                delete_flag = True  # make the delete flag true
            if delete_flag:  # if delete flag is true append the row to rows_to_delete array
                rows_to_delete.append(i)  # append the row to rows_to_delete array
            if stopText in str(sheet[f"{refcolumn}{i}"].value):  # if stop text is in the reference column cell it's the last row
                delete_flag = False  # make the delete flag false
        for x in reversed(rows_to_delete):  # iterate the array in reversed order to avoid the index problem while deleting the rows
            sheet.delete_rows(x)  # delete the row
        return wb

    def delete_column(wb, refText):  # deleting an existing column
        sheet = wb.active
        column_index = None
        for col in range(1, sheet.max_column + 1):  # iterating through all the columns in the sheet
            if refText in sheet.cell(row=1, column=col).value:  # if the reference text is in the column header
                column_index = col  # store the column index
                break
        if column_index is None:  # if column index is none do nothing
            return wb
        sheet.delete_cols(column_index)  # delete the column using column index
        return wb

    def get_header(wb):  # get all the 8 column headers by using cell address
        sheet = wb.active
        header = [sheet["A1"].value, sheet["B1"].value, sheet["C1"].value, sheet["D1"].value, sheet["E1"].value,
                  sheet["F1"].value, sheet["G1"].value, sheet["H1"].value]
        return header

    def find_column_index_by_header(wb, header):  # find column index using header text
        sheet = wb.active
        column_index = None
        for column in range(65, 65 + sheet.max_column):  # iterating through all the columns using ascii values
            if header in str(sheet[f"{chr(column)}1"].value):  # if header text is in the 1st cell of the column
                column_index = column  # store the column index
                break
        return column_index

    def check_neagativeValue_by_column(wb, header):  #
        sheet = wb.active
        column = Excel.find_column_index_by_header(wb, header)
        for i in range(2, sheet.max_row + 1):
            value = sheet[f"{chr(column)}{i}"].value
            if isinstance(value, str) and value.strip() != '' and float(value.replace(',', '')) < 0.0:
                temp = str(sheet[f"{chr(column)}{i}"].value).replace(',', '')
                sheet[f"{chr(column + 1)}{i}"].value = float(temp.replace("-", ""))
                sheet[f"{chr(column)}{i}"].value = None
        return wb

    def empty_cell_to_none(wb, start, end, header):
        sheet = wb.active
        column = Excel.find_column_index_by_header(wb, header)
        for x in range(start, end):
            if len(str(sheet[f"{chr(column)}{x}"].value)) < 1:
                sheet[f"{chr(column)}{x}"].value = None
        return wb

    def remove_string(wb, start, end, refString, column):
        sheet = wb.active
        for x in range(start, end):
            if refString in str(sheet[f"{column}{x}"].value):
                sheet[f"{column}{x}"].value = str(sheet[f"{column}{x}"].value).replace(refString, "")
        return wb

    def replace_to_none(wb, start, end, refText, column):
        sheet = wb.active
        for x in range(start, end):
            if refText in str(sheet[f"{column}{x}"].value):
                sheet[f"{column}{x}"].value = None
        return wb

    # def transaction_type_column(wb):
    #     sheet = wb.active
    #     Excel.creat_column(wb, header="Transaction_Type")
    #     trans_type_column = chr(Excel.find_column_index_by_header(wb, header="Transaction_Type"))
    #     withdrawal_column = chr(Excel.find_column_index_by_header(wb, header="Withdrawal"))
    #     deposit_column = chr(Excel.find_column_index_by_header(wb, header="Deposit"))
    #     for i in range(2, sheet.max_row + 1):
    #         if sheet[f"{withdrawal_column}{i}"].value is not None and float(str(sheet[f"{withdrawal_column}{i}"].value)) >= 1:
    #             sheet[f"{trans_type_column}{i}"].value = "Debit"
    #
    #         if sheet[f"{deposit_column}{i}"].value is not None and float(str(sheet[f"{deposit_column}{i}"].value)) >= 1:
    #             sheet[f"{trans_type_column}{i}"].value = "Credit"
    #     return wb

    def transaction_type_column(wb):
        sheet = wb.active
        Excel.creat_column(wb, header="Transaction_Type")
        trans_type_column = chr(Excel.find_column_index_by_header(wb, header="Transaction_Type"))
        withdrawal_column = chr(Excel.find_column_index_by_header(wb, header="Withdrawal"))
        deposit_column = chr(Excel.find_column_index_by_header(wb, header="Deposit"))

        for i in range(2, sheet.max_row + 1):
            withdrawal_value = sheet[f"{withdrawal_column}{i}"].value
            deposit_value = sheet[f"{deposit_column}{i}"].value

            if withdrawal_value is not None:
                try:
                    withdrawal_float = float(str(withdrawal_value))
                    if withdrawal_float >= 1:
                        sheet[f"{trans_type_column}{i}"].value = "Debit"
                except ValueError as e:
                    print(f"Error processing withdrawal value at row {i}: {e}")

            if deposit_value is not None:
                try:
                    deposit_float = float(str(deposit_value))
                    if deposit_float >= 1:
                        sheet[f"{trans_type_column}{i}"].value = "Credit"
                except ValueError as e:
                    print(f"Error processing deposit value at row {i}: {e}")

        return wb


    def minio_upload_pdf(file_path, bucket_name, folder_path):
        config = configparser.ConfigParser()
        config.read(".env")
        client = Minio('ksvca-server-01:3502', access_key=config.get("DEFAULT", "MINIO_ACCESS_KEY"), secret_key=config.get("DEFAULT", "MINIO_SECRET_KEY"), secure=False)
        if r"/" in file_path:
            file_name = file_path.split(r'/')[-1]
        if R"Downloads\temp_" in file_path:
            file_name = file_path.split(r"Downloads\temp_")[-1]
        client.fput_object(bucket_name, folder_path + file_name, file_path)
        url = client.presigned_get_object(bucket_name, folder_path + file_name, response_headers={'response-content-type': 'application/pdf'})
        return url




