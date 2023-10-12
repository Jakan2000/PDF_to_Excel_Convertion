import configparser
import os
import os.path
from datetime import datetime
from distutils.command.config import config
from io import BytesIO

import aspose.pdf as ap
import pandas
import psycopg2
import requests
from PyPDF2 import PdfReader, PdfWriter
from openpyxl import Workbook, load_workbook

from KSV.FormatingExcelFiles.AXIS1 import axis1_main
from KSV.FormatingExcelFiles.CANARA1 import canara1_main
from KSV.FormatingExcelFiles.CITY_UNION1 import cityunion1_main
from KSV.FormatingExcelFiles.DBS1 import dbs1_main
from KSV.FormatingExcelFiles.EQUITAS1 import equitas1_main
from KSV.FormatingExcelFiles.FEDERAL1 import federal1_main
from KSV.FormatingExcelFiles.HDFC1 import hdfc1_main
from KSV.FormatingExcelFiles.ICICI1 import icici1_main
from KSV.FormatingExcelFiles.ICICI2 import icici2_main
from KSV.FormatingExcelFiles.ICICI3 import icici3_main
from KSV.FormatingExcelFiles.INDIAN_BANK1 import indian_bank1_main
from KSV.FormatingExcelFiles.INDUSIND1 import indusind1_main
from KSV.FormatingExcelFiles.INDUSIND2 import indusind2_main
from KSV.FormatingExcelFiles.IOB1 import iob1_main
from KSV.FormatingExcelFiles.KOTAK1 import kotak1_main
from KSV.FormatingExcelFiles.KOTAK2 import kotak2_main
from KSV.FormatingExcelFiles.SBI1 import sbi1_main
from KSV.FormatingExcelFiles.TMB1 import tmb1_main
from KSV.FormatingExcelFiles.YES1 import yes1_main


def driver(work_book, bank, type, path):
    banks = {"axis": {"type1": axis1_main},
             "canara": {"type1": canara1_main},
             "city union": {"type1": cityunion1_main},
             "dbs": {"type1": dbs1_main},
             "equitas": {"type1": equitas1_main},
             "federal": {"type1": federal1_main},
             "hdfc": {"type1": hdfc1_main},
             "icici": {"type1": icici1_main,
                       "type2": icici2_main,
                       "type3": icici3_main},
             "indian bank": {"type1": indian_bank1_main},
             "indusind": {"type1": indusind1_main,
                          "type2": indusind2_main},
             "iob": {"type1": iob1_main},
             "kotak": {"type1": kotak1_main,
                       "type2": kotak2_main},
             "sbi": {"type1": sbi1_main},
             "tmb": {"type1": tmb1_main},
             "yes": {"type1": yes1_main}}
    if bank in banks and type in banks[bank]:
        wb = banks[bank][type](work_book)
        temp = str(os.path.basename(path)).replace(".pdf", ".xlsx")
        file = temp.replace(".PDF", ".xlsx")
        wb.save(f"C:/Users/Admin/Desktop/FinalOutput/{file}")
        sheet = wb.active
        config = configparser.ConfigParser()
        config.read(".env")
        dbname = config.get("DEFAULT", "DATABASE")
        user = config.get("DEFAULT", "USER")
        password = config.get("DEFAULT", "PASSWORD")
        host = config.get("DEFAULT", "HOST")
        port = config.get("DEFAULT", "PORT")
        conn = None
        # Connect to the PostgreSQL database
        try:
            connection = psycopg2.connect(
                dbname=dbname,
                user=user,
                password=password,
                host=host,
                port=port
            )
            cursor = connection.cursor()
            table_name = "ksv.example_table"
            slno = "Undefined"
            transaction_date = "Undefined"
            value_date = "Undefined"
            chequeno_refno = "Undefined"
            narration = "Undefined"
            withdrawal = "Undefined"
            deposit = "Undefined"
            balance = "Undefined"
            for row in range(2, sheet.max_row + 1):
                data = [1, 2, 3, 4, 5, 6, 7, 8]
                for column in range(65, 65 + sheet.max_column):
                    if "Sl.No." in str(sheet[f"{chr(column)}1"].value):
                        slno = sheet[f"{chr(column)}{row}"].value
                        data[0] = slno
                    if "Transaction_Date" in str(sheet[f"{chr(column)}1"].value):
                        transaction_date = sheet[f"{chr(column)}{row}"].value
                        data[1] = transaction_date
                    if "Value_Date" in str(sheet[f"{chr(column)}1"].value):
                        value_date = sheet[f"{chr(column)}{row}"].value
                        data[2] = value_date
                    if "ChequeNo_RefNo" in str(sheet[f"{chr(column)}1"].value):
                        chequeno_refno = sheet[f"{chr(column)}{row}"].value
                        data[3] = chequeno_refno
                    if "Narration" in str(sheet[f"{chr(column)}1"].value):
                        narration = sheet[f"{chr(column)}{row}"].value
                        data[4] = narration
                    if "Withdrawal" in str(sheet[f"{chr(column)}1"].value):
                        withdrawal = sheet[f"{chr(column)}{row}"].value
                        data[5] = withdrawal
                    if "Deposit" in str(sheet[f"{chr(column)}1"].value):
                        deposit = sheet[f"{chr(column)}{row}"].value
                        data[6] = deposit
                    if "Balance" in str(sheet[f"{chr(column)}1"].value):
                        balance = sheet[f"{chr(column)}{row}"].value
                        data[7] = balance
                print(data)
                # Insert a record into the table
                insert_query = f"""
                        INSERT INTO {table_name} (slno, transaction_date, value_date, chequeno_refno, narration, withdrawal, deposit, balance)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s);
                    """
                data_to_insert = (data[0], data[1], data[2], data[3], data[4], data[5], data[6], data[7])
                cursor.execute(insert_query, data_to_insert)
            connection.commit()
            cursor.close()
            connection.close()
            print("Record inserted successfully.")

        except psycopg2.Error as e:
            print(f"Error: {e}")

    else:
        raise Exception(f"<Bank or type not found in the dictionary>")

def convert_url_to_bytes(pdf_url):
    bytes_list = []
    response = requests.get(pdf_url)
    response.raise_for_status()  # Check for any request errors

    bytes_stream = BytesIO(response.content)
    reader = PdfReader(bytes_stream)

    for page in reader.pages:
        writer = PdfWriter()
        writer.add_page(page)
        with BytesIO() as bytes_stream:
            writer.write(bytes_stream)
            bytes_stream.seek(0)
            bytes_list.append(bytes_stream.getvalue())

    return bytes_list


def convert_bytes_to_excel(pdf_bytes):
    def create_output_excel(output_xlsx):
        workbook = Workbook()
        sheet = workbook.active
        sheet['A1'] = ''
        workbook.save(output_xlsx)

    def join_data(temp_xlsx, output_xlsx):
        source_workbook_1 = load_workbook(output_xlsx)
        sheet1 = source_workbook_1.active
        source_workbook_2 = load_workbook(temp_xlsx)
        sheet2 = source_workbook_2['Sheet1']
        for row in sheet2.iter_rows(min_row=2, values_only=True):
            sheet1.append(row)
        source_workbook_1.save(output_xlsx)

    def convert_excel(pdf, temp_xlsx, output_xlsx):
        with open(temp_xlsx, 'wb') as f:
            document = ap.Document(BytesIO(pdf))
            save_option = ap.ExcelSaveOptions()
            save_option.minimize_the_number_of_worksheets = True
            document.save(f, options=save_option)
        join_data(temp_xlsx, output_xlsx)

    now = datetime.now()
    t = now.strftime("__%d-%m-%Y-%H-%M-%S")
    output_xlsx = 'output' + t + '.xlsx'
    temp_xlsx = 'temp' + t + '.xlsx'
    create_output_excel(output_xlsx)
    for page_bytes in pdf_bytes:
        convert_excel(page_bytes, temp_xlsx, output_xlsx)
    os.remove(temp_xlsx)

    # Load and return the final workbook
    return load_workbook(output_xlsx), output_xlsx


def pdf_to_excel_main(pdf_url, bank, type):
    pdf_bytes = convert_url_to_bytes(pdf_url)
    output_workbook, output_workbook_xlsx = convert_bytes_to_excel(pdf_bytes)
    sheet = output_workbook.active
    max_column = sheet.max_column
    if max_column < 2:
        raise Exception("Insufficient Data In convert_bytes_to_excel To Process Driver Dictionary")
    else:
        driver(output_workbook, bank, type, pdf_url)
    os.remove(output_workbook_xlsx)  # Remove the output workbook XLSX file


if __name__ == "__main__":
    pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/1.Axis_-_8874-PW_-_GNAN842166790_unlocked.pdf", "axis", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/1.Canara_-_6183.pdf", "canara", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/CITY_UNION_BANK_-_SB-500101012199098.pdf", "city union", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/LVB_-_0145P.W_-_1L1675876_unlocked.pdf", "dbs", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/Equitas_-_6802_unlocked.pdf", "equitas", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/2.%20R%20RAVICHANDRAN%20-%20Federal%20-%202416%20Pass%20-%20RAVI016%20.pdf", "federal", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/HDFC_-_7768.pdf", "hdfc", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/ICICI_-_3281.pdf", "icici", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/ICICI_-_2207PW-088601502207_unlocked.pdf", "icici", "type2")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/ilovepdf_merged_7.pdf", "icici", "type3")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/SRT_-_INDIAN_BANK_-_6096825697_.pdf", "indian bank", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/Senthil_indusind_pdf.io.pdf", "indusind", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/1._Indusind_-_2673.pdf", "indusind", "type2")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/IOB_-_8713.pdf", "iob", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/Kotak1._Apr-22_637102.pdf", "kotak", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/Kotak_-_5887.PDF", "kotak", "type2")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/SBI12._March_-_2023.pdf", "sbi", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/TMB_-_2333.pdf", "tmb", "type1")
    # pdf_to_excel_main("http://ksvca-server-01:3502/ksv/bank_statements/8._YES_bank_-_8241_Aug-Oct.pdf", "yes", "type1")
