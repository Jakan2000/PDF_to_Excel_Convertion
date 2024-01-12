from datetime import datetime

import openpyxl

from CommonClass import Excel


def removeNoneRows(wb, start, end, column):
    """
        Remove unwanted rows from the Excel workbook based on a specified reference column.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data.
        - column (str): The column letter representing the reference column.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object after removing unwanted rows.

        Note:
        This function iterates through the specified range of rows (from 'end' to 'start') in the Excel workbook ('wb').
        If the cell value in the specified reference column ('column') is None, it deletes the corresponding row.

    """
    sheet = wb.active
    for x in range(end, start, -1):  # iterating through table data end row to table data start row
        if sheet[f"{column}{x}"].value is None:  # if reference column cell value is none
            sheet.delete_rows(x)  # delete row
    return wb


def align_balance_column(wb, start, end, column):
    """
        Align balance column data in an Excel workbook by moving non-empty values to the balance column.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data.
        - column (str): The column letter representing the balance column.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object after aligning balance column data.

        Note:
        This function iterates through the specified range of rows (from 'start' to 'end') in the Excel workbook ('wb').
        If the balance column cell value is None, it moves the non-empty value from the previous column to the balance column.
        The previous column cell value is then set to None.

    """
    sheet = wb.active
    for row in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{column}{row}"].value is None:  # if balance column cell value is none
            sheet[f"{column}{row}"].value = sheet[f"{chr(ord(column)-1)}{row}"].value  # assigning previous column cell value to balance column cell value
            sheet[f"{chr(ord(column) - 1)}{row}"].value = None  # assigning none to previous column cell value
    return wb


def dateConvertion(wb, start, end, column):
    """
        Convert date values in a specified column of an Excel workbook to a standard date format.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data (exclusive).
        - column (str): The column letter representing the column containing date values.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object after converting date values.

        Note:
        This function iterates through the specified range of rows (from 'start' to 'end') in the Excel workbook ('wb').
        It assumes that the date values in the specified column are in the format "%d-%m-%Y".
        The date values are converted to a standard date format and updated in the same column.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to table data end row
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%m-%Y").date()  # converting to standard date formate
    return wb


# make the cell None if there is only year in the cell
def make_none_date_column(wb, start, end, refColumn):
    """
        Make cells in a specified date column None based on the length of the date values.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data (exclusive).
        - refColumn (str): The column letter representing the date column.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object after updating date cells to None.

        Note:
        This function iterates through the specified range of rows (from 'start' to 'end') in the Excel workbook ('wb').
        It checks the length of the date values in the specified column ('refColumn').
        If the length is less than 5, the cell value is set to None.

    """
    sheet = wb.active
    for row in range(start, end):  # iterating through table data start row to table data end row
        if len(str(sheet[f"{refColumn}{row}"].value)) < 5:  # if length of date column cell value < 5
            sheet[f"{refColumn}{row}"].value = None  # assign none to cell value
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    """
        Merge consecutive rows in a specified column based on a reference column's values.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data (exclusive).
        - refColumn (str): The column letter representing the reference column used to identify starting rows.
        - mergingColumn (str): The column letter representing the column to merge consecutive rows.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object after merging rows in the specified column.

        Note:
        This function iterates through the specified range of rows (from 'start' to 'end') in the Excel workbook ('wb').
        It merges consecutive rows in the specified column ('mergingColumn') based on the values in the reference column ('refColumn').
        The reference column is used to identify starting rows, and consecutive rows are merged until a new starting row is encountered.

    """
    sheet = wb.active
    dataToMerge = []  # array to store row data
    for i in range(start, end):  # iterating through table data start row to table data end row
        slno = sheet[f"{refColumn}{i}"].value  # getting reference column cell value
        if slno is not None:  # if reference column cell is not none then it's the starting row
            if len(dataToMerge) == 0:  # if dataToMerge is empty this is the starting row
                dataToMerge.append(f"{mergingColumn}{i}")  # store cell address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # store data from the 1 index
            else:  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                s = ""  # if refColumn is not none and dataToMerge is not empty -> it is next starting row
                for j in range(1, len(dataToMerge)):  # iterate the dataToMerge array
                    s += str(dataToMerge[j])  # concat the row data
                cell_address = dataToMerge[0]  # take current cell address from 0 index
                sheet[str(cell_address)].value = s  # assign conceited data to the cell
                dataToMerge = []  # emptying dataToMerge ot find the next row starting
                dataToMerge.append(f"{mergingColumn}{i}")  # appending next starting row address in the 0 index
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # appending row data to the corresponding index
        if slno is None:  # if date is none this is not the starting row
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)  # append data to the corresponding index
    # while iterating through the loop the last row will be skipped, so merge the last row by this set of code
    st1 = ""  # empty string to merge the row data
    for m in range(1, len(dataToMerge)):  # iterate dataToMerge array
        st1 += str(dataToMerge[m])  # concat the row data
    cell_address = dataToMerge[0]  # take current cell address from 0 index
    sheet[str(cell_address)].value = st1  # assign conceited data to the cell
    dataToMerge = []  # emptying the dataToMerge
    return wb  # return work book by merging the corresponding rows in the column


def align_date_column(wb, start, end, column):
    """
        Align date column in an Excel workbook by copying the next row's value if the current row's value is within a specified length range.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data (exclusive).
        - column (str): The column letter representing the date column to align.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object after aligning the date column.

        Note:
        This function iterates through the specified range of rows (from 'start' to 'end') in the Excel workbook ('wb').
        It checks the length of each cell value in the specified date column ('column'). If the length is between 5 and 9 (exclusive),
        it copies the value from the next row to align the date data.

    """
    sheet = wb.active
    for row in range(start, end):  # iterating through table data start row to table data end row
        if sheet[f"{column}{row}"].value is not None:  # if cell value is not none
            if len(str(sheet[f"{column}{row}"].value)) < 9 and len(str(sheet[f"{column}{row}"].value)) > 4:  # if length of cell value < 9 and length of cell value is > 4
                sheet[f"{column}{row}"].value = sheet[f"{column}{row + 1}"].value  # assign next row cell value to current row cell value
    return wb


def deleteHeader(wb, start):
    """
        Delete header rows from the specified Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.
        - start (int): The row index indicating the start of the table data.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object after deleting header rows.

        Note:
        This function iterates in reverse order through the rows from the specified 'start' row to the first row (1st row of the sheet),
        deleting each row along the way. It is designed to remove header rows from the Excel workbook.

    """
    sheet = wb.active
    for x in range(start, 0, -1):  # iterating through table data start row to 1st row of sheet
        sheet.delete_rows(x)  # delete row
    return wb


def deleteFooter(wb, end):
    """
        Delete footer rows from the specified Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.
        - end (int): The row index indicating the end of the table data.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object after deleting footer rows.

        Note:
        This function iterates in reverse order through the rows from the last row of the sheet to the specified 'end' row,
        deleting each row along the way. It is designed to remove footer rows from the Excel workbook.

    """
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):  # iterating through max row of sheet to table data end row
        sheet.delete_rows(x)  # delete row
    return wb



def aligningAllColumns(wb, start, end, refColumn):
    """
        Align misaligned column data in the specified Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.
        - start (int): The row index indicating the start of the table data.
        - end (int): The row index indicating the end of the table data.
        - refColumn (str): The reference column letter to determine alignment.

        Returns:
        - openpyxl.Workbook: The modified openpyxl Workbook object after aligning misaligned columns.

        Note:
        This function aligns misaligned data in columns C to G based on the data in the specified reference column.
        It iterates through the table data rows, aligning the data in each corresponding column (C to G) based on the
        provided reference column. It is designed to correct misalignments in the Excel workbook.

    """
    sheet = wb.active
    for i in range(start, end):  # iterating through table data start row to
        if sheet[f"{refColumn}{i}"].value is not None:  # if reference column cell value is not none
            sheet[f'C{i}'].value = sheet[f'D{i}'].value  # assigning D column cell value to C column cell value
            sheet[f'D{i}'].value = sheet[f'E{i}'].value  # assigning E column cell value to D column cell value
            sheet[f'E{i}'].value = sheet[f'F{i}'].value  # assigning F column cell value to E column cell value
            sheet[f'F{i}'].value = sheet[f'G{i}'].value  # assigning G column cell value to F column cell value
            sheet[f'G{i}'].value = None  # assigning None to G column cell value
    return wb


def kotak3_validation(wb):
    """
        Validate the column count for the core logic in the provided Excel workbook.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.

        Returns:
        - bool: True if the column count does not match the expected count, False otherwise.

        Note:
        This function validates the column count for a specific core logic design. It compares the actual column count
        in the workbook with the expected count (max_column). If the counts do not match, it returns True, indicating
        a mismatch. Otherwise, it returns False, indicating a successful validation.

    """
    sheet = wb.active
    max_column = 6  # column count of designed core logic
    countOfColumn = Excel.column_count(wb)  # getting count of columns using column_count() function
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def kotak3_main(wb):
    """
        Process an Excel workbook with specific steps for Kotak 3 data formatting.

        Parameters:
        - wb (openpyxl.Workbook): The openpyxl Workbook object representing the Excel workbook.

        Returns:
        - dict: A dictionary containing the processed workbook and a status message.
          The dictionary has the following structure:
          {"data": openpyxl.Workbook, "msg": str or None}

        Note:
        This function processes the provided Excel workbook with a series of steps to format Kotak 3 data.
        The steps include deleting unwanted rows, aligning misaligned columns, deleting header and footer rows,
        validating the column count, creating new columns, and standardizing column names.

    """
    sheet = wb.active
    countOfColumn = 6
    startText = "Date"  # header text to define table data start row
    stopText = "Statement Summary"  # text defining end row of table data
    startEndRefColumn = "A"  # column containing start end text
    deleteFlagStartText = "Period"  # starting row reference text to delete rows by range
    deleteFlagEndText = "Narration"  # ending row reference text to delete rows by range
    deleteFlagRefColumn = "B"  # column contains starting row reference text and ending row reference text to delete the rows by range
    refColumnToAlignAllColumns = "G"  # column to align balance data
    dateColumnAlignRefcolumn = "A"  # colun to align date data
    refTextToRemoveRow1 = "B/F"  # reference text to remove row
    refColumnToRemoveRow1 = "B"  # column containing reference text to remove tow
    refColumnToRemoveInvalidDate = "A"  # column to remove invalid date
    refColumnToMerg = "A"  # reference column to merge other column misaligned rows
    columnToMerg1 = "B"  # column to merge misaligned rows
    columnToMerg2 = "C"  # column to merge misaligned rows
    refStringToRemove1 = "None"  # reference string to remove by column
    refStringToRemove2 = "(Cr)"  # reference string to remove by column
    refColumnToRemoveString2 = "F"  # column containing reference string to remove
    dateConversionColumn1 = "A"  # column to convert date to standard date formate
    refHeaderText1 = "Date"  # header text to replace with standardised column name
    refHeaderText2 = "Narration"  # header text to replace with standardised column name
    refHeaderText3 = "Chq/Ref No"  # header text to replace with standardised column name
    refHeaderText4 = "Withdrawal (Dr)"  # header text to replace with standardised column name
    refHeaderText5 = "Deposit(Cr)"  # header text to replace with standardised column name
    refHeaderText6 = "Balance"  # header text to replace with standardised column name
    headerText1 = "Transaction_Date"  # standard column name
    headerText2 = "Narration"  # standard column name
    headerText3 = "ChequeNo_RefNo"  # standard column name
    headerText4 = "Withdrawal"  # standard column name
    headerText5 = "Deposit"  # standard column name
    headerText6 = "Balance"  # standard column name
    balance_column = "F"  # colum to align balance data
    columns = ["Sl.No.", "Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]  # standard columns to be present in the file
    start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)  # get start and end row index to specify table data with in
    Excel.delete_rows_by_range(wb, start + 1, end + 1, deleteFlagStartText, deleteFlagEndText, deleteFlagRefColumn)  # deleting unwanted rows by range
    start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)  # get start and end row index to specify table data with in
    aligningAllColumns(wb, start, end + 1, refColumnToAlignAllColumns)  # aligning misaligned column data
    deleteFooter(wb, end - 1)  # deleting footer rows
    start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)  # get start and end row index to specify table data with in
    deleteHeader(wb, start - 1)  # deleting header data
    if kotak3_validation(wb):  # validating column count for core logic
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response  # returning response with error msg
    start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)  # get start and end row index to specify table data with in
    Excel.remove_row(wb, start, end + 1, refTextToRemoveRow1, refColumnToRemoveRow1)  # remove a single row by checking the referance text is in the column cell
    start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)  # get start and end row index to specify table data with in
    align_date_column(wb, start, end + 1, dateColumnAlignRefcolumn)  # aligning date column rows
    make_none_date_column(wb, start + 1, end + 1, refColumnToRemoveInvalidDate)  # making cells to none in date column by date length
    mergingRows(wb, start, end + 1, refColumnToMerg, columnToMerg1)  # merging misaligned rows of B column
    mergingRows(wb, start, end + 1, refColumnToMerg, columnToMerg2)  # merging misaligned rows of C column
    removeNoneRows(wb, start, end + 1, refColumnToMerg)  # removing the unwanted rows, when the reference column cell value is none
    start, end = Excel.get_start_end_row_index(wb, startText, stopText, startEndRefColumn)  # get start and end row index to specify table data with in
    align_balance_column(wb, start, end, balance_column)  # aligning balance column data
    Excel.remove_string(wb, start, end + 1, refStringToRemove1, columnToMerg1)  # removing string "None" from B column
    Excel.remove_string(wb, start, end + 1, refStringToRemove1, columnToMerg2)  # removing string "None" from C column
    Excel.remove_string(wb, start, end + 1, refStringToRemove2, refColumnToRemoveString2)  # removing reference string from F column
    dateConvertion(wb, start + 1, end + 1, dateConversionColumn1)  # converting date to standard date formate
    columnToCreateSlNo = 65 + Excel.column_count(wb)  # 65 -> ASCII value
    Excel.create_slno_column(wb, start, end + 1, chr(columnToCreateSlNo))  # creating new slno column
    lastCol = 65 + sheet.max_column  # 65 => ASCII value "A"
    transdate = Excel.alter_header_name(wb, refHeaderText1, headerText1, lastCol)  # alter header name from the excel file to the standard column name
    narration = Excel.alter_header_name(wb, refHeaderText2, headerText2, lastCol)  # alter header name from the excel file to the standard column name
    chqNo = Excel.alter_header_name(wb, refHeaderText3, headerText3, lastCol)  # alter header name from the excel file to the standard column name
    debit = Excel.alter_header_name(wb, refHeaderText4, headerText4, lastCol)  # alter header name from the excel file to the standard column name
    credit = Excel.alter_header_name(wb, refHeaderText5, headerText5, lastCol)  # alter header name from the excel file to the standard column name
    balance = Excel.alter_header_name(wb, refHeaderText6, headerText6, lastCol)  # alter header name from the excel file to the standard column name
    Excel.finalise_column(wb, columns)  # standardizing count of column
    Excel.transaction_type_column(wb)  # creating new transaction type column
    response = {"data": wb,
                "msg": None}
    return response


if __name__ == "__main__":
    # path = "C:/Users/Admin/Desktop/KSV/source_excel_files/sasikala_kotak__14-12-2023-10-37-39.xlsx"
    # path = "C:/Users/Admin/Downloads/R_JAYARAMAN__27-12-2023-13-29-48.xlsx"
    # path = "C:/Users/Admin/Desktop/NKP Roadways/Kasthuri/kotak-6311839988__28-12-2023-15-26-24.xlsx"
    path = ""
    wb = openpyxl.load_workbook(path)
    result = kotak3_main(wb)
    # result["data"].save("C:/Users/Admin/Desktop/Kotak3output.xlsx")