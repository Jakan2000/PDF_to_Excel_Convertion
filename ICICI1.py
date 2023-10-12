import os
from datetime import datetime
import openpyxl
from openpyxl.utils import column_index_from_string

from KSV.FormatingExcelFiles.CommonClass import Excel


def dateConvertion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d/%b/%Y").date()
    return wb


def delete_header(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return wb


def delete_footer(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


def delete_nonerows(wb, start, end, column):
    sheet = wb.active
    for x in range(end, start, -1):
        if sheet[f"{column}{x}"].value is None:
            sheet.delete_rows(x)
    return wb


def remove_none(wb, start, end, Column):
    sheet = wb.active
    for x in range(start, end):
        if "None" in str(sheet[f"{Column}{x}"].value):
            sheet[f"{Column}{x}"].value = str(sheet[f"{Column}{x}"].value).replace("None", "")
    return wb


def merging_rows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def delete_pagenorow(wb, start, end, Column, refText):
    sheet = wb.active
    for x in range(end, start, -1):
        if refText in str(sheet[f"{Column}{x}"].value):
            end -= 1
            sheet.delete_rows(x)
    return wb


def icici1_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 10
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def icici1_main(wb):
    sheet = wb.active
    if icici1_validation(wb):
        raise Exception(f"<= INVALID FORMATE =>  <Count Of Column Mismatch>")
    else:
        startText = "Sl"
        endText = "Opening Bal:"
        startEndDefColumn = "A"
        deletePageNoRowFromColumn = "A"
        referanceTextToDelete = "Page"
        refColumn = "A"
        column1 = "B"
        column2 = "C"
        column3 = "D"
        column4 = "E"
        column5 = "G"
        refTextToDeleteColumn1 = "SlNo"
        refTextToDeleteColumn2 = "TranId"
        refTextToDeleteColumn3 = "TransactionPosted Date"
        refHeaderText1 = "ValueDate"
        refHeaderText2 = "TransactionDate"
        refHeaderText3 = "Cheque no /Ref No"
        refHeaderText4 = "TransactionRemarks"
        refHeaderText5 = "Withdrawal (Dr)"
        refHeaderText6 = "Deposit(Cr)"
        refHeaderText7 = "Balance"
        headerText1 = "Value_Date"
        headerText2 = "Transaction_Date"
        headerText3 = "ChequeNo_RefNo"
        headerText4 = "Narration"
        headerText5 = "Withdrawal"
        headerText6 = "Deposit"
        headerText7 = "Balance"
        stringAlignColumn4 = "A"
        stringAlignColumn5 = "B"
        dateStringAlignColumn1 = "C"
        dateStringAlignColumn2 = "D"
        stringAlignColumn6 = "E"
        stringAlignColumn7 = "F"
        stringAlignColumn3 = "G"
        stringAlignColumn8 = "H"
        stringAlignColumn9 = "I"
        stringAlignColumn10 = "J"
        dateConversionColumn1 = "A"
        dateConversionColumn2 = "B"
        noneRemoveColumn1 = "A"
        noneRemoveColumn2 = "B"
        noneRemoveColumn3 = "C"
        noneRemoveColumn4 = "D"
        noneRemoveColumn5 = "E"
        noneRemoveColumn6 = "F"
        noneRemoveColumn7 = "G"
        negativeValueColumnRefText1 = "Withdrawal"
        headerTextToMakeEmptyCellsToNone1 = "Withdrawal"
        headerTextToMakeEmptyCellsToNone2 = "Deposit"
        headerTextToMakeEmptyCellsToNone3 = "ChequeNo_RefNo"
        columnToRemoveString1 = "E"
        columnToRemoveString2 = "F"
        columnToRemoveString3 = "G"
        stringToRemove1 = ","
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)
        pageNoRowDeletedwb = delete_pagenorow(wb, start, end, deletePageNoRowFromColumn, referanceTextToDelete)
        start, end = Excel.get_start_end_row_index(pageNoRowDeletedwb, startText, endText, startEndDefColumn)
        mergcolumnB = merging_rows(pageNoRowDeletedwb, start, end, refColumn, column1)
        mergColumnC = merging_rows(mergcolumnB, start, end, refColumn, column2)
        mergColumnD = merging_rows(mergColumnC, start, end, refColumn, column3)
        mergColumnE = merging_rows(mergColumnD, start, end, refColumn, column4)
        mergColumnG = merging_rows(mergColumnD, start, end, refColumn, column5)
        removeNoneB = remove_none(mergColumnG, start, end, column1)
        removeNoneC = remove_none(removeNoneB, start, end, column2)
        removeNoneD = remove_none(removeNoneC, start, end, column3)
        removeNoneE = remove_none(removeNoneD, start, end, column4)
        deletedNoneRows = delete_nonerows(removeNoneE, start, end, refColumn)
        start, end = Excel.get_start_end_row_index(deletedNoneRows, startText, endText, startEndDefColumn)
        footerDeleted = delete_footer(deletedNoneRows, end - 1)  # end-1 to Include End Footer
        headerDeleted = delete_header(footerDeleted, start - 1)  # start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndDefColumn)
        stringAlignedA = Excel.string_align(headerDeleted, start, end + 1, stringAlignColumn4)  # end+1 to Include Last Row
        stringAlignedB = Excel.string_align(headerDeleted, start, end + 1, stringAlignColumn5)  # end+1 to Include Last Row
        stringAlignedC = Excel.string_align(stringAlignedB, start, end + 1, dateStringAlignColumn1)  # end+1 to Include Last Row
        stringAlignedD = Excel.string_align(stringAlignedC, start, end + 1, dateStringAlignColumn2)  # end+1 to Include Last Row
        stringAlignedE = Excel.string_align(stringAlignedD, start, end + 1, stringAlignColumn6)  # end+1 to Include Last Row
        stringAlignedF = Excel.string_align(stringAlignedE, start, end + 1, stringAlignColumn7)  # end+1 to Include Last Row
        stringAlignedG = Excel.string_align(stringAlignedF, start, end + 1, stringAlignColumn3)  # end+1 to Include Last Row
        stringAlignedH = Excel.string_align(stringAlignedG, start, end + 1, stringAlignColumn8)  # end+1 to Include Last Row
        stringAlignedI = Excel.string_align(stringAlignedH, start, end + 1, stringAlignColumn9)  # end+1 to Include Last Row
        stringAlignedJ = Excel.string_align(stringAlignedI, start, end + 1, stringAlignColumn10)  # end+1 to Include Last Row
        deletedColumn1 = Excel.delete_column(stringAlignedD, refTextToDeleteColumn1)
        deletedColumn2 = Excel.delete_column(deletedColumn1, refTextToDeleteColumn2)
        deletedColumn3 = Excel.delete_column(deletedColumn2, refTextToDeleteColumn3)
        lastCol = 65 + Excel.column_count(wb)
        valueDate = Excel.alter_header_name(deletedColumn2, refHeaderText1, headerText1, lastCol)
        transDate = Excel.alter_header_name(valueDate, refHeaderText2, headerText2, lastCol)
        chequeNo = Excel.alter_header_name(transDate, refHeaderText3, headerText3, lastCol)
        naration = Excel.alter_header_name(chequeNo, refHeaderText4, headerText4, lastCol)
        debit = Excel.alter_header_name(naration, refHeaderText5, headerText5, lastCol)
        credit = Excel.alter_header_name(debit, refHeaderText6, headerText6, lastCol)
        balance = Excel.alter_header_name(credit, refHeaderText7, headerText7, lastCol)
        dateConvertedA = dateConvertion(balance, start + 1, end + 1, dateConversionColumn1)  # start+1 to Skip Header, end-1 to Include Last Row
        dateConvertedB = dateConvertion(dateConvertedA, start + 1, end + 1, dateConversionColumn2)  # start+1 to Skip Header, end-1 to Include Last Row
        noneRemovedA = remove_none(dateConvertedB, start, end + 1, noneRemoveColumn1)  # end+1 to Include Last Row
        noneRemovedB = remove_none(noneRemovedA, start, end + 1, noneRemoveColumn2)  # end+1 to Include Last Row
        noneRemovedC = remove_none(noneRemovedB, start, end + 1, noneRemoveColumn3)  # end+1 to Include Last Row
        noneRemovedD = remove_none(noneRemovedC, start, end + 1, noneRemoveColumn4)  # end+1 to Include Last Row
        noneRemovedE = remove_none(noneRemovedD, start, end + 1, noneRemoveColumn5)  # end+1 to Include Last Row
        noneRemovedF = remove_none(noneRemovedE, start, end + 1, noneRemoveColumn6)  # end+1 to Include Last Row
        noneRemovedG = remove_none(noneRemovedF, start, end + 1, noneRemoveColumn7)  # end+1 to Include Last RowcolumnToCreateSlNo = 65 + column_count(wb)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        slCreated = Excel.create_slno_column(noneRemovedG, start, end + 1, chr(columnToCreateSlNo))
        negativeValueChecked = Excel.check_neagativeValue_by_column(slCreated, negativeValueColumnRefText1)
        withdrawalNoneReplaced = Excel.empty_cell_to_none(negativeValueChecked, start, end + 1, headerTextToMakeEmptyCellsToNone1)
        depositNoneReplaced = Excel.empty_cell_to_none(withdrawalNoneReplaced, start, end + 1, headerTextToMakeEmptyCellsToNone2)
        chqnoNoneReplaced = Excel.empty_cell_to_none(depositNoneReplaced, start, end + 1, headerTextToMakeEmptyCellsToNone3)
        stringRemovedE = Excel.remove_string(chqnoNoneReplaced, start, end + 1, stringToRemove1, columnToRemoveString1)
        stringRemovedF = Excel.remove_string(stringRemovedE, start, end + 1, stringToRemove1, columnToRemoveString2)
        stringRemovedG = Excel.remove_string(stringRemovedF, start, end + 1, stringToRemove1, columnToRemoveString3)
        return wb


if __name__ == "__main__":
    path = "C:/Users/Admin/Downloads/ICICI_-_3281__05-09-2023-15-28-06.xlsx"
    wb = openpyxl.load_workbook(path)
    result = icici1_main(wb)
    result.save('C:/Users/Admin/Desktop/FinalOutput/ICICI1output.xlsx')
