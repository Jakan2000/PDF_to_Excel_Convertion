from datetime import datetime

import openpyxl

from CommonClass import Excel


def aligningAllColumns(wb, start, end, refColumn):
    sheet = wb.active
    for i in range(start, end):
        if sheet[f"{refColumn}{i}"].value is None:
            sheet[f'F{i}'].value = sheet[f'E{i}'].value
            sheet[f'E{i}'].value = sheet[f'D{i}'].value
            sheet[f'D{i}'].value = sheet[f'C{i}'].value
            sheet[f'C{i}'].value = sheet[f'B{i}'].value
            sheet[f'B{i}'].value = None
    return wb


def dateConvertion(wb, start, end, column):
    sheet = wb.active
    for i in range(start, end):
        sheet[f"{column}{i}"].value = datetime.strptime(str(sheet[f"{column}{i}"].value), "%d-%m-%Y").date()
    return wb


def deleteHeader(wb, start):
    sheet = wb.active
    for x in range(start, 0, -1):
        sheet.delete_rows(x)
    return (wb)


def deleteFooter(wb, end):
    sheet = wb.active
    for x in range(sheet.max_row, end, -1):
        sheet.delete_rows(x)
    return wb


def removeNoneRows(wb, start, end, column):
    sheet = wb.active
    for x in range(end - 1, start, -1):
        if sheet[f"{column}{x}"].value is None:
            sheet.delete_rows(x)
    return wb


def mergingRows(wb, start, end, refColumn, mergingColumn):
    sheet = wb.active
    dataToMerge = []
    for i in range(start, end):
        slno = sheet[f"{refColumn}{i}"].value
        if slno is not None:
            if len(dataToMerge) == 0:
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
            else:
                s = ""
                for j in range(1, len(dataToMerge)):
                    s += str(dataToMerge[j])
                cell_address = dataToMerge[0]
                sheet[str(cell_address)].value = s
                dataToMerge = []
                dataToMerge.append(f"{mergingColumn}{i}")
                dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
        if slno is None:
            dataToMerge.append(sheet[f"{mergingColumn}{i}"].value)
    st1 = ""
    for m in range(1, len(dataToMerge)):
        st1 += str(dataToMerge[m])
    cell_address = dataToMerge[0]
    sheet[str(cell_address)].value = st1
    dataToMerge = []
    return wb


def icici2_validation(wb):
    sheet = wb.active
    max_column = sheet.max_column
    countOfColumn = 6
    if max_column < countOfColumn or max_column > countOfColumn:
        return True
    else:
        return False


def icici2_main(wb):
    sheet = wb.active
    if icici2_validation(wb):
        print(f"<= INVALID FORMATE : Count Of Column Mismatch =>")
        response = {"data": None,
                    "msg": "<= INVALID FORMATE : Count Of Column Mismatch =>"}
        return response
    else:
        startText = "DATE"
        endText = "Account Related Other Information"
        startEndDefColumn = "A"
        deleteFlagStartText = "Page"
        deleteFlagEndText = "MODE"
        refColumn1 = "B"
        refColumnToMerg = "A"
        columnToMerg1 = "C"
        openBalRefText = "B/F"
        openBalRefColumn = "C"
        refColumnToAlignAllColumns = "F"
        dateConversionColumn = "A"
        refTextDeleteColumn = "MODE**"
        refHeaderText1 = "DATE"
        refHeaderText2 = "PARTICULARS"
        refHeaderText3 = "DEPOSITS"
        refHeaderText4 = "WITHDRAWALS"
        refHeaderText5 = "BALANCE"
        lastCol = 65 + sheet.max_column  # 65 => ASCII value "A"
        headerText1 = "Transaction_Date"
        headerText2 = "Narration"
        headerText3 = "Deposit"
        headerText4 = "Withdrawal"
        headerText5 = "Balance"
        stringAlignColumn1 = "A"
        stringAlignColumn2 = "B"
        stringAlignColumn3 = "C"
        stringAlignColumn4 = "D"
        stringAlignColumn5 = "E"
        headerToMakeEmptyCellToNOne1 = "Deposit"
        headerToMakeEmptyCellToNOne2 = "Withdrawal"
        refStringToRemoveFromColumn1 = "None"
        columnToRemoveNone = "C"
        columns = ["Transaction_Date", "Value_Date", "ChequeNo_RefNo", "Narration", "Deposit", "Withdrawal", "Balance"]
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)
        rowsRemoveD = Excel.delete_rows_by_range(wb, start, end, deleteFlagStartText, deleteFlagEndText, refColumn1)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)
        mergColumnC = mergingRows(rowsRemoveD, start, end, refColumnToMerg, columnToMerg1)
        removeNull = removeNoneRows(mergColumnC, start, end, refColumnToMerg)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)
        start, end = Excel.get_start_end_row_index(wb, startText, endText, startEndDefColumn)
        footerDeleted = deleteFooter(removeNull, end - 1)  # end-1 to Include Last Row
        headerDeleted = deleteHeader(footerDeleted, start - 1)  # start-1 to Skip Header
        start, end = Excel.get_start_end_row_index(headerDeleted, startText, endText, startEndDefColumn)
        removeOpenBal = Excel.remove_row(headerDeleted, start, end, openBalRefText, openBalRefColumn)
        start, end = Excel.get_start_end_row_index(removeOpenBal, startText, endText, startEndDefColumn)
        alignedAllColumns = aligningAllColumns(removeOpenBal, start, end+1, refColumnToAlignAllColumns)
        dateConvertedA = dateConvertion(alignedAllColumns, start + 1, end + 1, dateConversionColumn)  # start+1 to Skip Header, end+1 to Include Last Row
        deletedModeColumn = Excel.delete_column(wb, refTextDeleteColumn)
        date = Excel.alter_header_name(deletedModeColumn, refHeaderText1, headerText1, lastCol - 1)
        naration = Excel.alter_header_name(date, refHeaderText2, headerText2, lastCol - 1)
        deposits = Excel.alter_header_name(naration, refHeaderText3, headerText3, lastCol - 1)
        withdrawal = Excel.alter_header_name(deposits, refHeaderText4, headerText4, lastCol - 1)
        balance = Excel.alter_header_name(withdrawal, refHeaderText5, headerText5, lastCol - 1)
        alignedA = Excel.string_align(balance, start, end, stringAlignColumn1)
        alignedB = Excel.string_align(alignedA, start, end, stringAlignColumn2)
        columnToCreateSlNo = 65 + Excel.column_count(wb)
        slCreated = Excel.create_slno_column(alignedB, start, end + 1, chr(columnToCreateSlNo))
        columnFinalised = Excel.finalise_column(slCreated, columns)
        noneRemovedDeposit = Excel.remove_string(columnFinalised, start, end+1, refStringToRemoveFromColumn1, columnToRemoveNone)
        depositMadeNone = Excel.empty_cell_to_none(noneRemovedDeposit, start, end+1, headerToMakeEmptyCellToNOne1)
        withdrawalMadeNone = Excel.empty_cell_to_none(depositMadeNone, start, end+1, headerToMakeEmptyCellToNOne2)
        createdTransTypeColumn = Excel.transaction_type_column(withdrawalMadeNone)
        response = {"data": wb,
                    "msg": None}
        return response


if __name__ == "__main__":
    path = "C:/Users/Admin/Desktop/KSV/source_excel_files/ICICI_-_2207PW-088601502207_unlocked__15-09-2023-12-58-00.xlsx"
    # path = "C:/Users/Admin/Downloads/2._Rajamani_-_ICICI_8226 (1)__23-11-2023-13-21-31.xlsx"
    wb = openpyxl.load_workbook(path)
    result = icici2_main(wb)
    # result.save('C:/Users/Admin/Desktop/FinalOutput/ICICI2output.xlsx')
    result.save('C:/Users/Admin/Desktop/ICICI2output.xlsx')
